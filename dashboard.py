"""
dashboard.py — ApexScan Streamlit Dashboard v13
"""

import sys
import os

# Ensure project root is on path (required for Streamlit Cloud)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import yfinance as yf
from pathlib import Path
from datetime import datetime, timedelta
import json

from scanner import load_config, run_scan, save_report
import time as _time
from datetime import timezone as _timezone
import threading as _threading
import hashlib as _hashlib

# Import each module independently so one failure doesn't break everything
def _safe_import(module_path, names):
    """Import names from module, returning None for each if import fails."""
    try:
        import importlib
        mod = importlib.import_module(module_path)
        return {n: getattr(mod, n, None) for n in names}
    except Exception as _e:
        return {n: None for n in names}

_of   = _safe_import("modules.options_flow",      ["scan_options_flow", "scan_multiple"])
_it   = _safe_import("modules.insider_tracker",   ["fetch_insider_trades", "get_insider_summary"])
_bt   = _safe_import("modules.backtester",        ["backtest_ticker", "backtest_portfolio"])
_rc   = _safe_import("modules.risk_calc",         ["TradeSetup", "calculate_position", "pyramiding_plan"])
_ab   = _safe_import("modules.ai_briefing",       ["generate_briefing", "load_latest_briefing"])
_wm   = _safe_import("modules.watchlist_manager", ["load_watchlists", "save_watchlists", "add_ticker",
                                                    "remove_ticker", "create_list", "delete_list",
                                                    "scan_watchlist", "import_tickers", "export_watchlist"])
_al   = _safe_import("modules.alerts",            ["load_alert_settings", "save_alert_settings",
                                                    "test_telegram", "send_email", "dispatch_alert",
                                                    "check_and_fire_alerts", "build_daily_briefing_alert"])
_av   = _safe_import("modules.alpha_vantage",     ["analyse_eps", "get_upcoming_earnings_for_watchlist",
                                                    "_cache_path", "_cache_valid"])

# Expose as module-level names (None if module missing)
scan_options_flow      = _of["scan_options_flow"]
scan_multiple          = _of["scan_multiple"]
fetch_insider_trades   = _it["fetch_insider_trades"]
get_insider_summary    = _it["get_insider_summary"]
backtest_ticker        = _bt["backtest_ticker"]
backtest_portfolio     = _bt["backtest_portfolio"]
TradeSetup             = _rc["TradeSetup"]
calculate_position     = _rc["calculate_position"]
pyramiding_plan        = _rc["pyramiding_plan"]
generate_briefing      = _ab["generate_briefing"]
load_latest_briefing   = _ab["load_latest_briefing"]
load_watchlists        = _wm["load_watchlists"]
save_watchlists        = _wm["save_watchlists"]
add_ticker             = _wm["add_ticker"]
remove_ticker          = _wm["remove_ticker"]
create_list            = _wm["create_list"]
delete_list            = _wm["delete_list"]
scan_watchlist         = _wm["scan_watchlist"]
import_tickers         = _wm["import_tickers"]
export_watchlist       = _wm["export_watchlist"]
load_alert_settings    = _al["load_alert_settings"]
save_alert_settings    = _al["save_alert_settings"]
test_telegram          = _al["test_telegram"]
send_email             = _al["send_email"]
dispatch_alert         = _al["dispatch_alert"]
check_and_fire_alerts  = _al["check_and_fire_alerts"]
build_daily_briefing_alert = _al["build_daily_briefing_alert"]
get_upcoming_earnings_for_watchlist = _av["get_upcoming_earnings_for_watchlist"]

# Watchlist helper stubs — only define if module didn't load them
if scan_watchlist is None:
    def scan_watchlist(list_name, tickers, cfg, analyze_fn):
        import pandas as pd
        rows = []
        for tk in tickers:
            try:
                r = analyze_fn(tk, cfg)
                if r: rows.append(r)
            except Exception:
                pass
        return pd.DataFrame(rows) if rows else pd.DataFrame()

# Fallback: if alert module missing, define minimal stubs so app doesn't crash
if load_alert_settings is None:
    def load_alert_settings(): return {"alerts_enabled": False, "telegram_token": "", "telegram_chat_id": "", "email_from": "", "email_password": "", "email_to": "", "alert_breakouts": True, "alert_stop_breach": True, "alert_earnings": True, "alert_sfp_setup": True, "alert_persistent_flow": True, "alert_vwap_imbalance": True, "min_score_alert": 60}

# ── AI Briefing stubs — module may not exist; our built-in engine handles it ──
if generate_briefing is None:
    def generate_briefing(*args, **kwargs): return ""
if load_latest_briefing is None:
    def load_latest_briefing():
        """Load last saved briefing from disk (all fallback paths)."""
        _paths = [
            Path(__file__).resolve().parent / "data" / "last_briefing.md",
            Path("/tmp/apexscan_briefing.md"),
        ]
        for _p in _paths:
            try:
                if _p.exists() and _p.stat().st_size > 10:
                    return _p.read_text(encoding="utf-8")
            except Exception:
                pass
        return ""
if save_alert_settings is None:
    def save_alert_settings(s): pass
if dispatch_alert is None:
    def dispatch_alert(settings: dict, message: str, title: str = "") -> dict:
        """Stub: send via Telegram if token configured, else return failure silently."""
        result = {"telegram": False, "email": False}
        try:
            import urllib.request, urllib.parse
            _tok = settings.get("telegram_token","")
            _cid = settings.get("telegram_chat_id","")
            if _tok and _cid:
                _text = (f"*{title}*\n\n{message}" if title else message)
                _body = urllib.parse.urlencode({
                    "chat_id": _cid,
                    "text": _text[:4096],
                    "parse_mode": "Markdown",
                }).encode()
                _req = urllib.request.Request(
                    f"https://api.telegram.org/bot{_tok}/sendMessage",
                    data=_body, method="POST"
                )
                _req.add_header("Content-Type","application/x-www-form-urlencoded")
                with urllib.request.urlopen(_req, timeout=8) as _resp:
                    _js = json.loads(_resp.read().decode())
                    result["telegram"] = _js.get("ok", False)
        except Exception as _te:
            pass
        return result

# ── Trade Journal storage (separate from trade log — captures checklist data) ──
_JOURNAL_FILE = _PORT_DIR / "trade_journal.json"  if "_PORT_DIR" in dir() else Path("data/trade_journal.json")
_JOURNAL_TMP  = Path("/tmp/apexscan_journal.json")

def load_journal() -> list:
    for _p in (_JOURNAL_FILE, _JOURNAL_TMP):
        try:
            if _p.exists() and _p.stat().st_size > 2:
                with open(_p) as _f:
                    _d = json.load(_f)
                if isinstance(_d, list): return _d
        except Exception: pass
    return []

def save_journal(journal: list):
    for _p in (_JOURNAL_FILE, _JOURNAL_TMP):
        try:
            _p.parent.mkdir(parents=True, exist_ok=True)
            _tmp = _p.with_suffix(".tmp")
            with open(_tmp,"w") as _f: json.dump(journal, _f, indent=2, default=str)
            import shutil; shutil.move(str(_tmp), str(_p))
        except Exception: pass

# ── Checklist watchlist storage (setups to monitor for status change) ──────────
_CHKWATCH_FILE = _PORT_DIR / "checklist_watchlist.json" if "_PORT_DIR" in dir() else Path("data/checklist_watchlist.json")
_CHKWATCH_TMP  = Path("/tmp/apexscan_chkwatch.json")

def load_chk_watchlist() -> list:
    for _p in (_CHKWATCH_FILE, _CHKWATCH_TMP):
        try:
            if _p.exists() and _p.stat().st_size > 2:
                with open(_p) as _f:
                    _d = json.load(_f)
                if isinstance(_d, list): return _d
        except Exception: pass
    return []

def save_chk_watchlist(items: list):
    for _p in (_CHKWATCH_FILE, _CHKWATCH_TMP):
        try:
            _p.parent.mkdir(parents=True, exist_ok=True)
            _tmp = _p.with_suffix(".tmp")
            with open(_tmp,"w") as _f: json.dump(items, _f, indent=2, default=str)
            import shutil; shutil.move(str(_tmp), str(_p))
        except Exception: pass
# ── Persistent Watchlist Engine ───────────────────────────────────────────────
# Always override the module stubs with a robust JSON-backed implementation.
# Streamlit Cloud: uses /tmp/apexscan_watchlists.json (survives rerenders,
# clears on dyno restart — acceptable; users must re-add after server restarts).
# Local / self-hosted: uses watchlists.json next to dashboard.py (permanent).
import json as _json, os as _os

_WL_LOCAL  = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "watchlists.json")
_WL_TMP    = "/tmp/apexscan_watchlists.json"
_WL_DEFAULTS = {
    "High Conviction": [],
    "Monitoring":      [],
    "Earnings Soon":   [],
    "Swing Trades":    [],
}

def _wl_path():
    """Return a writable path — prefer local (persistent), fall back to /tmp."""
    try:
        _os.makedirs(_os.path.dirname(_WL_LOCAL), exist_ok=True)
        # Test write access
        with open(_WL_LOCAL, "a") as _f:
            pass
        return _WL_LOCAL
    except Exception:
        return _WL_TMP

def load_watchlists():
    """Load watchlists from JSON file, merging with defaults so new lists always appear."""
    path = _wl_path()
    data = dict(_WL_DEFAULTS)
    if _os.path.exists(path):
        try:
            with open(path, "r") as _f:
                saved = _json.load(_f)
            if isinstance(saved, dict):
                for k, v in saved.items():
                    if isinstance(v, list):
                        data[k] = v
        except Exception:
            pass
    return data

def save_watchlists(wls):
    """Save watchlists to JSON file. Always writes to ensure persistence."""
    path = _wl_path()
    try:
        with open(path, "w") as _f:
            _json.dump(wls, _f, indent=2)
    except Exception as _e:
        import streamlit as _st
        _st.error(f"⚠️ Could not save watchlists: {_e}")

def add_ticker(wls, list_name, ticker):
    ticker = ticker.upper().strip()
    if list_name in wls and ticker and ticker not in wls[list_name]:
        wls[list_name].append(ticker)
    return wls

def remove_ticker(wls, list_name, ticker):
    if list_name in wls and ticker in wls[list_name]:
        wls[list_name].remove(ticker)
    return wls

def create_list(wls, list_name):
    if list_name and list_name not in wls:
        wls[list_name] = []
    return wls

def delete_list(wls, list_name):
    if list_name in wls:
        del wls[list_name]
    return wls

def import_tickers(wls, list_name, comma_str):
    tickers = [t.strip().upper() for t in comma_str.replace("\n", ",").split(",") if t.strip()]
    for tk in tickers:
        if tk and tk not in wls.get(list_name, []):
            wls.setdefault(list_name, []).append(tk)
    return wls

def export_watchlist(wls, list_name):
    return ", ".join(wls.get(list_name, []))

# ── Page Config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ApexScan — US Stock Scanner",
    page_icon="📡",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
[data-testid="stAppViewContainer"] { background:#0d1117; color:#e6edf3; }
[data-testid="stSidebar"] { background:#161b22; border-right:1px solid #30363d; }
.metric-card { background:#161b22; border:1px solid #30363d; border-radius:8px;
               padding:16px 20px; margin-bottom:8px; }
.metric-card h3 { margin:0; font-size:0.75rem; color:#8b949e;
                  text-transform:uppercase; letter-spacing:.08em; }
.metric-card .value { font-size:1.7rem; font-weight:700; font-family:'SF Mono',monospace; }
.green{color:#3fb950;} .red{color:#f85149;} .amber{color:#d29922;}
.blue{color:#388bfd;}  .white{color:#e6edf3;}
.alert-box { background:#1a2a1a; border:1px solid #3fb950; border-radius:8px;
             padding:12px 16px; margin:6px 0; }
.warn-box  { background:#2a2200; border:1px solid #d29922; border-radius:8px;
             padding:12px 16px; margin:6px 0; }
.danger-box{ background:#2a1010; border:1px solid #f85149; border-radius:8px;
             padding:12px 16px; margin:6px 0; }
div.stButton > button { background:#21262d; color:#e6edf3;
                        border:1px solid #30363d; border-radius:6px; }
div.stButton > button:hover { border-color:#388bfd; color:#79c0ff; }
</style>
""", unsafe_allow_html=True)

PORTFOLIO_FILE = "data/portfolio.json"
Path("data").mkdir(exist_ok=True)
Path("reports").mkdir(exist_ok=True)
Path("logs").mkdir(exist_ok=True)


# ══════════════════════════════════════════════════════════════════════════════
# SHARED HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def pct_fmt(v):
    try:
        sign = "+" if float(v) > 0 else ""
        return f"{sign}{float(v):.1f}%"
    except:
        return str(v)

def color_score(v):
    try:
        v = float(v)
        if v >= 70: return "color:#3fb950;font-weight:700"
        if v >= 40: return "color:#d29922;font-weight:600"
        return "color:#8b949e"
    except: return ""

def color_perf(v):
    try:
        v = float(v)
        return "color:#3fb950" if v > 0 else ("color:#f85149" if v < 0 else "")
    except: return ""

def color_rs(v):
    try:
        v = float(v)
        if v >= 100: return "color:#3fb950;font-weight:700"
        if v >= 70:  return "color:#d29922"
        if v > 0:    return "color:#8b949e"
        return "color:#f85149"
    except: return ""

def format_mcap(v):
    try:
        v = float(v)
        if v >= 1e9: return f"${v/1e9:.1f}B"
        if v >= 1e6: return f"${v/1e6:.0f}M"
        return f"${v:,.0f}"
    except: return "–"

def color_mcap_cat(v):
    v = str(v)
    if "Micro" in v: return "color:#f85149;font-size:0.8rem"
    if "Small" in v: return "color:#d29922;font-size:0.8rem"
    if "Mid"   in v: return "color:#388bfd;font-size:0.8rem"
    return "color:#8b949e;font-size:0.8rem"

GEM_DISCLAIMER = """<div style="background:#1a1500;border:1px solid #d29922;border-radius:8px;
padding:10px 16px;margin:6px 0;font-size:0.83rem;">
⚠️ <b style="color:#d29922;">Emerging Gems — Higher Risk / Higher Volatility</b><br>
<span style="color:#c9d1d9;">Small/micro-cap stocks can drop 30–50% on a single bad quarter.
Use <b>0.5–1% max risk per trade</b> (half your normal size).
These are long-term conviction plays — only hold what you can stomach losing short-term.</span>
</div>"""

def load_latest_report() -> pd.DataFrame:
    reports = sorted(Path("reports").glob("scan_*.csv"), reverse=True)
    if reports:
        return pd.read_csv(reports[0], index_col="rank")
    return pd.DataFrame()


# ══════════════════════════════════════════════════════════════════════════════
# CORRECTION WATCHLIST — William O'Neil's basing-during-corrections lesson
# "New bases form during market corrections. Don't spend corrections trying
#  to predict the bottom. Spend them building your watchlist."
# Purely additive: does not touch the Pre-Buy Checklist's own local market
# context check, and does not alter any existing watchlist logic.
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(ttl=900)  # 15 min cache, independent of the Pre-Buy Checklist's own cache
def get_broad_market_condition() -> dict:
    """Lightweight S&P 500 stage check used only by the Correction Watchlist feature."""
    try:
        import yfinance as _yf
        _spx  = _yf.Ticker("^GSPC").history(period="1y")["Close"]
        _cur   = float(_spx.iloc[-1])
        _ma50  = float(_spx.rolling(50).mean().iloc[-1])
        _ma200 = float(_spx.rolling(200).mean().iloc[-1])
        _uptrend = _cur > _ma50 and _cur > _ma200 and _ma50 > _ma200
        if _uptrend:
            _stage = "Stage 2 ✅ Uptrend"
        elif _cur > _ma200:
            _stage = "Stage 1 ⏳ Basing"
        elif _cur < _ma50 < _ma200:
            _stage = "Stage 4 🔴 Downtrend"
        else:
            _stage = "Stage 3 ⚠️ Topping"
        return {"uptrend": _uptrend, "stage": _stage}
    except Exception:
        # Fail safe: assume uptrend so the banner never falsely alarms on a data hiccup
        return {"uptrend": True, "stage": "Unknown"}


def find_correction_watchlist_candidates(scan_df: pd.DataFrame) -> pd.DataFrame:
    """
    Surface Stage 1 (basing) stocks with a tight/low-volatility base — the
    stocks O'Neil says institutions quietly accumulate during corrections,
    and which are often first to break out once the market confirms a new
    uptrend. Read-only: never mutates scan_df.
    """
    if scan_df is None or scan_df.empty or "stage" not in scan_df.columns:
        return pd.DataFrame()

    cand = scan_df[scan_df["stage"].astype(str).str.contains("1 ⏳", na=False)].copy()
    if cand.empty:
        return cand

    tight_mask = pd.Series(False, index=cand.index)
    if "low_adr_base" in cand.columns:
        tight_mask = tight_mask | cand["low_adr_base"].astype(str).str.lower().isin(["true", "1"])
    if "weekly_base_tight" in cand.columns:
        tight_mask = tight_mask | cand["weekly_base_tight"].astype(str).str.lower().isin(["true", "1"])
    if "pattern" in cand.columns:
        tight_mask = tight_mask | cand["pattern"].astype(str).str.contains("Tight|Handle", case=False, na=False)

    cand = cand[tight_mask]
    if "apex_score" in cand.columns:
        cand = cand.sort_values("apex_score", ascending=False)
    return cand


# ══════════════════════════════════════════════════════════════════════════════
# AUTO-SCAN ENGINE — no API key required, pure Streamlit state + time logic
# ══════════════════════════════════════════════════════════════════════════════

_AUTOSCAN_FILE   = Path(__file__).resolve().parent / "data" / "autoscan_state.json"
_AUTOSCAN_FILE.parent.mkdir(parents=True, exist_ok=True)

# US market open/close in UTC
_MARKET_OPEN_UTC  = {"hour": 14, "minute": 30}   # 9:30 AM EST = 14:30 UTC
_MARKET_CLOSE_UTC = {"hour": 20, "minute": 30}   # 3:30 PM EST = 20:30 UTC (30 min before close)

def _autoscan_load() -> dict:
    try:
        if _AUTOSCAN_FILE.exists():
            with open(_AUTOSCAN_FILE) as f:
                return json.load(f)
    except Exception:
        pass
    return {"last_open_scan": "", "last_close_scan": "", "enabled": False,
            "universe": "theme", "next_scan": ""}

def _autoscan_save(state: dict):
    try:
        with open(_AUTOSCAN_FILE, "w") as f:
            json.dump(state, f, indent=2, default=str)
    except Exception:
        pass

def _is_market_day() -> bool:
    """Return True if today is Mon–Fri (US market day). Does not check holidays."""
    return datetime.now(_timezone.utc).replace(tzinfo=None).weekday() < 5   # 0=Mon … 4=Fri

def _minutes_until(target_h: int, target_m: int) -> float:
    """Minutes until the next occurrence of target UTC time today or tomorrow."""
    now = datetime.now(_timezone.utc).replace(tzinfo=None)
    target = now.replace(hour=target_h, minute=target_m, second=0, microsecond=0)
    if target <= now:
        target += timedelta(days=1)
    return (target - now).total_seconds() / 60

def check_autoscan_trigger(state: dict) -> str | None:
    """
    Returns "open" or "close" if a scan should fire now, else None.
    Uses a ±4-minute window around each target time to tolerate Streamlit's
    rerun timing — the scan date is stamped so it never fires twice in one window.
    """
    if not state.get("enabled") or not _is_market_day():
        return None
    now     = datetime.now(_timezone.utc).replace(tzinfo=None)
    today   = now.strftime("%Y-%m-%d")
    h, m    = now.hour, now.minute

    def _in_window(target_h, target_m):
        target_total = target_h * 60 + target_m
        now_total    = h * 60 + m
        return abs(now_total - target_total) <= 4   # ±4 min window

    if _in_window(_MARKET_OPEN_UTC["hour"], _MARKET_OPEN_UTC["minute"]):
        key = f"{today}_open"
        if state.get("last_open_scan") != key:
            state["last_open_scan"] = key
            _autoscan_save(state)
            return "open"

    if _in_window(_MARKET_CLOSE_UTC["hour"], _MARKET_CLOSE_UTC["minute"]):
        key = f"{today}_close"
        if state.get("last_close_scan") != key:
            state["last_close_scan"] = key
            _autoscan_save(state)
            return "close"

    return None


# ══════════════════════════════════════════════════════════════════════════════
# NARRATIVE GENERATOR — pure rule-based, zero API calls
# Produces professional-quality plain-English analysis from scan data fields
# ══════════════════════════════════════════════════════════════════════════════

def _pct(v, decimals=1):
    try: return f"{float(v):+.{decimals}f}%"
    except: return "–"

def _val(v, fmt=".1f"):
    try: return f"{float(v):{fmt}}"
    except: return "–"

def _bool(v):
    if v is None: return False
    if isinstance(v, bool): return v
    return str(v).lower() in ("true","1","yes")

def generate_narrative(row: pd.Series) -> str:
    """
    Generate a full plain-English deep-read narrative for a single ticker
    from its scan result row. No API required.
    """
    tk          = str(row.get("ticker","–"))
    price       = row.get("price")
    stage       = str(row.get("stage","–"))
    apex        = row.get("apex_score", 0)
    perf_1m     = row.get("perf_1m_%")
    perf_3m     = row.get("perf_3m_%")
    perf_6m     = row.get("perf_6m_%")
    rs_3m       = row.get("rs_3m")
    rs_r2500    = row.get("rs_r2500_3m")
    rs_r3000g   = row.get("rs_r3000g_3m")
    rs_multi    = _bool(row.get("rs_multi_leader"))
    adr         = row.get("adr_%")
    vs_50       = row.get("vs_50ma_%")
    vs_200      = row.get("vs_200ma_%")
    above_50    = _bool(row.get("above_50ma"))
    above_200   = _bool(row.get("above_200ma"))
    ma50_gt_200 = _bool(row.get("ma50_gt_ma200"))
    near_52wh   = _bool(row.get("near_52wh"))
    pct_off_hi  = row.get("pct_off_high_%")
    pattern     = str(row.get("pattern","") or "")
    breaking    = _bool(row.get("breaking_out"))
    sentiment   = str(row.get("sentiment","") or "")
    news_count  = row.get("news_count", 0)
    earn_mom    = str(row.get("earn_momentum","") or "")
    eps_growth  = row.get("eps_growth_%")
    eps_surp    = row.get("eps_surprise_%")
    eps_accel   = _bool(row.get("eps_accel"))
    consec_beat = row.get("consec_beats", 0)
    rev_growth  = row.get("rev_growth_%")
    eps_score   = row.get("eps_score", 0)
    analyst_tgt = row.get("analyst_target")
    pe          = row.get("pe_ratio")
    peg         = row.get("peg_ratio")
    next_earn   = str(row.get("next_earnings","") or "")
    of_bias     = str(row.get("of_bias","") or "Neutral")
    of_ratio    = row.get("of_up_vol_ratio")
    of_bull_days= row.get("of_bullish_days")
    of_score    = row.get("of_score", 0)
    vwap        = row.get("vwap")
    vwap_pos    = str(row.get("vwap_position","") or "")
    vwap_slope  = str(row.get("vwap_slope","") or "")
    vs_vwap     = row.get("vs_vwap_%")
    ms_struct   = str(row.get("ms_structure","") or "")
    ms_hh_hl    = _bool(row.get("ms_hh_hl"))
    ms_bos      = _bool(row.get("ms_bos"))
    pa_patterns = str(row.get("pa_patterns","") or "")
    pa_engulf   = str(row.get("pa_engulfing","") or "")
    pa_sfp      = str(row.get("pa_sfp","") or "")
    pa_inside   = _bool(row.get("pa_inside_day"))
    pa_score_v  = row.get("pa_score", 0)
    is_gem      = _bool(row.get("is_gem"))
    mcap_cat    = str(row.get("mcap_category","") or "")
    liq_warn    = _bool(row.get("liquidity_warn"))
    theme       = str(row.get("theme","") or "")
    changes     = str(row.get("changes","") or "")
    delta_score = row.get("delta_score")

    lines = []

    # ── 1. HEADLINE ──────────────────────────────────────────────────────────
    if try_f(apex) >= 70:
        conviction = "HIGH CONVICTION"
    elif try_f(apex) >= 50:
        conviction = "MODERATE SETUP"
    else:
        conviction = "LOW PRIORITY"

    stage_label = "Stage 2 ✅ (Uptrend)" if "2" in stage else                   "Stage 1 (Basing)"   if "1" in stage else                   "Stage 3 (Topping)"  if "3" in stage else                   "Stage 4 🔴 (Downtrend)"
    price_str   = f"${float(price):.2f}" if price else "–"
    lines.append(f"## {tk} — {conviction} | Apex Score: {_val(apex, '.0f')}/100")
    lines.append(f"**Price:** {price_str}  |  **Stage:** {stage_label}  |  **Sector:** {theme}  |  **Size:** {mcap_cat}")
    lines.append("")

    # ── 2. MOMENTUM ──────────────────────────────────────────────────────────
    lines.append("### 📈 Price Momentum")
    m_sentences = []
    if try_f(perf_3m) >= 20:
        m_sentences.append(f"{tk} has surged **{_pct(perf_3m)}** over the last 3 months — this is exceptional momentum that puts it in the top tier of the market.")
    elif try_f(perf_3m) >= 10:
        m_sentences.append(f"{tk} has gained **{_pct(perf_3m)}** over the last 3 months, showing solid positive momentum.")
    elif try_f(perf_3m) > 0:
        m_sentences.append(f"{tk} is up **{_pct(perf_3m)}** over 3 months — modest but positive.")
    else:
        m_sentences.append(f"{tk} is down **{_pct(perf_3m)}** over 3 months — momentum is negative.")

    if perf_6m is not None:
        if try_f(perf_6m) > 0 and try_f(perf_3m) > 0:
            m_sentences.append(f"The 6-month return of **{_pct(perf_6m)}** confirms the trend has durability beyond a single-month spike.")
        elif try_f(perf_6m) < 0 and try_f(perf_3m) > 0:
            m_sentences.append(f"Note: the 6-month return is **{_pct(perf_6m)}**, meaning recent strength is a recovery from a prior decline — not a new breakout from strength.")

    if try_f(adr) > 5:
        m_sentences.append(f"The Average Daily Range of **{_val(adr)}%** means {tk} typically moves ${float(price or 0)*float(adr or 0)/100:.2f} per day — size positions accordingly.")
    lines.append(" ".join(m_sentences))
    lines.append("")

    # ── 3. RELATIVE STRENGTH ─────────────────────────────────────────────────
    lines.append("### 💪 Relative Strength")
    rs_sentences = []
    if try_f(rs_3m) >= 150:
        rs_sentences.append(f"RS vs S&P 500 is **{_val(rs_3m, '.0f')}** — this stock is massively outperforming the index. This level of RS is rare and is one of the strongest signals of institutional sponsorship.")
    elif try_f(rs_3m) >= 100:
        rs_sentences.append(f"RS vs S&P 500 is **{_val(rs_3m, '.0f')}** — outperforming the index. Leaders lead the market; followers follow it.")
    elif try_f(rs_3m) >= 70:
        rs_sentences.append(f"RS vs S&P 500 is **{_val(rs_3m, '.0f')}** — slightly below benchmark-beating threshold of 100, but still in acceptable range.")
    else:
        rs_sentences.append(f"RS vs S&P 500 is **{_val(rs_3m, '.0f')}** — underperforming the index. Caution: money is going elsewhere.")

    if rs_r2500 is not None:
        if try_f(rs_r2500) >= 100:
            rs_sentences.append(f"Beating the Russell 2500 small/mid-cap benchmark (RS: **{_val(rs_r2500, '.0f')}**) confirms leadership within its natural peer group.")
        else:
            rs_sentences.append(f"RS vs Russell 2500 is **{_val(rs_r2500, '.0f')}** — lagging its small/mid-cap peers.")

    if rs_r3000g is not None:
        if try_f(rs_r3000g) >= 100:
            rs_sentences.append(f"It is also outperforming the Russell 3000 Growth index (RS: **{_val(rs_r3000g, '.0f')}**) — the toughest growth benchmark — which places it among the elite growth leaders in the entire market.")

    if rs_multi:
        rs_sentences.append(f"**Multi-Benchmark Leader ⭐** — {tk} is simultaneously beating the S&P 500, Russell 2500, and Russell 3000 Growth. This is extremely rare and highly significant.")

    lines.append(" ".join(rs_sentences))
    lines.append("")

    # ── 4. TREND & STRUCTURE ─────────────────────────────────────────────────
    lines.append("### 🏗 Trend & Market Structure")
    t_sentences = []
    if above_50 and above_200 and ma50_gt_200:
        t_sentences.append(f"{tk} is in a confirmed **Stage 2 uptrend** — price is above both the 50-day and 200-day moving averages, and the 50MA is above the 200MA. This is the only Weinstein stage where you should be holding long positions.")
        if try_f(vs_50) and try_f(vs_50) > 0:
            t_sentences.append(f"It is currently **{_pct(vs_50)}** above the 50MA and **{_pct(vs_200)}** above the 200MA.")
        if try_f(vs_50) and try_f(vs_50) > 20:
            t_sentences.append(f"At over 20% above the 50MA, the stock is extended — wait for a pullback to the 50MA or a tight base before adding exposure.")
    elif above_200 and not ma50_gt_200:
        t_sentences.append(f"{tk} is above the 200MA but the 50MA has not yet crossed above it — this is Stage 1 basing territory. The trend is improving but not yet confirmed for aggressive long positions.")
    else:
        t_sentences.append(f"⚠️ {tk} is below the 200-day moving average — this is a Stage 3 or 4 stock. Avoid new long entries; any bounce is likely a dead-cat until price reclaims the 200MA.")

    if near_52wh:
        t_sentences.append(f"Trading near its 52-week high (**{_val(pct_off_hi, '.1f')}%** off the high) — the best breakouts come from stocks near highs, not bottoms.")
    elif try_f(pct_off_hi) and try_f(pct_off_hi) < -30:
        t_sentences.append(f"Currently **{_val(pct_off_hi, '.1f')}%** below its 52-week high — still in significant recovery territory.")

    if ms_hh_hl:
        t_sentences.append(f"Market structure confirms an uptrend with **Higher Highs and Higher Lows** intact.")
    if ms_bos:
        t_sentences.append(f"A recent **Break of Structure** has occurred, signalling acceleration of the current trend direction.")
    lines.append(" ".join(t_sentences))
    lines.append("")

    # ── 5. BREAKOUT / PATTERN ────────────────────────────────────────────────
    if pattern or breaking:
        lines.append("### 🚀 Pattern & Breakout")
        p_sentences = []
        if breaking:
            p_sentences.append(f"🚨 **ACTIVE BREAKOUT** — {tk} is breaking out right now. This is the highest-priority actionable signal in the scan.")
        if pattern and pattern not in ("None","nan",""):
            p_sentences.append(f"Pattern detected: **{pattern}**.")
            if "Breakout" in pattern:
                p_sentences.append("A breakout from a proper base on volume is the classic Mark Minervini / William O'Neil entry signal. The key is to buy as close to the pivot as possible and not chase if price is more than 5% past the breakout point.")
            elif "Handle" in pattern or "Tight" in pattern:
                p_sentences.append("This is a high-quality base still forming. Watch for the pivot breakout — it has not triggered yet. Set a price alert at the base highs.")
            elif "Pullback" in pattern or "50MA" in pattern:
                p_sentences.append("A pullback to the 50MA in an uptrend is a lower-risk entry point than a breakout — closer to the stop, better R:R ratio.")
        lines.append(" ".join(p_sentences))
        lines.append("")

    # ── 6. ORDER FLOW ────────────────────────────────────────────────────────
    lines.append("### 🌊 Institutional Order Flow")
    o_sentences = []
    if "Strong Bullish" in of_bias:
        o_sentences.append(f"Order flow is **Strong Bullish** — institutions are persistently and aggressively accumulating {tk}.")
    elif "Bullish" in of_bias:
        o_sentences.append(f"Order flow is **Bullish** — there is consistent buying pressure over the recent sessions.")
    elif "Bearish" in of_bias:
        o_sentences.append(f"Order flow is **Bearish** — selling pressure is dominating. Institutions may be distributing.")
    else:
        o_sentences.append(f"Order flow is **Neutral** — no clear directional bias from institutional activity.")

    if try_f(of_ratio) >= 1.5:
        o_sentences.append(f"Up-day volume is **{_val(of_ratio, '.2f')}×** higher than down-day volume — a ratio above 1.5× indicates institutional-level accumulation, not retail buying.")
    if try_f(of_bull_days) >= 60:
        o_sentences.append(f"**{_val(of_bull_days, '.0f')}%** of recent sessions closed higher — persistent directional bias is one of the most reliable leading indicators of a sustained move.")
    lines.append(" ".join(o_sentences))
    lines.append("")

    # ── 6b. WEEKLY TIMEFRAME ─────────────────────────────────────────────────
    _wk_conf_n   = _bool(row.get("weekly_confirmed"))
    _wk_contra_n = _bool(row.get("weekly_contradicts"))
    _wk_stage_n  = str(row.get("weekly_stage","") or "")
    _wk_tight_n  = _bool(row.get("weekly_base_tight"))
    _wk_rs_n     = row.get("weekly_rs")
    _wk_score_n  = row.get("weekly_score", 0)
    _wk_hh_n     = _bool(row.get("weekly_hh_hl"))
    _wk_depth_n  = row.get("weekly_base_depth_%")

    lines.append("### 📅 Weekly Chart Confirmation")
    wk_sentences = []

    if _wk_conf_n:
        wk_sentences.append(
            f"✅ **Weekly chart confirmed.** The weekly timeframe is in {_wk_stage_n} — "
            f"price is above both the 10-week and 40-week MAs with the 10WMA above the 40WMA. "
            f"This is the single most important filter. A daily signal backed by a confirmed weekly "
            f"uptrend has a dramatically higher success rate than one that isn't."
        )
    elif _wk_contra_n:
        wk_sentences.append(
            f"⚠️ **Weekly chart contradicts the daily setup.** The weekly is showing {_wk_stage_n}. "
            f"Price is below the 40-week MA in a weekly downtrend. This is the trap — retail traders see "
            f"a daily breakout and buy, not realising the weekly trend is still against them. "
            f"If you trade this, use maximum half-size and be ready to exit quickly."
        )
    else:
        wk_sentences.append(
            f"📊 Weekly chart is **transitioning** ({_wk_stage_n}). "
            f"Not yet confirmed but not contradicting either. Wait for the weekly to resolve above "
            f"the 40WMA before committing full size."
        )

    if _wk_tight_n:
        wk_sentences.append(
            f"The weekly base is **tight** ({_val(_wk_depth_n,'.0f')}% depth over 8 weeks) — "
            f"this is the compression before the explosion. Tight weekly bases precede the biggest moves."
        )

    if _wk_hh_n:
        wk_sentences.append("The weekly chart is making **Higher Highs and Higher Lows** — the uptrend is healthy at the most important timeframe.")

    if _wk_rs_n and try_f(_wk_rs_n) > 100:
        wk_sentences.append(f"Weekly RS of **{_val(_wk_rs_n,'.0f')}** confirms outperformance is sustained over the medium term, not just a short-term spike.")

    lines.append(" ".join(wk_sentences))
    lines.append("")

    # ── 7. VWAP ──────────────────────────────────────────────────────────────
    lines.append("### 💧 VWAP Analysis")
    v_sentences = []
    if "Above" in vwap_pos and "Extended" not in vwap_pos:
        v_sentences.append(f"{tk} is trading **above a {'rising' if 'Rising' in vwap_slope else 'flat'} VWAP** — the ideal condition for long positions. Institutional algorithms anchor to VWAP; being above it means buyers are in control of the auction.")
    elif "Extended Above" in vwap_pos:
        v_sentences.append(f"{tk} is **extended above VWAP** ({_pct(vs_vwap)} above) — overbought relative to short-term fair value. Consider waiting for a VWAP pullback before adding.")
    elif "Below" in vwap_pos:
        v_sentences.append(f"{tk} is trading **below VWAP** — sellers are currently in control of the short-term auction. Watch for a VWAP reclaim on volume as a potential entry trigger.")
    if vwap:
        v_sentences.append(f"VWAP is at **${float(vwap):.2f}** and the slope is **{vwap_slope.lower()}**.")
    lines.append(" ".join(v_sentences))
    lines.append("")

    # ── 8. PRICE ACTION ──────────────────────────────────────────────────────
    if pa_patterns and pa_patterns not in ("None","nan",""):
        lines.append("### 🕯 Price Action Signals")
        pa_sentences = []
        if "Bullish Engulfing" in pa_engulf:
            pa_sentences.append("A **Bullish Engulfing** candle has formed — a large up candle fully engulfs the previous down candle, signalling strong reversal/continuation conviction from buyers.")
        if "Bullish SFP" in pa_sfp or "Bear Trap" in pa_sfp:
            pa_sentences.append("A **Swing Failure Pattern (Bear Trap)** has been detected — price briefly dipped below a prior swing low but closed back above it, trapping short sellers and typically triggering a sharp move higher. This is one of the highest-probability price action setups.")
        if pa_inside:
            pa_sentences.append("An **Inside Day** (price compression) is present — today's range is entirely within yesterday's, signalling a coiled spring. The breakout direction from this compression determines the next move.")
        if pa_sentences:
            lines.append(" ".join(pa_sentences))
        lines.append("")

    # ── 9. FUNDAMENTALS ──────────────────────────────────────────────────────
    lines.append("### 📊 Fundamental Backdrop")
    f_sentences = []
    if earn_mom and earn_mom not in ("–","None","nan",""):
        if "Strong" in earn_mom:
            f_sentences.append(f"Earnings momentum is **Strong** — the fundamental engine is firing.")
        elif "Moderate" in earn_mom:
            f_sentences.append(f"Earnings momentum is **Moderate** — decent but not exceptional fundamentals.")
        else:
            f_sentences.append(f"Earnings momentum is **{earn_mom}** — fundamentals are not a tailwind here.")

    if eps_growth is not None and not (isinstance(eps_growth, float) and pd.isna(eps_growth)):
        if try_f(eps_growth) >= 50:
            f_sentences.append(f"EPS is growing at **{_pct(eps_growth)}** YoY — this is the kind of earnings acceleration that attracts institutional fund managers.")
        elif try_f(eps_growth) >= 25:
            f_sentences.append(f"EPS growth of **{_pct(eps_growth)}** YoY meets the minimum threshold for a quality growth stock.")
        elif try_f(eps_growth) < 0:
            f_sentences.append(f"⚠️ EPS is declining at **{_pct(eps_growth)}** — earnings are moving in the wrong direction.")

    if eps_accel:
        f_sentences.append(f"Earnings are **accelerating** quarter over quarter — this is the single most powerful fundamental signal, as acceleration drives institutional re-rating and multiple expansion.")

    if try_f(consec_beat) >= 3:
        f_sentences.append(f"**{int(float(consec_beat))} consecutive earnings beats** — management is consistently under-promising and over-delivering, a hallmark of high-quality growth companies.")

    if eps_surp is not None and not (isinstance(eps_surp, float) and pd.isna(eps_surp)):
        if try_f(eps_surp) >= 10:
            f_sentences.append(f"The last quarter came in **{_pct(eps_surp)}** above consensus estimates — a significant positive surprise.")

    if analyst_tgt is not None and price is not None:
        try:
            upside = (float(analyst_tgt) / float(price) - 1) * 100
            f_sentences.append(f"Analyst consensus target is **${float(analyst_tgt):.2f}**, implying **{upside:+.1f}%** upside from current levels.")
        except: pass

    if next_earn and next_earn not in ("–","None","nan",""):
        f_sentences.append(f"⚠️ **Next earnings: {next_earn}** — be aware of binary event risk. Do not hold through earnings unless your position size accounts for a potential 10–30% gap.")

    lines.append(" ".join(f_sentences) if f_sentences else "Fundamental data not yet available for this ticker (Alpha Vantage enrichment may not have run for this stock).")
    lines.append("")

    # ── 10. RISK & POSITION SIZING ───────────────────────────────────────────
    lines.append("### ⚠️ Risk & Position Sizing")
    r_sentences = []
    if is_gem:
        r_sentences.append(f"🔶 **Emerging Gem** — {tk} is classified as a small/mid-cap emerging growth stock. Use a maximum position size of 0.5–1% of portfolio risk. Higher potential reward comes with higher volatility.")
    else:
        r_sentences.append(f"Standard position sizing applies: risk no more than 0.5–2% of total portfolio on any single trade.")
    if liq_warn:
        r_sentences.append(f"⚠️ **Liquidity Warning** — average daily volume is below 300,000 shares. Use limit orders only; avoid market orders. Large positions may be difficult to exit quickly.")
    if try_f(adr) > 6:
        r_sentences.append(f"High ADR of **{_val(adr)}%** means wide intraday swings — set your stop below a meaningful level (swing low or base), not just a fixed percentage.")
    if next_earn and next_earn not in ("–","None","nan",""):
        r_sentences.append(f"Earnings risk on **{next_earn}** — consider reducing position size to half before the report.")
    lines.append(" ".join(r_sentences))
    lines.append("")

    # ── 11. VERDICT ──────────────────────────────────────────────────────────
    lines.append("### 🎯 ApexScan Verdict")
    if try_f(apex) >= 70 and above_200 and ma50_gt_200:
        if breaking:
            verdict = f"**BUY ALERT** — {tk} is actively breaking out with an Apex Score of {_val(apex,'.0f')}/100. This is a high-conviction, time-sensitive setup. Enter as close to the breakout pivot as possible. Set your stop below the base lows."
        elif near_52wh:
            verdict = f"**STRONG WATCH** — {tk} scores {_val(apex,'.0f')}/100 and is near 52-week highs in a confirmed Stage 2 uptrend. Wait for the breakout trigger or a low-risk pullback entry."
        else:
            verdict = f"**ADD TO WATCHLIST** — {tk} scores {_val(apex,'.0f')}/100 and the technical structure is sound. Monitor for a cleaner entry trigger."
    elif try_f(apex) >= 50:
        verdict = f"**WATCHLIST CANDIDATE** — {tk} scores {_val(apex,'.0f')}/100. The setup has merit but lacks one or more key confirmations. Revisit after the next scan."
    else:
        verdict = f"**SKIP / AVOID** — {tk} scores {_val(apex,'.0f')}/100. Insufficient technical and/or fundamental alignment. Better opportunities exist elsewhere in the scan."

    if changes and changes not in ("–","↔ No change","nan"):
        verdict += f" **Recent changes:** {changes}."

    lines.append(verdict)

    return "\n".join(lines)


def try_f(v, default=0.0):
    """Safe float conversion for comparison."""
    try:
        f = float(v)
        return f if not (f != f) else default   # NaN check
    except: return default


def generate_scan_briefing(df: pd.DataFrame) -> str:
    """
    Generate a full market briefing from a scan result DataFrame.
    Covers: market summary, top setups, active breakouts, sector rotation,
    risk reminders. No API required.
    """
    if df.empty:
        return "No scan data available. Run a Live Scan first."

    lines = []
    today = datetime.now().strftime("%A, %B %d %Y")
    lines.append(f"# 📡 ApexScan Morning Briefing — {today}")
    lines.append("")

    # ── Market overview ───────────────────────────────────────────────────────
    total       = len(df)
    stage2      = df[df.get("stage","").astype(str).str.contains("2") if "stage" in df.columns else pd.Series([True]*total)].shape[0]
    breakouts   = df[df["breaking_out"].astype(str).str.lower()=="true"].shape[0] if "breaking_out" in df.columns else 0
    avg_score   = round(df["apex_score"].mean(), 1) if "apex_score" in df.columns else 0
    high_conv   = (df["apex_score"] >= 70).sum() if "apex_score" in df.columns else 0
    bull_flow   = df[df.get("of_bias","").astype(str).str.contains("Bullish")].shape[0] if "of_bias" in df.columns else 0

    lines.append("## 📊 Market Snapshot")
    lines.append(
        f"Today's scan identified **{total} qualifying setups** across the watchlist. "
        f"**{stage2}** are in confirmed Stage 2 uptrends. "
        f"Average Apex Score across all results is **{avg_score}/100**. "
        f"**{high_conv}** stocks score 70+, indicating high-conviction setups. "
        f"**{breakouts}** active breakouts detected. "
        f"**{bull_flow}** stocks show bullish or strong-bullish institutional order flow."
    )
    lines.append("")

    # ── Top 5 by Apex Score ───────────────────────────────────────────────────
    lines.append("## 🏆 Top 5 Setups (by Apex Score)")
    top5 = df.nlargest(5, "apex_score") if "apex_score" in df.columns else df.head(5)
    for i, (_, r) in enumerate(top5.iterrows(), 1):
        tk     = r.get("ticker","–")
        sc     = r.get("apex_score","–")
        stg    = r.get("stage","–")
        p3m    = r.get("perf_3m_%")
        rs     = r.get("rs_3m")
        bo     = _bool(r.get("breaking_out"))
        pat    = str(r.get("pattern","") or "")
        of_b   = str(r.get("of_bias","") or "")
        emoji  = "🚨" if bo else ("⭐" if try_f(sc) >= 70 else "📌")
        bo_str = " **— BREAKING OUT NOW**" if bo else ""
        lines.append(
            f"{i}. {emoji} **{tk}** (Score: {_val(sc,'.0f')}/100 | Stage: {stg} | "
            f"3M: {_pct(p3m)} | RS: {_val(rs,'.0f')} | Flow: {of_b}){bo_str}"
        )
        if pat and pat not in ("None","nan",""):
            lines.append(f"   ↳ Pattern: {pat}")
    lines.append("")

    # ── Active breakouts ─────────────────────────────────────────────────────
    if "breaking_out" in df.columns:
        bo_df = df[df["breaking_out"].astype(str).str.lower()=="true"]
        if not bo_df.empty:
            lines.append("## 🚨 Active Breakouts — Time-Sensitive")
            for _, r in bo_df.iterrows():
                tk   = r.get("ticker","–")
                sc   = r.get("apex_score","–")
                p    = r.get("price")
                pat  = r.get("pattern","")
                vs   = r.get("vol_surge_x")
                lines.append(
                    f"- **{tk}** @ ${float(p):.2f if p else '–'} | Score: {_val(sc,'.0f')} | "
                    f"Pattern: {pat} | Vol surge: {_val(vs,'.1f')}×"
                )
            lines.append("")

    # ── Sector rotation ───────────────────────────────────────────────────────
    if "theme" in df.columns and "apex_score" in df.columns:
        lines.append("## 🔄 Sector Rotation — Where is Money Flowing?")
        sec_grp = df.groupby("theme")["apex_score"].agg(["mean","count"]).sort_values("mean", ascending=False)
        for sec, row_s in sec_grp.head(5).iterrows():
            avg_s = round(row_s["mean"], 1)
            cnt   = int(row_s["count"])
            bar   = "█" * int(avg_s // 10)
            lines.append(f"- **{sec}**: avg score {avg_s}/100 ({cnt} stocks)  {bar}")
        lines.append("")

    # ── Risk reminders ────────────────────────────────────────────────────────
    lines.append("## ⚠️ Risk Reminders")
    earn_soon = []
    if "next_earnings" in df.columns:
        for _, r in df.iterrows():
            ne = str(r.get("next_earnings","") or "")
            if ne and ne not in ("–","None","nan",""):
                try:
                    days_away = (pd.to_datetime(ne) - datetime.now()).days
                    if 0 <= days_away <= 14:
                        earn_soon.append((r.get("ticker","–"), ne, days_away))
                except: pass
    if earn_soon:
        lines.append("**Earnings in the next 14 days — manage risk:**")
        for tk, ne, d in sorted(earn_soon, key=lambda x: x[2]):
            lines.append(f"- **{tk}** reports in {d} days ({ne}) — consider half-size positions")
    else:
        lines.append("No earnings events flagged within the next 14 days for scanned stocks.")

    liq_warns = df[df.get("liquidity_warn","").astype(str).str.lower()=="true"]["ticker"].tolist() if "liquidity_warn" in df.columns else []
    if liq_warns:
        lines.append(f"**Low liquidity — use limit orders only:** {', '.join(liq_warns[:5])}")

    lines.append("")
    lines.append("---")
    lines.append("*Generated by ApexScan rule-based narrative engine. Not financial advice.*")

    return "\n".join(lines)


# ── Column metadata: display label + plain-English interpretation ──────────────
COLUMN_META = {
    "rank":            ("Rank",               "Position in this scan, sorted by Apex Score descending. #1 is the strongest setup right now."),
    "ticker":          ("Ticker",             "Stock symbol. The unique identifier for this company on the exchange."),
    "market":          ("Market",             "Market the ticker trades on. Currently US-only."),
    "theme":           ("Sector / Theme",     "GICS sector or config theme this stock belongs to. GICS sectors: Energy, Materials, Industrials, Utilities, Healthcare, Financials, Consumer Discretionary, Consumer Staples, Information Technology, Communication Services, Real Estate. Config themes (ai_semis, cybersecurity etc.) take priority for stocks in your watchlist themes. Sector rotation: when a full GICS sector starts outperforming, it lifts all stocks within it."),
    "price":           ("Price ($)",          "Last closing price in USD."),
    "stage":           ("Stage",              "Weinstein Stage. Stage 2 ✅ = only buyable stage (price above both MAs, 50MA > 200MA). Stage 1 = basing. Stage 3 = topping. Stage 4 🔴 = downtrend — avoid."),
    "perf_1m_%":       ("1M Return %",        "Price performance over the last 21 trading days (≈1 month). Captures recent momentum. >5% is positive."),
    "perf_3m_%":       ("3M Return %",        "Price performance over the last 63 trading days (≈3 months). Core momentum filter. >15% is strong; >30% is exceptional."),
    "perf_6m_%":       ("6M Return %",        "Price performance over 126 trading days (≈6 months). Confirms the trend has durability, not just a one-month spike."),
    "rs_3m":           ("RS 3M",              "Relative Strength vs S&P 500 over 3 months. >100 = outperforming the index. >150 = massively outperforming. Buy leaders (>100), avoid laggards (<70)."),
    "rs_6m":           ("RS 6M",              "Relative Strength vs S&P 500 over 6 months. Confirms the outperformance is sustained, not just a recent fluke."),
    "rs_r2500_3m":     ("RS vs R2500 (3M)",   "Relative Strength vs the Russell 2500 index over 3 months. The Russell 2500 covers small and mid-cap stocks — this tells you if the stock is beating its natural peer group. >100 = beating small/mid peers. Critical for emerging growth stocks."),
    "rs_r2500_6m":     ("RS vs R2500 (6M)",   "Relative Strength vs Russell 2500 over 6 months. Confirms the stock has been a sustained leader within the small/mid-cap universe, not just a short-term spike."),
    "rs_r3000g_3m":    ("RS vs R3000G (3M)",  "Relative Strength vs the Russell 3000 Growth index over 3 months. The R3000 Growth is the broadest growth benchmark covering all cap sizes. >100 = outperforming the entire growth universe. This is the toughest RS test."),
    "rs_r3000g_6m":    ("RS vs R3000G (6M)",  "Relative Strength vs Russell 3000 Growth over 6 months. A stock beating the R3000G over 6 months is a genuine sustained growth leader across the entire market."),
    "rs_multi_leader": ("Multi-Bench Leader",  "True if the stock is outperforming ALL three benchmarks (S&P 500, Russell 2500, Russell 3000 Growth) simultaneously on a 3-month basis. This is the highest RS signal — the stock is leading across every peer group. Extremely rare and extremely bullish."),
    "adr_%":           ("ADR %",              "Average Daily Range % over the last 20 days. Measures volatility/movement potential. Higher ADR = bigger daily swings = wider stops needed but larger profit potential."),
    "vs_50ma_%":       ("vs 50MA %",          "How far the current price is above or below the 50-day moving average. >10% above = extended, risk of pullback. Below = potential support or weakness."),
    "vs_200ma_%":      ("vs 200MA %",         "How far price is above or below the 200-day moving average. The 200MA is the long-term trend line. Below 200MA = avoid for long entries."),
    "volume":          ("Volume (Today)",     "Today's raw share volume. High volume on up days confirms institutional participation."),
    "vol_filter":      ("Volume (Filter)",    "Max of today's volume and 20-day average volume. Used as the liquidity filter threshold to avoid illiquid setups."),
    "vol_surge_x":     ("Vol Surge X",        "Ratio of 5-day average volume to 50-day average volume. >1.4x = elevated interest. >2x = significant surge, often accompanies breakouts."),
    "above_50ma":      ("Above 50MA",         "True/False — price is above the 50-day moving average. A basic trend filter. False means the stock is below near-term trend support."),
    "above_200ma":     ("Above 200MA",        "True/False — price is above the 200-day moving average. Core long-term trend filter. False = avoid for swing/position trades."),
    "ma50_gt_ma200":   ("50MA > 200MA",       "True/False — the 50-day MA is above the 200-day MA (Golden Cross condition). Required for Stage 2 confirmation. False = trend not yet established."),
    "near_52wh":       ("Near 52W High",      "True/False — price is within 15% of its 52-week high. Breakouts happen near highs, not at the bottom. True = stock is in the right zone for a breakout."),
    "pct_off_high_%":  ("% Off 52W High",     "How far below the 52-week high the stock currently is. 0% = at all-time highs. -10% = 10% below highs. The best breakouts come from stocks less than 10–15% below highs."),
    "pattern":         ("Pattern",            "Base/breakout pattern detected. E.g. 'Flat Base Breakout', 'Cup Breakout', 'Handle Forming', 'Tight Base'. Breakout = active trigger. Handle/Tight = watch for entry."),
    "breaking_out":    ("Breaking Out",       "True/False — the stock is actively breaking out of a base on volume. True is the highest-priority actionable signal in the entire scan."),
    "news_count":      ("News Count",         "Number of news articles in the last 7 days (via Finnhub). Higher count can indicate a catalyst event (earnings, product launch, analyst upgrade)."),
    "sentiment":       ("Sentiment",          "News sentiment from Finnhub: Positive, Neutral, or N/A. Positive sentiment alongside technical strength is a confluence signal."),
    "earn_momentum":   ("Earnings Momentum",  "Earnings quality signal: Strong / Moderate / Weak. Derived from EPS growth, surprise %, and acceleration (Alpha Vantage) or proxied from news + price when AV not available."),
    "eps_growth_%":    ("EPS Growth %",       "Year-over-year EPS growth from Alpha Vantage. >25% = strong growth stock. >50% = exceptional. Negative = earnings declining — be cautious."),
    "eps_surprise_%":  ("EPS Surprise %",     "How much the last quarterly EPS beat or missed analyst consensus. >5% beat = positive signal. Misses weigh on price even in uptrends."),
    "eps_accel":       ("EPS Accelerating",   "True/False — earnings growth rate is accelerating quarter over quarter. Acceleration is the most powerful fundamental signal; it drives institutional re-rating and multiple expansion."),
    "consec_beats":    ("Consec. Beats",      "Number of consecutive quarters where the company beat EPS estimates. 3+ beats = management is consistently under-promising and over-delivering — a trust signal for institutions."),
    "rev_growth_%":    ("Revenue Growth %",   "Year-over-year revenue growth. Confirms earnings improvement is driven by real business expansion, not just cost cuts or financial engineering."),
    "eps_score":       ("EPS Score /15",      "Composite earnings score from 0–15. Combines growth, surprise %, acceleration, consecutive beats, and revenue growth. >10 = very strong fundamental backdrop."),
    "eps_trend":       ("EPS Trend",          "List of recent quarterly EPS values (most recent first). Shows whether earnings are growing, flat, or declining over recent quarters."),
    "analyst_target":  ("Analyst Target",     "Consensus analyst price target from Alpha Vantage. Compare to current price to see implied upside. Use as a reference, not gospel."),
    "pe_ratio":        ("P/E Ratio",          "Price-to-Earnings ratio. Growth stocks often trade at high P/Es (40–100x). A high P/E is fine if earnings are growing fast; what matters is whether the growth justifies the premium."),
    "peg_ratio":       ("PEG Ratio",          "Price/Earnings-to-Growth ratio. PEG < 1.0 = potentially undervalued relative to growth rate. PEG > 2.0 = growth is fully or over-priced. The Goldilocks zone is 1.0–1.5."),
    "eps_details":     ("EPS Details",        "Raw detail string from Alpha Vantage showing the last few quarters of EPS. Useful for verifying the trend behind the score."),
    "next_earnings":   ("Next Earnings",      "Upcoming earnings date from Alpha Vantage or yfinance. Critical for risk management — stocks can gap 10–30% on earnings. Do not hold through earnings unless sized appropriately."),
    "of_bias":         ("Order Flow Bias",    "Directional order flow assessment: Strong Bullish / Bullish / Neutral / Bearish / Strong Bearish. Measures whether institutions are persistently buying or selling over the last 10 sessions."),
    "of_up_vol_ratio": ("OF Up/Down Vol",     "Ratio of total volume on up days vs down days over the last 10 sessions. >1.5x = institutional-level buying. >2.0x = heavy accumulation. <0.7x = distribution pattern."),
    "of_bullish_days": ("OF Bullish Days %",  "Percentage of the last 10 sessions that closed higher than the previous day. >60% = persistent buying. >70% = strong institutional flow. <40% = bearish pressure."),
    "of_consec_up":    ("OF Consec. Up",      "Maximum consecutive up-closes in the last 10 sessions. 4+ consecutive up-closes suggests an active TWAP/VWAP algorithm systematically working a large buy order."),
    "of_score":        ("OF Score /8",        "Order Flow Persistence Score from 0–8. Combines bullish day %, up/down volume ratio, and consecutive up-closes. Score ≥6 = strong institutional flow. Added to Apex Score."),
    "vwap":            ("VWAP ($)",           "Volume Weighted Average Price over the last 20 days. The fairest measure of where the market has agreed to transact. Institutional algorithms often anchor to VWAP for large order execution."),
    "vwap_upper":      ("VWAP Upper Band",    "VWAP + 1 standard deviation. Acts as short-term resistance. Price extended above this band is overextended and likely to mean-revert. Consider tightening stops above this level."),
    "vwap_lower":      ("VWAP Lower Band",    "VWAP - 1 standard deviation. Acts as short-term support. Strong stocks often find buyers at this level — it is a potential low-risk entry zone in an uptrend."),
    "vs_vwap_%":       ("vs VWAP %",          "How far the current price is above or below VWAP. +5% = extended above. -5% = below VWAP. Ideal long entries are +1% to +4% above a rising VWAP."),
    "vwap_position":   ("VWAP Position",      "Categorical VWAP relationship: 'Above VWAP', 'Extended Above VWAP' (overbought zone), 'Below VWAP' (weak), 'Extended Below VWAP' (avoid). Above a rising VWAP = strongest long scenario."),
    "vwap_slope":      ("VWAP Slope",         "Direction of the VWAP trend: Rising / Flat / Falling. A Rising VWAP confirms buyers are in control and value is being accepted higher — ideal for longs. Falling VWAP = sellers in control."),
    "vwap_score":      ("VWAP Score /4",      "VWAP contribution to Apex Score (0–4). Max score when price is above a rising VWAP with a recent VWAP reclaim. Added to total Apex Score."),
    "ms_structure":    ("Market Structure",   "Overall market structure assessment: 'Bullish (HH/HL)' = uptrend confirmed. 'Bearish (LH/LL)' = downtrend. 'Transitioning' = structure is shifting. Only trade longs in Bullish structure."),
    "ms_hh_hl":        ("HH/HL Confirmed",   "True/False — the stock is making Higher Highs and Higher Lows. This is the textbook definition of an uptrend. True = structure is intact for longs. False = do not buy the dip."),
    "ms_bos":          ("Break of Structure", "True/False — a recent Break of Structure occurred (price broke above a prior swing high in a downtrend, or below a prior swing low in an uptrend). In an uptrend, BOS confirms trend acceleration."),
    "ms_swing_high":   ("Last Swing High",    "Price of the most recent swing high pivot. Acts as near-term resistance. A close above this level on volume confirms the next leg of the uptrend."),
    "ms_swing_low":    ("Last Swing Low",     "Price of the most recent swing low pivot. Acts as near-term support. A close below this level in an uptrend is a warning sign — consider tightening stops."),
    "pa_patterns":     ("PA Patterns",        "Price Action patterns detected on the most recent candle(s). E.g. 'Bullish Engulfing', 'Bullish SFP (Bear Trap)', 'Inside Day (Compression)', 'PA Confluence'. Multiple signals = stronger setup."),
    "pa_engulfing":    ("PA Engulfing",       "Engulfing candle signal: 'Bullish' = a large up candle fully engulfs the prior down candle — strong reversal/continuation signal. 'Bearish' = the opposite. None = no engulfing pattern."),
    "pa_sfp":          ("PA SFP",            "Swing Failure Pattern: 'Bullish SFP (Bear Trap)' = price briefly dipped below a prior swing low but closed back above it, trapping short sellers. One of the most reliable reversal signals in technical analysis."),
    "pa_inside_day":   ("PA Inside Day",      "True/False — today's high is lower than yesterday's high AND today's low is higher than yesterday's low. An Inside Day signals price compression and a potential explosive move. Watch for the breakout direction."),
    "pa_context":      ("PA Context Candle",  "Context candle analysis: 'Bullish' = large-range up candle closing near the high on above-average volume. 'Bearish' = large-range down candle. Bullish context candle confirms strong buying intent."),
    "pa_score":        ("PA Score /5",        "Price Action Score from 0–5. Combines engulfing, SFP, inside day, and context candle signals. Score ≥3 = meaningful price action confirmation. Added to total Apex Score."),
    # ── Weekly Timeframe Confirmation ────────────────────────────────────────
    "weekly_stage":          ("Weekly Stage",         "Weinstein Stage on the WEEKLY chart. This is the most important filter of all — a daily breakout inside a weekly Stage 4 downtrend is a trap that fails 70%+ of the time. You MUST see weekly Stage 2 or at minimum Stage 1 basing before acting on daily signals."),
    "weekly_above_10wma":    ("Above 10WMA",          "True if price is above the 10-week moving average (equivalent to the 50-day MA on the weekly chart). Losing this level in an uptrend is the first warning sign of deteriorating momentum."),
    "weekly_above_40wma":    ("Above 40WMA",          "True if price is above the 40-week moving average (equivalent to the 200-day MA on the weekly chart). This is the most important long-term trend line. Being above it is a non-negotiable requirement for swing positions."),
    "weekly_10gt40":         ("10WMA > 40WMA",        "True if the 10-week MA is above the 40-week MA — the weekly golden cross. This confirms a proper weekly Stage 2 uptrend. When both MAs slope up and price is above both, the path of least resistance is higher."),
    "weekly_rs":             ("Weekly RS",            "Relative Strength vs S&P 500 on the weekly timeframe (13-week lookback). A stock with positive weekly RS is genuinely outperforming the market over the medium term, not just a daily spike. >100 = outperforming."),
    "weekly_base_tight":     ("Weekly Base Tight",    "True if the last 8 weeks show a price range of less than 15%. A tight weekly base = low volatility consolidation = energy coiling. The tightest weekly bases precede the most explosive moves. This is what you are looking for before a big breakout."),
    "weekly_base_depth_%":   ("Weekly Base Depth %",  "How deep the last 8-week price range is as a percentage. <10% = extremely tight (rare, very bullish). 10–20% = healthy base. 20–30% = deeper correction but workable. >30% = too volatile, avoid."),
    "weekly_hh_hl":          ("Weekly HH/HL",         "True if the weekly chart is making Higher Highs and Higher Lows over the last 6 weeks. This is the textbook definition of a healthy uptrend on the timeframe that matters most. When this is True on the weekly, daily dips are buying opportunities."),
    "weekly_trending_up":    ("Weekly Trending Up",   "True if the stock closed higher for 2 or more consecutive weeks. Multiple consecutive up weeks signals sustained buying, not a one-day spike. 3+ consecutive up weeks = institutional conviction."),
    "weekly_consec_up_wks":  ("Consec. Up Weeks",     "Number of consecutive weeks the stock closed higher. 1–2 = momentum building. 3–4 = strong trend in motion. 5+ = potentially extended — wait for a weekly pullback before adding."),
    "weekly_confirmed":      ("Weekly Confirmed ✅",   "True when the weekly chart fully supports the daily setup: above 40WMA, 10WMA > 40WMA, and positive weekly RS. This is the green light — daily signal + weekly confirmation = highest probability trade. Never take a daily signal without checking this."),
    "weekly_contradicts":    ("Weekly Contradicts ⚠️", "True when the weekly chart is in Stage 3 or 4 while the daily shows a Stage 2 setup. This is the most common trap for retail traders — a daily breakout inside a weekly downtrend. The weekly trend will win eventually. Avoid or use very small size."),
    "weekly_score":          ("Weekly Score /10",      "Weekly timeframe contribution to Apex Score (0–10). Added to score when weekly is confirmed, subtracted (-15) when weekly contradicts daily. Max score when weekly Stage 2 + HH/HL + tight base + RS leader all align simultaneously."),

    # ── Early Entry Signals ──────────────────────────────────────────────────
    "early_entry":             ("Early Entry",         "True if the stock shows one or more early-stage entry signals — fresh MA cross, pullback to 50MA, low-ADR base, or inside day compression. These are the setups to buy BEFORE the move is obvious."),
    "early_entry_type":        ("Entry Type",          "What type of early entry signal was detected: Fresh 200MA Cross (brand-new uptrend), Fresh 50MA Cross (momentum turning), Pullback to 50MA (low-risk add), Low-ADR Base (tight coiled spring), or Inside Day Compression (explosive move pending)."),
    "fresh_200ma_cross":       ("Fresh 200MA Cross",   "True if price crossed above the 200-day MA within the last 10 bars. This is the single most powerful early entry signal — the stock is literally just entering a new uptrend. Buy zones are typically within 5% of the 200MA."),
    "fresh_50ma_cross":        ("Fresh 50MA Cross",    "True if price crossed above the 50-day MA within the last 5 bars. Momentum is just turning. Earlier than waiting for a full Stage 2 confirmation."),
    "pullback_to_50ma":        ("Pullback to 50MA",    "True if price is within 3% of the 50MA while above the 200MA. The classic low-risk add-to-winner entry — you're buying the dip within an established uptrend with the stop just below the 50MA."),
    "low_adr_base":            ("Low-ADR Base",        "True if the 20-day Average Daily Range is below 3%. Low volatility during a base means the stock is coiling — energy is being stored for the eventual breakout. This is cheap entry in terms of stop distance."),
    "early_entry_score":       ("Early Entry Score /10","Composite early-entry score from 0–10. Each signal type adds points: Fresh 200MA Cross (+8), Pullback to 50MA (+5), Fresh 50MA Cross (+4), Low-ADR Base (+3), Inside Day Compression (+2). Score ≥8 = textbook early entry."),
    "days_since_200ma_cross":  ("Days Since 200MA Cross","How many days ago price crossed above the 200MA. 1–3 days = hottest possible signal. 4–10 days = still early. >10 days = momentum is established, no longer early entry."),
    "apex_score":      ("Apex Score /100",    "The master composite score (0–100). Combines momentum (40 pts), RS (25 pts), stage/MAs (15 pts), 52W high (10 pts), breakout (10 pts), order flow (8 pts), price action (5 pts), VWAP (4 pts), structure (3 pts), EPS (15 pts). >70 = high conviction. 40–70 = watchlist."),
    "scanned_at":      ("Scanned At",         "Timestamp of when this ticker was analysed. All results in a single scan share the same approximate timestamp."),
    "market_cap":      ("Market Cap ($)",     "Total market capitalisation in dollars. Larger caps are more liquid; smaller caps (micro/small) have higher volatility and risk."),
    "market_cap_bn":   ("Market Cap ($B)",    "Market cap in billions for easier reading. <$2B = small/micro cap. $2–10B = mid cap. $10–200B = large cap. >$200B = mega cap."),
    "mcap_category":   ("MCap Category",      "Categorical size label: Micro Cap (<$300M), Small Cap ($300M–$2B), Mid Cap ($2B–$10B), Large Cap ($10B–$200B), Mega Cap (>$200B). Smaller = higher risk/reward."),
    "is_gem":          ("Is Gem",             "True if the stock qualifies as an Emerging Gem ($100M–$5B market cap). Gems receive score boosts for strong OF and PA signals. Use smaller position sizes (0.5–1% max risk)."),
    "liquidity_score": ("Liquidity Score",    "Liquidity rating 0–3 based on 30-day average volume. Score 3 = >1M shares/day (highly liquid). Score 2 = 300K–1M. Score 1 = 100K–300K. Score 0 = <100K (very illiquid, avoid large positions)."),
    "liquidity_warn":  ("Liquidity Warning",  "True if 30-day average volume is below 300,000 shares/day. Low liquidity means wide bid-ask spreads and difficulty exiting large positions — size down significantly."),
    "avg_volume_30d":  ("Avg Volume 30D",     "30-day average daily share volume from yfinance fast_info. The most reliable liquidity measure. Below 300K = tread carefully; below 100K = avoid unless small position size."),
    "changes":         ("Changes",            "Plain-English summary of what changed since the last scan. E.g. 'Score ▲8 | Flow→Strong Bull | Reclaimed VWAP'. 🆕 = new entry. ↔ = no notable change."),
    "is_new":          ("Is New",             "True if this ticker was not present in the previous scan. New entries have no delta comparison — treat them as fresh signals requiring your own confirmation."),
    "delta_score":     ("Score Change",       "Apex Score change since the last scan (positive = improving, negative = deteriorating). A stock jumping +10 points between scans is gaining momentum and worth attention."),
}


def build_excel_download(ticker_row: pd.Series, ticker_name: str) -> bytes:
    """
    Build a fully-labelled Excel workbook for a single ticker deep read.
    Sheet 1: Raw data with every column labelled + interpreted.
    Sheet 2: Signal summary scorecard.
    Returns bytes ready for st.download_button.
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl not available
        buf = io.StringIO()
        ticker_row.to_frame().T.to_csv(buf)
        return buf.getvalue().encode("utf-8")

    wb = openpyxl.Workbook()

    # ── Colours ───────────────────────────────────────────────────────────────
    BG_DARK   = "0D1117"
    BG_CARD   = "161B22"
    BG_GREEN  = "1A3A2A"
    BG_AMBER  = "2A2200"
    BG_RED    = "2A1010"
    FG_GREEN  = "3FB950"
    FG_AMBER  = "D29922"
    FG_RED    = "F85149"
    FG_BLUE   = "388BFD"
    FG_WHITE  = "E6EDF3"
    FG_GREY   = "8B949E"
    BORDER_C  = "30363D"

    def fill(hex_): return PatternFill("solid", fgColor=hex_)
    def font(hex_, bold=False, sz=11): return Font(color=hex_, bold=bold, size=sz, name="Calibri")
    def thin_border():
        s = Side(border_style="thin", color=BORDER_C)
        return Border(left=s, right=s, top=s, bottom=s)
    def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left():   return Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Full Data Table
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = f"{ticker_name} — Data"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A3"

    # Title row
    ws1.merge_cells("A1:D1")
    t = ws1["A1"]
    t.value = f"ApexScan — {ticker_name} Deep Read    |    Scanned: {ticker_row.get('scanned_at','–')}"
    t.font      = font(FG_WHITE, bold=True, sz=14)
    t.fill      = fill(BG_CARD)
    t.alignment = center()
    ws1.row_dimensions[1].height = 30

    # Header row
    headers = ["Field", "Label", "Value", "Interpretation"]
    hdr_colours = [FG_BLUE, FG_BLUE, FG_AMBER, FG_WHITE]
    for col_i, (h, c) in enumerate(zip(headers, hdr_colours), 1):
        cell = ws1.cell(row=2, column=col_i, value=h)
        cell.font      = font(c, bold=True, sz=10)
        cell.fill      = fill(BG_CARD)
        cell.alignment = center()
        cell.border    = thin_border()
    ws1.row_dimensions[2].height = 20

    # Define column order (full list from your spec)
    ORDERED_COLS = [
        "rank","ticker","market","theme","price","stage",
        "perf_1m_%","perf_3m_%","perf_6m_%","rs_3m","rs_6m",
        "rs_r2500_3m","rs_r2500_6m","rs_r3000g_3m","rs_r3000g_6m","rs_multi_leader",
        "adr_%","vs_50ma_%","vs_200ma_%","volume","vol_filter","vol_surge_x",
        "above_50ma","above_200ma","ma50_gt_ma200","near_52wh","pct_off_high_%",
        "pattern","breaking_out","news_count","sentiment","earn_momentum",
        "eps_growth_%","eps_surprise_%","eps_accel","consec_beats","rev_growth_%",
        "eps_score","eps_trend","analyst_target","pe_ratio","peg_ratio","eps_details","next_earnings",
        "of_bias","of_up_vol_ratio","of_bullish_days","of_consec_up","of_score",
        "vwap","vwap_upper","vwap_lower","vs_vwap_%","vwap_position","vwap_slope","vwap_score",
        "ms_structure","ms_hh_hl","ms_bos","ms_swing_high","ms_swing_low",
        "pa_patterns","pa_engulfing","pa_sfp","pa_inside_day","pa_context","pa_score",
        "apex_score","scanned_at","market_cap","market_cap_bn","mcap_category",
        "is_gem","liquidity_score","liquidity_warn","avg_volume_30d",
        "changes","is_new","delta_score",
        "weekly_stage","weekly_above_40wma","weekly_10gt40","weekly_rs",
        "weekly_base_tight","weekly_base_depth_%","weekly_hh_hl",
        "weekly_confirmed","weekly_contradicts","weekly_score",
        "early_entry","early_entry_type","fresh_200ma_cross",
        "fresh_50ma_cross","pullback_to_50ma","low_adr_base",
        "early_entry_score","days_since_200ma_cross",
    ]

    def fmt_val(col, raw):
        """Format raw value for display."""
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            return "–"
        if col in ("price","vwap","vwap_upper","vwap_lower","ms_swing_high","ms_swing_low","analyst_target"):
            try: return f"${float(raw):.2f}"
            except: return str(raw)
        if col in ("perf_1m_%","perf_3m_%","perf_6m_%","vs_50ma_%","vs_200ma_%",
                   "vs_vwap_%","eps_growth_%","eps_surprise_%","rev_growth_%","pct_off_high_%"):
            try:
                v = float(raw)
                return f"{v:+.1f}%"
            except: return str(raw)
        if col in ("adr_%",):
            try: return f"{float(raw):.1f}%"
            except: return str(raw)
        if col in ("of_bullish_days",):
            try: return f"{float(raw):.0f}%"
            except: return str(raw)
        if col in ("rs_3m","rs_6m"):
            try: return f"{float(raw):.0f}"
            except: return str(raw)
        if col in ("vol_surge_x","of_up_vol_ratio"):
            try: return f"{float(raw):.2f}x"
            except: return str(raw)
        if col in ("market_cap",):
            try:
                v = float(raw)
                return f"${v/1e9:.2f}B" if v >= 1e9 else f"${v/1e6:.0f}M"
            except: return str(raw)
        if col in ("pe_ratio","peg_ratio"):
            try: return f"{float(raw):.2f}"
            except: return str(raw)
        if col in ("eps_trend",) and isinstance(raw, list):
            return " → ".join(str(x) for x in raw[:6])
        return str(raw)

    def row_bg(col, raw):
        """Pick a background fill based on signal quality."""
        col_green = {
            "stage": lambda v: "2 ✅" in str(v),
            "above_50ma": lambda v: v is True or str(v).lower() == "true",
            "above_200ma": lambda v: v is True or str(v).lower() == "true",
            "ma50_gt_ma200": lambda v: v is True or str(v).lower() == "true",
            "near_52wh": lambda v: v is True or str(v).lower() == "true",
            "breaking_out": lambda v: v is True or str(v).lower() == "true",
            "ms_hh_hl": lambda v: v is True or str(v).lower() == "true",
            "ms_bos": lambda v: v is True or str(v).lower() == "true",
            "pa_inside_day": lambda v: v is True or str(v).lower() == "true",
            "of_bias": lambda v: "Bullish" in str(v),
            "vwap_position": lambda v: "Above" in str(v) and "Extended" not in str(v),
            "ms_structure": lambda v: "Bullish" in str(v),
            "earn_momentum": lambda v: "Strong" in str(v),
            "eps_accel": lambda v: v is True or str(v).lower() == "true",
            "is_gem": lambda v: v is True or str(v).lower() == "true",
        }
        col_red = {
            "stage": lambda v: "4 🔴" in str(v),
            "above_50ma": lambda v: v is False or str(v).lower() == "false",
            "above_200ma": lambda v: v is False or str(v).lower() == "false",
            "of_bias": lambda v: "Bearish" in str(v),
            "vwap_position": lambda v: "Below" in str(v),
            "ms_structure": lambda v: "Bearish" in str(v),
            "liquidity_warn": lambda v: v is True or str(v).lower() == "true",
        }
        if col in col_green and col_green[col](raw):  return fill(BG_GREEN)
        if col in col_red   and col_red[col](raw):    return fill(BG_RED)
        return fill(BG_DARK)

    # Data rows
    for r_i, col_key in enumerate(ORDERED_COLS, 3):
        raw = ticker_row.get(col_key)
        if col_key == "rank":
            # rank comes from the index
            raw = ticker_row.name if hasattr(ticker_row, "name") else "–"

        meta  = COLUMN_META.get(col_key, (col_key, "–"))
        label = meta[0]
        interp= meta[1]
        val   = fmt_val(col_key, raw)
        bg    = row_bg(col_key, raw)

        row_data = [(col_key, FG_GREY), (label, FG_AMBER), (val, FG_WHITE), (interp, FG_GREY)]
        for c_i, (text, fg) in enumerate(row_data, 1):
            cell = ws1.cell(row=r_i, column=c_i, value=str(text))
            cell.font      = font(fg, bold=(c_i == 3), sz=10)
            cell.fill      = bg if c_i == 3 else fill(BG_DARK if r_i % 2 == 1 else BG_CARD)
            cell.alignment = left()
            cell.border    = thin_border()
        ws1.row_dimensions[r_i].height = 48

    # Column widths
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 80

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Signal Scorecard
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title=f"{ticker_name} — Scorecard")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:C1")
    t2 = ws2["A1"]
    t2.value    = f"ApexScan — {ticker_name} Signal Scorecard"
    t2.font     = font(FG_WHITE, bold=True, sz=14)
    t2.fill     = fill(BG_CARD)
    t2.alignment = center()
    ws2.row_dimensions[1].height = 30

    scorecard_items = [
        ("📊 APEX SCORE",     ticker_row.get("apex_score","–"),       ">70 = High Conviction | 40–70 = Watchlist | <40 = Not Ready"),
        ("📐 Stage",          ticker_row.get("stage","–"),             "Stage 2 ✅ is the only buyable stage"),
        ("📈 3M Return",      fmt_val("perf_3m_%", ticker_row.get("perf_3m_%")), ">15% = strong momentum threshold"),
        ("⚡ RS vs S&P500",   fmt_val("rs_3m", ticker_row.get("rs_3m")),        ">100 = outperforming the market"),
        ("🚀 Breaking Out",   ticker_row.get("breaking_out","–"),       "True = active breakout — highest priority"),
        ("🌊 Order Flow",     ticker_row.get("of_bias","–"),            "Strong Bullish = institutional accumulation"),
        ("💧 VWAP Position",  ticker_row.get("vwap_position","–"),      "Above Rising VWAP = ideal long zone"),
        ("🏗 MS Structure",   ticker_row.get("ms_structure","–"),       "Bullish HH/HL = uptrend confirmed"),
        ("🎯 Price Action",   ticker_row.get("pa_patterns","None"),     "SFP / Confluence = highest PA signal"),
        ("💎 Is Gem",         ticker_row.get("is_gem","–"),             "True = emerging gem — use half position size"),
        ("📅 Next Earnings",  ticker_row.get("next_earnings","–"),      "Do not hold through earnings unless sized for gap risk"),
        ("💰 EPS Growth",     fmt_val("eps_growth_%", ticker_row.get("eps_growth_%")), ">25% = strong growth stock"),
        ("🎯 Analyst Target", fmt_val("analyst_target", ticker_row.get("analyst_target")), "Consensus price target"),
        ("🔄 Since Last Scan",ticker_row.get("changes","–"),           "What changed vs previous scan"),
    ]

    for sc_i, (label, value, note) in enumerate(scorecard_items, 2):
        ws2.cell(sc_i, 1, label).font      = font(FG_AMBER, bold=True, sz=10)
        ws2.cell(sc_i, 1).fill             = fill(BG_CARD)
        ws2.cell(sc_i, 1).alignment        = left()
        ws2.cell(sc_i, 1).border           = thin_border()
        ws2.cell(sc_i, 2, str(value)).font  = font(FG_WHITE, bold=True, sz=11)
        ws2.cell(sc_i, 2).fill             = fill(BG_DARK)
        ws2.cell(sc_i, 2).alignment        = center()
        ws2.cell(sc_i, 2).border           = thin_border()
        ws2.cell(sc_i, 3, note).font       = font(FG_GREY, sz=9)
        ws2.cell(sc_i, 3).fill             = fill(BG_DARK)
        ws2.cell(sc_i, 3).alignment        = left()
        ws2.cell(sc_i, 3).border           = thin_border()
        ws2.row_dimensions[sc_i].height    = 22

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 55

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def load_previous_report() -> pd.DataFrame:
    """Load the second-most-recent scan report for delta comparison."""
    reports = sorted(Path("reports").glob("scan_*.csv"), reverse=True)
    if len(reports) >= 2:
        return pd.read_csv(reports[1], index_col="rank")
    return pd.DataFrame()


def compute_deltas(current: pd.DataFrame, previous: pd.DataFrame) -> pd.DataFrame:
    """
    Compare current scan to previous scan and produce a changes summary.
    Returns current df with added delta columns:
      - delta_score:    Apex Score change
      - delta_rs:       RS 3m change
      - delta_3m:       3m performance change
      - delta_vwap:     VWAP position change (text)
      - delta_of:       Order Flow bias change (text)
      - delta_stage:    Stage change (text)
      - delta_pattern:  Pattern change (text)
      - is_new:         True if ticker wasn't in last scan
      - is_gone:        detected via absence (handled separately)
      - change_summary: human-readable summary of what changed
    """
    if previous.empty or current.empty:
        current["delta_score"]   = None
        current["delta_rs"]      = None
        current["delta_3m"]      = None
        current["changes"]       = "No prior scan"
        current["is_new"]        = False
        return current

    prev_idx = previous.set_index("ticker") if "ticker" in previous.columns else previous
    curr     = current.copy()

    delta_scores   = []
    delta_rs_vals  = []
    delta_3m_vals  = []
    change_summaries = []
    is_new_flags   = []

    for _, row in curr.iterrows():
        tk = row.get("ticker","")

        if tk not in prev_idx.index:
            delta_scores.append(None)
            delta_rs_vals.append(None)
            delta_3m_vals.append(None)
            change_summaries.append("🆕 New entry")
            is_new_flags.append(True)
            continue

        prev = prev_idx.loc[tk]
        changes = []
        is_new_flags.append(False)

        # Score delta
        try:
            ds = round(float(row.get("apex_score",0)) - float(prev.get("apex_score",0)), 1)
            delta_scores.append(ds)
            if ds >= 5:   changes.append(f"Score ▲{ds:+.0f}")
            elif ds <= -5: changes.append(f"Score ▼{ds:+.0f}")
        except:
            delta_scores.append(None)

        # RS delta
        try:
            dr = round(float(row.get("rs_3m",0)) - float(prev.get("rs_3m",0)), 0)
            delta_rs_vals.append(dr)
            if dr >= 20:   changes.append(f"RS ▲{dr:+.0f}")
            elif dr <= -20: changes.append(f"RS ▼{dr:+.0f}")
        except:
            delta_rs_vals.append(None)

        # 3m perf delta
        try:
            dp = round(float(row.get("perf_3m_%",0)) - float(prev.get("perf_3m_%",0)), 1)
            delta_3m_vals.append(dp)
            if abs(dp) >= 2: changes.append(f"3m {dp:+.1f}%")
        except:
            delta_3m_vals.append(None)

        # Stage change
        curr_stage = str(row.get("stage",""))
        prev_stage = str(prev.get("stage",""))
        if curr_stage != prev_stage:
            if "2 ✅" in curr_stage:
                changes.append("⬆️ Entered Stage 2")
            elif "4 🔴" in curr_stage:
                changes.append("⬇️ Dropped to Stage 4")
            elif "1 ⏳" in curr_stage and "2 ✅" in prev_stage:
                changes.append("⚠️ Lost Stage 2")
            else:
                changes.append(f"Stage: {prev_stage[:6]}→{curr_stage[:6]}")

        # Order Flow change
        curr_of = str(row.get("of_bias",""))
        prev_of = str(prev.get("of_bias",""))
        if curr_of != prev_of:
            if "Strong Bullish" in curr_of:
                changes.append("📈 Flow→Strong Bull")
            elif "Bullish" in curr_of and "Bearish" in prev_of:
                changes.append("📈 Flow flipped Bull")
            elif "Bearish" in curr_of and "Bullish" in prev_of:
                changes.append("📉 Flow flipped Bear")

        # VWAP position change
        curr_vwap = str(row.get("vwap_position",""))
        prev_vwap = str(prev.get("vwap_position",""))
        if curr_vwap != prev_vwap:
            if "Above" in curr_vwap and "Below" in prev_vwap:
                changes.append("💧 Reclaimed VWAP")
            elif "Below" in curr_vwap and "Above" in prev_vwap:
                changes.append("💧 Lost VWAP")
            elif "Extended Above" in curr_vwap and "Extended Above" not in prev_vwap:
                changes.append("⚡ Extended above VWAP")

        # Pattern change
        curr_pat = str(row.get("pattern",""))
        prev_pat = str(prev.get("pattern",""))
        if curr_pat != prev_pat:
            if "Breakout" in curr_pat:
                changes.append("🚀 Breakout!")
            elif "Handle" in curr_pat and "Handle" not in prev_pat:
                changes.append("📐 Handle forming")
            elif "Tight" in curr_pat and "Tight" not in prev_pat:
                changes.append("🔄 Tightening")

        # Breakout newly fired
        curr_brk = bool(row.get("breaking_out", False))
        prev_brk = bool(prev.get("breaking_out", False))
        if curr_brk and not prev_brk:
            changes.append("🚀 NEW Breakout!")

        # SFP appeared
        curr_sfp = str(row.get("pa_sfp",""))
        prev_sfp = str(prev.get("pa_sfp",""))
        if curr_sfp and curr_sfp != prev_sfp and curr_sfp not in ["","nan","None"]:
            changes.append(f"🎯 {curr_sfp}")

        summary = " | ".join(changes) if changes else "↔ No change"
        change_summaries.append(summary)

    curr["delta_score"]  = delta_scores
    curr["delta_rs"]     = delta_rs_vals
    curr["delta_3m"]     = delta_3m_vals
    curr["changes"]      = change_summaries
    curr["is_new"]       = is_new_flags

    # Tickers that disappeared since last scan
    if "ticker" in previous.columns:
        prev_tickers = set(previous["ticker"].tolist())
        curr_tickers = set(curr["ticker"].tolist())
        gone = prev_tickers - curr_tickers
    else:
        gone = set()

    return curr, gone

@st.cache_data(ttl=300)
def fetch_hist(ticker: str, period: str = "1y") -> pd.DataFrame:
    try:
        h = yf.Ticker(ticker).history(period=period)
        h["MA50"]  = h["Close"].rolling(50).mean()
        h["MA200"] = h["Close"].rolling(200).mean()
        return h
    except:
        return pd.DataFrame()

@st.cache_data(ttl=600)
def fetch_price(ticker: str) -> dict:
    """Return current price, 52w high, MA50, MA200 for a ticker."""
    try:
        h = yf.Ticker(ticker).history(period="1y")
        if h.empty:
            return {}
        close   = h["Close"]
        price   = close.iloc[-1]
        ma50    = close.rolling(50).mean().iloc[-1]
        ma200   = close.rolling(200).mean().iloc[-1]
        high52  = close.rolling(252).max().iloc[-1]
        prev    = close.iloc[-2] if len(close) > 1 else price
        return {
            "price":  round(price, 2),
            "prev":   round(prev, 2),
            "ma50":   round(ma50, 2),
            "ma200":  round(ma200, 2),
            "high52": round(high52, 2),
            "chg_pct": round((price / prev - 1) * 100, 2),
        }
    except:
        return {}

@st.cache_data(ttl=3600)
def fetch_earnings(ticker: str) -> dict:
    """Get next earnings date from yfinance."""
    try:
        info = yf.Ticker(ticker).info
        cal  = yf.Ticker(ticker).calendar
        date = None
        if cal is not None and not cal.empty:
            if "Earnings Date" in cal.index:
                val = cal.loc["Earnings Date"].iloc[0]
                date = pd.to_datetime(val).strftime("%Y-%m-%d") if pd.notna(val) else None
        return {
            "next_earnings": date,
            "eps_est":  info.get("forwardEps"),
            "rev_est":  info.get("revenueEstimate"),
        }
    except:
        return {"next_earnings": None}

@st.cache_data(ttl=600)
@st.cache_data(ttl=600)
def sector_performance() -> pd.DataFrame:
    """Fetch weekly/monthly performance for major US sector ETFs."""
    etfs = {
        "Technology":    "XLK",
        "Financials":    "XLF",
        "Healthcare":    "XLV",
        "Energy":        "XLE",
        "Consumer Disc": "XLY",
        "Industrials":   "XLI",
        "Materials":     "XLB",
        "Real Estate":   "XLRE",
        "Utilities":     "XLU",
        "Comm Services": "XLC",
        "Consumer Stap": "XLP",
    }
    rows = []
    for sector, sym in etfs.items():
        try:
            h = yf.Ticker(sym).history(period="3mo")["Close"].dropna()
            if len(h) < 5:
                continue
            w1 = round((h.iloc[-1] / h.iloc[-5]  - 1) * 100, 2) if len(h) >= 5  else None
            m1 = round((h.iloc[-1] / h.iloc[-21] - 1) * 100, 2) if len(h) >= 21 else None
            m3 = round((h.iloc[-1] / h.iloc[0]   - 1) * 100, 2)
            rows.append({
                "Sector": sector, "ETF": sym,
                "1W %":  w1, "1M %": m1, "3M %": m3,
                "Price": round(h.iloc[-1], 2)
            })
        except Exception as _e:
            continue

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    # Ensure numeric columns are float, not object
    for col in ["1W %", "1M %", "3M %"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# PORTFOLIO PERSISTENCE
# ══════════════════════════════════════════════════════════════════════════════

# ── Portfolio file paths (absolute, same pattern as watchlist_manager) ────────
_PORT_DIR     = Path(__file__).resolve().parent / "data"
_PORT_FILE    = _PORT_DIR / "portfolio.json"
_PORT_BACKUP  = _PORT_DIR / "portfolio_backup.json"
_PORT_TMP     = Path("/tmp/apexscan_portfolio.json")
_PORT_DIR.mkdir(parents=True, exist_ok=True)

# Trade log (closed positions history)
_TRADELOG_FILE   = _PORT_DIR / "trade_log.json"
_TRADELOG_TMP    = Path("/tmp/apexscan_tradelog.json")

def _port_write(path: Path, data) -> bool:
    """Atomic write — write to .tmp then rename."""
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp = path.with_suffix(".tmp")
        with open(tmp, "w") as f:
            json.dump(data, f, indent=2, default=str)
        import shutil; shutil.move(str(tmp), str(path))
        return True
    except Exception:
        return False

def _port_read(path: Path):
    try:
        if path.exists() and path.stat().st_size > 2:
            with open(path) as f:
                return json.load(f)
    except Exception:
        pass
    return None

def load_portfolio() -> list:
    """Load holdings from primary file, /tmp mirror, or backup — never silently lose data."""
    for src_path in (_PORT_FILE, _PORT_TMP, _PORT_BACKUP):
        data = _port_read(src_path)
        if data is not None and isinstance(data, list):
            if src_path != _PORT_FILE:
                _port_write(_PORT_FILE, data)   # restore primary
            return data
    return []

def save_portfolio(holdings: list):
    """Write to primary + /tmp mirror + backup simultaneously."""
    _port_write(_PORT_FILE,   holdings)
    _port_write(_PORT_TMP,    holdings)
    _port_write(_PORT_BACKUP, holdings)

def load_trade_log() -> list:
    """Load closed-position trade history."""
    for src_path in (_TRADELOG_FILE, _TRADELOG_TMP):
        data = _port_read(src_path)
        if data is not None and isinstance(data, list):
            return data
    return []

def save_trade_log(log: list):
    _port_write(_TRADELOG_FILE, log)
    _port_write(_TRADELOG_TMP,  log)

def close_position(holdings: list, trade_log: list, ticker: str,
                   close_price: float, close_date: str, qty_close: float = None) -> tuple:
    """
    Partially or fully close a position.
    Returns (updated_holdings, updated_trade_log).
    """
    new_holdings = []
    for h in holdings:
        if h["ticker"] != ticker:
            new_holdings.append(h)
            continue
        qty_total = h["qty"]
        qty_sold  = qty_close if qty_close and qty_close < qty_total else qty_total
        pnl_pct   = round((close_price / h["buy_price"] - 1) * 100, 2)
        pnl_dol   = round((close_price - h["buy_price"]) * qty_sold, 2)
        hold_days = (pd.to_datetime(close_date) - pd.to_datetime(h["buy_date"])).days

        trade_log.append({
            "ticker":      ticker,
            "qty":         qty_sold,
            "buy_price":   h["buy_price"],
            "buy_date":    h["buy_date"],
            "close_price": close_price,
            "close_date":  close_date,
            "pnl_dol":     pnl_dol,
            "pnl_pct":     pnl_pct,
            "hold_days":   hold_days,
            "setup":       h.get("setup", ""),
            "notes":       h.get("notes", ""),
        })
        # If partial close, keep remainder
        remaining = qty_total - qty_sold
        if remaining > 0.001:
            h_rem = dict(h)
            h_rem["qty"] = round(remaining, 4)
            new_holdings.append(h_rem)

    return new_holdings, trade_log


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## 📡 ApexScan")
    st.markdown("*momentum · Stage · Theme Rotation*")
    st.divider()
    min_score = st.slider("Min Apex Score", 0, 100, 30, 5)
    min_3m    = st.slider("Min 3m Return %", -50, 100, 0, 5)
    st.divider()
    st.markdown("**🌐 Scan Universe**")
    universe_mode = st.radio(
        "Ticker Universe",
        [
            "📋 Theme Watchlist (Config)",
            "💎 Gems Only (Small/Mid-cap)",
            "🌐 Extended Universe (NASDAQ + NYSE + NYSE American)",
        ],
        index=0, key="universe_mode",
        help=(
            "Theme Watchlist = tickers in config.yaml (~59, ~2 min). "
            "Gems Only = micro/small/mid-cap growth stocks $100M–$5B mcap (~150 tickers, early entry focus). "
            "Extended Universe = full NASDAQ + NYSE + NYSE American (~500 tickers, ~8 min)."
        )
    )
    st.divider()
    # ── MANUAL SCAN ───────────────────────────────────────────────────────────
    st.markdown("**▶️ Manual Scan**")
    run_btn = st.button("🚀 Run Live Scan Now", use_container_width=True,
                        help="Run a scan immediately using the universe selected above.")
    st.caption("Runs instantly on demand regardless of auto-scan schedule.")

    # ── AUTO-SCAN STATUS INDICATOR ────────────────────────────────────────────
    _as_state_sidebar = _autoscan_load()
    if _as_state_sidebar.get("enabled"):
        _min_o = _minutes_until(_MARKET_OPEN_UTC["hour"],  _MARKET_OPEN_UTC["minute"])
        _min_c = _minutes_until(_MARKET_CLOSE_UTC["hour"], _MARKET_CLOSE_UTC["minute"])
        st.success(
            f"🤖 Auto-scan **ON**\n\n"
            f"⏰ Next open scan: **{_min_o:.0f} min**\n\n"
            f"⏰ Next pre-close: **{_min_c:.0f} min**"
        )
        st.caption("Configure in 🤖 AI Briefing → ⏰ Auto-Scan Schedule")
    else:
        st.info("🤖 Auto-scan **OFF** — configure in 🤖 AI Briefing tab.")
    load_btn = st.button("📂 Load Last Report", use_container_width=True)
    st.divider()

    # API Key Status — load_config reads from Streamlit secrets automatically
    _cfg_check = load_config("config.yaml")

    def _key_ok(k):
        v = _cfg_check.get(k, "")
        return bool(v and not str(v).startswith("YOUR_"))

    _av_ok = _key_ok("alpha_vantage_key")
    _fh_ok = _key_ok("finnhub_key")
    _td_ok = _key_ok("twelve_data_key")
    _ms_ok = _key_ok("marketstack_key")

    st.markdown("**API Status**")
    st.markdown(f"{'🟢' if _av_ok else '🔴'} Alpha Vantage {'✓ EPS data'  if _av_ok else '✗ Not set'}")
    st.markdown(f"{'🟢' if _td_ok else '🔴'} Twelve Data {'✓ Indicators' if _td_ok else '✗ Not set'}")
    st.markdown(f"{'🟢' if _ms_ok else '🟡'} MarketStack {'✓ Backup'     if _ms_ok else 'Optional'}")
    st.markdown(f"{'🟢' if _fh_ok else '🟡'} Finnhub {'✓ News'          if _fh_ok else 'Optional'}")
    if not _av_ok:
        st.caption("Add alpha_vantage_key to Streamlit Secrets")
    if not _td_ok:
        st.caption("Add twelve_data_key to Streamlit Secrets")

    st.divider()
    st.caption("Data: yfinance · Finnhub · Alpha Vantage")
    st.caption(f"Updated: {datetime.now().strftime('%H:%M:%S')}")

    # ── 🌍 Always-visible market status badge ────────────────────────────
    # Independent confirmation that the Correction Watchlist logic is live:
    # this shows the S&P 500's stage every time, whether or not the
    # Correction Watchlist banner itself is currently showing on the
    # Leaderboard tab (that banner only appears during an actual downtrend).
    _bmc_sidebar = get_broad_market_condition()
    _bmc_color = "🟢" if _bmc_sidebar.get("uptrend") else "🟡"
    st.caption(f"{_bmc_color} S&P 500: {_bmc_sidebar.get('stage','–')}")

st.markdown("""
<h1 style="margin:0 0 16px 0;font-size:1.5rem;">
  📡 ApexScan
  <span style="font-size:1rem;color:#8b949e;font-weight:400">
    — US Market Intelligence
  </span>
</h1>
""", unsafe_allow_html=True)

tabs = st.tabs([
    "🏆 Leaderboard",
    "📈 Chart Viewer",
    "🌍 Theme Heatmap",
    "💼 Portfolio Tracker",
    "📅 Earnings Calendar",
    "🔄 Sector Rotation",
    "🔍 Stock Deep Dive",
    "🎯 Options Flow",
    "🕵️ Insider Tracker",
    "📊 Dividend Calculator",
    "⏱ Backtester",
    "⚖️ Risk Calculator",
    "🤖 AI Briefing",
    "📋 Watchlists",
    "🔔 Alert Settings",
    "🧠 Interpretation",
    "📊 Scan Delta",
    "✅ Pre-Buy Checklist",
    "📓 Trade Journal",
    "👁 Setup Monitor",
    "📖 Guide",
])


# ══════════════════════════════════════════════════════════════════════════════
# LOAD / SCAN DATA
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# LOAD / SCAN DATA
# ══════════════════════════════════════════════════════════════════════════════

df_raw       = pd.DataFrame()
prev_df      = pd.DataFrame()
gone_tickers = set()

# ══════════════════════════════════════════════════════════════════════════════
# SCAN TRIGGER LOGIC — manual and auto coexist independently
# ══════════════════════════════════════════════════════════════════════════════
_autoscan_state   = _autoscan_load()
_autoscan_trigger = check_autoscan_trigger(_autoscan_state)
_auto_fired       = False   # flag: did auto-scan fire this rerun?

# Auto-scan fires ONLY when:
#   1. Auto-scan is enabled in settings
#   2. The time trigger window is hit (±4 min of 9:30 or 15:30 EST)
#   3. The user did NOT also click Run manually (manual takes priority)
if _autoscan_trigger and not run_btn:
    _auto_fired = True
    # Use the universe saved in auto-scan settings (independent of sidebar)
    _as_saved_universe = _autoscan_state.get("universe", "theme")
    if _as_saved_universe == "extended":
        universe_mode = "🌐 Extended Universe (NASDAQ + NYSE + NYSE American)"
    else:
        universe_mode = "📋 Theme Watchlist (Config)"
    st.toast(
        f"🤖 Auto-scan firing — {_autoscan_trigger.upper()} SESSION "
        f"| Universe: {universe_mode}",
        icon="⏰"
    )

# Auto-refresh polling: checks trigger every 5 minutes while auto-scan is on.
# Only active when auto-scan is enabled — no unnecessary reruns otherwise.
if _autoscan_state.get("enabled"):
    _now_u  = datetime.now(_timezone.utc).replace(tzinfo=None)
    _min_5  = _now_u.minute % 5
    _sec    = _now_u.second
    # Rerun once at the top of each 5-minute boundary (seconds 0–8)
    if _min_5 == 0 and _sec < 9:
        _time.sleep(1)
        st.rerun()

if run_btn or _auto_fired:
    _scan_label = (
        f"🤖 Auto-scan running ({_autoscan_trigger} session)… "
        if _auto_fired else "🚀 Running live scan…"
    )
    _scan_universe_label = universe_mode

    with st.spinner(f"{_scan_label} Universe: {_scan_universe_label}"):
        cfg     = load_config("config.yaml")
        prev_df = load_latest_report()   # capture BEFORE saving new one

        # ── Universal ticker universe ─────────────────────────────────────
        _universe_override = None
        # ── 💎 GEMS ONLY universe ─────────────────────────────────────────────
        if universe_mode == "💎 Gems Only (Small/Mid-cap)":
            _GEMS_UNIVERSE = [
                # ── Fintech / Crypto Gems ─────────────────────────────────────
                "HOOD","SOFI","AFRM","UPST","DAVE","OPEN","UWMC","MSTR","COIN","DKNG",
                "PSFE","RPAY","FLYW","RELY","STEP","PRCT","LPRO","QFIN","CURO","NU",
                # ── Biotech / Health Gems ─────────────────────────────────────
                "HIMS","TMDX","RXRX","ACAD","SAGE","AUPH","AVXL","KRTX","VRNA","AKRO",
                "TARS","ARQT","GOSS","PRAX","IMVT","DNLI","SNDX","NKTR","BEAM","NTLA",
                "CRSP","EDIT","RARE","EXAS","INVA","PRTA","KYMR","ITCI","ALKS","JAZZ",
                "HALO","FOLD","VCEL","MGNX","AGIO","FATE","RCUS","ALLO","APLS","KPTI",
                "VERV","SGMO","BLUE","ONCE","QURE","IRON","APGE","BPMC","KURA","MRUS",
                # ── Space / Deep Tech Gems ────────────────────────────────────
                "RKLB","IONQ","ASTS","ACHR","LUNR","JOBY","SOUN","KTOS","DRS","SPIR",
                "BWXT","MNTS","ARQQ","QUBT","RGTI","QBTS","IQST","SATL","LOAR","FTAI",
                # ── Software / SaaS Gems ──────────────────────────────────────
                "APPN","DOCN","GTLB","PATH","BRZE","DOMO","WEAV","PCOR","ASAN","FRSH",
                "TASK","ACMR","CFLT","ESTC","JAMF","BAND","SPSC","ALTR","TDNS","TOST",
                "MAPS","RSKD","SMAR","CCSI","ALNT","PGNY","VNET","WIX","GETY","SHLX",
                # ── Consumer / Lifestyle Gems ─────────────────────────────────
                "CAVA","RDDT","BIRK","BROS","WING","SHAK","PLNT","DNUT","FAT","TXRH",
                "EWCZ","ELF","SKIN","GOLI","PLTK","LOVE","CURV","FTCH","RENT","REAL",
                "RVLV","SOSF","PTON","LAZY","XPOF","FORM","GIII","POWL","TILE","BOOT",
                # ── Energy Transition Gems ────────────────────────────────────
                "ENPH","FSLR","RUN","PLUG","NOVA","ARRY","SHLS","SPWR","CWEN","CLNE",
                "EVGO","CHPT","BLNK","VLTA","AMPX","STEM","BSFC","FLUX","DCFC","AMRC",
                "MP","LAC","LTHM","SQM","ALB","ALTM","OROCF","PMET","LIT","SGML",
                # ── Industrials / Defence Micro-caps ─────────────────────────
                "KTOS","RCAT","AVAV","HXSW","ATRO","AIR","CDRE","VSEC","HAYW","TDW",
                "GNSS","MOOG","ARCB","ODFL","SAIA","HUBG","HTLD","LSTR","MRTN","PTSI",
                # ── Mining / Commodities Gems ─────────────────────────────────
                "AG","EXK","SILV","CDE","HL","MUX","GATO","MAG","FSM","ERO",
                "VZLA","SAND","AUMN","USAS","LGDTF","SSRM","MRDDF","AUMN","GPL","SVM",
                # ── Emerging Market / International Growth ────────────────────
                "NU","MELI","SE","GRAB","BEKE","LU","BTBT","MARA","RIOT","HUT",
                "BITF","CIFR","CLSK","IREN","WULF","ARBK","SDIG","BTDR","CORZ","AULT",
            ]
            # Deduplicate
            _seen_g = set()
            _GEMS_UNIVERSE = [t for t in _GEMS_UNIVERSE if not (_seen_g.add(t) or t in _seen_g - {t})]
            _universe_override = _GEMS_UNIVERSE
            st.info(
                f"💎 Gems Only: scanning {len(_GEMS_UNIVERSE)} small/mid-cap growth stocks "
                f"with relaxed thresholds to surface early-stage setups before they move."
            )

        elif universe_mode == "🌐 Extended Universe (NASDAQ + NYSE + NYSE American)":
            # ── Already in list (NASDAQ + mixed) ──────────────────────────
            _NASDAQ_NAMES = [
                # Mega-cap tech / NASDAQ 100 core
                "AAPL","MSFT","AMZN","NVDA","GOOGL","GOOG","META","TSLA","AVGO","COST",
                "NFLX","ORCL","ADBE","QCOM","TXN","INTU","AMD","ARM","ASML","TSM",
                # Semiconductors
                "AMAT","LRCX","KLAC","MU","MRVL","SMCI","CDNS","SNPS","ON","MPWR",
                # Software / Cloud
                "CRM","DDOG","SNOW","NET","ZS","PANW","FTNT","CRWD","PLTR","VEEV",
                "WDAY","TEAM","HUBS","NOW","MDB","GTLB","BILL","PATH","AI","APPN",
                # Fintech / Crypto
                "PYPL","COIN","HOOD","SOFI","AFRM","MSTR","SQ","UPST","DAVE","SMAR",
                # Biotech / Health (NASDAQ-listed)
                "MRNA","BNTX","REGN","BIIB","GILD","IDXX","DXCM","ISRG","ILMN","VRTX",
                "ALNY","SGEN","BMRN","INCY","EXAS","RARE","NTLA","BEAM","CRSP","EDIT",
                # Consumer / Retail (NASDAQ)
                "MNST","CELH","LULU","ONON","DUOL","ROST","DLTR","FAST","ODFL","CTAS",
                # Growth / Emerging (NASDAQ)
                "RKLB","IONQ","ASTS","ACHR","SOUN","RXRX","HIMS","RDDT","CAVA","TMDX",
                "LUNR","BTDR","DOCN","OPEN","UWMC","JOBY","ABNB","DASH","LYFT","UBER",
                "SHOP","SPOT","ROKU","TTD","MTCH","MELI","SE","GRAB","DKNG","RBLX",
            ]

            # ── NYSE — Blue-chip, Industrials, Financials, Energy, Healthcare ─
            _NYSE_NAMES = [
                # Financials (NYSE)
                "JPM","GS","MS","BAC","WFC","C","AXP","BLK","SCHW","ICE","CME",
                "SPGI","MCO","AMP","PGR","MET","TRV","AFL","ALL","CB","HIG","L",
                "BX","KKR","APO","CG","ARES","TPG","BN","BAM","TROW","IVZ","BEN",
                "WTW","AON","MMC","USB","PNC","TFC","FITB","KEY","CFG","RF","HBAN",
                # Healthcare (NYSE)
                "UNH","CI","CVS","HCA","MCK","CAH","DHR","TMO","ABT","MDT","SYK",
                "BSX","EW","ZBH","BDX","BAX","STE","HOLX","IQV","CRL","MTD","WAT",
                "LH","DGX","CTLT","PKI","VTRS","RPRX","JAZZ","ALKS","ITCI","ACAD",
                "LLY","ABBV","BMY","PFE","JNJ","MRK","AZN","NVO","GSK","SNY",
                # Energy (NYSE)
                "XOM","CVX","COP","SLB","BKR","HAL","PSX","VLO","MPC","EOG",
                "PXD","DVN","OXY","FANG","HES","APA","NOV","WHD","TRGP","KMI",
                "WMB","OKE","EPD","ET","PAA","MMP","LNG","AR","EQT","RRC",
                # Industrials / Defence (NYSE)
                "BA","RTX","LMT","NOC","GD","HII","TDG","HWM","GE","HON","MMM",
                "CAT","DE","EMR","ETN","PH","ITW","ROK","AME","ROP","CPRT","EXPD",
                "UPS","FDX","GXO","XPO","CHRW","JBHT","SAIA","TFII","ZTO","DAL",
                "UAL","AAL","LUV","ALK","SAVE","H","MAR","HLT","WH","CHH","NCLH",
                # Materials / Metals (NYSE)
                "LIN","APD","SHW","ECL","IFF","PPG","RPM","FMC","CF","MOS","NTR",
                "NUE","STLD","CMC","RS","ATI","FCX","SCCO","AA","CLF","MP","ALB",
                "LAC","LTHM","SQM","VALE","RIO","BHP","GOLD","NEM","AEM","PAAS",
                # Chemicals / Specialty Materials (NYSE)
                "DOW","DD","LYB","HUN","CE","EMN","OLN","ASH","TROX","IOSP",
                # Utilities (NYSE)
                "NEE","D","SO","DUK","AEP","SRE","PCG","XEL","AWK","ES","EXC",
                "ED","PPL","ETR","FE","AEE","CMS","DTE","LNT","PNW","WEC","NI",
                # Real Estate / REITs (NYSE)
                "PLD","AMT","CCI","SBAC","EQIX","DLR","O","SPG","PSA","EXR",
                "AVB","EQR","UDR","ESS","MAA","CPT","NNN","VICI","MGM","WYNN","LVS",
                "HST","RHP","PK","SHO","PLYA","APLE","CLDT","CPLG","RLJ","XHR",
                # Consumer Staples (NYSE)
                "WMT","PG","KO","PEP","PM","MO","CL","KMB","CHD","CLX","HRL",
                "SJM","CAG","CPB","GIS","K","MKC","HSY","TR","MDLZ","KHC","STZ",
                "BF-B","TAP","SAM","BUD","DEO","BTI","BURL","TJX","COST","DG","DLTR",
                # Consumer Discretionary (NYSE)
                "HD","TGT","LOW","MCD","SBUX","YUM","CMG","DPZ","QSR","EAT","DRI",
                "TXRH","BLMN","BJRI","CAKE","SHAK","WING","PLNT","BJ","FIVE","OLLI",
                "F","GM","STLA","HOG","RACE","TM","HMC","MGA","LEA","BWA","ALV",
                "NKE","DECK","SKX","CROX","PVH","RL","TPR","TIF","VFC","HBI","UA",
                # Technology (NYSE-listed)
                "IBM","ORCL","HPQ","HPE","DELL","NCR","CDW","LDOS","SAIC","BAH",
                "ACN","WIT","INFY","CTSH","EPAM","GLOB","MFAC","DXC","CSC","CACI",
            ]

            # ── NYSE American (AMEX) — small/mid growth and mining ────────
            _NYSE_AMERICAN_NAMES = [
                # Growth / Emerging (NYSE American)
                "LUNR","ACHR","JOBY","ASTS","SOUN","RXRX","IONQ","BTBT","MARA",
                "RIOT","HUT","BITF","CIFR","CLSK","IREN","WULF","BTDR",
                # Mining / Resources (NYSE American)
                "AG","EXK","PAAS","SILV","CDE","HL","GPL","MUX","AUY","KGC",
                "GATO","MAG","SVM","FSM","ERO","ATX","VZLA","SAND","WPM","OR",
                # Biotech / Pharma (NYSE American)
                "ACAD","SAGE","INVA","PRTA","KYMR","ARQT","GOSS","AUPH","NKTR",
                "AVXL","SNDX","PRAX","IMVT","DNLI","KRTX","VRNA","AKRO","TARS",
                # Energy (NYSE American)
                "CRC","SM","CIVI","MGY","ESTE","REX","FLNG","GMLP","SLNG",
                # Special Situations / Growth (NYSE American)
                "OPEN","UWMC","DAVE","HIMS","RDDT","CAVA","TMDX","MSTR",
            ]

            # Merge all, deduplicate, preserve order
            _seen = set()
            _EXTENDED_UNIVERSE = []
            for _tk in (_NASDAQ_NAMES + _NYSE_NAMES + _NYSE_AMERICAN_NAMES):
                if _tk not in _seen:
                    _seen.add(_tk)
                    _EXTENDED_UNIVERSE.append(_tk)

            _universe_override = _EXTENDED_UNIVERSE
            st.info(
                f"🌐 Extended Universe: scanning {len(_EXTENDED_UNIVERSE)} tickers "
                f"(NASDAQ + NYSE + NYSE American)…"
            )

        df_raw  = run_scan(cfg, universe_override=_universe_override)
        if not df_raw.empty:
            save_report(df_raw)
            try:
                alert_settings = load_alert_settings()
                if alert_settings.get("alerts_enabled"):
                    portfolio_data = load_portfolio()
                    fired = check_and_fire_alerts(df_raw, portfolio_data, alert_settings, fetch_price)
                    if fired:
                        st.info(f"🔔 {len(fired)} alert(s) sent to your configured channels.")
            except Exception as ae:
                pass
            if _auto_fired:
                st.success(
                    f"🤖 Auto-scan complete ({_autoscan_trigger} session) — "
                    f"{len(df_raw)} setups found. Results auto-saved."
                )
            else:
                st.success(f"✅ Manual scan complete — {len(df_raw)} setups found!")
            st.rerun()  # reload so Leaderboard tab populates from saved CSV
        else:
            st.warning("No setups found. Try lowering the Score or Volume thresholds in config.yaml.")
else:
    prev_df = load_previous_report()
    df_raw  = load_latest_report()

# Compute deltas (changes vs previous scan)
if not df_raw.empty and not prev_df.empty:
    try:
        df_raw, gone_tickers = compute_deltas(df_raw, prev_df)
    except Exception as _de:
        df_raw["changes"]     = "–"
        df_raw["is_new"]      = False
        df_raw["delta_score"] = None
        gone_tickers          = set()
elif not df_raw.empty:
    df_raw["changes"]     = "First scan"
    df_raw["is_new"]      = False
    df_raw["delta_score"] = None

df = df_raw.copy()
if not df.empty:
    if "apex_score" in df.columns:
        df = df[pd.to_numeric(df["apex_score"], errors="coerce") >= min_score]
    if "perf_3m_%" in df.columns:
        df = df[pd.to_numeric(df["perf_3m_%"], errors="coerce") >= min_3m]



# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — LEADERBOARD
# ══════════════════════════════════════════════════════════════════════════════

with tabs[0]:
    if df.empty:
        st.info("Click **🚀 Run Live Scan** or **📂 Load Last Report** in the sidebar.")
    else:
        c1,c2,c3,c4,c5 = st.columns(5)
        with c1:
            st.markdown(f'<div class="metric-card"><h3>Tickers Passing</h3><div class="value white">{len(df)}</div></div>', unsafe_allow_html=True)
        with c2:
            top = pd.to_numeric(df["apex_score"], errors="coerce").max()
            st.markdown(f'<div class="metric-card"><h3>Top Score</h3><div class="value green">{top:.0f}</div></div>', unsafe_allow_html=True)
        with c3:
            bo = int(df.get("breaking_out", pd.Series([False]*len(df))).sum()) if "breaking_out" in df.columns else 0
            st.markdown(f'<div class="metric-card"><h3>Breakouts</h3><div class="value amber">{bo}</div></div>', unsafe_allow_html=True)
        with c4:
            stage2 = int((df.get("stage", pd.Series([""] * len(df))).str.contains("2 ✅", na=False)).sum()) if "stage" in df.columns else 0
            st.markdown(f'<div class="metric-card"><h3>Stage 2 Stocks</h3><div class="value blue">{stage2}</div></div>', unsafe_allow_html=True)
        with c5:
            themes_n = df["theme"].nunique() if "theme" in df.columns else 0
            st.markdown(f'<div class="metric-card"><h3>Themes Active</h3><div class="value green">{themes_n}</div></div>', unsafe_allow_html=True)

        st.markdown("---")

        # ── 🌱 Correction Watchlist (William O'Neil basing lesson) ─────────
        # "New bases form during market corrections. Don't spend corrections
        #  trying to predict the bottom. Spend them building your watchlist."
        _bmc = get_broad_market_condition()
        _corr_candidates = find_correction_watchlist_candidates(df)

        if not _bmc.get("uptrend", True):
            st.markdown(
                f'<div style="background:#1a1500;border:1px solid #d29922;border-radius:10px;'
                f'padding:14px 18px;margin-bottom:12px;">'
                f'<b style="color:#d29922;">🌱 Market Correction Detected</b> '
                f'<span style="color:#8b949e;font-size:0.85rem;">— S&P 500: {_bmc.get("stage","–")}</span><br>'
                f'<span style="color:#c9d1d9;font-size:0.88rem;">'
                f'William O\'Neil\'s lesson: new bases form during corrections — institutions '
                f'quietly accumulate leading stocks while weak holders exit. These are often the '
                f'first to break out once the market confirms a new uptrend. '
                f'Don\'t try to predict the bottom — build your watchlist instead.'
                f'</span></div>',
                unsafe_allow_html=True
            )
            if _corr_candidates.empty:
                st.caption("No Stage 1 tight-base candidates found in the current scan.")
            else:
                st.markdown(f"**{len(_corr_candidates)} Stage 1 base-builder(s) found** — tight bases forming while the market corrects:")
                _corr_show = [c for c in ["ticker","theme","price","perf_3m_%","adr_%","pattern","apex_score"]
                              if c in _corr_candidates.columns]
                _corr_disp = _corr_candidates[_corr_show].head(20).copy()
                for _cc in ["apex_score","perf_3m_%","adr_%"]:
                    if _cc in _corr_disp.columns:
                        _corr_disp[_cc] = pd.to_numeric(_corr_disp[_cc], errors="coerce")
                st.dataframe(
                    _corr_disp.style.format({
                        "price": "${:.2f}", "perf_3m_%": pct_fmt,
                        "adr_%": lambda v: f"{v:.1f}%" if pd.notna(v) else "–",
                        "apex_score": "{:.0f}",
                    }, na_rep="–"),
                    use_container_width=True, hide_index=True, height=min(300, 45 + 35*len(_corr_disp))
                )
                if st.button("🌱 Add All to 'Correction Watchlist'", key="add_correction_wl"):
                    _wls_corr = load_watchlists()
                    _wls_corr = create_list(_wls_corr, "Correction Watchlist")
                    for _tk in _corr_candidates["ticker"].head(20).tolist():
                        _wls_corr = add_ticker(_wls_corr, "Correction Watchlist", _tk)
                    save_watchlists(_wls_corr)
                    st.success(
                        f"✅ Added {min(20, len(_corr_candidates))} ticker(s) to 'Correction Watchlist'. "
                        f"Find it in the 📋 Watchlists tab."
                    )
            st.markdown("---")

        # ── New signal filters ─────────────────────────────────────────────
        with st.expander("🔬 Advanced Signal Filters", expanded=False):
            af1, af2, af3, af4 = st.columns(4)
            with af1:
                filter_of = st.selectbox("Order Flow Bias",
                    ["All", "Strong Bullish", "Bullish", "Neutral", "Bearish"], index=0)
            with af2:
                filter_vwap = st.selectbox("VWAP Position",
                    ["All", "Above VWAP", "Extended Above VWAP",
                     "Below VWAP", "Extended Below VWAP"], index=0)
            with af3:
                filter_ms = st.selectbox("Market Structure",
                    ["All", "Bullish (HH/HL)", "Bearish (LH/LL)", "Transitioning"], index=0)
            with af4:
                filter_pa = st.selectbox("PA Pattern",
                    ["All", "Bullish SFP", "Bullish Engulfing",
                     "Inside Day", "Bullish Context Candle", "PA Confluence"], index=0)

        # Apply advanced filters
        df_filtered = df.copy()
        if filter_of != "All" and "of_bias" in df_filtered.columns:
            df_filtered = df_filtered[df_filtered["of_bias"] == filter_of]
        if filter_vwap != "All" and "vwap_position" in df_filtered.columns:
            df_filtered = df_filtered[df_filtered["vwap_position"] == filter_vwap]
        if filter_ms != "All" and "ms_structure" in df_filtered.columns:
            df_filtered = df_filtered[df_filtered["ms_structure"] == filter_ms]
        if filter_pa != "All" and "pa_patterns" in df_filtered.columns:
            df_filtered = df_filtered[df_filtered["pa_patterns"].str.contains(filter_pa, na=False)]

        # ── Sector / Theme Filter — GICS 11 sectors + config themes ─────
        _GICS_SECTORS_11 = [
            "Energy","Materials","Industrials","Utilities","Healthcare",
            "Financials","Consumer Discretionary","Consumer Staples",
            "Information Technology","Communication Services","Real Estate",
        ]
        _all_raw_themes   = sorted(df_filtered["theme"].dropna().unique().tolist()) if "theme" in df_filtered.columns else []
        _sectors_present  = [s for s in _GICS_SECTORS_11 if s in _all_raw_themes]
        _cfg_themes_pres  = [t for t in _all_raw_themes if t not in _GICS_SECTORS_11]

        # Build a clean selectbox: All → GICS sectors → Config themes
        _sector_options   = (
            ["🌐 All Sectors"]
            + [f"📊 {s}" for s in _sectors_present]
            + ([f"── {t}" for t in _cfg_themes_pres] if _cfg_themes_pres else [])
        )
        sel_sector = st.selectbox(
            "🏭 Sector / Theme",
            _sector_options,
            key="sector_theme_filter",
            help="Filter by GICS sector (the 11 official market sectors) or by your config themes (ai_semis, cybersecurity etc.)"
        )
        # Apply sector filter (strip emoji prefix before matching)
        if sel_sector != "🌐 All Sectors":
            _sel_clean = sel_sector.lstrip("📊 ").lstrip("── ").strip()
            if "theme" in df_filtered.columns:
                df_filtered = df_filtered[df_filtered["theme"] == _sel_clean]

        # Quick-view buttons preserved
        theme_filter = st.radio(
            "📊 Quick Filter",
            ["🌐 All", "🚀 Growth Leaders", "💎 Emerging Gems", "🎯 Early Entry"],
            horizontal=True, key="lb_theme_filter"
        )
        if theme_filter == "💎 Emerging Gems":
            st.markdown(GEM_DISCLAIMER, unsafe_allow_html=True)
            if "is_gem" in df_filtered.columns:
                df_filtered = df_filtered[df_filtered["is_gem"] == True]
            elif "theme" in df_filtered.columns:
                df_filtered = df_filtered[df_filtered["theme"] == "emerging_gems"]
        elif theme_filter == "🚀 Growth Leaders":
            if "market_cap" in df_filtered.columns:
                mcap_num = pd.to_numeric(df_filtered["market_cap"], errors="coerce")
                df_filtered = df_filtered[mcap_num >= 10_000_000_000]
        elif theme_filter == "🎯 Early Entry":
            # Show stocks at the START of a move — cheap entry zone
            st.markdown(
                '<div style="background:#1a2a1a;border:1px solid #3fb950;border-radius:8px;'
                'padding:10px 16px;margin-bottom:10px;font-size:0.85rem;color:#3fb950;">'
                '🎯 <b>Early Entry Filter</b> — stocks just crossing above MAs, '
                'pulling back to 50MA, or forming tight low-ADR bases. '
                'These are cheap entries BEFORE the move, not after.'
                '</div>',
                unsafe_allow_html=True
            )
            if "early_entry" in df_filtered.columns:
                df_filtered = df_filtered[
                    df_filtered["early_entry"].astype(str).str.lower().isin(["true","1"])
                ]
            else:
                # Fallback: show stocks within 5% of 50MA or with fresh MA cross
                if "vs_50ma_%" in df_filtered.columns:
                    _vs50 = pd.to_numeric(df_filtered["vs_50ma_%"], errors="coerce")
                    df_filtered = df_filtered[(_vs50.abs() <= 5) | (_vs50 >= 0)]

        # Column view toggle
        col_view = st.radio("Column View", ["Standard", "Order Flow", "VWAP & Structure", "Price Action", "Fundamentals"], horizontal=True)

        if col_view == "Standard":
            want = ["ticker","theme","price","mcap_category","stage",
                    "perf_1m_%","perf_3m_%","perf_6m_%",
                    "rs_3m","rs_r2500_3m","rs_r3000g_3m","rs_multi_leader",
                    "vol_surge_x","near_52wh","pattern",
                    "earn_momentum","eps_growth_%","eps_surprise_%","consec_beats","apex_score"]
        elif col_view == "Order Flow":
            want = ["ticker","price","of_bias","of_up_vol_ratio",
                    "of_bullish_days","of_consec_up","of_score",
                    "vol_surge_x","ms_structure","apex_score"]
        elif col_view == "VWAP & Structure":
            want = ["ticker","price","vwap","vs_vwap_%","vwap_position",
                    "vwap_slope","vwap_score","ms_structure","ms_hh_hl","ms_bos",
                    "ms_swing_high","ms_swing_low","apex_score"]
        elif col_view == "Price Action":
            want = ["ticker","price","pa_patterns","pa_engulfing","pa_sfp",
                    "pa_inside_day","pa_context","pa_score",
                    "of_bias","vwap_position","apex_score"]
        else:  # Fundamentals
            want = ["ticker","price","earn_momentum",
                    "eps_growth_%","eps_surprise_%","eps_accel",
                    "consec_beats","rev_growth_%","eps_score",
                    "analyst_target","pe_ratio","peg_ratio","apex_score"]

        show_cols = [c for c in want if c in df_filtered.columns]
        disp = df_filtered[show_cols].head(30).copy()
        for col in ["apex_score","perf_1m_%","perf_3m_%","perf_6m_%","rs_3m","rs_6m"]:
            if col in disp.columns:
                disp[col] = pd.to_numeric(disp[col], errors="coerce")

        # Color helpers for new columns
        def color_of_bias(v):
            if "Strong Bullish" in str(v): return "color:#3fb950;font-weight:700"
            if "Bullish" in str(v):        return "color:#3fb950"
            if "Strong Bearish" in str(v): return "color:#f85149;font-weight:700"
            if "Bearish" in str(v):        return "color:#f85149"
            return "color:#8b949e"

        def color_vwap_pos(v):
            if "Extended Above" in str(v): return "color:#d29922"
            if "Above" in str(v):          return "color:#3fb950"
            if "Extended Below" in str(v): return "color:#f85149;font-weight:700"
            if "Below" in str(v):          return "color:#f85149"
            return ""

        def color_ms(v):
            if "Bullish" in str(v): return "color:#3fb950"
            if "Bearish" in str(v): return "color:#f85149"
            return "color:#d29922"

        def color_pa(v):
            if "SFP" in str(v) or "Confluence" in str(v): return "color:#d29922;font-weight:700"
            if "Bullish" in str(v): return "color:#3fb950"
            if "Bearish" in str(v): return "color:#f85149"
            return ""

        fmt_dict = {
            "price": "{:.2f}", "apex_score": "{:.0f}",
            "perf_1m_%": pct_fmt, "perf_3m_%": pct_fmt, "perf_6m_%": pct_fmt,
            "rs_3m":        lambda v: f"{v:.0f}" if pd.notna(v) and v != 0 else "–",
            "rs_r2500_3m":  lambda v: f"{v:.0f}" if pd.notna(v) and v not in (0, None) else "–",
            "rs_r3000g_3m": lambda v: f"{v:.0f}" if pd.notna(v) and v not in (0, None) else "–",
            "rs_multi_leader": lambda v: "✅" if v is True or str(v).lower()=="true" else "–",
            "adr_%": lambda v: f"{v:.1f}%" if pd.notna(v) else "–",
            "vs_50ma_%": pct_fmt, "vs_200ma_%": pct_fmt,
            "vol_surge_x": "{:.1f}x",
            "of_up_vol_ratio": "{:.2f}x",
            "of_bullish_days": "{:.0f}%",
            "vs_vwap_%": pct_fmt,
            "vwap": "${:.2f}",
            "ms_swing_high": lambda v: f"${v:.2f}" if pd.notna(v) else "–",
            "ms_swing_low":  lambda v: f"${v:.2f}" if pd.notna(v) else "–",
            # AV fundamentals
            "eps_growth_%":   lambda v: f"{v:+.1f}%" if pd.notna(v) else "–",
            "eps_surprise_%": lambda v: f"{v:+.1f}%" if pd.notna(v) else "–",
            "rev_growth_%":   lambda v: f"{v:+.1f}%" if pd.notna(v) else "–",
            "eps_score":      lambda v: f"{v}/15"    if pd.notna(v) else "–",
            "analyst_target": lambda v: f"${v:.2f}"  if pd.notna(v) and v else "–",
            "pe_ratio":       lambda v: f"{v:.1f}x"  if pd.notna(v) and v else "–",
            "peg_ratio":      lambda v: f"{v:.2f}"   if pd.notna(v) and v else "–",
            "consec_beats":   lambda v: f"{int(v)}Q" if pd.notna(v) else "–",
        }
        active_fmt = {k: v for k, v in fmt_dict.items() if k in disp.columns}

        map_cols_of   = [c for c in ["of_bias"] if c in disp.columns]
        map_cols_vwap = [c for c in ["vwap_position"] if c in disp.columns]
        map_cols_ms   = [c for c in ["ms_structure"] if c in disp.columns]
        map_cols_pa   = [c for c in ["pa_patterns","pa_sfp","pa_engulfing","pa_context"] if c in disp.columns]

        styled = disp.style.map(color_score, subset=["apex_score"])
        if [c for c in ["perf_1m_%","perf_3m_%","perf_6m_%","vs_50ma_%","vs_200ma_%","vs_vwap_%"] if c in disp.columns]:
            styled = styled.map(color_perf, subset=[c for c in ["perf_1m_%","perf_3m_%","perf_6m_%","vs_50ma_%","vs_200ma_%","vs_vwap_%"] if c in disp.columns])
        if "rs_3m" in disp.columns:
            styled = styled.map(color_rs, subset=["rs_3m"])
        for _rs_col in ["rs_r2500_3m","rs_r3000g_3m"]:
            if _rs_col in disp.columns:
                styled = styled.map(color_rs, subset=[_rs_col])
        if map_cols_of:   styled = styled.map(color_of_bias,  subset=map_cols_of)
        if map_cols_vwap: styled = styled.map(color_vwap_pos, subset=map_cols_vwap)
        if map_cols_ms:   styled = styled.map(color_ms,       subset=map_cols_ms)
        if map_cols_pa:   styled = styled.map(color_pa,       subset=map_cols_pa)
        styled = styled.format(active_fmt, na_rep="–")

        st.dataframe(styled, use_container_width=True, height=520)

        # New signal summary badges
        if not df_filtered.empty:
            sfp_count  = df_filtered["pa_sfp"].notna().sum() if "pa_sfp" in df_filtered.columns else 0
            of_bull    = (df_filtered.get("of_bias","").str.contains("Bullish", na=False)).sum() if "of_bias" in df_filtered.columns else 0
            vwap_above = (df_filtered.get("vwap_position","").str.contains("Above", na=False)).sum() if "vwap_position" in df_filtered.columns else 0
            hh_hl      = df_filtered["ms_hh_hl"].sum() if "ms_hh_hl" in df_filtered.columns else 0
            st.markdown(
                f"**Signal Summary:** "
                f"🎯 {sfp_count} SFP setups &nbsp;|&nbsp; "
                f"📈 {of_bull} persistent bull flow &nbsp;|&nbsp; "
                f"💧 {vwap_above} above VWAP &nbsp;|&nbsp; "
                f"🏗 {hh_hl} HH/HL structure",
                unsafe_allow_html=True
            )

        top15 = df_filtered.head(15).copy()
        top15["apex_score"] = pd.to_numeric(top15["apex_score"], errors="coerce")
        colors = ["#3fb950" for _ in top15["ticker"]]
        fig = go.Figure(go.Bar(
            x=top15["apex_score"], y=top15["ticker"], orientation="h",
            marker_color=colors, text=top15["apex_score"].round(0).astype("Int64"),
            textposition="outside",
        ))
        fig.update_layout(
            title="Top 15 — Apex Score",
            paper_bgcolor="#0d1117", plot_bgcolor="#0d1117", font_color="#e6edf3",
            yaxis=dict(autorange="reversed", gridcolor="#21262d"),
            xaxis=dict(range=[0,115], gridcolor="#21262d"),
            height=400, margin=dict(l=10,r=60,t=40,b=20),
        )
        st.plotly_chart(fig, use_container_width=True)

        # ── Changes Since Last Scan ───────────────────────────────────────
        st.markdown("---")
        st.markdown("### 🔄 Changes Since Last Scan")

        if "changes" not in df_filtered.columns or df_filtered["changes"].eq("No prior scan").all():
            st.info("Run a second scan to start seeing changes between sessions.")
        else:
            # Newly appeared tickers
            new_entries = df_filtered[df_filtered.get("is_new", pd.Series([False]*len(df_filtered))) == True] if "is_new" in df_filtered.columns else pd.DataFrame()
            if not new_entries.empty:
                st.markdown(f"**🆕 New entries this scan:** " +
                    " &nbsp; ".join([f'<span style="background:#1a3a2a;color:#3fb950;padding:2px 8px;border-radius:4px;font-weight:700;">{t}</span>'
                    for t in new_entries["ticker"].tolist()]),
                    unsafe_allow_html=True)

            # Tickers that dropped out
            if gone_tickers:
                st.markdown(f"**❌ Dropped from scan:** " +
                    " &nbsp; ".join([f'<span style="background:#2a1010;color:#f85149;padding:2px 8px;border-radius:4px;">{t}</span>'
                    for t in sorted(gone_tickers)]),
                    unsafe_allow_html=True)

            # Changes table
            change_cols = ["ticker", "price", "apex_score", "delta_score",
                           "stage", "of_bias", "vwap_position", "pa_patterns", "changes"]
            chg_show = [c for c in change_cols if c in df_filtered.columns]
            chg_df   = df_filtered[chg_show].copy()

            # Only show rows that actually changed or are new
            has_change = chg_df["changes"].apply(
                lambda x: x not in ["↔ No change", "No prior scan", "First scan", "–", None, ""]
            ) if "changes" in chg_df.columns else pd.Series([True]*len(chg_df))

            changed_df  = chg_df[has_change]
            unchanged_df = chg_df[~has_change]

            def color_delta(v):
                try:
                    v = float(v)
                    return "color:#3fb950;font-weight:700" if v > 0 else ("color:#f85149;font-weight:700" if v < 0 else "")
                except: return ""

            def color_changes(v):
                v = str(v)
                if "🆕" in v or "▲" in v or "Stage 2" in v or "Reclaimed" in v or "Breakout" in v:
                    return "color:#3fb950"
                if "▼" in v or "Lost" in v or "Stage 4" in v or "Bear" in v:
                    return "color:#f85149"
                if "SFP" in v or "Tighten" in v or "Handle" in v:
                    return "color:#d29922"
                return "color:#8b949e"

            if not changed_df.empty:
                st.markdown(f"**{len(changed_df)} tickers with notable changes:**")

                # Display as styled cards for easier reading
                for _, row in changed_df.head(20).iterrows():
                    tk       = row.get("ticker","")
                    chg_txt  = str(row.get("changes","–"))
                    ds       = row.get("delta_score")
                    score    = row.get("apex_score","–")
                    stage    = str(row.get("stage","–"))
                    of_bias  = str(row.get("of_bias","–"))
                    vwap_p   = str(row.get("vwap_position","–"))

                    # Card border colour based on change direction
                    if any(x in chg_txt for x in ["▲","🆕","Stage 2","Reclaimed","Breakout","🚀","📈"]):
                        border = "#3fb950"
                    elif any(x in chg_txt for x in ["▼","Lost","Stage 4","📉"]):
                        border = "#f85149"
                    elif any(x in chg_txt for x in ["🎯","📐","🔄"]):
                        border = "#d29922"
                    else:
                        border = "#30363d"

                    ds_str = f"({'+' if (ds or 0)>0 else ''}{ds:.0f} pts)" if pd.notna(ds) else ""
                    ds_color = "#3fb950" if (ds or 0) > 0 else ("#f85149" if (ds or 0) < 0 else "#8b949e")

                    st.markdown(
                        f'<div style="background:#0d1117;border-left:3px solid {border};'
                        f'border-radius:0 8px 8px 0;padding:12px 16px;margin:5px 0;'
                        f'display:flex;align-items:center;gap:16px;">'
                        f'<div style="min-width:60px;">'
                        f'<span style="font-weight:800;font-size:1rem;color:#e6edf3;">{tk}</span></div>'
                        f'<div style="min-width:80px;">'
                        f'<span style="color:#8b949e;font-size:0.75rem;">SCORE</span><br>'
                        f'<span style="font-weight:700;">{score}</span> '
                        f'<span style="font-size:0.8rem;color:{ds_color};">{ds_str}</span></div>'
                        f'<div style="min-width:130px;">'
                        f'<span style="color:#8b949e;font-size:0.75rem;">STAGE / OF / VWAP</span><br>'
                        f'<span style="font-size:0.8rem;color:#c9d1d9;">{stage[:10]} · {of_bias[:8]} · {vwap_p[:12]}</span></div>'
                        f'<div style="flex:1;">'
                        f'<span style="color:#8b949e;font-size:0.75rem;">WHAT CHANGED</span><br>'
                        f'<span style="color:{border};font-size:0.88rem;font-weight:600;">{chg_txt}</span></div>'
                        f'</div>',
                        unsafe_allow_html=True
                    )
            else:
                st.markdown('<div class="metric-card" style="text-align:center;color:#8b949e;">↔ No significant changes since last scan — market conditions are stable.</div>', unsafe_allow_html=True)

            if not unchanged_df.empty:
                with st.expander(f"↔ {len(unchanged_df)} tickers with no significant change"):
                    st.dataframe(
                        unchanged_df[["ticker","apex_score","stage","of_bias","vwap_position"]].style
                        .format({"apex_score": "{:.0f}"}, na_rep="–"),
                        use_container_width=True, hide_index=True
                    )

        st.download_button("⬇ Download CSV", df_filtered.to_csv().encode("utf-8"),
            file_name=f"apexscan_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — CHART VIEWER
# ══════════════════════════════════════════════════════════════════════════════

with tabs[1]:
    st.markdown("### 📈 Price Chart + Moving Averages + VWAP")
    ticker_opts = df["ticker"].tolist() if not df.empty else ["NVDA","AAPL","TSM","ASML"]
    ca, cb, cc = st.columns([2, 1, 1])
    with ca: sel = st.selectbox("Ticker", ticker_opts)
    with cb: period = st.selectbox("Period", ["3mo","6mo","1y","2y"], index=1)
    with cc:
        show_vwap   = st.checkbox("VWAP", value=True)
        show_swings = st.checkbox("Swing Levels", value=True)

    if sel:
        hist = fetch_hist(sel, period)
        if not hist.empty:
            # Compute VWAP for chart overlay
            hist_vwap = hist.copy()
            hist_vwap["typical"] = (hist_vwap["High"] + hist_vwap["Low"] + hist_vwap["Close"]) / 3
            hist_vwap["tp_vol"]  = hist_vwap["typical"] * hist_vwap["Volume"]
            # Rolling 20-day VWAP
            hist_vwap["VWAP"]   = (hist_vwap["tp_vol"].rolling(20).sum() /
                                    hist_vwap["Volume"].rolling(20).sum())
            hist_vwap["VWAP_U"] = hist_vwap["VWAP"] + hist_vwap["typical"].rolling(20).std()
            hist_vwap["VWAP_L"] = hist_vwap["VWAP"] - hist_vwap["typical"].rolling(20).std()

            fig2 = make_subplots(rows=2, cols=1, shared_xaxes=True,
                                 row_heights=[0.72,0.28], vertical_spacing=0.04)

            # Candlesticks
            fig2.add_trace(go.Candlestick(
                x=hist_vwap.index, open=hist_vwap["Open"], high=hist_vwap["High"],
                low=hist_vwap["Low"], close=hist_vwap["Close"], name="Price",
                increasing_line_color="#3fb950", decreasing_line_color="#f85149",
            ), row=1, col=1)

            # MAs
            fig2.add_trace(go.Scatter(x=hist_vwap.index, y=hist_vwap["MA50"],
                line=dict(color="#d29922", width=1.5), name="50 MA"), row=1, col=1)
            fig2.add_trace(go.Scatter(x=hist_vwap.index, y=hist_vwap["MA200"],
                line=dict(color="#388bfd", width=1.5, dash="dot"), name="200 MA"), row=1, col=1)

            # VWAP overlay
            if show_vwap:
                fig2.add_trace(go.Scatter(
                    x=hist_vwap.index, y=hist_vwap["VWAP"],
                    line=dict(color="#c084fc", width=1.8),
                    name="VWAP (20d)", opacity=0.9,
                ), row=1, col=1)
                fig2.add_trace(go.Scatter(
                    x=hist_vwap.index, y=hist_vwap["VWAP_U"],
                    line=dict(color="#c084fc", width=0.8, dash="dot"),
                    name="VWAP +1σ", opacity=0.5,
                ), row=1, col=1)
                fig2.add_trace(go.Scatter(
                    x=hist_vwap.index, y=hist_vwap["VWAP_L"],
                    line=dict(color="#c084fc", width=0.8, dash="dot"),
                    fill="tonexty", fillcolor="rgba(192,132,252,0.05)",
                    name="VWAP -1σ", opacity=0.5,
                ), row=1, col=1)

            # Swing high/low levels from scan data
            if show_swings and not df.empty and sel in df["ticker"].values:
                row_data = df[df["ticker"] == sel].iloc[0]
                sh = row_data.get("ms_swing_high")
                sl = row_data.get("ms_swing_low")
                if sh and pd.notna(sh):
                    fig2.add_hline(y=sh, line_color="#f85149", line_dash="dash",
                                   line_width=1, opacity=0.7,
                                   annotation_text=f"Swing High ${sh:.2f}",
                                   annotation_position="right", row=1, col=1)
                if sl and pd.notna(sl):
                    fig2.add_hline(y=sl, line_color="#3fb950", line_dash="dash",
                                   line_width=1, opacity=0.7,
                                   annotation_text=f"Swing Low ${sl:.2f}",
                                   annotation_position="right", row=1, col=1)

            # PA pattern annotations on last candle
            if not df.empty and sel in df["ticker"].values:
                row_data = df[df["ticker"] == sel].iloc[0]
                _pa_raw = row_data.get("pa_patterns", "")
                # Guard: CSV reload can return NaN (float) for missing PA values
                pa = (
                    str(_pa_raw)
                    if _pa_raw is not None
                    and not (isinstance(_pa_raw, float) and pd.isna(_pa_raw))
                    else ""
                )
                if pa and pa not in ("None", "nan", ""):
                    last_date  = hist_vwap.index[-1]
                    last_price = hist_vwap["High"].iloc[-1]
                    fig2.add_annotation(
                        x=last_date, y=last_price * 1.02,
                        text=f"📍 {pa[:40]}",
                        showarrow=True, arrowhead=2,
                        font=dict(color="#d29922", size=10),
                        bgcolor="#1a1a2e", bordercolor="#d29922",
                        row=1, col=1
                    )

            # Volume bars
            vol_colors = ["#3fb950" if hist_vwap["Close"].iloc[i] >= hist_vwap["Open"].iloc[i]
                          else "#f85149" for i in range(len(hist_vwap))]
            fig2.add_trace(go.Bar(x=hist_vwap.index, y=hist_vwap["Volume"],
                marker_color=vol_colors, name="Volume", opacity=0.7), row=2, col=1)

            fig2.update_layout(
                paper_bgcolor="#0d1117", plot_bgcolor="#0d1117", font_color="#e6edf3",
                xaxis_rangeslider_visible=False, height=600,
                margin=dict(l=10,r=10,t=30,b=20), legend=dict(orientation="h",y=1.02),
            )
            fig2.update_yaxes(gridcolor="#21262d")
            fig2.update_xaxes(gridcolor="#21262d")
            st.plotly_chart(fig2, use_container_width=True)

            if not df.empty and sel in df["ticker"].values:
                row_data = df[df["ticker"]==sel].iloc[0]
                # Row 1: existing metrics
                s1,s2,s3,s4,s5 = st.columns(5)
                s1.metric("3m Return",  pct_fmt(row_data.get("perf_3m_%",0)))
                s2.metric("Apex Score", f"{float(row_data.get('apex_score',0)):.0f}")
                s3.metric("RS (3m)",    f"{float(row_data.get('rs_3m',0)):.0f}" if row_data.get("rs_3m",0) else "–")
                s4.metric("Stage",      str(row_data.get("stage","–")))
                s5.metric("ADR %",      f"{float(row_data.get('adr_%',0)):.1f}%")
                # Row 2: new metrics
                n1,n2,n3,n4,n5 = st.columns(5)
                n1.metric("Order Flow",   str(row_data.get("of_bias","–")))
                n2.metric("VWAP Position",str(row_data.get("vwap_position","–")))
                n3.metric("vs VWAP",      pct_fmt(row_data.get("vs_vwap_%",0)))
                n4.metric("Structure",    str(row_data.get("ms_structure","–")))
                n5.metric("PA Patterns",  str(row_data.get("pa_patterns","None"))[:30])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — THEME HEATMAP
# ══════════════════════════════════════════════════════════════════════════════

with tabs[2]:
    st.markdown("### 🌍 Theme Rotation Heatmap")
    if df.empty:
        st.info("Run a scan first.")
    else:
        agg = df.groupby(["theme","market"]).agg(
            avg_score=("apex_score","mean"),
            avg_3m=("perf_3m_%","mean"),
            count=("ticker","count"),
            breakouts=("breaking_out","sum"),
        ).reset_index()

        pivot = agg.pivot_table(index="theme", columns="market",
                                values="avg_score", fill_value=0)
        fig3 = px.imshow(pivot, text_auto=".0f", aspect="auto",
            color_continuous_scale=[[0,"#0d1117"],[0.4,"#1f4e79"],
                                    [0.7,"#d29922"],[1,"#3fb950"]],
            title="Avg Apex Score by Theme & Market")
        fig3.update_layout(paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                           font_color="#e6edf3", height=360)
        st.plotly_chart(fig3, use_container_width=True)

        h1,h2 = st.columns(2)
        with h1:
            fig4 = px.scatter(agg, x="avg_3m", y="avg_score", size="count",
                color="market", text="theme", title="Performance vs Score",
                color_discrete_map={"US":"#388bfd"})
            fig4.update_traces(textposition="top center")
            fig4.update_layout(paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                font_color="#e6edf3", height=380,
                xaxis=dict(title="Avg 3m %",gridcolor="#21262d"),
                yaxis=dict(title="Avg Score",gridcolor="#21262d"))
            st.plotly_chart(fig4, use_container_width=True)
        with h2:
            fig5 = px.bar(agg.sort_values("breakouts",ascending=False),
                x="theme", y="breakouts", color="market", barmode="group",
                title="Active Breakouts by Theme",
                color_discrete_map={"US":"#388bfd"})
            fig5.update_layout(paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                font_color="#e6edf3", height=380,
                xaxis=dict(tickangle=-30,gridcolor="#21262d"),
                yaxis=dict(gridcolor="#21262d"))
            st.plotly_chart(fig5, use_container_width=True)

        # ── Emerging Gems Spotlight ───────────────────────────────────────
        st.markdown("---")
        st.markdown("### 💎 Emerging Gems Spotlight")
        st.markdown(GEM_DISCLAIMER, unsafe_allow_html=True)

        gems_df = pd.DataFrame()
        if "theme" in df.columns:
            gems_df = df[df["theme"] == "emerging_gems"].copy()
        if "is_gem" in df.columns:
            extra = df[df["is_gem"] == True].copy()
            gems_df = pd.concat([gems_df, extra]).drop_duplicates(subset=["ticker"]) if not gems_df.empty else extra

        if gems_df.empty:
            st.info("No emerging gems in current scan results. Run a Live Scan — gems are automatically detected from your config.")
        else:
            g1, g2, g3, g4 = st.columns(4)
            with g1:
                st.markdown(f'<div class="metric-card"><h3>💎 Gems Found</h3><div class="value amber">{len(gems_df)}</div></div>', unsafe_allow_html=True)
            gem_s2 = int(gems_df["stage"].str.contains("2 ✅", na=False).sum()) if "stage" in gems_df.columns else 0
            with g2:
                st.markdown(f'<div class="metric-card"><h3>Stage 2 Gems</h3><div class="value green">{gem_s2}</div></div>', unsafe_allow_html=True)
            gem_brk = int((gems_df.get("breaking_out", pd.Series([False]*len(gems_df))) == True).sum()) if "breaking_out" in gems_df.columns else 0
            with g3:
                st.markdown(f'<div class="metric-card"><h3>Gem Breakouts 🚀</h3><div class="value amber">{gem_brk}</div></div>', unsafe_allow_html=True)
            top_gem_score = pd.to_numeric(gems_df["apex_score"], errors="coerce").max() if not gems_df.empty else 0
            with g4:
                st.markdown(f'<div class="metric-card"><h3>Top Gem Score</h3><div class="value green">{top_gem_score:.0f}</div></div>', unsafe_allow_html=True)

            gem_show = [c for c in ["ticker","mcap_category","market_cap_bn","price","stage",
                                    "perf_3m_%","rs_3m","of_bias","vwap_position",
                                    "pa_patterns","apex_score"] if c in gems_df.columns]
            gem_disp = gems_df[gem_show].copy()
            gem_fmt  = {
                "price":         "${:.2f}",
                "market_cap_bn": lambda v: f"${v:.2f}B" if pd.notna(v) and v else "–",
                "apex_score":    "{:.0f}",
                "perf_3m_%":     pct_fmt,
                "rs_3m":         lambda v: f"{v:.0f}" if pd.notna(v) and v else "–",
            }
            active_gem_fmt = {k: v for k, v in gem_fmt.items() if k in gem_disp.columns}
            st.dataframe(
                gem_disp.style.map(color_score, subset=["apex_score"]).format(active_gem_fmt, na_rep="–"),
                use_container_width=True, hide_index=True
            )

            if len(gems_df) >= 2:
                gems_df["apex_score"] = pd.to_numeric(gems_df["apex_score"], errors="coerce")
                fig_gems = go.Figure(go.Bar(
                    x=gems_df["apex_score"].head(10),
                    y=gems_df["ticker"].head(10),
                    orientation="h", marker_color="#d29922",
                    text=gems_df["apex_score"].head(10).round(0).astype("Int64"),
                    textposition="outside",
                ))
                fig_gems.update_layout(
                    title="💎 Emerging Gems — Apex Score Ranking",
                    paper_bgcolor="#0d1117", plot_bgcolor="#0d1117", font_color="#e6edf3",
                    yaxis=dict(autorange="reversed", gridcolor="#21262d"),
                    xaxis=dict(range=[0, 115], gridcolor="#21262d"),
                    height=320, margin=dict(l=10, r=60, t=40, b=20),
                )
                st.plotly_chart(fig_gems, use_container_width=True)
# ══════════════════════════════════════════════════════════════════════════════

with tabs[3]:
    st.markdown("### 💼 Portfolio Tracker")
    st.caption("Positions, P&L, stop management, and full trade history — saved permanently.")

    holdings  = load_portfolio()
    trade_log = load_trade_log()

    # ══════════════════════════════════════════════════════════════════════════
    # SUB-TABS: Open Positions | Close / Manage | Trade History | Analytics
    # ══════════════════════════════════════════════════════════════════════════
    pt1, pt2, pt3, pt4 = st.tabs([
        "📂 Open Positions", "➕ Add / Manage", "📜 Trade History", "📊 Analytics"
    ])

    # ─────────────────────────────────────────────────────────────────────────
    # SUB-TAB 1 — OPEN POSITIONS
    # ─────────────────────────────────────────────────────────────────────────
    with pt1:
        if not holdings:
            st.info("No open positions. Go to ➕ Add / Manage to add your first trade.")
        else:
            rows   = []
            alerts = []

            for h in holdings:
                tk        = h["ticker"]
                qty       = float(h["qty"])
                cost      = float(h["buy_price"])
                stop      = float(h.get("stop_loss", 0)) or None
                target    = float(h.get("target", 0)) or None
                live      = fetch_price(tk)

                if not live:
                    rows.append({
                        "Ticker": tk, "Qty": qty, "Buy $": cost,
                        "Current $": None, "Day %": None,
                        "P&L $": None, "P&L %": None, "Value $": None,
                        "Stop $": stop, "Target $": target,
                        "R Multiple": None, "vs 50MA %": None,
                        "vs 200MA %": None, "Days Held": None, "Signal": "No data",
                    })
                    continue

                price      = live["price"]
                ma50       = live["ma50"]
                ma200      = live["ma200"]
                chg        = live["chg_pct"]
                pnl_pct    = round((price / cost - 1) * 100, 2)
                pnl_dol    = round((price - cost) * qty, 2)
                value      = round(price * qty, 2)
                vs50       = round((price / ma50 - 1) * 100, 1)   if ma50  else None
                vs200      = round((price / ma200 - 1) * 100, 1)  if ma200 else None
                hold_days  = (datetime.today() - pd.to_datetime(h["buy_date"])).days

                # R Multiple — how many R's of profit/loss relative to initial risk
                risk_per_sh = (cost - stop) if stop else None
                r_mult      = round(pnl_dol / (risk_per_sh * qty), 2) if risk_per_sh and risk_per_sh > 0 else None

                # ── Automated signal logic ─────────────────────────────────
                if stop and price <= stop:
                    signal = "🛑 STOP HIT"
                    alerts.append(("danger", tk, f"Price ${price:.2f} has hit/breached stop ${stop:.2f}"))
                elif target and price >= target:
                    signal = "🎯 TARGET HIT"
                    alerts.append(("success", tk, f"Price ${price:.2f} has reached target ${target:.2f}"))
                elif ma50 and ma200 and price < ma50 and price < ma200:
                    signal = "🔴 Below Both MAs"
                    alerts.append(("danger", tk, f"${price:.2f} below 50MA (${ma50:.2f}) & 200MA (${ma200:.2f})"))
                elif ma50 and price < ma50:
                    signal = "⚠️ Below 50MA"
                    alerts.append(("warn", tk, f"${price:.2f} dropped below 50MA (${ma50:.2f})"))
                elif vs50 and vs50 > 25:
                    signal = "⚡ Extended +25% above 50MA"
                elif pnl_pct >= 20:
                    signal = "🚀 +20%+ — Trail stop"
                else:
                    signal = "✅ Hold"

                rows.append({
                    "Ticker":    tk,
                    "Qty":       qty,
                    "Buy $":     cost,
                    "Current $": price,
                    "Day %":     chg,
                    "P&L $":     pnl_dol,
                    "P&L %":     pnl_pct,
                    "Value $":   value,
                    "Stop $":    stop,
                    "Target $":  target,
                    "R Mult":    r_mult,
                    "vs 50MA %": vs50,
                    "Days Held": hold_days,
                    "Signal":    signal,
                    "Setup":     h.get("setup",""),
                })

            port_df = pd.DataFrame(rows)

            # ── 🚨 Alerts bar ──────────────────────────────────────────────
            if alerts:
                st.markdown("#### 🚨 Position Alerts")
                for level, tk, msg in alerts:
                    if level == "danger":
                        st.error(f"**{tk}** — {msg}")
                    elif level == "success":
                        st.success(f"**{tk}** — {msg}")
                    else:
                        st.warning(f"**{tk}** — {msg}")
                st.markdown("---")

            # ── Summary KPIs ───────────────────────────────────────────────
            _num = port_df.copy()
            for _c in ["P&L $","P&L %","Value $","Day %"]:
                _num[_c] = pd.to_numeric(_num[_c], errors="coerce")

            total_value   = _num["Value $"].sum()
            total_pnl     = _num["P&L $"].sum()
            total_cost    = sum(float(h["qty"])*float(h["buy_price"]) for h in holdings)
            total_pnl_pct = round((total_pnl / total_cost * 100), 2) if total_cost else 0
            winners       = int((_num["P&L %"] > 0).sum())
            losers        = int((_num["P&L %"] < 0).sum())
            day_pnl       = sum(
                float(r["Day %"] or 0)/100 * float(r["Value $"] or 0)
                for r in rows if isinstance(r.get("Day %"),(int,float))
            )

            k1,k2,k3,k4,k5 = st.columns(5)
            gc = "green" if total_pnl >= 0 else "red"
            dc = "green" if day_pnl  >= 0 else "red"
            k1.metric("Portfolio Value",  f"${total_value:,.0f}")
            k2.metric("Total P&L",        f"${total_pnl:+,.0f}",  f"{total_pnl_pct:+.1f}%")
            k3.metric("Today's P&L",      f"${day_pnl:+,.0f}")
            k4.metric("Win / Loss",       f"{winners}W / {losers}L")
            k5.metric("Open Positions",   len(holdings))

            st.markdown("---")

            # ── Holdings table ─────────────────────────────────────────────
            def _cpnl(v):
                try: return "color:#3fb950;font-weight:700" if float(v)>0 else ("color:#f85149;font-weight:700" if float(v)<0 else "")
                except: return ""

            disp_cols = ["Ticker","Qty","Buy $","Current $","Day %",
                         "P&L $","P&L %","Value $","Stop $","Target $",
                         "R Mult","vs 50MA %","Days Held","Signal","Setup"]
            disp_df = port_df[[c for c in disp_cols if c in port_df.columns]]

            styled_p = disp_df.style \
                .map(_cpnl, subset=[c for c in ["P&L $","P&L %","Day %","R Mult"] if c in disp_df.columns]) \
                .format({
                    "Buy $":      lambda v: f"${v:.2f}" if isinstance(v,(int,float)) else "–",
                    "Current $":  lambda v: f"${v:.2f}" if isinstance(v,(int,float)) else "–",
                    "Stop $":     lambda v: f"${v:.2f}" if isinstance(v,(int,float)) else "–",
                    "Target $":   lambda v: f"${v:.2f}" if isinstance(v,(int,float)) else "–",
                    "Day %":      lambda v: f"{v:+.2f}%" if isinstance(v,(int,float)) else "–",
                    "P&L $":      lambda v: f"${v:+,.2f}" if isinstance(v,(int,float)) else "–",
                    "P&L %":      lambda v: f"{v:+.1f}%" if isinstance(v,(int,float)) else "–",
                    "Value $":    lambda v: f"${v:,.2f}" if isinstance(v,(int,float)) else "–",
                    "vs 50MA %":  lambda v: f"{v:+.1f}%" if isinstance(v,(int,float)) else "–",
                    "R Mult":     lambda v: f"{v:+.2f}R" if isinstance(v,(int,float)) else "–",
                    "Days Held":  lambda v: f"{int(v)}d"   if isinstance(v,(int,float)) else "–",
                }, na_rep="–")
            st.dataframe(styled_p, use_container_width=True, height=420)

            # ── Allocation pie + Sector breakdown ─────────────────────────
            pie_data = _num.dropna(subset=["Value $"])
            if not pie_data.empty:
                col_pie, col_sec = st.columns(2)
                with col_pie:
                    fig_pie = px.pie(
                        pie_data, names="Ticker", values="Value $",
                        title="Position Allocation",
                        color_discrete_sequence=px.colors.sequential.Viridis,
                        hole=0.35,
                    )
                    fig_pie.update_layout(
                        paper_bgcolor="#0d1117", font_color="#e6edf3", height=340,
                        margin=dict(t=40,b=10,l=10,r=10),
                    )
                    st.plotly_chart(fig_pie, use_container_width=True)

                with col_sec:
                    # P&L bar chart per position
                    pnl_bar = _num.dropna(subset=["P&L $"]).copy()
                    pnl_bar["Color"] = pnl_bar["P&L $"].apply(lambda v: "#3fb950" if v>=0 else "#f85149")
                    fig_bar = px.bar(
                        pnl_bar, x="Ticker", y="P&L $",
                        title="P&L by Position ($)",
                        color="Color", color_discrete_map="identity",
                    )
                    fig_bar.update_layout(
                        paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                        font_color="#e6edf3", height=340, showlegend=False,
                        margin=dict(t=40,b=10,l=10,r=10),
                    )
                    fig_bar.update_traces(marker_line_width=0)
                    st.plotly_chart(fig_bar, use_container_width=True)

    # ─────────────────────────────────────────────────────────────────────────
    # SUB-TAB 2 — ADD / MANAGE
    # ─────────────────────────────────────────────────────────────────────────
    with pt2:
        manage_mode = st.radio(
            "Action", ["➕ Add Position", "✏️ Edit Stop/Target", "📤 Close Position"],
            horizontal=True
        )

        # ── ADD POSITION ───────────────────────────────────────────────────
        if manage_mode == "➕ Add Position":
            with st.form("add_position_form"):
                st.markdown("#### ➕ Add a New Position")
                c1,c2,c3 = st.columns(3)
                with c1:
                    n_tk    = st.text_input("Ticker *", placeholder="NVDA").upper().strip()
                    n_qty   = st.number_input("Shares *", min_value=0.001, value=1.0, step=1.0, format="%.3f")
                with c2:
                    n_price = st.number_input("Buy Price ($) *", min_value=0.01, value=100.0, step=0.01)
                    n_stop  = st.number_input("Stop Loss ($)", min_value=0.0, value=0.0, step=0.01,
                                              help="Price at which you will exit if wrong. 0 = no stop set.")
                with c3:
                    n_target= st.number_input("Target Price ($)", min_value=0.0, value=0.0, step=0.01,
                                              help="Price target for this trade. 0 = no target set.")
                    n_date  = st.date_input("Buy Date *", value=datetime.today())

                c4,c5 = st.columns(2)
                with c4:
                    n_setup = st.selectbox("Setup Type",
                        ["", "Breakout", "Pullback to 50MA", "VWAP Reclaim",
                         "SFP / Bear Trap", "Flat Base", "Cup & Handle",
                         "Earnings Play", "Sector Rotation", "Other"])
                with c5:
                    n_notes = st.text_input("Notes", placeholder="e.g. Breaking out on 2x volume")

                # Live risk preview
                if n_price > 0 and n_stop > 0 and n_stop < n_price:
                    risk_sh   = round(n_price - n_stop, 2)
                    risk_pct  = round(risk_sh / n_price * 100, 2)
                    risk_tot  = round(risk_sh * n_qty, 2)
                    rr        = round((n_target - n_price) / risk_sh, 2) if n_target > n_price else None
                    st.info(
                        f"**Risk per share:** ${risk_sh:.2f} ({risk_pct:.1f}%)  |  "
                        f"**Total risk:** ${risk_tot:,.2f}  |  "
                        f"**R:R ratio:** {rr:.1f}:1" if rr else
                        f"**Risk per share:** ${risk_sh:.2f} ({risk_pct:.1f}%)  |  "
                        f"**Total risk:** ${risk_tot:,.2f}"
                    )

                submitted = st.form_submit_button("Add Position", use_container_width=True)
                if submitted and n_tk:
                    holdings.append({
                        "ticker":    n_tk,
                        "qty":       n_qty,
                        "buy_price": n_price,
                        "buy_date":  str(n_date),
                        "stop_loss": n_stop if n_stop > 0 else None,
                        "target":    n_target if n_target > 0 else None,
                        "setup":     n_setup,
                        "notes":     n_notes,
                    })
                    save_portfolio(holdings)
                    st.success(f"✅ Added {n_tk} — {n_qty} shares @ ${n_price:.2f}")
                    st.rerun()

        # ── EDIT STOP / TARGET ─────────────────────────────────────────────
        elif manage_mode == "✏️ Edit Stop/Target":
            if not holdings:
                st.info("No open positions to edit.")
            else:
                st.markdown("#### ✏️ Update Stop Loss or Target")
                edit_tk = st.selectbox("Select Position", [h["ticker"] for h in holdings])
                h_edit  = next((h for h in holdings if h["ticker"] == edit_tk), None)
                if h_edit:
                    live_e  = fetch_price(edit_tk)
                    cur_p   = live_e["price"] if live_e else h_edit["buy_price"]
                    e1,e2   = st.columns(2)
                    with e1:
                        new_stop = st.number_input(
                            "New Stop Loss ($)",
                            min_value=0.0,
                            value=float(h_edit.get("stop_loss") or 0),
                            step=0.01,
                        )
                    with e2:
                        new_tgt  = st.number_input(
                            "New Target ($)",
                            min_value=0.0,
                            value=float(h_edit.get("target") or 0),
                            step=0.01,
                        )
                    new_notes = st.text_input("Update Notes", value=h_edit.get("notes",""))

                    # Show updated R:R
                    if new_stop > 0 and new_stop < cur_p:
                        _rs = cur_p - new_stop
                        _rr = round((new_tgt - cur_p) / _rs, 2) if new_tgt > cur_p else None
                        pnl_now = round((cur_p / h_edit["buy_price"] - 1) * 100, 2)
                        st.info(
                            f"Current Price: **${cur_p:.2f}**  |  "
                            f"Open P&L: **{pnl_now:+.1f}%**  |  "
                            f"New risk/share: **${_rs:.2f}**  |  "
                            + (f"R:R: **{_rr:.1f}:1**" if _rr else "Set target for R:R")
                        )

                    if st.button("💾 Save Changes"):
                        for h in holdings:
                            if h["ticker"] == edit_tk:
                                h["stop_loss"] = new_stop if new_stop > 0 else None
                                h["target"]    = new_tgt  if new_tgt  > 0 else None
                                h["notes"]     = new_notes
                        save_portfolio(holdings)
                        st.success(f"Updated {edit_tk}")
                        st.rerun()

        # ── CLOSE POSITION ─────────────────────────────────────────────────
        elif manage_mode == "📤 Close Position":
            if not holdings:
                st.info("No open positions to close.")
            else:
                st.markdown("#### 📤 Close a Position")
                close_tk = st.selectbox("Position to Close", [h["ticker"] for h in holdings])
                h_close  = next((h for h in holdings if h["ticker"] == close_tk), None)
                if h_close:
                    live_c   = fetch_price(close_tk)
                    def_price= live_c["price"] if live_c else float(h_close["buy_price"])
                    max_qty  = float(h_close["qty"])

                    cl1,cl2,cl3 = st.columns(3)
                    with cl1:
                        close_price = st.number_input(
                            "Exit Price ($)", min_value=0.01, value=def_price, step=0.01
                        )
                    with cl2:
                        close_qty = st.number_input(
                            "Shares to Close",
                            min_value=0.001, max_value=max_qty,
                            value=max_qty, step=1.0, format="%.3f",
                            help="Partial close: enter less than full qty"
                        )
                    with cl3:
                        close_date = st.date_input("Exit Date", value=datetime.today())

                    # P&L preview
                    _pnl_prev = round((close_price / float(h_close["buy_price"]) - 1)*100, 2)
                    _pnl_dol  = round((close_price - float(h_close["buy_price"])) * close_qty, 2)
                    _hold_d   = (pd.to_datetime(str(close_date)) - pd.to_datetime(h_close["buy_date"])).days
                    _stop_r   = float(h_close.get("stop_loss") or 0)
                    _risk_sh  = (float(h_close["buy_price"]) - _stop_r) if _stop_r else None
                    _r_mult   = round(_pnl_dol / (_risk_sh * close_qty), 2) if _risk_sh and _risk_sh > 0 else None

                    col = "green" if _pnl_dol >= 0 else "red"
                    st.markdown(
                        f'<div style="background:#161b22;padding:14px;border-radius:8px;margin:8px 0;">'
                        f'<b>Exit P&L:</b> <span style="color:{"#3fb950" if _pnl_dol>=0 else "#f85149"};font-size:1.2rem;font-weight:700;">'
                        f'${_pnl_dol:+,.2f} ({_pnl_prev:+.1f}%)</span>'
                        f'&nbsp;&nbsp;|&nbsp;&nbsp;<b>Days held:</b> {_hold_d}d'
                        + (f'&nbsp;&nbsp;|&nbsp;&nbsp;<b>R multiple:</b> {_r_mult:+.2f}R' if _r_mult else '')
                        + f'&nbsp;&nbsp;|&nbsp;&nbsp;<b>Closing:</b> {close_qty:.3f} of {max_qty:.3f} shares'
                        + '</div>',
                        unsafe_allow_html=True
                    )

                    close_notes = st.text_input("Exit Notes", placeholder="e.g. Stop hit / Target reached / Cutting loss")

                    if st.button("📤 Confirm Close", type="primary"):
                        h_close["notes"] = (h_close.get("notes","") + " | Exit: " + close_notes).strip(" | ")
                        new_holdings, new_log = close_position(
                            holdings, trade_log, close_tk,
                            close_price, str(close_date), close_qty
                        )
                        save_portfolio(new_holdings)
                        save_trade_log(new_log)
                        partial = close_qty < max_qty
                        msg = f"Partially closed {close_qty:.0f} shares of {close_tk}" if partial else f"Closed {close_tk}"
                        st.success(f"✅ {msg} — P&L: ${_pnl_dol:+,.2f} ({_pnl_prev:+.1f}%)")
                        st.rerun()

    # ─────────────────────────────────────────────────────────────────────────
    # SUB-TAB 3 — TRADE HISTORY
    # ─────────────────────────────────────────────────────────────────────────
    with pt3:
        st.markdown("#### 📜 Closed Trade History")
        if not trade_log:
            st.info("No closed trades yet. Close your first position in ➕ Add / Manage.")
        else:
            tlog_df = pd.DataFrame(trade_log)
            # Summary stats
            wins    = tlog_df[tlog_df["pnl_dol"] > 0]
            losses  = tlog_df[tlog_df["pnl_dol"] <= 0]
            win_rt  = round(len(wins)/len(tlog_df)*100,1) if tlog_df.shape[0] > 0 else 0
            avg_win = wins["pnl_pct"].mean()  if not wins.empty  else 0
            avg_los = losses["pnl_pct"].mean()if not losses.empty else 0
            expectancy = round((win_rt/100 * avg_win) + ((1-win_rt/100) * avg_los), 2)
            avg_r   = tlog_df["pnl_dol"].mean() / \
                      ((tlog_df["buy_price"] - 0).mean()) if "stop" not in tlog_df.columns else None
            total_closed_pnl = tlog_df["pnl_dol"].sum()
            avg_hold = round(tlog_df["hold_days"].mean(), 1)

            h1,h2,h3,h4,h5 = st.columns(5)
            h1.metric("Total Closed P&L",  f"${total_closed_pnl:+,.2f}")
            h2.metric("Win Rate",           f"{win_rt:.1f}%")
            h3.metric("Avg Winner",         f"{avg_win:+.1f}%")
            h4.metric("Avg Loser",          f"{avg_los:+.1f}%")
            h5.metric("Avg Hold",           f"{avg_hold}d")

            st.markdown("---")
            # Expectancy
            exp_color = "#3fb950" if expectancy > 0 else "#f85149"
            st.markdown(
                f'<div style="background:#161b22;padding:12px;border-radius:8px;margin-bottom:12px;">'
                f'📐 <b>System Expectancy:</b> '
                f'<span style="color:{exp_color};font-size:1.1rem;font-weight:700;">{expectancy:+.2f}%</span>'
                f' per trade &nbsp;|&nbsp; '
                f'<span style="color:#8b949e;font-size:0.85rem;">'
                f'Expectancy = (Win% × Avg Win) + (Loss% × Avg Loss). '
                f'Positive = edge. Target: >0.5%</span></div>',
                unsafe_allow_html=True
            )

            # Trade log table
            def _cl(v):
                try: return "color:#3fb950;font-weight:700" if float(v)>0 else "color:#f85149;font-weight:700"
                except: return ""

            display_log = tlog_df[["ticker","qty","buy_price","buy_date",
                                    "close_price","close_date","pnl_dol","pnl_pct",
                                    "hold_days","setup","notes"]].copy()
            display_log.columns = ["Ticker","Qty","Buy $","Buy Date",
                                    "Exit $","Exit Date","P&L $","P&L %",
                                    "Days","Setup","Notes"]

            styled_log = display_log.style \
                .map(_cl, subset=["P&L $","P&L %"]) \
                .format({
                    "Buy $":  "${:.2f}",
                    "Exit $": "${:.2f}",
                    "P&L $":  lambda v: f"${v:+,.2f}" if isinstance(v,(int,float)) else v,
                    "P&L %":  lambda v: f"{v:+.1f}%"  if isinstance(v,(int,float)) else v,
                    "Days":   lambda v: f"{int(v)}d"   if isinstance(v,(int,float)) else v,
                }, na_rep="–")
            st.dataframe(styled_log, use_container_width=True, height=400)

            # Cumulative P&L chart
            st.markdown("---")
            tlog_df_sorted = tlog_df.sort_values("close_date")
            tlog_df_sorted["cumulative_pnl"] = tlog_df_sorted["pnl_dol"].cumsum()
            fig_cum = px.area(
                tlog_df_sorted, x="close_date", y="cumulative_pnl",
                title="Cumulative Closed P&L ($)",
                color_discrete_sequence=["#3fb950"],
            )
            fig_cum.add_hline(y=0, line_dash="dot", line_color="#8b949e")
            fig_cum.update_layout(
                paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                font_color="#e6edf3", height=320,
                xaxis_title="Date", yaxis_title="Cumulative P&L ($)",
                margin=dict(t=40,b=20,l=20,r=20),
            )
            fig_cum.update_traces(fill="tozeroy", fillcolor="rgba(63,185,80,0.15)")
            st.plotly_chart(fig_cum, use_container_width=True)

            # Export trade log
            st.download_button(
                "⬇ Export Trade Log (CSV)",
                data=display_log.to_csv(index=False).encode("utf-8"),
                file_name=f"apexscan_trades_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
            )

    # ─────────────────────────────────────────────────────────────────────────
    # SUB-TAB 4 — ANALYTICS
    # ─────────────────────────────────────────────────────────────────────────
    with pt4:
        st.markdown("#### 📊 Portfolio Analytics")

        if not trade_log and not holdings:
            st.info("Add positions and close some trades to see analytics.")
        else:
            all_trades = pd.DataFrame(trade_log) if trade_log else pd.DataFrame()

            an1, an2 = st.columns(2)

            with an1:
                # Win/Loss distribution
                if not all_trades.empty:
                    fig_dist = px.histogram(
                        all_trades, x="pnl_pct", nbins=20,
                        title="P&L % Distribution (Closed Trades)",
                        color_discrete_sequence=["#388bfd"],
                    )
                    fig_dist.add_vline(x=0, line_dash="dot", line_color="#f85149")
                    fig_dist.update_layout(
                        paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                        font_color="#e6edf3", height=300,
                        margin=dict(t=40,b=20,l=20,r=20),
                    )
                    st.plotly_chart(fig_dist, use_container_width=True)

                    # Best/Worst trades
                    best  = all_trades.nlargest(3,"pnl_pct")[["ticker","pnl_pct","pnl_dol","hold_days"]]
                    worst = all_trades.nsmallest(3,"pnl_pct")[["ticker","pnl_pct","pnl_dol","hold_days"]]
                    st.markdown("**🏆 Best Trades**")
                    st.dataframe(best.style.format({"pnl_pct":"{:+.1f}%","pnl_dol":"${:+,.2f}","hold_days":"{:.0f}d"}),
                                 use_container_width=True, height=140)
                    st.markdown("**💀 Worst Trades**")
                    st.dataframe(worst.style.format({"pnl_pct":"{:+.1f}%","pnl_dol":"${:+,.2f}","hold_days":"{:.0f}d"}),
                                 use_container_width=True, height=140)

            with an2:
                # Setup type breakdown
                if not all_trades.empty and "setup" in all_trades.columns:
                    setup_perf = all_trades.groupby("setup").agg(
                        Trades=("pnl_dol","count"),
                        Avg_PnL_pct=("pnl_pct","mean"),
                        Total_PnL=("pnl_dol","sum"),
                        Win_Rate=("pnl_dol", lambda x: (x>0).mean()*100)
                    ).reset_index().sort_values("Avg_PnL_pct", ascending=False)
                    setup_perf.columns = ["Setup","Trades","Avg P&L %","Total P&L $","Win %"]
                    st.markdown("**📋 Performance by Setup Type**")
                    st.dataframe(
                        setup_perf.style.format({
                            "Avg P&L %": "{:+.1f}%",
                            "Total P&L $": "${:+,.2f}",
                            "Win %": "{:.0f}%",
                        }),
                        use_container_width=True, height=280
                    )

                # Open positions R multiple chart
                open_r = [(h["ticker"], h.get("stop_loss"), h.get("buy_price"))
                          for h in holdings if h.get("stop_loss") and h.get("buy_price")]
                if open_r:
                    r_rows = []
                    for tk, sl, bp in open_r:
                        lv = fetch_price(tk)
                        if lv:
                            cur = lv["price"]
                            risk = float(bp) - float(sl)
                            if risk > 0:
                                r_mult = round((cur - float(bp)) / risk, 2)
                                r_rows.append({"Ticker": tk, "R Multiple": r_mult,
                                               "Color": "#3fb950" if r_mult >= 0 else "#f85149"})
                    if r_rows:
                        r_df = pd.DataFrame(r_rows)
                        fig_r = px.bar(r_df, x="Ticker", y="R Multiple",
                            title="Open Positions — R Multiple",
                            color="Color", color_discrete_map="identity")
                        fig_r.add_hline(y=0, line_dash="dot", line_color="#8b949e")
                        fig_r.update_layout(
                            paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                            font_color="#e6edf3", height=280, showlegend=False,
                            margin=dict(t=40,b=20,l=20,r=20),
                        )
                        st.plotly_chart(fig_r, use_container_width=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — EARNINGS CALENDAR
# ══════════════════════════════════════════════════════════════════════════════

with tabs[4]:
    st.markdown("### 📅 Earnings Calendar")
    st.caption("Tracks upcoming earnings for your watchlist. Data from yfinance — may not cover all tickers.")

    cfg_e = load_config("config.yaml")
    all_us = list(set(t for theme in cfg_e["us_themes"].values() for t in theme))

    # Let user pick which tickers to track
    default_watch = df["ticker"].tolist()[:10] if not df.empty else all_us[:10]
    watch_sel = st.multiselect(
        "Select tickers to track",
        options=all_us,
        default=[t for t in default_watch if t in all_us][:8],
    )

    if st.button("🔄 Fetch Earnings Dates") and watch_sel:
        cal_rows = []
        prog = st.progress(0)

        # Try Alpha Vantage earnings calendar first (1 call for all tickers)
        av_key_cal = load_config("config.yaml").get("alpha_vantage_key","")
        av_dates   = {}
        av_eps_map = {}
        if av_key_cal and not av_key_cal.startswith("YOUR_"):
            try:
                # alpha_vantage already imported at top level
                av_dates = get_upcoming_earnings_for_watchlist(watch_sel, av_key_cal)
                st.caption(f"📊 Alpha Vantage: found upcoming dates for {len(av_dates)} tickers")
            except Exception as av_e:
                st.caption(f"AV calendar unavailable: {av_e}")

        for i, tk in enumerate(watch_sel):
            prog.progress((i+1)/len(watch_sel))

            # Get next earnings date — AV first, yfinance fallback
            next_e = av_dates.get(tk)
            eps_est = "–"
            eps_growth_str = "–"

            if not next_e:
                yf_data = fetch_earnings(tk)
                next_e  = yf_data.get("next_earnings")
                eps_est = yf_data.get("eps_est") or "–"

            # Get real EPS data if AV available
            if av_key_cal and not av_key_cal.startswith("YOUR_") and tk not in av_eps_map:
                try:
                    av_fund = analyse_eps(tk, av_key_cal, cache_hours=24)
                    av_eps_map[tk] = av_fund
                    eg = av_fund.get("eps_growth_pct")
                    es = av_fund.get("eps_surprise_pct")
                    cb = av_fund.get("consecutive_beats",0)
                    if eg is not None:
                        eps_growth_str = f"{eg:+.1f}%"
                    if es is not None:
                        eps_est = f"Last surprise: {es:+.1f}% | {cb}Q beats"
                except:
                    pass

            if next_e:
                try:
                    dt = datetime.strptime(next_e, "%Y-%m-%d")
                    days_away = (dt - datetime.now()).days
                except:
                    dt = None
                    days_away = None
            else:
                dt = None
                days_away = None

            urgency = "–"
            if days_away is not None:
                if days_away <= 2:   urgency = "🔴 THIS WEEK"
                elif days_away <= 7: urgency = "🟡 NEXT WEEK"
                elif days_away <= 30:urgency = "🟢 THIS MONTH"
                else:                urgency = f"📅 {days_away}d away"

            cal_rows.append({
                "Ticker":        tk,
                "Next Earnings": next_e or "Unknown",
                "Days Away":     days_away if days_away is not None else "–",
                "Urgency":       urgency,
                "EPS YoY Growth":eps_growth_str,
                "EPS Est/Surprise": eps_est,
            })
        prog.empty()

        cal_df = pd.DataFrame(cal_rows)
        # Sort: known dates first, soonest first
        cal_df["_sort"] = pd.to_numeric(cal_df["Days Away"], errors="coerce")
        cal_df = cal_df.sort_values("_sort").drop(columns=["_sort"])

        # Highlight upcoming
        st.markdown("---")
        urgent = cal_df[cal_df["Urgency"].str.contains("THIS WEEK|NEXT WEEK", na=False)]
        if not urgent.empty:
            st.markdown("#### 🚨 Earnings This Week / Next Week")
            for _, row in urgent.iterrows():
                st.markdown(
                    f'<div class="warn-box">⚠️ <b>{row["Ticker"]}</b> reports on '
                    f'<b>{row["Next Earnings"]}</b> — {row["Days Away"]} days away</div>',
                    unsafe_allow_html=True
                )
            st.markdown("---")

        st.markdown("#### Full Calendar")
        st.dataframe(cal_df, use_container_width=True, hide_index=True)
    else:
        st.info("Select tickers above and click **Fetch Earnings Dates**.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 — SECTOR ROTATION
# ══════════════════════════════════════════════════════════════════════════════

with tabs[5]:
    st.markdown("### 🔄 Sector Rotation Dashboard")
    st.caption("Tracks US sector ETF performance to show where institutional money is flowing.")

    if st.button("🔄 Refresh Sector Data"):
        fetch_sector = sector_performance.__wrapped__ if hasattr(sector_performance,"__wrapped__") else sector_performance
        sector_df = sector_performance()
    else:
        sector_df = sector_performance()

    if sector_df.empty:
        st.warning("Could not load sector data — yfinance may be temporarily unavailable. Try the Refresh button in 60 seconds.")
    else:
        # Safe idxmax — drop NAs first to avoid "all NA values" crash
        w_valid  = sector_df["1W %"].dropna()
        m_valid  = sector_df["1M %"].dropna()
        m3_valid = sector_df["3M %"].dropna()

        if w_valid.empty or m3_valid.empty:
            st.warning("Sector ETF data returned empty — yfinance may be rate-limiting. Wait 60 seconds and click Refresh.")
        else:
            best_1w  = sector_df.loc[w_valid.idxmax()]
            worst_1w = sector_df.loc[w_valid.idxmin()]
            best_1m  = sector_df.loc[m_valid.idxmax()] if not m_valid.empty else best_1w
            best_3m  = sector_df.loc[m3_valid.idxmax()]

            r1,r2,r3,r4 = st.columns(4)
            with r1: st.markdown(f'<div class="metric-card"><h3>🏆 Best This Week</h3><div class="value green">{best_1w["Sector"]}</div><div style="color:#8b949e;font-size:.85rem">{best_1w["1W %"]:+.1f}%</div></div>', unsafe_allow_html=True)
            with r2: st.markdown(f'<div class="metric-card"><h3>📉 Worst This Week</h3><div class="value red">{worst_1w["Sector"]}</div><div style="color:#8b949e;font-size:.85rem">{worst_1w["1W %"]:+.1f}%</div></div>', unsafe_allow_html=True)
            with r3: st.markdown(f'<div class="metric-card"><h3>🥇 Best This Month</h3><div class="value amber">{best_1m["Sector"]}</div><div style="color:#8b949e;font-size:.85rem">{best_1m["1M %"]:+.1f}%</div></div>', unsafe_allow_html=True)
            with r4: st.markdown(f'<div class="metric-card"><h3>🚀 Best 3 Months</h3><div class="value blue">{best_3m["Sector"]}</div><div style="color:#8b949e;font-size:.85rem">{best_3m["3M %"]:+.1f}%</div></div>', unsafe_allow_html=True)

        st.markdown("---")

        # ── Bar chart: 1W, 1M, 3M performance ────────────────────────────
        sector_melt = sector_df.melt(
            id_vars="Sector", value_vars=["1W %","1M %","3M %"],
            var_name="Period", value_name="Return %"
        )
        fig_sec = px.bar(
            sector_melt, x="Sector", y="Return %", color="Period",
            barmode="group", title="Sector Performance — 1W / 1M / 3M",
            color_discrete_map={"1W %":"#388bfd","1M %":"#d29922","3M %":"#3fb950"},
        )
        fig_sec.update_layout(
            paper_bgcolor="#0d1117", plot_bgcolor="#0d1117", font_color="#e6edf3",
            xaxis=dict(tickangle=-30, gridcolor="#21262d"),
            yaxis=dict(gridcolor="#21262d", ticksuffix="%"),
            height=420, margin=dict(l=10,r=10,t=40,b=80),
            legend=dict(orientation="h", y=1.05),
        )
        fig_sec.add_hline(y=0, line_color="#30363d", line_width=1)
        st.plotly_chart(fig_sec, use_container_width=True)

        # ── Heatmap: sectors by timeframe ─────────────────────────────────
        heat_data = sector_df.set_index("Sector")[["1W %","1M %","3M %"]]
        fig_heat = px.imshow(
            heat_data, text_auto=".1f", aspect="auto",
            color_continuous_scale=[[0,"#f85149"],[0.5,"#21262d"],[1,"#3fb950"]],
            title="Sector Rotation Heatmap",
            color_continuous_midpoint=0,
        )
        fig_heat.update_layout(
            paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
            font_color="#e6edf3", height=380,
        )
        st.plotly_chart(fig_heat, use_container_width=True)

        # ── Raw table ─────────────────────────────────────────────────────
        st.markdown("#### Sector Detail Table")
        def _color_ret(v):
            try: return "color:#3fb950;font-weight:600" if float(v)>0 else "color:#f85149;font-weight:600"
            except: return ""

        styled_s = sector_df.style \
            .map(_color_ret, subset=["1W %","1M %","3M %"]) \
            .format({"1W %": "{:+.2f}%","1M %": "{:+.2f}%","3M %": "{:+.2f}%",
                     "Price": "${:.2f}"}, na_rep="–")
        st.dataframe(styled_s, use_container_width=True, hide_index=True)

        # ── Rotation insight ──────────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### 💡 Rotation Insight")
        top3 = sector_df.nlargest(3,"1W %")["Sector"].tolist()
        bot3 = sector_df.nsmallest(3,"1W %")["Sector"].tolist()
        st.markdown(
            f'<div class="alert-box">🟢 <b>Money flowing INTO:</b> {", ".join(top3)}</div>',
            unsafe_allow_html=True)
        st.markdown(
            f'<div class="danger-box">🔴 <b>Money flowing OUT OF:</b> {", ".join(bot3)}</div>',
            unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 — STOCK DEEP DIVE
# ══════════════════════════════════════════════════════════════════════════════

with tabs[6]:
    st.markdown("### 🔍 Stock Deep Dive")
    st.caption("Analyse any US stock — same scoring engine as the live scan.")

    d1, d2 = st.columns([2, 1])
    with d1:
        dive_ticker = st.text_input(
            "Enter any ticker", placeholder="e.g. HIMS, ARM, CELH, HOOD",
            key="dive_input"
        ).upper().strip()
    with d2:
        dive_btn = st.button("🔍 Analyse", use_container_width=True)

    if dive_btn and dive_ticker:
        with st.spinner(f"Analysing {dive_ticker}…"):
            cfg_d = load_config("config.yaml")
            result = None
            try:
                from scanner import analyze_stock
                result = analyze_stock(dive_ticker, cfg_d)
            except Exception as e:
                st.error(f"Error: {e}")

        if result is None:
            st.error(f"Could not analyse **{dive_ticker}**. Check the ticker is correct and has enough price history.")
        else:
            # ── Score banner ──────────────────────────────────────────────
            score = result["apex_score"]
            score_color = "#3fb950" if score >= 70 else ("#d29922" if score >= 40 else "#f85149")
            verdict = "Strong Setup 🟢" if score >= 70 else ("Watch List 🟡" if score >= 40 else "Skip / Weak 🔴")

            st.markdown(f"""
            <div style="background:#161b22;border:1px solid {score_color};border-radius:12px;
                        padding:20px 28px;margin-bottom:20px;">
              <div style="display:flex;justify-content:space-between;align-items:center;">
                <div>
                  <div style="font-size:2rem;font-weight:800;color:#e6edf3;">{dive_ticker}</div>
                  <div style="color:#8b949e;font-size:0.9rem;">
                    Market: {result['market']} &nbsp;|&nbsp; Theme: {result['theme']} &nbsp;|&nbsp;
                    Stage: {result.get('stage','–')}
                  </div>
                </div>
                <div style="text-align:right;">
                  <div style="font-size:3rem;font-weight:900;color:{score_color};">{score:.0f}</div>
                  <div style="color:{score_color};font-size:0.9rem;font-weight:600;">
                    Apex Score — {verdict}
                  </div>
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            # ── KPI row 1: Price & Performance ───────────────────────────
            st.markdown("#### 📊 Price & Performance")
            k1,k2,k3,k4,k5 = st.columns(5)
            k1.metric("Price",        f"${result['price']:.2f}")
            k2.metric("1M Return",    pct_fmt(result['perf_1m_%']))
            k3.metric("3M Return",    pct_fmt(result['perf_3m_%']))
            k4.metric("6M Return",    pct_fmt(result['perf_6m_%']))
            k5.metric("% Off 52W High", pct_fmt(result['pct_off_high_%']))

            st.markdown("---")

            # ── KPI row 2: Strength & Structure ──────────────────────────
            st.markdown("#### 💪 Relative Strength & Structure")
            k6,k7,k8,k9,k10 = st.columns(5)
            rs3  = result.get("rs_3m", 0)
            rs6  = result.get("rs_6m", 0)
            k6.metric("RS vs S&P500 (3m)",  f"{rs3:.0f}" if rs3 else "–")
            k7.metric("RS vs S&P500 (6m)",  f"{rs6:.0f}" if rs6 else "–")
            k8.metric("vs 50MA",               pct_fmt(result.get("vs_50ma_%", 0)))
            k9.metric("vs 200MA",              pct_fmt(result.get("vs_200ma_%", 0)))
            k10.metric("ADR %",                f"{result.get('adr_%', 0):.1f}%")

            # Russell benchmark RS row
            _rr25_3  = result.get("rs_r2500_3m")
            _rr25_6  = result.get("rs_r2500_6m")
            _rr3g_3  = result.get("rs_r3000g_3m")
            _rr3g_6  = result.get("rs_r3000g_6m")
            _multi   = result.get("rs_multi_leader", False)

            def _rs_fmt(v):
                if v is None or (isinstance(v, float) and pd.isna(v)): return "–"
                return f"{float(v):.0f}"

            def _rs_delta(v):
                if v is None or (isinstance(v, float) and pd.isna(v)): return None
                return "▲ Leader" if float(v) > 100 else ("▼ Lagging" if float(v) < 70 else "≈ In-line")

            rk1,rk2,rk3,rk4,rk5 = st.columns(5)
            rk1.metric(
                "RS vs R2500 (3m)",
                _rs_fmt(_rr25_3),
                delta=_rs_delta(_rr25_3),
                help="Relative Strength vs Russell 2500 (small/mid-cap benchmark). >100 = beating small/mid peers."
            )
            rk2.metric(
                "RS vs R2500 (6m)",
                _rs_fmt(_rr25_6),
                delta=_rs_delta(_rr25_6),
                help="RS vs Russell 2500 over 6 months — sustained small/mid outperformance."
            )
            rk3.metric(
                "RS vs R3000G (3m)",
                _rs_fmt(_rr3g_3),
                delta=_rs_delta(_rr3g_3),
                help="Relative Strength vs Russell 3000 Growth (broadest growth benchmark). >100 = beating all growth stocks."
            )
            rk4.metric(
                "RS vs R3000G (6m)",
                _rs_fmt(_rr3g_6),
                delta=_rs_delta(_rr3g_6),
                help="RS vs Russell 3000 Growth over 6 months — elite sustained growth leadership."
            )
            _multi_val = bool(_multi) if _multi is not None and not (isinstance(_multi, float) and pd.isna(_multi)) else False
            rk5.metric(
                "Multi-Bench Leader",
                "✅ YES" if _multi_val else "❌ NO",
                delta="Beats ALL 3 benchmarks" if _multi_val else None,
                help="True = outperforming S&P 500, Russell 2500 AND Russell 3000 Growth simultaneously. Extremely rare and highest-conviction RS signal."
            )
            if _multi_val:
                st.success("⭐ **Multi-Benchmark Leader** — this stock is outperforming the S&P 500, Russell 2500, and Russell 3000 Growth simultaneously. Extremely rare, highest-conviction RS signal.")

            st.markdown("---")

            # ── WEEKLY TIMEFRAME CONFIRMATION PANEL ──────────────────────────
            st.markdown("#### 📅 Weekly Timeframe Confirmation")
            _wk_conf   = result.get("weekly_confirmed", False)
            _wk_contra = result.get("weekly_contradicts", False)
            _wk_stage  = str(result.get("weekly_stage", "–") or "–")
            _wk_score  = result.get("weekly_score", 0)
            _wk_rs     = result.get("weekly_rs")
            _wk_tight  = result.get("weekly_base_tight", False)
            _wk_depth  = result.get("weekly_base_depth_%")
            _wk_hh_hl  = result.get("weekly_hh_hl", False)
            _wk_consec = result.get("weekly_consec_up_wks", 0)
            _wk_10gt40 = result.get("weekly_10gt40", False)
            _wk40      = result.get("weekly_above_40wma", False)

            # Weekly verdict banner
            if _wk_conf and not _wk_contra:
                st.success(
                    f"✅ **Weekly Confirmed** — the weekly chart supports this setup. "
                    f"Weekly Stage: **{_wk_stage}** | Weekly Score: **{_wk_score}/10**. "
                    f"Daily signal + weekly alignment = highest probability trade."
                )
            elif _wk_contra:
                st.error(
                    f"⚠️ **Weekly Contradiction** — the weekly chart is in Stage 3/4 downtrend "
                    f"while the daily shows a setup. This is the most common retail trap. "
                    f"**Avoid or use very small size (0.25% risk max).** Weekly Stage: {_wk_stage}"
                )
            else:
                st.warning(
                    f"📊 **Weekly Neutral/Transitioning** — weekly chart is not yet confirmed. "
                    f"Stage: **{_wk_stage}**. Wait for weekly to resolve before sizing up."
                )

            # Weekly KPIs
            ww1,ww2,ww3,ww4,ww5 = st.columns(5)
            ww1.metric("Weekly Stage",    _wk_stage.split(" ")[0] + " " + " ".join(_wk_stage.split(" ")[1:2]) if _wk_stage != "–" else "–",
                       help="Weinstein Stage on the weekly chart. Stage 2 = only buyable stage.")
            ww2.metric("Weekly RS",       _rs_fmt(_wk_rs) if _wk_rs else "–",
                       delta="▲ Leader" if (_wk_rs and float(_wk_rs)>100) else ("▼ Lagging" if (_wk_rs and float(_wk_rs)<70) else None),
                       help="RS vs S&P 500 on the weekly chart (13-week lookback).")
            ww3.metric("10WMA > 40WMA",   "✅ Yes" if _wk_10gt40 else "❌ No",
                       help="Weekly golden cross — 10WMA above 40WMA = confirmed weekly uptrend.")
            ww4.metric("Weekly Base",     f"{_wk_depth:.0f}% deep" if _wk_depth else "–",
                       delta="Tight ✅" if _wk_tight else ("Deep ⚠️" if _wk_depth and _wk_depth > 25 else None),
                       help="Depth of the last 8-week price range. <15% = tight base = pre-breakout coiling.")
            ww5.metric("Consec. Up Wks",  f"{int(_wk_consec or 0)}",
                       delta="Strong 🚀" if int(_wk_consec or 0) >= 3 else None,
                       help="Consecutive weeks closing higher. 3+ = sustained institutional buying.")

            ww6,ww7,ww8 = st.columns(3)
            ww6.metric("Weekly HH/HL",    "✅ Yes" if _wk_hh_hl else "No",
                       help="Weekly Higher Highs and Higher Lows — textbook uptrend on the important timeframe.")
            ww7.metric("Weekly Score",    f"{int(_wk_score or 0)}/10",
                       help="Weekly contribution to Apex Score. +0–10 when confirmed, –15 when contradicting.")
            ww8.metric("Above 40WMA",     "✅ Yes" if _wk40 else "❌ No",
                       help="Price above 40-week MA (equivalent to 200-day MA on weekly). Non-negotiable for swing longs.")

            st.markdown("---")

            # ── KPI row 3: Stage & Volume ─────────────────────────────────
            st.markdown("#### 📐 Stage Analysis & Volume")
            k11,k12,k13,k14,k15 = st.columns(5)
            k11.metric("Stage",          result.get("stage","–"))
            k12.metric("Above 50MA",     "✅ Yes" if result.get("above_50ma") else "❌ No")
            k13.metric("Above 200MA",    "✅ Yes" if result.get("above_200ma") else "❌ No")
            k14.metric("50MA > 200MA",   "✅ Yes" if result.get("ma50_gt_ma200") else "❌ No")
            k15.metric("Vol Surge",      f"{result.get('vol_surge_x',1):.1f}x")

            st.markdown("---")

            # ── KPI row 4: Order Flow Persistence (NEW) ───────────────────
            st.markdown("#### 🌊 Order Flow Persistence")
            o1,o2,o3,o4,o5 = st.columns(5)
            o1.metric("Directional Bias",   result.get("of_bias","–"))
            o2.metric("Up Vol Ratio",       f"{result.get('of_up_vol_ratio',1):.2f}x")
            o3.metric("Bullish Days (10d)", f"{result.get('of_bullish_days',0):.0f}%")
            o4.metric("Consec Up Closes",   str(result.get("of_consec_up",0)))
            o5.metric("OF Score",           f"{result.get('of_score',0)}/8")
            st.caption("Up Vol Ratio > 1.5x = sustained institutional buying pressure across multiple sessions")

            st.markdown("---")

            # ── KPI row 5: VWAP / Auction Market Theory (NEW) ────────────
            st.markdown("#### 💧 VWAP & Auction Market Theory")
            v1,v2,v3,v4,v5 = st.columns(5)
            vwap_val = result.get("vwap")
            v1.metric("VWAP (20d)",      f"${vwap_val:.2f}" if vwap_val else "–")
            v2.metric("vs VWAP",         pct_fmt(result.get("vs_vwap_%",0)))
            v3.metric("VWAP Position",   result.get("vwap_position","–"))
            v4.metric("VWAP Slope",      result.get("vwap_slope","–"))
            v5.metric("VWAP Score",      f"{result.get('vwap_score',0)}/4")
            vwap_pos = result.get("vwap_position","")
            if "Above" in vwap_pos and "Extended" not in vwap_pos:
                st.markdown('<div class="alert-box">✅ Price above VWAP — buyers in control, value accepted higher</div>', unsafe_allow_html=True)
            elif "Extended Above" in vwap_pos:
                st.markdown('<div class="warn-box">⚡ Extended above VWAP — momentum strong but watch for mean reversion</div>', unsafe_allow_html=True)
            elif "Extended Below" in vwap_pos:
                st.markdown('<div class="danger-box">🔴 Extended below VWAP — sellers in control, avoid new longs</div>', unsafe_allow_html=True)

            st.markdown("---")

            # ── KPI row 6: Market Structure (NEW) ────────────────────────
            st.markdown("#### 🏗 Market Structure")
            m1,m2,m3,m4,m5 = st.columns(5)
            m1.metric("Structure",          result.get("ms_structure","–"))
            m2.metric("HH / HL",            "✅ Yes" if result.get("ms_hh_hl") else "❌ No")
            m3.metric("Break of Structure",  "🚨 Yes" if result.get("ms_bos") else "No")
            m4.metric("Last Swing High",     f"${result.get('ms_swing_high',0):.2f}" if result.get("ms_swing_high") else "–")
            m5.metric("Last Swing Low",      f"${result.get('ms_swing_low',0):.2f}"  if result.get("ms_swing_low")  else "–")

            st.markdown("---")

            # ── KPI row 7: Price Action Patterns (NEW) ───────────────────
            st.markdown("#### 🕯 Price Action Patterns")
            _pa_raw2 = result.get("pa_patterns", "None")
            pa_patterns = (
                str(_pa_raw2)
                if _pa_raw2 is not None
                and not (isinstance(_pa_raw2, float) and pd.isna(_pa_raw2))
                else "None"
            )
            if pa_patterns and pa_patterns not in ("None", "nan", ""):
                pa_list = [p.strip() for p in pa_patterns.split("|")]
                pa_cols = st.columns(min(len(pa_list),4))
                pa_colors = {
                    "Bullish SFP (Bear Trap)":  "#3fb950",
                    "Bullish Engulfing":         "#3fb950",
                    "PA Confluence":             "#d29922",
                    "Inside Day (Compression)":  "#388bfd",
                    "Bullish Context Candle":    "#3fb950",
                    "Bearish SFP (Bull Trap)":   "#f85149",
                    "Bearish Engulfing":         "#f85149",
                    "Bearish Context Candle":    "#f85149",
                }
                for i, p in enumerate(pa_list[:4]):
                    col_c = pa_colors.get(p, "#8b949e")
                    pa_cols[i].markdown(
                        f'<div style="background:#161b22;border:1px solid {col_c};'
                        f'border-radius:8px;padding:12px;text-align:center;">'
                        f'<div style="color:{col_c};font-weight:700;font-size:0.9rem;">{p}</div>'
                        f'</div>', unsafe_allow_html=True)
                if "Bullish SFP" in pa_patterns:
                    st.markdown('<div class="alert-box">🎯 <b>Bullish SFP:</b> Price wicked below a swing low trapping shorts, then closed above — bear trap, smart money reversal signal.</div>', unsafe_allow_html=True)
                elif "Bearish SFP" in pa_patterns:
                    st.markdown('<div class="danger-box">🎯 <b>Bearish SFP:</b> Price wicked above a swing high trapping longs, then closed below — bull trap, potential reversal lower.</div>', unsafe_allow_html=True)
                if "PA Confluence" in pa_patterns:
                    st.markdown('<div class="warn-box">⚡ <b>PA Confluence:</b> Multiple signals aligning — higher probability setup.</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div style="background:#161b22;border:1px solid #30363d;border-radius:8px;padding:16px;color:#8b949e;text-align:center;">No significant PA patterns on last candle</div>', unsafe_allow_html=True)
            st.metric("PA Score", f"{result.get('pa_score',0)}/5")

            st.markdown("---")

            # ── KPI row 8: Alpha Vantage Fundamentals (NEW) ─────────────
            st.markdown("#### 📊 Real Earnings & Fundamentals (Alpha Vantage)")
            eps_growth  = result.get("eps_growth_%")
            eps_surp    = result.get("eps_surprise_%")
            eps_accel   = result.get("eps_accel")
            beats       = result.get("consec_beats")
            rev_g       = result.get("rev_growth_%")
            eps_sc      = result.get("eps_score", 0)
            analyst_t   = result.get("analyst_target")
            pe          = result.get("pe_ratio")
            peg         = result.get("peg_ratio")
            eps_det     = result.get("eps_details","–")
            eps_trend   = result.get("eps_trend",[])

            if eps_growth is None and eps_surp is None:
                st.markdown('<div class="warn-box">⚠️ Alpha Vantage key not configured or quota reached. '
                            'Add your key to config.yaml under <code>alpha_vantage_key</code>.</div>',
                            unsafe_allow_html=True)
            else:
                av1,av2,av3,av4,av5 = st.columns(5)
                eps_c = "#3fb950" if (eps_growth or 0) > 15 else ("#d29922" if (eps_growth or 0) > 0 else "#f85149")
                av1.metric("EPS Growth YoY",   f"{eps_growth:+.1f}%" if eps_growth is not None else "–")
                av2.metric("Last Surprise",     f"{eps_surp:+.1f}%"  if eps_surp  is not None else "–")
                av3.metric("Accelerating",      "✅ Yes" if eps_accel else "❌ No")
                av4.metric("Consec Beats",      f"{beats}Q" if beats is not None else "–")
                av5.metric("Revenue Growth",    f"{rev_g:+.1f}%"     if rev_g     is not None else "–")

                av6,av7,av8,av9,av10 = st.columns(5)
                av6.metric("EPS Score",  f"{eps_sc}/15")
                av7.metric("Analyst Target", f"${analyst_t:.2f}" if analyst_t else "–")
                av8.metric("PE Ratio",   f"{pe:.1f}x" if pe else "–")
                av9.metric("PEG Ratio",  f"{peg:.2f}"  if peg else "–")
                av10.metric("Earn Momentum", result.get("earn_momentum","–"))

                # EPS trend sparkline
                if eps_trend:
                    pass  # go already imported at top
                    fig_eps = go.Figure(go.Scatter(
                        x=[f"Q-{len(eps_trend)-i}" for i in range(len(eps_trend))],
                        y=list(reversed(eps_trend)),
                        mode="lines+markers+text",
                        text=[f"${v:.2f}" for v in reversed(eps_trend)],
                        textposition="top center",
                        line=dict(color="#3fb950", width=2),
                        marker=dict(size=8, color="#3fb950"),
                    ))
                    fig_eps.update_layout(
                        title="EPS Trend (last 4 quarters)",
                        paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                        font_color="#e6edf3", height=200,
                        xaxis=dict(gridcolor="#21262d"),
                        yaxis=dict(gridcolor="#21262d", tickprefix="$"),
                        margin=dict(l=10,r=10,t=40,b=20),
                    )
                    st.plotly_chart(fig_eps, use_container_width=True)

                st.caption(f"Details: {eps_det}")

                # EPS quality signals
                if eps_accel and (beats or 0) >= 3 and (eps_growth or 0) > 25:
                    st.markdown('<div class="alert-box">🏆 <b>Exceptional EPS quality:</b> '
                                'Accelerating growth + 3+ consecutive beats + >25% YoY growth. '
                                'This is the kind of fundamental profile that precedes major price moves.</div>',
                                unsafe_allow_html=True)
                elif (eps_surp or 0) > 20:
                    st.markdown(f'<div class="alert-box">⚡ <b>Large earnings beat:</b> '
                                f'Beat estimates by {eps_surp:.1f}% last quarter — '
                                f'institutional re-rating often follows large positive surprises.</div>',
                                unsafe_allow_html=True)
                elif (eps_growth or 0) < 0:
                    st.markdown('<div class="danger-box">🔴 <b>Declining EPS:</b> '
                                'Earnings are shrinking year-over-year. '
                                'Strong price momentum with declining fundamentals is a yellow flag.</div>',
                                unsafe_allow_html=True)

                # Upside to analyst target
                if analyst_t and result.get("price"):
                    upside = (analyst_t / result["price"] - 1) * 100
                    target_color = "alert-box" if upside > 10 else ("warn-box" if upside > 0 else "danger-box")
                    st.markdown(f'<div class="{target_color}">📍 <b>Analyst Target:</b> ${analyst_t:.2f} '
                                f'= <b>{upside:+.1f}% upside</b> from current price ${result["price"]:.2f}</div>',
                                unsafe_allow_html=True)

            st.markdown("---")

            # ── Pattern + Earnings ────────────────────────────────────────
            p1, p2 = st.columns(2)
            with p1:
                pattern = result.get("pattern","–")
                breaking = result.get("breaking_out", False)
                pat_color = "#3fb950" if breaking else "#d29922"
                st.markdown(f"""
                <div style="background:#161b22;border:1px solid {pat_color};
                            border-radius:8px;padding:16px 20px;">
                  <div style="color:#8b949e;font-size:0.75rem;text-transform:uppercase;">
                    Chart Pattern
                  </div>
                  <div style="font-size:1.3rem;font-weight:700;color:{pat_color};margin-top:6px;">
                    {pattern}
                  </div>
                  <div style="color:#8b949e;font-size:0.8rem;margin-top:4px;">
                    {'Active breakout — watch closely' if breaking else 'Not yet breaking out'}
                  </div>
                </div>
                """, unsafe_allow_html=True)
            with p2:
                earn = result.get("earn_momentum","–")
                earn_color = "#3fb950" if earn=="Strong" else ("#d29922" if earn=="Moderate" else "#8b949e")
                near = "✅ Near 52W High" if result.get("near_52wh") else "❌ Below 52W High"
                st.markdown(f"""
                <div style="background:#161b22;border:1px solid #30363d;
                            border-radius:8px;padding:16px 20px;">
                  <div style="color:#8b949e;font-size:0.75rem;text-transform:uppercase;">
                    Earnings Momentum
                  </div>
                  <div style="font-size:1.3rem;font-weight:700;color:{earn_color};margin-top:6px;">
                    {earn}
                  </div>
                  <div style="color:#8b949e;font-size:0.8rem;margin-top:4px;">{near}</div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("---")

            # ── Chart ─────────────────────────────────────────────────────
            st.markdown("#### 📈 Price Chart")
            hist_d = fetch_hist(dive_ticker, "1y")
            if not hist_d.empty:
                fig_d = make_subplots(rows=2, cols=1, shared_xaxes=True,
                                      row_heights=[0.72, 0.28], vertical_spacing=0.04)
                fig_d.add_trace(go.Candlestick(
                    x=hist_d.index,
                    open=hist_d["Open"], high=hist_d["High"],
                    low=hist_d["Low"],   close=hist_d["Close"],
                    name="Price",
                    increasing_line_color="#3fb950",
                    decreasing_line_color="#f85149",
                ), row=1, col=1)
                fig_d.add_trace(go.Scatter(
                    x=hist_d.index, y=hist_d["MA50"],
                    line=dict(color="#d29922", width=1.5), name="50 MA"
                ), row=1, col=1)
                fig_d.add_trace(go.Scatter(
                    x=hist_d.index, y=hist_d["MA200"],
                    line=dict(color="#388bfd", width=1.5, dash="dot"), name="200 MA"
                ), row=1, col=1)
                vol_c = ["#3fb950" if hist_d["Close"].iloc[i] >= hist_d["Open"].iloc[i]
                         else "#f85149" for i in range(len(hist_d))]
                fig_d.add_trace(go.Bar(
                    x=hist_d.index, y=hist_d["Volume"],
                    marker_color=vol_c, name="Volume", opacity=0.7
                ), row=2, col=1)
                fig_d.update_layout(
                    paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                    font_color="#e6edf3", xaxis_rangeslider_visible=False,
                    height=550, margin=dict(l=10,r=10,t=20,b=20),
                    legend=dict(orientation="h", y=1.02),
                )
                fig_d.update_yaxes(gridcolor="#21262d")
                fig_d.update_xaxes(gridcolor="#21262d")
                st.plotly_chart(fig_d, use_container_width=True)

            # ── Plain English Summary ─────────────────────────────────────
            st.markdown("---")
            st.markdown("#### 💬 Plain English Summary")

            bullish = []
            bearish = []

            if result.get("above_200ma"):   bullish.append("Trading above the 200-day moving average (long-term uptrend)")
            else:                           bearish.append("Below the 200-day moving average — not in a confirmed uptrend")
            if result.get("ma50_gt_ma200"): bullish.append("50MA is above 200MA — classic Stage 2 uptrend structure")
            else:                           bearish.append("50MA is below 200MA — not yet in Stage 2")
            if rs3 and rs3 > 100:           bullish.append(f"RS score of {rs3:.0f} — beating the benchmark strongly")
            elif rs3 and rs3 > 70:          bullish.append(f"RS score of {rs3:.0f} — keeping pace with the market")
            else:                           bearish.append(f"RS score of {rs3:.0f} — lagging behind the market")
            if result.get("near_52wh"):     bullish.append("Near its 52-week high — price strength confirmed")
            else:                           bearish.append(f"{abs(result.get('pct_off_high_%',0)):.0f}% off 52-week high — needs to reclaim highs")
            if result.get("breaking_out"):  bullish.append(f"Active breakout pattern: {pattern}")
            if result.get("vol_surge_x",1) >= 1.4: bullish.append(f"Volume surge of {result.get('vol_surge_x',1):.1f}x — institutional interest")
            if result.get("perf_3m_%",0) > 15: bullish.append(f"3-month return of +{result.get('perf_3m_%',0):.1f}% — strong momentum")
            else:                           bearish.append(f"3-month return of {result.get('perf_3m_%',0):.1f}% — below momentum threshold")

            if bullish:
                for b in bullish:
                    st.markdown(f'<div class="alert-box">✅ {b}</div>', unsafe_allow_html=True)
            if bearish:
                for b in bearish:
                    st.markdown(f'<div class="danger-box">❌ {b}</div>', unsafe_allow_html=True)

    elif not dive_btn:
        st.markdown("""
        <div style="background:#161b22;border:1px solid #30363d;border-radius:8px;
                    padding:32px;text-align:center;color:#8b949e;">
          <div style="font-size:2rem;margin-bottom:8px;">🔍</div>
          <div style="font-size:1rem;">Type any ticker above and click <b>Analyse</b></div>
          <div style="font-size:0.85rem;margin-top:8px;">
            Works for any US stock: HIMS, HOOD, CELH, SOFI, DUOL, ONON…
          </div>
        </div>
        """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 8 — OPTIONS FLOW
# ══════════════════════════════════════════════════════════════════════════════

with tabs[7]:
    st.markdown("### 🎯 Options Flow Scanner")
    st.caption("Scans for unusual options activity — large bets placed by smart money via yfinance options chain.")


    opt_tickers_default = df["ticker"].tolist()[:8] if not df.empty else ["NVDA","AAPL","TSLA","AMZN"]
    cfg_o = load_config("config.yaml")
    all_tickers_o = list(set(t for theme in cfg_o["us_themes"].values() for t in theme))

    oc1, oc2, oc3 = st.columns([2, 1, 1])
    with oc1:
        opt_mode = st.radio("Scan mode", ["Single ticker", "Scan my watchlist"], horizontal=True)
    with oc2:
        opt_min_vol = st.number_input("Min contract volume", value=200, step=100)
    with oc3:
        opt_btn = st.button("🎯 Scan Options Flow", use_container_width=True)

    if opt_mode == "Single ticker":
        opt_single = st.text_input("Ticker", placeholder="e.g. NVDA", key="opt_single").upper().strip()
    else:
        opt_single = None

    if opt_btn:
        with st.spinner("Fetching options chain data…"):
            if opt_mode == "Single ticker" and opt_single:
                opt_df = scan_options_flow(opt_single, min_volume=int(opt_min_vol))
            else:
                scan_tix = opt_tickers_default[:6]
                opt_df = scan_multiple(scan_tix, min_volume=int(opt_min_vol))

        if opt_df is None or opt_df.empty:
            st.warning("No unusual options activity found. Try lowering Min Volume or choosing a different ticker.")
        else:
            unusual = opt_df[opt_df["unusual"] == True]
            normal  = opt_df[opt_df["unusual"] == False]

            # KPIs
            ok1, ok2, ok3, ok4 = st.columns(4)
            with ok1: st.markdown(f'<div class="metric-card"><h3>Total Contracts</h3><div class="value white">{len(opt_df)}</div></div>', unsafe_allow_html=True)
            with ok2: st.markdown(f'<div class="metric-card"><h3>Unusual Activity</h3><div class="value amber">{len(unusual)}</div></div>', unsafe_allow_html=True)
            with ok3:
                calls = int((opt_df["type"] == "CALL").sum())
                st.markdown(f'<div class="metric-card"><h3>Calls (Bullish)</h3><div class="value green">{calls}</div></div>', unsafe_allow_html=True)
            with ok4:
                puts = int((opt_df["type"] == "PUT").sum())
                st.markdown(f'<div class="metric-card"><h3>Puts (Bearish)</h3><div class="value red">{puts}</div></div>', unsafe_allow_html=True)

            st.markdown("---")

            if not unusual.empty:
                st.markdown("#### 🚨 Unusual Activity — Possible Smart Money Moves")
                for _, row in unusual.head(10).iterrows():
                    color = "#3fb950" if "Bullish" in str(row.get("sentiment","")) else "#f85149"
                    tk_label = f" [{row['ticker']}]" if "ticker" in row and row["ticker"] else ""
                    st.markdown(
                        f'<div style="background:#161b22;border:1px solid {color};border-radius:8px;'
                        f'padding:12px 16px;margin:4px 0;">'
                        f'<b>{row["type"]}{tk_label}</b> — Strike ${row["strike"]} | '
                        f'Exp: {row.get("expiry","?")} | '
                        f'Vol: {row["volume"]:,} | OI: {row["open_interest"]:,} | '
                        f'Vol/OI: {row["vol/OI"]}x | IV: {row["IV_%"]}% | '
                        f'Notional: ${row["notional_$"]:,.0f} — <b style="color:{color}">{row["sentiment"]}</b>'
                        f'</div>',
                        unsafe_allow_html=True
                    )
                st.markdown("---")

            st.markdown("#### Full Options Chain")
            display_cols = [c for c in ["ticker","type","expiry","strike","spot","moneyness_%",
                            "volume","open_interest","vol/OI","IV_%","last_price",
                            "notional_$","unusual","sentiment"] if c in opt_df.columns]
            st.dataframe(opt_df[display_cols], use_container_width=True, hide_index=True, height=400)
    else:
        st.info("Select a ticker or watchlist scan and click **🎯 Scan Options Flow**.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 9 — INSIDER TRACKER
# ══════════════════════════════════════════════════════════════════════════════

with tabs[8]:
    st.markdown("### 🕵️ Insider Trading Tracker")
    st.caption("SEC Form 4 filings via OpenInsider — when executives buy their own stock, pay attention.")


    ic1, ic2, ic3 = st.columns([2, 1, 1])
    with ic1:
        ins_mode = st.radio("Mode", ["Single ticker", "Scan watchlist"], horizontal=True, key="ins_mode")
    with ic2:
        ins_days = st.selectbox("Look back", [7, 14, 30, 60, 90], index=2, key="ins_days")
    with ic3:
        ins_btn = st.button("🕵️ Fetch Insider Data", use_container_width=True)

    if ins_mode == "Single ticker":
        ins_ticker = st.text_input("Ticker", placeholder="e.g. TSM", key="ins_ticker").upper().strip()
    else:
        ins_ticker = None

    if ins_btn:
        with st.spinner("Fetching SEC Form 4 data from OpenInsider…"):
            if ins_mode == "Single ticker" and ins_ticker:
                ins_df = fetch_insider_trades(ticker=ins_ticker, days_back=ins_days, trade_type="P")
                if not ins_df.empty:
                    st.success(f"Found {len(ins_df)} insider purchase(s) for {ins_ticker}")
            else:
                cfg_i = load_config("config.yaml")
                watch_i = list(set(t for theme in cfg_i["us_themes"].values() for t in theme))
                ins_df = get_insider_summary(watch_i[:20], days_back=ins_days)
                if not ins_df.empty:
                    st.success(f"Insider summary across {len(ins_df)} tickers")

        if ins_df is None or ins_df.empty:
            st.warning("No insider purchases found in this period. Try extending the look-back window.")
        else:
            # Highlight cluster buys
            if "cluster_buy" in ins_df.columns:
                clusters = ins_df[ins_df["cluster_buy"] == True]
                if not clusters.empty:
                    st.markdown("#### 🔥 Cluster Buys — Multiple Insiders Buying Same Stock")
                    for _, row in clusters.iterrows():
                        tk = row.get("ticker","")
                        val = row.get("total_$", row.get("value", 0))
                        n   = row.get("insiders", row.get("insider_count", "?"))
                        st.markdown(
                            f'<div class="alert-box">🔥 <b>{tk}</b> — '
                            f'{n} insiders bought | Total: ${val:,.0f}</div>',
                            unsafe_allow_html=True
                        )
                    st.markdown("---")

            st.markdown("#### All Insider Purchases")
            st.dataframe(ins_df, use_container_width=True, hide_index=True, height=450)

            st.markdown("""
            ---
            **How to use this:**
            - 🔥 **Cluster buys** (2+ insiders) = strongest possible signal — they rarely all buy at the same time unless they see value
            - Single executive buys = worth noting, especially C-suite (CEO, CFO, COO)
            - Insider *sales* are less meaningful — they sell for many reasons (taxes, diversification)
            - Combine with a high Apex Score for maximum conviction
            """)
    else:
        st.info("Select a mode and click **🕵️ Fetch Insider Data**.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 10 — DIVIDEND REINVESTMENT CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════

with tabs[9]:
    st.markdown("### 📊 Dividend Reinvestment Calculator (DRIP)")
    st.caption("Shows the compounding power of reinvesting dividends over time.")

    @st.cache_data(ttl=3600)
    def get_dividend_info(ticker: str) -> dict:
        try:
            info = yf.Ticker(ticker).info
            return {
                "div_yield":    info.get("dividendYield", 0) or 0,
                "div_annual":   info.get("dividendRate", 0) or 0,
                "payout_ratio": info.get("payoutRatio", 0) or 0,
                "name":         info.get("longName", ticker),
                "sector":       info.get("sector", "–"),
            }
        except:
            return {}

    da1, da2 = st.columns(2)
    with da1:
        drip_ticker  = st.text_input("Ticker (optional — for live yield)", placeholder="e.g. AAPL, JNJ, KO", key="drip_tk").upper().strip()
        initial_inv  = st.number_input("Initial Investment ($)", min_value=100.0, value=10000.0, step=500.0)
        monthly_add  = st.number_input("Monthly Contribution ($)", min_value=0.0, value=200.0, step=50.0)
        years        = st.slider("Investment Period (Years)", 1, 40, 10)
    with da2:
        if drip_ticker:
            dinfo = get_dividend_info(drip_ticker)
            default_yield = round(dinfo.get("div_yield", 0.03) * 100, 2)
            default_growth = 5.0
            if dinfo.get("name"):
                st.info(f"**{dinfo['name']}** | Sector: {dinfo.get('sector','–')} | "
                        f"Div Yield: {default_yield:.2f}% | Payout: {dinfo.get('payout_ratio',0)*100:.0f}%")
        else:
            default_yield  = 3.0
            default_growth = 5.0

        div_yield_pct  = st.number_input("Annual Dividend Yield (%)", min_value=0.0, max_value=20.0,
                                          value=float(default_yield), step=0.1)
        price_growth   = st.number_input("Expected Annual Price Growth (%)", min_value=-5.0, max_value=30.0,
                                          value=float(default_growth), step=0.5)
        drip_on        = st.checkbox("Reinvest dividends (DRIP)", value=True)
        tax_rate       = st.slider("Dividend Tax Rate (%)", 0, 40, 15)

    calc_btn = st.button("📊 Calculate", use_container_width=True)

    if calc_btn:
        # ── Simulation ────────────────────────────────────────────────────
        div_rate     = div_yield_pct / 100
        growth_rate  = price_growth / 100
        tax          = tax_rate / 100
        months       = years * 12
        monthly_rate = growth_rate / 12
        monthly_div  = div_rate / 12

        records    = []
        value      = initial_inv
        shares     = initial_inv  # treat as $ units for simplicity
        total_divs = 0
        total_cont = initial_inv

        for m in range(1, months + 1):
            # Price appreciation
            value *= (1 + monthly_rate)
            # Dividend
            div_income = value * monthly_div
            after_tax  = div_income * (1 - tax)
            total_divs += after_tax
            if drip_on:
                value += after_tax
            # Monthly contribution
            value      += monthly_add
            total_cont += monthly_add

            if m % 12 == 0:
                records.append({
                    "Year":            m // 12,
                    "Portfolio Value": round(value, 2),
                    "Total Invested":  round(total_cont, 2),
                    "Total Dividends": round(total_divs, 2),
                    "Gain":            round(value - total_cont, 2),
                    "Gain %":          round((value / total_cont - 1) * 100, 1),
                })

        result_df = pd.DataFrame(records)

        # KPIs
        final_val  = result_df["Portfolio Value"].iloc[-1]
        total_inv  = result_df["Total Invested"].iloc[-1]
        total_div  = result_df["Total Dividends"].iloc[-1]
        total_gain = final_val - total_inv

        k1,k2,k3,k4 = st.columns(4)
        with k1: st.markdown(f'<div class="metric-card"><h3>Final Value</h3><div class="value green">${final_val:,.0f}</div></div>', unsafe_allow_html=True)
        with k2: st.markdown(f'<div class="metric-card"><h3>Total Invested</h3><div class="value white">${total_inv:,.0f}</div></div>', unsafe_allow_html=True)
        with k3: st.markdown(f'<div class="metric-card"><h3>Total Dividends</h3><div class="value amber">${total_div:,.0f}</div></div>', unsafe_allow_html=True)
        with k4: st.markdown(f'<div class="metric-card"><h3>Total Gain</h3><div class="value green">${total_gain:,.0f} ({(final_val/total_inv-1)*100:.0f}%)</div></div>', unsafe_allow_html=True)

        st.markdown("---")

        # Chart
        fig_drip = go.Figure()
        fig_drip.add_trace(go.Scatter(x=result_df["Year"], y=result_df["Portfolio Value"],
            name="Portfolio Value", line=dict(color="#3fb950", width=2.5), fill="tozeroy",
            fillcolor="rgba(63,185,80,0.08)"))
        fig_drip.add_trace(go.Scatter(x=result_df["Year"], y=result_df["Total Invested"],
            name="Total Invested", line=dict(color="#388bfd", width=2, dash="dot")))
        fig_drip.add_trace(go.Bar(x=result_df["Year"], y=result_df["Total Dividends"],
            name="Cumulative Dividends", marker_color="#d29922", opacity=0.5))
        fig_drip.update_layout(
            title=f"{'DRIP' if drip_on else 'No DRIP'} — {years}-Year Projection",
            paper_bgcolor="#0d1117", plot_bgcolor="#0d1117", font_color="#e6edf3",
            xaxis=dict(title="Year", gridcolor="#21262d"),
            yaxis=dict(title="Value ($)", gridcolor="#21262d", tickprefix="$"),
            height=400, legend=dict(orientation="h", y=1.05),
            margin=dict(l=10,r=10,t=50,b=20),
        )
        st.plotly_chart(fig_drip, use_container_width=True)

        st.markdown("#### Year-by-Year Breakdown")
        styled_d = result_df.style.format({
            "Portfolio Value": "${:,.0f}",
            "Total Invested":  "${:,.0f}",
            "Total Dividends": "${:,.0f}",
            "Gain":            "${:,.0f}",
            "Gain %":          "{:.1f}%",
        })
        st.dataframe(styled_d, use_container_width=True, hide_index=True)

        # DRIP vs No-DRIP comparison
        if drip_on:
            st.info(f"💡 By reinvesting dividends, your portfolio reaches **${final_val:,.0f}** — "
                    f"dividends contributed **${total_div:,.0f}** to your final wealth through compounding.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 11 — BACKTESTER
# ══════════════════════════════════════════════════════════════════════════════

with tabs[10]:
    st.markdown("### ⏱ Strategy Backtester")
    st.caption("Test the Apex Score strategy on historical data. Did it work? By how much?")


    ba1, ba2 = st.columns(2)
    with ba1:
        bt_mode       = st.radio("Mode", ["Single ticker", "Full watchlist"], horizontal=True)
        bt_ticker_in  = st.text_input("Ticker (single mode)", value="NVDA", key="bt_tk").upper().strip()
        bt_start      = st.date_input("Start Date", value=datetime(2021,1,1))
        bt_end        = st.date_input("End Date",   value=datetime(2024,12,31))
    with ba2:
        bt_entry_score = st.slider("Entry Score Threshold", 20, 80, 40, 5)
        bt_hold_days   = st.slider("Max Hold Days", 10, 120, 60, 5)
        st.markdown("""
        **Entry:** Apex Score ≥ threshold + Stage 2 confirmed
        **Exit:** Price closes below 50 MA or max hold days reached
        """)

    st.markdown("##### 🔬 Optional New Signal Filters")
    bf1, bf2, bf3, bf4 = st.columns(4)
    with bf1: bt_req_of    = st.checkbox("Require Persistent OF Bias",   value=False)
    with bf2: bt_req_vwap  = st.checkbox("Require Price Above VWAP",     value=False)
    with bf3: bt_req_hh_hl = st.checkbox("Require HH/HL Structure",      value=False)
    with bf4: bt_req_pa    = st.checkbox("Require PA Pattern (Engulf/Context)", value=False)
    st.caption("Tick these to test how much each new signal improves results vs base strategy alone.")

    bt_btn = st.button("⏱ Run Backtest", use_container_width=True)

    if bt_btn:
        start_str = bt_start.strftime("%Y-%m-%d")
        end_str   = bt_end.strftime("%Y-%m-%d")

        with st.spinner("Running backtest… (may take 1–2 min for full watchlist)"):
            if bt_mode == "Single ticker":
                bt_result = backtest_ticker(bt_ticker_in, start_str, end_str,
                                             bt_entry_score, True, bt_hold_days,
                                             bt_req_of, bt_req_vwap, bt_req_hh_hl, bt_req_pa)
                if "error" in bt_result:
                    st.error(bt_result["error"])
                elif not bt_result.get("trades"):
                    st.warning("No trades triggered. Try lowering the entry score or widening the date range.")
                else:
                    s = bt_result["summary"]
                    trades_df = pd.DataFrame(bt_result["trades"])

                    bk1,bk2,bk3,bk4,bk5 = st.columns(5)
                    with bk1: st.markdown(f'<div class="metric-card"><h3>Total Trades</h3><div class="value white">{s["total_trades"]}</div></div>', unsafe_allow_html=True)
                    wc = "green" if s["win_rate_%"] >= 50 else "red"
                    with bk2: st.markdown(f'<div class="metric-card"><h3>Win Rate</h3><div class="value {wc}">{s["win_rate_%"]}%</div></div>', unsafe_allow_html=True)
                    ac = "green" if s["avg_return_%"] >= 0 else "red"
                    with bk3: st.markdown(f'<div class="metric-card"><h3>Avg Return</h3><div class="value {ac}">{s["avg_return_%"]:+.1f}%</div></div>', unsafe_allow_html=True)
                    with bk4: st.markdown(f'<div class="metric-card"><h3>Best Trade</h3><div class="value green">{s["best_trade_%"]:+.1f}%</div></div>', unsafe_allow_html=True)
                    with bk5: st.markdown(f'<div class="metric-card"><h3>Worst Trade</h3><div class="value red">{s["worst_trade_%"]:+.1f}%</div></div>', unsafe_allow_html=True)

                    st.markdown("---")

                    # Equity curve
                    trades_df["cumulative_return_%"] = (1 + trades_df["return_%"] / 100).cumprod() * 100 - 100
                    fig_bt = go.Figure()
                    fig_bt.add_trace(go.Scatter(
                        x=trades_df["exit_date"], y=trades_df["cumulative_return_%"],
                        name="Cumulative Return", line=dict(color="#3fb950", width=2),
                        fill="tozeroy", fillcolor="rgba(63,185,80,0.08)"
                    ))
                    fig_bt.add_hline(y=0, line_color="#30363d")
                    fig_bt.update_layout(
                        title=f"{bt_ticker_in} — Venu Strategy Backtest ({start_str} to {end_str})",
                        paper_bgcolor="#0d1117", plot_bgcolor="#0d1117", font_color="#e6edf3",
                        xaxis=dict(gridcolor="#21262d"),
                        yaxis=dict(gridcolor="#21262d", ticksuffix="%"),
                        height=380, margin=dict(l=10,r=10,t=40,b=20),
                    )
                    st.plotly_chart(fig_bt, use_container_width=True)

                    # Individual trades
                    st.markdown("#### Trade Log")
                    def _color_ret_bt(v):
                        try: return "color:#3fb950" if float(v)>0 else "color:#f85149"
                        except: return ""
                    styled_bt = trades_df.style \
                        .map(_color_ret_bt, subset=["return_%"]) \
                        .format({"entry_price":"${:.2f}","exit_price":"${:.2f}",
                                 "return_%":"{:+.1f}%"})
                    st.dataframe(styled_bt, use_container_width=True, hide_index=True)

            else:
                cfg_bt = load_config("config.yaml")
                tix_bt = list(set(t for th in cfg_bt["us_themes"].values() for t in th))[:15]
                trades_df, agg = backtest_portfolio(tix_bt, start_str, end_str,
                                                     bt_entry_score, bt_hold_days,
                                                     bt_req_of, bt_req_vwap,
                                                     bt_req_hh_hl, bt_req_pa)
                if trades_df.empty:
                    st.warning("No trades triggered across the watchlist.")
                else:
                    bk1,bk2,bk3,bk4 = st.columns(4)
                    with bk1: st.markdown(f'<div class="metric-card"><h3>Total Trades</h3><div class="value white">{agg["total_trades"]}</div></div>', unsafe_allow_html=True)
                    wc = "green" if agg["win_rate_%"] >= 50 else "red"
                    with bk2: st.markdown(f'<div class="metric-card"><h3>Win Rate</h3><div class="value {wc}">{agg["win_rate_%"]}%</div></div>', unsafe_allow_html=True)
                    ac = "green" if agg["avg_return_%"] >= 0 else "red"
                    with bk3: st.markdown(f'<div class="metric-card"><h3>Avg Return/Trade</h3><div class="value {ac}">{agg["avg_return_%"]:+.1f}%</div></div>', unsafe_allow_html=True)
                    with bk4: st.markdown(f'<div class="metric-card"><h3>Best Trade</h3><div class="value green">{agg["best_trade"]}</div></div>', unsafe_allow_html=True)

                    st.markdown("---")
                    st.markdown("#### All Trades")
                    st.dataframe(trades_df, use_container_width=True, hide_index=True, height=400)
    else:
        st.info("Configure your backtest above and click **⏱ Run Backtest**.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 12 — RISK CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════

with tabs[11]:
    st.markdown("### ⚖️ Position Size & Risk Calculator")
    st.caption("Never risk more than 1–2% of your account on any single trade. This calculator tells you exactly how many shares to buy.")


    rc1, rc2 = st.columns(2)
    with rc1:
        st.markdown("#### Trade Parameters")
        rc_account  = st.number_input("Account Size ($)", min_value=100.0, value=10000.0, step=500.0)
        rc_risk_pct = st.slider("Max Risk per Trade (%)", 0.25, 5.0, 1.0, 0.25)
        rc_entry    = st.number_input("Entry Price ($)", min_value=0.01, value=100.0, step=0.01)
        rc_stop     = st.number_input("Stop Loss Price ($)", min_value=0.01, value=93.0, step=0.01)
        rc_target   = st.number_input("Profit Target ($ — optional, 0 to skip)", min_value=0.0, value=115.0, step=0.01)
        rc_comm     = st.number_input("Commission per trade ($)", min_value=0.0, value=0.0, step=0.5)

    with rc2:
        st.markdown("#### What is this?")
        st.markdown("""
        **The #1 rule of trading:** Control how much you lose, not just how much you gain.

        This calculator uses the formula:
        > **Shares = (Account × Risk%) ÷ (Entry − Stop)**

        **Example:** $10,000 account, 1% risk, entry $100, stop $95:
        > Shares = ($10,000 × 1%) ÷ ($100 − $95) = $100 ÷ $5 = **20 shares**

        If the trade hits your stop, you lose exactly $100 — 1% of your account.

        **Reward:Risk (R:R):**
        - Minimum acceptable: **2:1** (risk $1 to make $2)
        - Ideal: **3:1 or higher**
        """)

    calc_r_btn = st.button("⚖️ Calculate Position", use_container_width=True)

    if calc_r_btn:
        if rc_stop >= rc_entry:
            st.error("Stop loss must be BELOW entry price for a long trade.")
        else:
            try:
                setup = TradeSetup(
                    account_size  = rc_account,
                    risk_pct      = rc_risk_pct,
                    entry_price   = rc_entry,
                    stop_price    = rc_stop,
                    target_price  = rc_target if rc_target > 0 else None,
                    commission    = rc_comm,
                )
                result = calculate_position(setup)

                # Warning banner
                if result.warning:
                    st.markdown(f'<div class="warn-box">⚠️ {result.warning}</div>', unsafe_allow_html=True)
                    st.markdown("")

                # Main result
                rr_color = "#3fb950" if (result.reward_risk or 0) >= 2 else "#f85149"
                st.markdown(f"""
                <div style="background:#161b22;border:2px solid #3fb950;border-radius:12px;
                            padding:24px;margin-bottom:20px;text-align:center;">
                  <div style="font-size:0.85rem;color:#8b949e;text-transform:uppercase;letter-spacing:.1em;">
                    Buy Exactly
                  </div>
                  <div style="font-size:4rem;font-weight:900;color:#3fb950;line-height:1.1;">
                    {result.shares} shares
                  </div>
                  <div style="font-size:1rem;color:#e6edf3;">
                    @ ${rc_entry:.2f} = <b>${result.position_size:,.2f}</b> total position
                  </div>
                </div>
                """, unsafe_allow_html=True)

                # KPI row
                k1,k2,k3,k4,k5 = st.columns(5)
                with k1: st.markdown(f'<div class="metric-card"><h3>Max Loss $</h3><div class="value red">${result.risk_per_trade:,.2f}</div></div>', unsafe_allow_html=True)
                with k2: st.markdown(f'<div class="metric-card"><h3>Risk % of Account</h3><div class="value amber">{result.risk_pct_actual:.2f}%</div></div>', unsafe_allow_html=True)
                with k3: st.markdown(f'<div class="metric-card"><h3>Stop Distance</h3><div class="value white">${result.stop_distance:.2f} ({result.stop_pct:.1f}%)</div></div>', unsafe_allow_html=True)
                with k4:
                    rr_val = f"{result.reward_risk:.1f}:1" if result.reward_risk else "–"
                    st.markdown(f'<div class="metric-card"><h3>Reward:Risk</h3><div class="value" style="color:{rr_color}">{rr_val}</div></div>', unsafe_allow_html=True)
                with k5:
                    gain_val = f"+${result.target_gain:,.0f} ({result.target_gain_pct:.1f}%)" if result.target_gain else "–"
                    st.markdown(f'<div class="metric-card"><h3>Target Gain</h3><div class="value green">{gain_val}</div></div>', unsafe_allow_html=True)

                st.markdown("---")

                # Pyramiding plan
                st.markdown("#### 📐 Pyramiding Plan (Adding to a Winner)")
                st.caption("How to add to your position as the stock moves in your favour.")
                pyr_add_pct = st.slider("Add when price rises by (%)", 3, 15, 5, 1)
                pyr_adds    = st.slider("Number of add-ons", 1, 3, 2, 1)

                pyr_plan = pyramiding_plan(setup, add_pct=pyr_add_pct, n_adds=pyr_adds)
                pyr_df   = pd.DataFrame(pyr_plan)
                styled_pyr = pyr_df.style.format({
                    "price":         "${:.2f}",
                    "position_$":    "${:,.2f}",
                    "cumulative_$":  "${:,.2f}",
                })
                st.dataframe(styled_pyr, use_container_width=True, hide_index=True)

                st.markdown(f"""
                ---
                💡 **Breakeven price** (after commissions): **${result.breakeven_price:.4f}**
                
                Set your stop at **${rc_stop:.2f}** immediately after your order fills.
                If {rc_ticker_in if 'rc_ticker_in' in dir() else 'the stock'} hits that level, exit without hesitation — the math protects your account.
                """)

            except ValueError as e:
                st.error(str(e))

    else:
        # Show example
        st.markdown("""
        ---
        #### 💡 Quick Example
        | Parameter | Value |
        |---|---|
        | Account | $10,000 |
        | Risk per trade | 1% = $100 max loss |
        | Entry price | $50.00 |
        | Stop loss | $47.50 (5% below entry) |
        | **Shares to buy** | **$100 ÷ $2.50 = 40 shares** |
        | Position size | 40 × $50 = $2,000 (20% of account) |

        If stop is hit → lose $100 (1%). If target at $57 is hit → gain $280 (2.8:1 R:R) ✅
        """)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 13 — AI DAILY BRIEFING
# ══════════════════════════════════════════════════════════════════════════════

with tabs[12]:
    st.markdown("### 🤖 AI Daily Briefing")
    st.caption("Professional-grade market briefing and per-stock narrative — generated entirely from scan data. No API key required.")

    ai_mode = st.radio(
        "Briefing Type",
        ["📊 Full Scan Briefing", "🔍 Single Stock Narrative"],
        horizontal=True, key="ai_mode_sel"
    )
    st.markdown("---")

    # ── FULL SCAN BRIEFING ────────────────────────────────────────────────────
    if ai_mode == "📊 Full Scan Briefing":
        ab1, ab2 = st.columns([3,1])
        with ab1:
            st.markdown("""
            Generates a complete morning briefing from your latest scan:
            - Market snapshot (total setups, avg score, active breakouts)
            - Top 5 highest-conviction setups with signal summary
            - Active breakouts — time-sensitive entries
            - Sector rotation heatmap — where money is flowing
            - Earnings risk calendar for the next 14 days
            - Liquidity warnings
            """)
        with ab2:
            gen_briefing_btn = st.button("📊 Generate Briefing", use_container_width=True, key="gen_briefing")
            load_briefing_btn = st.button("📂 Load Last Briefing", use_container_width=True, key="load_briefing")

        briefing_text = ""

        if gen_briefing_btn:
            if df.empty:
                st.warning("No scan data loaded. Run a Live Scan or Load Last Report first.")
            else:
                with st.spinner("Analysing scan results…"):
                    briefing_text = generate_scan_briefing(df)
                    # Save briefing to disk
                    _brief_path = Path(__file__).resolve().parent / "data" / "last_briefing.md"
                    _brief_path.parent.mkdir(parents=True, exist_ok=True)
                    try:
                        _brief_path.write_text(briefing_text, encoding="utf-8")
                    except Exception:
                        pass
                st.success("✅ Briefing generated!")

        elif load_briefing_btn:
            _brief_path = Path(__file__).resolve().parent / "data" / "last_briefing.md"
            _tmp_brief  = Path("/tmp/apexscan_briefing.md")
            for _bp in (_brief_path, _tmp_brief):
                if _bp.exists():
                    try:
                        briefing_text = _bp.read_text(encoding="utf-8")
                        break
                    except Exception:
                        pass
            if not briefing_text:
                try:
                    briefing_text = load_latest_briefing() or ""
                except Exception:
                    briefing_text = ""
            if not briefing_text:
                st.info("No saved briefing found. Generate one first.")

        if briefing_text:
            # Render as styled markdown card
            st.markdown(
                f'<div style="background:#0d1117;border:1px solid #30363d;border-radius:12px;'
                f'padding:28px 32px;line-height:1.85;font-size:0.95rem;color:#e6edf3;">'
                f'{briefing_text.replace(chr(10),"<br>").replace("## ","<h3 style=\"color:#388bfd;margin-top:18px;\">").replace("### ","<h4 style=\"color:#d29922;\">").replace("# ","<h2 style=\"color:#e6edf3;\">")}'
                f'</div>',
                unsafe_allow_html=True
            )
            st.markdown("---")
            col_dl1, col_dl2 = st.columns(2)
            with col_dl1:
                st.download_button(
                    "⬇ Download Briefing (Markdown)",
                    data=briefing_text.encode("utf-8"),
                    file_name=f"apexscan_briefing_{datetime.now().strftime('%Y%m%d')}.md",
                    mime="text/markdown",
                )
            with col_dl2:
                if st.button("📲 Send to Telegram", key="send_briefing_tg"):
                    _cur_settings = load_alert_settings()
                    if _cur_settings.get("telegram_token") and _cur_settings.get("telegram_chat_id"):
                        # Send first 4000 chars (Telegram message limit)
                        _msg = briefing_text[:4000] + ("\n…[truncated]" if len(briefing_text) > 4000 else "")
                        _res = dispatch_alert(_cur_settings, _msg, "ApexScan Briefing")
                        if _res.get("telegram"):
                            st.success("✅ Sent to Telegram!")
                        else:
                            st.error("❌ Telegram send failed. Check Alert Settings.")
                    else:
                        st.warning("Configure Telegram in 🔔 Alert Settings first.")

    # ── SINGLE STOCK NARRATIVE ────────────────────────────────────────────────
    else:
        if df.empty:
            st.warning("No scan data loaded. Run a Live Scan or Load Last Report first.")
        else:
            narr_col1, narr_col2 = st.columns([2,1])
            with narr_col1:
                narr_ticker = st.selectbox(
                    "Select Stock",
                    sorted(df["ticker"].dropna().unique().tolist()) if "ticker" in df.columns else [],
                    key="narr_ticker_sel"
                )
            with narr_col2:
                gen_narr_btn = st.button("🔍 Generate Narrative", use_container_width=True, key="gen_narr")

            if gen_narr_btn and narr_ticker:
                narr_row = df[df["ticker"] == narr_ticker]
                if narr_row.empty:
                    st.warning(f"{narr_ticker} not found in scan results.")
                else:
                    with st.spinner(f"Building narrative for {narr_ticker}…"):
                        narr_text = generate_narrative(narr_row.iloc[0])

                    # Display
                    st.markdown(
                        f'<div style="background:#0d1117;border:1px solid #30363d;border-radius:12px;'
                        f'padding:28px 32px;line-height:1.85;font-size:0.95rem;color:#e6edf3;">'
                        f'{narr_text.replace(chr(10),"<br>").replace("## ","<h3 style=\"color:#388bfd;margin-top:18px;\">").replace("### ","<h4 style=\"color:#d29922;\">").replace("# ","<h2 style=\"color:#e6edf3;\">")}'
                        f'</div>',
                        unsafe_allow_html=True
                    )
                    st.markdown("---")

                    col_n1, col_n2 = st.columns(2)
                    with col_n1:
                        st.download_button(
                            f"⬇ Download {narr_ticker} Narrative",
                            data=narr_text.encode("utf-8"),
                            file_name=f"apexscan_{narr_ticker}_{datetime.now().strftime('%Y%m%d')}.md",
                            mime="text/markdown",
                        )
                    with col_n2:
                        if st.button(f"📲 Send {narr_ticker} to Telegram", key="narr_tg"):
                            _cur_s = load_alert_settings()
                            if _cur_s.get("telegram_token"):
                                _msg = f"📊 ApexScan — {narr_ticker} Deep Read\n\n" + narr_text[:3800]
                                _r   = dispatch_alert(_cur_s, _msg, f"ApexScan — {narr_ticker}")
                                st.success("✅ Sent!") if _r.get("telegram") else st.error("❌ Failed")
                            else:
                                st.warning("Configure Telegram in 🔔 Alert Settings first.")

    # ── AUTO-SCAN SCHEDULE SETTINGS ───────────────────────────────────────────
    st.markdown("---")
    with st.expander("⏰ Auto-Scan Schedule", expanded=False):
        st.markdown("""
        **How it works:** When enabled, ApexScan automatically runs a full scan
        at **9:30 AM EST** (market open) and **3:30 PM EST** (30 min before close)
        every trading day — Mon to Fri. Keep this browser tab open for it to fire.
        Results save to the reports folder and are immediately available in all tabs.
        """)
        _as = _autoscan_load()
        a1, a2, a3 = st.columns(3)
        with a1:
            as_enabled = st.toggle(
                "Enable Auto-Scan",
                value=_as.get("enabled", False),
                key="as_toggle",
                help="Fires at 9:30 AM and 3:30 PM EST Mon–Fri while this tab is open"
            )
        with a2:
            as_universe = st.radio(
                "Universe",
                ["📋 Theme Watchlist", "🌐 Extended Universe"],
                index=0 if _as.get("universe","theme")=="theme" else 1,
                key="as_universe",
                horizontal=True,
            )
        with a3:
            st.markdown("**Next scans (EST):**")
            _now_est = datetime.now(_timezone.utc).replace(tzinfo=None) - timedelta(hours=5)
            st.caption(f"🕤 Market open: 9:30 AM")
            st.caption(f"🕞 Pre-close:   3:30 PM")
            st.caption(f"🕐 Now (EST):   {_now_est.strftime('%I:%M %p')}")

        last_open  = _as.get("last_open_scan","Never")
        last_close = _as.get("last_close_scan","Never")
        st.caption(f"Last open scan: **{last_open}** | Last close scan: **{last_close}**")

        if st.button("💾 Save Schedule Settings", key="save_sched"):
            _as["enabled"]  = as_enabled
            _as["universe"] = "theme" if "Theme" in as_universe else "extended"
            _autoscan_save(_as)
            if as_enabled:
                st.success("✅ Auto-scan enabled. Keep this browser tab open for scans to fire automatically.")
            else:
                st.info("Auto-scan disabled.")
            st.rerun()

with tabs[13]:
    st.markdown("### 📋 Watchlist Manager")
    st.caption("Save named watchlists and scan each one independently. Your lists persist between sessions.")

    from scanner import analyze_stock as _analyze_stock

    wls = load_watchlists()

    wl1, wl2 = st.columns([1, 2])

    with wl1:
        st.markdown("#### Your Watchlists")

        # Create new list
        with st.expander("➕ Create New List"):
            new_list_name = st.text_input("List name", placeholder="e.g. Earnings Plays", key="new_wl_name")
            if st.button("Create") and new_list_name:
                wls = create_list(wls, new_list_name)
                save_watchlists(wls)
                st.success(f"Created '{new_list_name}'")
                st.rerun()

        # Select active list
        list_names = list(wls.keys())
        active_list = st.selectbox("Select watchlist", list_names, key="active_wl")

        if active_list:
            tickers_in_list = wls.get(active_list, [])

            # Add ticker
            wl_add_col, wl_btn_col = st.columns([3, 1])
            with wl_add_col:
                new_tk_wl = st.text_input("Add ticker", placeholder="e.g. HIMS", key="wl_add_tk").upper().strip()
            with wl_btn_col:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("Add", key="wl_add_btn") and new_tk_wl:
                    wls = add_ticker(wls, active_list, new_tk_wl)
                    save_watchlists(wls)
                    st.rerun()

            # Import from comma list
            with st.expander("📥 Import tickers (comma separated)"):
                import_str = st.text_area("Paste tickers", placeholder="NVDA, AAPL, TSLA, MSFT", key="wl_import")
                if st.button("Import") and import_str:
                    wls = import_tickers(wls, active_list, import_str)
                    save_watchlists(wls)
                    st.success(f"Imported to '{active_list}'")
                    st.rerun()

            st.markdown(f"**{active_list}** — {len(tickers_in_list)} tickers")

            # Show tickers with remove buttons
            if tickers_in_list:
                for tk in tickers_in_list:
                    tc1, tc2 = st.columns([4, 1])
                    with tc1: st.markdown(f"• **{tk}**")
                    with tc2:
                        if st.button("✕", key=f"rm_{tk}_{active_list}"):
                            wls = remove_ticker(wls, active_list, tk)
                            save_watchlists(wls)
                            st.rerun()

                # Export
                st.markdown("---")
                export_str = export_watchlist(wls, active_list)
                st.text_area("Export (copy this)", value=export_str, height=60, key="wl_export")

                # Delete list (not default ones)
                if active_list not in ["High Conviction", "Monitoring", "Earnings Soon", "Swing Trades"]:
                    if st.button(f"🗑 Delete '{active_list}'", type="secondary"):
                        wls = delete_list(wls, active_list)
                        save_watchlists(wls)
                        st.rerun()
            else:
                st.info("No tickers yet. Add some above or import a comma-separated list.")

    with wl2:
        st.markdown("#### Scan This Watchlist")

        if active_list and wls.get(active_list):
            scan_wl_btn = st.button(f"🚀 Scan '{active_list}'", use_container_width=True)

            if scan_wl_btn:
                tix = wls[active_list]
                cfg_wl = load_config("config.yaml")
                with st.spinner(f"Scanning {len(tix)} tickers in '{active_list}'…"):
                    wl_df = scan_watchlist(active_list, tix, cfg_wl, _analyze_stock)

                if wl_df.empty:
                    st.warning("No results. Tickers may not have enough history or didn't pass filters.")
                else:
                    st.success(f"{len(wl_df)} results from '{active_list}'")

                    # Quick KPIs
                    wk1, wk2, wk3 = st.columns(3)
                    with wk1:
                        top_s = pd.to_numeric(wl_df["apex_score"], errors="coerce").max()
                        st.markdown(f'<div class="metric-card"><h3>Top Score</h3><div class="value green">{top_s:.0f}</div></div>', unsafe_allow_html=True)
                    with wk2:
                        bo_wl = int(wl_df.get("breaking_out", pd.Series([False]*len(wl_df))).sum()) if "breaking_out" in wl_df.columns else 0
                        st.markdown(f'<div class="metric-card"><h3>Breakouts</h3><div class="value amber">{bo_wl}</div></div>', unsafe_allow_html=True)
                    with wk3:
                        s2_wl = int(wl_df.get("stage", pd.Series([""]*len(wl_df))).str.contains("2 ✅", na=False).sum()) if "stage" in wl_df.columns else 0
                        st.markdown(f'<div class="metric-card"><h3>Stage 2</h3><div class="value blue">{s2_wl}</div></div>', unsafe_allow_html=True)

                    st.markdown("---")

                    # Results table
                    wl_show = [c for c in ["ticker","price","stage","perf_3m_%","rs_3m",
                                           "vol_surge_x","pattern","apex_score"] if c in wl_df.columns]
                    wl_disp = wl_df[wl_show].copy()
                    for col in ["apex_score","perf_3m_%","rs_3m"]:
                        if col in wl_disp.columns:
                            wl_disp[col] = pd.to_numeric(wl_disp[col], errors="coerce")

                    styled_wl = wl_disp.style \
                        .map(color_score, subset=["apex_score"]) \
                        .map(color_perf,  subset=["perf_3m_%"] if "perf_3m_%" in wl_disp.columns else []) \
                        .map(color_rs,    subset=["rs_3m"] if "rs_3m" in wl_disp.columns else []) \
                        .format({
                            "price":       "${:.2f}",
                            "perf_3m_%":   pct_fmt,
                            "rs_3m":       lambda v: f"{v:.0f}" if pd.notna(v) and v != 0 else "–",
                            "vol_surge_x": "{:.1f}x",
                            "apex_score":  "{:.0f}",
                        }, na_rep="–")
                    st.dataframe(styled_wl, use_container_width=True, height=420)

                    # Add top results to another watchlist
                    st.markdown("---")
                    st.markdown("**Promote top results to another list:**")
                    promote_target = st.selectbox("Target list", [l for l in list_names if l != active_list], key="wl_promote_target")
                    promote_n = st.slider("Promote top N", 1, min(10, len(wl_df)), 3, key="wl_promote_n")
                    if st.button("Promote"):
                        top_tix = wl_df["ticker"].head(promote_n).tolist()
                        for t in top_tix:
                            wls = add_ticker(wls, promote_target, t)
                        save_watchlists(wls)
                        st.success(f"Added {top_tix} → '{promote_target}'")
        else:
            st.info("Add tickers to your watchlist on the left, then scan here.")

        # ── Summary across all watchlists ─────────────────────────────────
        st.markdown("---")
        st.markdown("#### All Watchlists Overview")
        summary_rows = []
        for name, tix in wls.items():
            summary_rows.append({"Watchlist": name, "Tickers": len(tix),
                                  "Contents": ", ".join(tix[:5]) + ("…" if len(tix) > 5 else "")})
        st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 15 — ALERT SETTINGS
# ══════════════════════════════════════════════════════════════════════════════

with tabs[14]:
    st.markdown("### 🔔 Alert Settings")
    st.caption("Configure Telegram and Email alerts. Both work simultaneously — alerts fire automatically after every scan.")

    settings = load_alert_settings()

    # ── TELEGRAM ──────────────────────────────────────────────────────────
    st.markdown("#### 📲 Telegram Alerts")
    tg_col1, tg_col2 = st.columns(2)

    with tg_col1:
        tg_token = st.text_input(
            "Bot Token",
            value=settings.get("telegram_token", ""),
            type="password",
            placeholder="8972969518:AAGAmdRD...",
            key="tg_token"
        )
        tg_chat_id = st.text_input(
            "Chat ID",
            value=settings.get("telegram_chat_id", ""),
            placeholder="Find this at step 4 below",
            key="tg_chat"
        )

    with tg_col2:
        st.markdown("""
        **Your bot (ApexV8bot) is already created ✅**

        Now get your Chat ID:
        1. Open Telegram → find **ApexV8bot**
        2. Send it any message (e.g. "hello")
        3. Visit this URL in your browser:
           `https://api.telegram.org/bot8972969518:AAGAmdRDtyyXdv4pzqJc6og4YqyCsnBxXUg/getUpdates`
        4. Look for `"chat":{"id":` — the number after it is your Chat ID
        5. Paste that number above and save
        """)

    tc1, tc2 = st.columns(2)
    with tc1:
        if st.button("📲 Test Telegram", use_container_width=True):
            if tg_token and tg_chat_id:
                ok = test_telegram(tg_token, tg_chat_id)
                st.success("✅ Telegram test sent! Check your bot.") if ok else st.error("❌ Failed — double check your Chat ID.")
            else:
                st.warning("Enter both Bot Token and Chat ID first.")

    st.markdown("---")

    # ── EMAIL ─────────────────────────────────────────────────────────────
    st.markdown("#### 📧 Email Alerts (Gmail)")
    em_col1, em_col2 = st.columns(2)

    with em_col1:
        em_from = st.text_input(
            "Your Gmail address",
            value=settings.get("email_from", ""),
            placeholder="yourname@gmail.com",
            key="em_from"
        )
        em_pass = st.text_input(
            "Gmail App Password",
            value=settings.get("email_password", ""),
            type="password",
            placeholder="xxxx xxxx xxxx xxxx",
            key="em_pass"
        )
        em_to = st.text_input(
            "Send alerts to (can be same address)",
            value=settings.get("email_to", ""),
            placeholder="yourname@gmail.com",
            key="em_to"
        )

    with em_col2:
        st.markdown("""
        **How to get a Gmail App Password:**
        1. Go to **myaccount.google.com**
        2. Click **Security** on the left
        3. Enable **2-Step Verification** if not already on
        4. Search for **App Passwords** in the search bar
        5. Select **Mail** → **Generate**
        6. Copy the 16-character password (e.g. `abcd efgh ijkl mnop`)
        7. Paste it above — spaces are fine

        ⚠️ Use an App Password, NOT your regular Gmail password
        """)

    ec1, ec2 = st.columns(2)
    with ec1:
        if st.button("📧 Test Email", use_container_width=True):
            if em_from and em_pass and em_to:
                ok = send_email(
                    em_from, em_pass, em_to,
                    "ApexScan — Test Alert",
                    "✅ Your ApexScan email alerts are working correctly!"
                )
                st.success("✅ Test email sent! Check your inbox.") if ok else st.error("❌ Failed — check your Gmail and App Password.")
            else:
                st.warning("Fill in all three email fields first.")

    st.markdown("---")

    # ── PREFERENCES ───────────────────────────────────────────────────────
    st.markdown("#### ⚙️ Alert Preferences")
    pr1, pr2, pr3, pr4 = st.columns(4)
    with pr1:
        alerts_on      = st.toggle("Enable All Alerts", value=settings.get("alerts_enabled", False))
        alert_breakout = st.checkbox("Breakout Alerts",      value=settings.get("alert_breakouts", True))
        alert_stop     = st.checkbox("Stop Loss Breach",     value=settings.get("alert_stop_breach", True))
    with pr2:
        alert_earn     = st.checkbox("Earnings Warnings",    value=settings.get("alert_earnings", True))
        alert_sfp      = st.checkbox("SFP Setup Alerts",     value=settings.get("alert_sfp_setup", True),
                                      help="Alert when a Swing Failure Pattern is detected — trap setup")
    with pr3:
        alert_of       = st.checkbox("Persistent Flow Alert",value=settings.get("alert_persistent_flow", True),
                                      help="Alert when Strong Bullish order flow persistence detected")
        alert_vwap_imb = st.checkbox("VWAP Imbalance Alert", value=settings.get("alert_vwap_imbalance", True),
                                      help="Alert when price is extended above or below VWAP")
    with pr4:
        min_score_alert = st.slider("Min Score to Alert", 30, 90,
                                     settings.get("min_score_alert", 60), 5)
        st.caption("Only alert on setups scoring above this threshold")

    st.markdown("---")

    if st.button("💾 Save All Settings", use_container_width=True, type="primary"):
        new_settings = {
            "telegram_token":        tg_token,
            "telegram_chat_id":      tg_chat_id,
            "email_from":            em_from,
            "email_password":        em_pass,
            "email_to":              em_to,
            "alerts_enabled":        alerts_on,
            "alert_breakouts":       alert_breakout,
            "alert_stop_breach":     alert_stop,
            "alert_earnings":        alert_earn,
            "alert_sfp_setup":       alert_sfp,
            "alert_persistent_flow": alert_of,
            "alert_vwap_imbalance":  alert_vwap_imb,
            "min_score_alert":       min_score_alert,
        }
        save_alert_settings(new_settings)
        st.success("✅ Settings saved! Both Telegram and Email alerts are configured.")

    # ── MANUAL TEST ───────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📨 Send a Manual Test Alert")
    st.caption("Tests both Telegram and Email simultaneously")
    test_msg = st.text_area(
        "Message",
        value="🔔 ApexScan Alert — Your breakout scanner is live and watching the market for you.",
        key="manual_alert_msg"
    )
    if st.button("Send to All Channels", use_container_width=True):
        cur = load_alert_settings()
        result = dispatch_alert(cur, test_msg, "ApexScan Alert")
        sent = []
        if result.get("telegram"): sent.append("Telegram ✅")
        if result.get("email"):    sent.append("Email ✅")
        if sent:
            st.success(f"Sent via: {', '.join(sent)}")
        elif not cur.get("alerts_enabled"):
            st.warning("Alerts are disabled — toggle 'Enable All Alerts' on and save first.")
        else:
            st.error("Nothing sent. Check your credentials and make sure settings are saved.")

    st.markdown("""
    ---
    **When alerts fire automatically:**
    - 🚀 **Breakout** — after every Live Scan, for high-scoring breakout setups
    - 🎯 **SFP Setup** — when a Swing Failure Pattern trap is detected
    - 📈 **Persistent Flow** — when Strong Bullish order flow is confirmed
    - 💧 **VWAP Imbalance** — when price is extended far from VWAP fair value
    - 💼 **Stop Breach** — after Portfolio Tracker loads, if holding drops below 50MA
    - 📅 **Earnings** — when you fetch the Earnings Calendar for upcoming reports
    - 🤖 **AI Briefing** — option to send morning briefing to Telegram after generation

    Telegram and Email both fire simultaneously when configured.
    Min Score filter applies to all scan-based alerts.
    """)




# ══════════════════════════════════════════════════════════════════════════════
# TAB 16 — INTERPRETATION
# ══════════════════════════════════════════════════════════════════════════════

with tabs[15]:
    st.markdown("### 🧠 Scan Interpretation")
    st.caption("Plain-English breakdown of your scan results across all four signal views.")

    if df.empty:
        st.info("Run a Live Scan or Load Last Report first, then come here for the interpretation.")
    else:
        # ── Ticker selector ───────────────────────────────────────────────
        interp_mode = st.radio(
            "Interpret",
            ["Full Scan Summary", "Single Ticker Deep Read"],
            horizontal=True
        )

        if interp_mode == "Single Ticker Deep Read":
            interp_ticker = st.selectbox("Choose ticker", df["ticker"].tolist(), key="interp_tk")
            interp_df = df[df["ticker"] == interp_ticker]
        else:
            interp_ticker = None
            interp_df = df.copy()

        st.markdown("---")

        # ════════════════════════════════════════════════════════════════
        # HELPER: colour-coded signal pill
        # ════════════════════════════════════════════════════════════════
        def pill(text, color="#3fb950"):
            return (f'<span style="background:{color}22;color:{color};'
                    f'border:1px solid {color};border-radius:4px;'
                    f'padding:2px 8px;font-size:0.8rem;font-weight:600;">'
                    f'{text}</span>')

        def green(t):  return pill(t, "#3fb950")
        def amber(t):  return pill(t, "#d29922")
        def red(t):    return pill(t, "#f85149")
        def blue(t):   return pill(t, "#388bfd")
        def purple(t): return pill(t, "#c084fc")

        # ════════════════════════════════════════════════════════════════
        # VIEW 1 — STANDARD
        # ════════════════════════════════════════════════════════════════
        with st.expander("📋 Standard View — Momentum & Stage", expanded=True):
            if interp_mode == "Full Scan Summary":
                total      = len(interp_df)
                stage2     = interp_df["stage"].str.contains("2 ✅", na=False).sum() if "stage" in interp_df.columns else 0
                stage4     = interp_df["stage"].str.contains("4 🔴", na=False).sum() if "stage" in interp_df.columns else 0
                avg_3m     = pd.to_numeric(interp_df.get("perf_3m_%", pd.Series()), errors="coerce").mean()
                avg_rs     = pd.to_numeric(interp_df.get("rs_3m", pd.Series()), errors="coerce").mean()
                breakouts  = interp_df.get("breaking_out", pd.Series([False]*total)).sum()
                top3       = interp_df.head(3)["ticker"].tolist()
                near_high  = interp_df.get("near_52wh", pd.Series([False]*total)).sum()

                # Market health verdict
                if stage2 / max(total,1) >= 0.6:
                    health = green("HEALTHY MARKET")
                    health_msg = "Most setups are in Stage 2 uptrends — conditions favour the bull side."
                elif stage2 / max(total,1) >= 0.4:
                    health = amber("MIXED CONDITIONS")
                    health_msg = "Market is split — be selective, stick to Stage 2 stocks only."
                else:
                    health = red("WEAK BREADTH")
                    health_msg = "Few Stage 2 setups — reduce position size and wait for clarity."

                st.markdown(f"""
                <div style="background:#161b22;border:1px solid #30363d;border-radius:10px;padding:20px 24px;margin-bottom:16px;">
                  <div style="font-size:1.1rem;font-weight:700;margin-bottom:12px;">
                    Market Condition: {health}
                  </div>
                  <p style="color:#e6edf3;margin:0 0 12px 0;">{health_msg}</p>
                  <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;">
                    <div><span style="color:#8b949e;font-size:0.75rem;">SETUPS PASSING</span><br>
                         <span style="font-size:1.4rem;font-weight:700;">{total}</span></div>
                    <div><span style="color:#8b949e;font-size:0.75rem;">STAGE 2 (BUY ZONE)</span><br>
                         <span style="font-size:1.4rem;font-weight:700;color:#3fb950;">{stage2}</span></div>
                    <div><span style="color:#8b949e;font-size:0.75rem;">AVG 3M RETURN</span><br>
                         <span style="font-size:1.4rem;font-weight:700;color:{'#3fb950' if avg_3m>0 else '#f85149'};">{avg_3m:+.1f}%</span></div>
                  </div>
                </div>
                """, unsafe_allow_html=True)

                # Per-stock standard interpretation
                st.markdown("#### Stock-by-Stock Reading")
                for _, row in interp_df.head(15).iterrows():
                    tk    = row["ticker"]
                    score = float(row.get("apex_score", 0))
                    stage = str(row.get("stage","?"))
                    p3m   = float(row.get("perf_3m_%", 0) or 0)
                    rs    = float(row.get("rs_3m", 0) or 0)
                    patt  = str(row.get("pattern","–"))
                    brk   = bool(row.get("breaking_out", False))

                    # Score badge
                    score_badge = green(f"Score {score:.0f}") if score>=70 else (amber(f"Score {score:.0f}") if score>=40 else red(f"Score {score:.0f}"))
                    stage_badge = green(stage) if "2 ✅" in stage else (amber(stage) if "1 ⏳" in stage else red(stage))

                    # Reading sentences
                    sentences = []
                    if "2 ✅" in stage:
                        sentences.append(f"In a confirmed Stage 2 uptrend — price is above both moving averages with the 50MA rising above the 200MA.")
                    elif "1 ⏳" in stage:
                        sentences.append(f"Still building a base (Stage 1) — not yet buyable, needs to clear the 50MA.")
                    elif "4 🔴" in stage:
                        sentences.append(f"In a Stage 4 downtrend — avoid entirely until structure repairs.")
                    else:
                        sentences.append(f"Mixed stage — structure is unclear, wait for cleaner setup.")

                    if p3m > 30:
                        sentences.append(f"Exceptional 3-month momentum of {p3m:+.1f}% — this is a leading stock in its theme.")
                    elif p3m > 15:
                        sentences.append(f"Solid 3-month return of {p3m:+.1f}% — above the minimum momentum threshold.")
                    else:
                        sentences.append(f"3-month return of {p3m:+.1f}% is below the momentum threshold — needs to accelerate.")

                    if rs > 150:
                        sentences.append(f"RS of {rs:.0f} means it's massively outperforming the S&P 500 — buy leaders like this.")
                    elif rs > 100:
                        sentences.append(f"RS of {rs:.0f} — beating the S&P 500, in line with what you want to own.")
                    elif rs > 70:
                        sentences.append(f"RS of {rs:.0f} — keeping pace with the market but not leading.")
                    else:
                        sentences.append(f"RS of {rs:.0f} is below the market — lagging stocks rarely lead the next move.")

                    if brk:
                        sentences.append(f"Active breakout in progress ({patt}) — this is the highest-priority actionable setup.")
                    elif "Handle" in patt or "Tight" in patt:
                        sentences.append(f"Pattern ({patt}) suggests it's coiling for a move — watch the pivot point.")
                    elif "Deep Correction" in patt:
                        sentences.append(f"Currently in a deep correction ({patt}) — let it base and tighten before considering entry.")

                    st.markdown(
                        f'<div style="background:#0d1117;border:1px solid #21262d;border-radius:8px;'
                        f'padding:14px 18px;margin:6px 0;">'
                        f'<div style="margin-bottom:8px;">'
                        f'<span style="font-size:1rem;font-weight:700;color:#e6edf3;">{tk}</span>'
                        f'&nbsp;&nbsp;{score_badge}&nbsp;{stage_badge}</div>'
                        f'<ul style="margin:0;padding-left:18px;color:#c9d1d9;font-size:0.88rem;line-height:1.8;">'
                        + "".join(f"<li>{s}</li>" for s in sentences) +
                        f'</ul></div>',
                        unsafe_allow_html=True
                    )

            else:
                # Single ticker standard
                row   = interp_df.iloc[0]
                score = float(row.get("apex_score",0))
                stage = str(row.get("stage","?"))
                p3m   = float(row.get("perf_3m_%",0) or 0)
                p6m   = float(row.get("perf_6m_%",0) or 0)
                rs3   = float(row.get("rs_3m",0) or 0)
                patt  = str(row.get("pattern","–"))
                brk   = bool(row.get("breaking_out",False))
                near  = bool(row.get("near_52wh",False))
                off_h = float(row.get("pct_off_high_%",0) or 0)

                verdict = "Strong Setup" if score>=70 else ("Watch List" if score>=40 else "Not Ready")
                vc      = "#3fb950" if score>=70 else ("#d29922" if score>=40 else "#f85149")

                st.markdown(f"""
                <div style="background:#161b22;border:2px solid {vc};border-radius:12px;padding:24px;margin-bottom:20px;">
                  <div style="font-size:2rem;font-weight:800;color:#e6edf3;">{interp_ticker}</div>
                  <div style="font-size:0.9rem;color:#8b949e;margin:4px 0 16px 0;">Standard Signal Reading</div>
                  <div style="font-size:0.95rem;color:#c9d1d9;line-height:1.9;">
                    {'✅' if '2 ✅' in stage else '❌'} <b>Stage:</b> {stage} —
                    {'Price is above both moving averages and the 50MA is rising above the 200MA. This is the only stage worth holding or buying.' if '2 ✅' in stage else 'Not in a confirmed uptrend. Stage 2 is required before considering entry.'}<br>
                    {'✅' if p3m>15 else '⚠️'} <b>3M Return:</b> {p3m:+.1f}% —
                    {'Strong momentum. The stock has outperformed most of the market over the past quarter.' if p3m>15 else 'Below the 15% momentum threshold. Needs to accelerate before it qualifies as a leading stock.'}<br>
                    {'✅' if p6m>20 else '⚠️'} <b>6M Return:</b> {p6m:+.1f}% —
                    {'Sustained trend over 6 months confirms this is not a one-month wonder.' if p6m>20 else 'Six-month performance is muted — the trend may be early or stalling.'}<br>
                    {'✅' if rs3>100 else '⚠️'} <b>RS Score:</b> {rs3:.0f} —
                    {'Outperforming the S&P 500. You want to own the leaders, not keep up with the index.' if rs3>100 else 'Underperforming or matching the index. Leaders should have RS well above 100.'}<br>
                    {'✅' if near else '⚠️'} <b>52-Week High:</b> {'Within 15% of highs — near the top of its range, which is where breakouts happen.' if near else f'{abs(off_h):.1f}% below its 52-week high — needs to reclaim ground before a breakout is possible.'}<br>
                    {'🚀' if brk else '⏳'} <b>Pattern:</b> {patt} —
                    {'Active breakout in progress. Highest priority actionable setup.' if brk else 'No active breakout yet. Monitor for the pivot point trigger.'}
                  </div>
                </div>
                """, unsafe_allow_html=True)

        # ════════════════════════════════════════════════════════════════
        # VIEW 2 — ORDER FLOW
        # ════════════════════════════════════════════════════════════════
        with st.expander("🌊 Order Flow Interpretation", expanded=True):
            if "of_bias" not in interp_df.columns:
                st.info("Run a fresh scan to get Order Flow data.")
            else:
                if interp_mode == "Full Scan Summary":
                    of_strong  = interp_df["of_bias"].str.contains("Strong Bullish", na=False).sum()
                    of_bull    = interp_df["of_bias"].str.contains("Bullish", na=False).sum()
                    of_bear    = interp_df["of_bias"].str.contains("Bearish", na=False).sum()
                    avg_ratio  = pd.to_numeric(interp_df.get("of_up_vol_ratio", pd.Series()), errors="coerce").mean()
                    max_consec = pd.to_numeric(interp_df.get("of_consec_up", pd.Series()), errors="coerce").max()

                    of_verdict = green("STRONG INSTITUTIONAL BUYING") if of_strong >= 3 else \
                                 amber("MODERATE BUYING PRESSURE") if of_bull > of_bear else \
                                 red("SELLING PRESSURE DOMINATES")

                    st.markdown(f"""
                    <div style="background:#161b22;border:1px solid #30363d;border-radius:10px;padding:20px 24px;margin-bottom:16px;">
                      <div style="font-size:1rem;font-weight:700;margin-bottom:10px;">
                        Order Flow Picture: {of_verdict}
                      </div>
                      <p style="color:#c9d1d9;font-size:0.9rem;margin:0 0 8px 0;">
                        Order flow persistence measures whether institutional money is consistently active on the buy side
                        over multiple sessions — the signature of TWAP/VWAP algorithms splitting large orders to avoid moving the market.
                      </p>
                      <p style="color:#c9d1d9;font-size:0.9rem;margin:0;">
                        <b>{of_strong}</b> stocks show Strong Bullish flow (the highest conviction signal) &nbsp;|&nbsp;
                        <b>{of_bull}</b> total with bullish bias &nbsp;|&nbsp;
                        <b>{of_bear}</b> with bearish bias &nbsp;|&nbsp;
                        Avg up/down vol ratio: <b>{avg_ratio:.2f}x</b> &nbsp;|&nbsp;
                        Longest consecutive up-close streak: <b>{int(max_consec) if pd.notna(max_consec) else 0} days</b>
                      </p>
                    </div>
                    """, unsafe_allow_html=True)

                    st.markdown("#### Order Flow Reading Per Stock")
                    for _, row in interp_df.head(15).iterrows():
                        tk       = row["ticker"]
                        bias     = str(row.get("of_bias","–"))
                        ratio    = float(row.get("of_up_vol_ratio",1) or 1)
                        bull_pct = float(row.get("of_bullish_days",50) or 50)
                        consec   = int(row.get("of_consec_up",0) or 0)
                        of_sc    = int(row.get("of_score",0) or 0)

                        bias_badge = green(bias) if "Strong Bullish" in bias else \
                                     amber(bias) if "Bullish" in bias else \
                                     red(bias) if "Bearish" in bias else blue(bias)

                        if "Strong Bullish" in bias:
                            reading = (f"Institutional buying is persistent and heavy. "
                                      f"{bull_pct:.0f}% of the last 10 sessions closed up, "
                                      f"with {ratio:.2f}x more volume on up days than down days. "
                                      f"{'Consecutive up closes of ' + str(consec) + ' sessions suggests an active algorithm is working a large buy order.' if consec >= 3 else ''} "
                                      f"This is the pattern left behind when a fund is accumulating a position quietly over time.")
                        elif "Bullish" in bias:
                            reading = (f"Moderate buying pressure detected. {bull_pct:.0f}% of sessions closed up "
                                      f"with {ratio:.2f}x up-volume ratio. "
                                      f"Directional bias is positive but not yet showing the heavy conviction of institutional accumulation.")
                        elif "Bearish" in bias:
                            reading = (f"Selling pressure is present. Only {bull_pct:.0f}% of sessions closed up "
                                      f"and down-day volume is outpacing up-day volume ({ratio:.2f}x). "
                                      f"This pattern can indicate distribution — institutions quietly reducing positions.")
                        else:
                            reading = (f"Flow is neutral — no clear directional conviction from either side over the last 10 sessions. "
                                      f"Wait for a clearer signal before committing capital.")

                        st.markdown(
                            f'<div style="background:#0d1117;border:1px solid #21262d;border-radius:8px;'
                            f'padding:14px 18px;margin:6px 0;">'
                            f'<div style="margin-bottom:6px;">'
                            f'<span style="font-weight:700;color:#e6edf3;">{tk}</span>&nbsp;&nbsp;'
                            f'{bias_badge}&nbsp;{purple(f"OF Score {of_sc}/8")}</div>'
                            f'<p style="margin:0;color:#c9d1d9;font-size:0.88rem;line-height:1.7;">{reading}</p>'
                            f'</div>',
                            unsafe_allow_html=True
                        )

                else:
                    row      = interp_df.iloc[0]
                    bias     = str(row.get("of_bias","–"))
                    ratio    = float(row.get("of_up_vol_ratio",1) or 1)
                    bull_pct = float(row.get("of_bullish_days",50) or 50)
                    consec   = int(row.get("of_consec_up",0) or 0)
                    of_sc    = int(row.get("of_score",0) or 0)

                    bias_color = "#3fb950" if "Bullish" in bias else ("#f85149" if "Bearish" in bias else "#d29922")

                    if "Strong Bullish" in bias:
                        of_summary = (
                            f"The order flow picture for {interp_ticker} is strongly bullish. "
                            f"Over the last 10 sessions, {bull_pct:.0f}% of days closed higher, "
                            f"and volume on up-days was {ratio:.2f}x the volume on down-days. "
                            f"{'A streak of ' + str(consec) + ' consecutive up-closes is a strong sign of an active institutional buyer systematically accumulating shares — typical of a fund using TWAP/VWAP execution.' if consec >= 3 else 'The volume-weighted bias confirms buyers are more active than sellers.'} "
                            f"This is the kind of persistent, quiet accumulation that often precedes a significant price move."
                        )
                    elif "Bullish" in bias:
                        of_summary = (
                            f"{interp_ticker} shows moderate bullish order flow. "
                            f"{bull_pct:.0f}% of the last 10 sessions closed up with a {ratio:.2f}x up-volume ratio. "
                            f"Buying pressure is real but not yet at the level that suggests heavy institutional accumulation. "
                            f"Monitor for the ratio to increase above 1.5x as confirmation."
                        )
                    elif "Bearish" in bias:
                        of_summary = (
                            f"{interp_ticker} is showing bearish order flow. "
                            f"Only {bull_pct:.0f}% of sessions closed up, and selling volume is dominating. "
                            f"This is a warning sign — even if the stock looks good technically, "
                            f"persistent down-day volume can signal that larger players are quietly exiting positions."
                        )
                    else:
                        of_summary = (
                            f"Order flow for {interp_ticker} is neutral — no clear directional bias from either side. "
                            f"The market is in equilibrium on this name. Wait for a directional signal before acting."
                        )

                    st.markdown(f"""
                    <div style="background:#161b22;border:1px solid {bias_color};border-radius:10px;padding:20px 24px;">
                      <div style="color:{bias_color};font-weight:700;font-size:1rem;margin-bottom:10px;">
                        {interp_ticker} — Order Flow: {bias} ({of_sc}/8 pts)
                      </div>
                      <p style="color:#c9d1d9;font-size:0.92rem;line-height:1.8;margin:0;">{of_summary}</p>
                      <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-top:16px;">
                        <div style="background:#0d1117;border-radius:6px;padding:10px;">
                          <div style="color:#8b949e;font-size:0.7rem;text-transform:uppercase;">Up/Down Vol Ratio</div>
                          <div style="font-size:1.3rem;font-weight:700;color:{bias_color};">{ratio:.2f}x</div>
                          <div style="color:#8b949e;font-size:0.75rem;">{'✅ Institutional level' if ratio>=1.5 else '⚠️ Below threshold (1.5x)'}</div>
                        </div>
                        <div style="background:#0d1117;border-radius:6px;padding:10px;">
                          <div style="color:#8b949e;font-size:0.7rem;text-transform:uppercase;">Bullish Sessions (10d)</div>
                          <div style="font-size:1.3rem;font-weight:700;color:{bias_color};">{bull_pct:.0f}%</div>
                          <div style="color:#8b949e;font-size:0.75rem;">{'✅ Persistent buying' if bull_pct>=60 else '⚠️ Needs >60% to qualify'}</div>
                        </div>
                        <div style="background:#0d1117;border-radius:6px;padding:10px;">
                          <div style="color:#8b949e;font-size:0.7rem;text-transform:uppercase;">Consecutive Up Closes</div>
                          <div style="font-size:1.3rem;font-weight:700;color:{bias_color};">{consec} days</div>
                          <div style="color:#8b949e;font-size:0.75rem;">{'🔥 Active algo detected' if consec>=4 else '📊 Normal range'}</div>
                        </div>
                      </div>
                    </div>
                    """, unsafe_allow_html=True)

        # ════════════════════════════════════════════════════════════════
        # VIEW 3 — VWAP & STRUCTURE
        # ════════════════════════════════════════════════════════════════
        with st.expander("💧 VWAP & Market Structure Interpretation", expanded=True):
            if "vwap_position" not in interp_df.columns:
                st.info("Run a fresh scan to get VWAP data.")
            else:
                if interp_mode == "Full Scan Summary":
                    above_vwap = interp_df["vwap_position"].str.contains("Above", na=False).sum()
                    ext_above  = interp_df["vwap_position"].str.contains("Extended Above", na=False).sum()
                    ext_below  = interp_df["vwap_position"].str.contains("Extended Below", na=False).sum()
                    hh_hl_ct   = interp_df.get("ms_hh_hl", pd.Series([False]*len(interp_df))).sum()
                    bos_ct     = interp_df.get("ms_bos", pd.Series([False]*len(interp_df))).sum()
                    rising_vwap= interp_df.get("vwap_slope", pd.Series()).str.contains("Rising", na=False).sum()

                    st.markdown(f"""
                    <div style="background:#161b22;border:1px solid #c084fc;border-radius:10px;padding:20px 24px;margin-bottom:16px;">
                      <div style="color:#c084fc;font-weight:700;font-size:1rem;margin-bottom:12px;">
                        VWAP & Auction Market Picture
                      </div>
                      <p style="color:#c9d1d9;font-size:0.9rem;line-height:1.7;margin:0 0 12px 0;">
                        VWAP (Volume Weighted Average Price) is the fairest measure of where the market agreed to transact.
                        Stocks above a rising VWAP have buyers in control and value is being accepted higher.
                        Stocks extended far above VWAP are at risk of mean reversion back to fair value.
                      </p>
                      <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;">
                        <div><span style="color:#8b949e;font-size:0.75rem;">ABOVE VWAP</span><br>
                             <span style="font-size:1.4rem;font-weight:700;color:#3fb950;">{above_vwap}</span>
                             <span style="color:#8b949e;font-size:0.8rem;"> of {len(interp_df)}</span></div>
                        <div><span style="color:#8b949e;font-size:0.75rem;">EXTENDED (RISKY)</span><br>
                             <span style="font-size:1.4rem;font-weight:700;color:#d29922;">{ext_above}</span></div>
                        <div><span style="color:#8b949e;font-size:0.75rem;">HH/HL STRUCTURE</span><br>
                             <span style="font-size:1.4rem;font-weight:700;color:#3fb950;">{hh_hl_ct}</span></div>
                      </div>
                    </div>
                    """, unsafe_allow_html=True)

                    st.markdown("#### VWAP & Structure Per Stock")
                    for _, row in interp_df.head(15).iterrows():
                        tk       = row["ticker"]
                        vwap_pos = str(row.get("vwap_position","–"))
                        vs_vwap  = float(row.get("vs_vwap_%",0) or 0)
                        slope    = str(row.get("vwap_slope","–"))
                        hh_hl    = bool(row.get("ms_hh_hl",False))
                        bos      = bool(row.get("ms_bos",False))
                        ms       = str(row.get("ms_structure","–"))
                        sh       = row.get("ms_swing_high")
                        sl       = row.get("ms_swing_low")

                        if "Extended Above" in vwap_pos:
                            vwap_read = (f"Trading {vs_vwap:+.1f}% above VWAP fair value — extended. "
                                        f"While momentum is strong, entering here means chasing. "
                                        f"Better to wait for a pullback toward VWAP before adding or initiating.")
                            vc = "#d29922"
                        elif "Above" in vwap_pos and slope == "Rising":
                            vwap_read = (f"Trading {vs_vwap:+.1f}% above a rising VWAP — ideal zone. "
                                        f"Buyers are in control, fair value is moving higher, "
                                        f"and this is where high-quality momentum entries are found.")
                            vc = "#3fb950"
                        elif "Above" in vwap_pos:
                            vwap_read = (f"Above VWAP by {vs_vwap:+.1f}% but VWAP slope is {slope.lower()}. "
                                        f"Position is acceptable but watch for VWAP to start declining, "
                                        f"which would signal a shift in auction control.")
                            vc = "#3fb950"
                        elif "Extended Below" in vwap_pos:
                            vwap_read = (f"Extended {vs_vwap:.1f}% below VWAP — sellers in full control. "
                                        f"Avoid new longs. Only consider watching for a VWAP reclaim on volume "
                                        f"as a potential reversal signal.")
                            vc = "#f85149"
                        else:
                            vwap_read = (f"Below VWAP by {abs(vs_vwap):.1f}%. The auction has not accepted value higher. "
                                        f"Wait for a confirmed close above VWAP before considering a long position.")
                            vc = "#f85149"

                        struct_read = ""
                        if hh_hl:
                            struct_read = f" Market structure confirms Higher Highs and Higher Lows — the uptrend is intact."
                            if sh: struct_read += f" Key resistance to watch: ${sh:.2f}."
                        elif "Bearish" in ms:
                            struct_read = f" Structure shows Lower Highs and Lower Lows — technically in a downtrend. Avoid."
                        else:
                            struct_read = f" Structure is transitional — no clear trend. Wait for HH/HL to establish."

                        if bos and hh_hl:
                            struct_read += f" A recent Break of Structure in an uptrend confirms the move is accelerating."

                        vwap_badge = green("Above VWAP ↑") if "Above" in vwap_pos and "Extended" not in vwap_pos \
                                else amber("Extended ↑") if "Extended Above" in vwap_pos \
                                else red("Below VWAP ↓")
                        ms_badge = green("HH/HL ✅") if hh_hl else red("No HH/HL")

                        st.markdown(
                            f'<div style="background:#0d1117;border:1px solid #21262d;border-radius:8px;'
                            f'padding:14px 18px;margin:6px 0;">'
                            f'<div style="margin-bottom:6px;">'
                            f'<span style="font-weight:700;color:#e6edf3;">{tk}</span>&nbsp;&nbsp;'
                            f'{vwap_badge}&nbsp;{ms_badge}</div>'
                            f'<p style="margin:0;color:#c9d1d9;font-size:0.88rem;line-height:1.7;">'
                            f'{vwap_read}{struct_read}</p>'
                            f'</div>',
                            unsafe_allow_html=True
                        )

                else:
                    row      = interp_df.iloc[0]
                    vwap_val = row.get("vwap")
                    vs_vwap  = float(row.get("vs_vwap_%",0) or 0)
                    vwap_pos = str(row.get("vwap_position","–"))
                    slope    = str(row.get("vwap_slope","–"))
                    vwap_u   = row.get("vwap_upper")
                    vwap_l   = row.get("vwap_lower")
                    hh_hl    = bool(row.get("ms_hh_hl",False))
                    bos      = bool(row.get("ms_bos",False))
                    ms       = str(row.get("ms_structure","–"))
                    sh       = row.get("ms_swing_high")
                    sl       = row.get("ms_swing_low")
                    price    = float(row.get("price",0))

                    vc = "#3fb950" if "Above" in vwap_pos and "Extended" not in vwap_pos \
                         else "#d29922" if "Extended Above" in vwap_pos else "#f85149"

                    st.markdown(f"""
                    <div style="background:#161b22;border:1px solid {vc};border-radius:10px;padding:20px 24px;margin-bottom:16px;">
                      <div style="color:{vc};font-weight:700;font-size:1rem;margin-bottom:14px;">
                        {interp_ticker} — VWAP & Market Structure
                      </div>
                      <div style="display:grid;grid-template-columns:repeat(2,1fr);gap:20px;">
                        <div>
                          <div style="color:#8b949e;font-size:0.75rem;text-transform:uppercase;margin-bottom:8px;">VWAP Analysis</div>
                          <p style="color:#c9d1d9;font-size:0.9rem;line-height:1.8;margin:0;">
                            Current price <b>${price:.2f}</b> sits <b style="color:{vc};">{vs_vwap:+.1f}%</b> {"above" if vs_vwap>0 else "below"} the 20-day VWAP of <b>${vwap_val:.2f}</b>.<br>
                            {'The VWAP is rising, meaning the institutional fair value reference is moving higher — this is the setup you want.' if slope=='Rising' else
                             'The VWAP is flat, meaning value is not yet being accepted higher — the auction is in balance.' if slope=='Flat' else
                             'The VWAP is declining — sellers are pushing the fair value reference lower, a bearish signal.'}<br>
                            {'Upper band at $' + f'{vwap_u:.2f}' + ' — this is resistance where price may pause or pull back.' if vwap_u else ''}
                            {'Lower band at $' + f'{vwap_l:.2f}' + ' — this is VWAP support where buyers typically step in.' if vwap_l else ''}
                          </p>
                        </div>
                        <div>
                          <div style="color:#8b949e;font-size:0.75rem;text-transform:uppercase;margin-bottom:8px;">Market Structure</div>
                          <p style="color:#c9d1d9;font-size:0.9rem;line-height:1.8;margin:0;">
                            Structure: <b style="color:{'#3fb950' if 'Bullish' in ms else '#f85149'};">{ms}</b><br>
                            {'✅ Higher Highs and Higher Lows confirmed — the stock is making progress and the uptrend is structurally sound.' if hh_hl else '❌ No HH/HL structure yet — the stock has not proven its uptrend with progressively higher pivot points.'}<br>
                            {'🚨 Recent Break of Structure detected — price has moved beyond a key swing level, confirming the directional move.' if bos else ''}<br>
                            {'Key swing high to watch: $' + f'{sh:.2f}' + ' — a close above this on volume would be a structural confirmation.' if sh else ''}
                            {'Key swing low to defend: $' + f'{sl:.2f}' + ' — a close below this would signal trend deterioration.' if sl else ''}
                          </p>
                        </div>
                      </div>
                    </div>
                    """, unsafe_allow_html=True)

        # ════════════════════════════════════════════════════════════════
        # VIEW 4 — PRICE ACTION
        # ════════════════════════════════════════════════════════════════
        with st.expander("🕯 Price Action Interpretation", expanded=True):
            if "pa_patterns" not in interp_df.columns:
                st.info("Run a fresh scan to get Price Action data.")
            else:
                pa_explanations = {
                    "Bullish SFP (Bear Trap)": (
                        "#3fb950",
                        "Swing Failure Pattern — Bullish",
                        "Price temporarily pierced below a recent swing low, triggering stop-losses and luring in short sellers, "
                        "then reversed and closed back above that level. The shorts are now trapped and forced to cover, "
                        "which fuels buying pressure. This is one of the highest-probability reversal setups in price action. "
                        "Entry is typically on the next candle's open if the close holds above the swing low."
                    ),
                    "Bearish SFP (Bull Trap)": (
                        "#f85149",
                        "Swing Failure Pattern — Bearish",
                        "Price spiked above a recent swing high, triggering breakout buyers and stop-losses on short positions, "
                        "then reversed and closed back below. The longs who chased the breakout are now trapped. "
                        "This signals the prior high is strong resistance and a move lower is likely."
                    ),
                    "Bullish Engulfing": (
                        "#3fb950",
                        "Bullish Engulfing Candle",
                        "A large bullish candle whose body completely engulfs the prior bearish candle. "
                        "This shows buyers decisively overwhelmed sellers in a single session. "
                        "Most powerful when occurring at a support level, above VWAP, or after a pullback in an uptrend."
                    ),
                    "Bearish Engulfing": (
                        "#f85149",
                        "Bearish Engulfing Candle",
                        "A large bearish candle engulfs the prior bullish candle — sellers took complete control. "
                        "Most significant at resistance levels or after an extended run-up."
                    ),
                    "Inside Day (Compression)": (
                        "#388bfd",
                        "Inside Day — Compression",
                        "Today's entire range fits within yesterday's high-low range. "
                        "The market is in equilibrium, with neither buyers nor sellers willing to extend the range. "
                        "Inside days signal compression before expansion — a directional move is building. "
                        "The breakout direction from the inside day often sets the short-term trend."
                    ),
                    "Bullish Context Candle": (
                        "#3fb950",
                        "Bullish Context Candle",
                        "A high-volume candle that closed in the top 25% of its range. "
                        "In auction market terms, the market tested lower prices but buyers rejected them decisively — "
                        "closing near the high shows the auction outcome was bullish. "
                        "Volume above average amplifies the significance of this signal."
                    ),
                    "Bearish Context Candle": (
                        "#f85149",
                        "Bearish Context Candle",
                        "High-volume candle closing in the bottom 25% of its range — sellers dominated the auction. "
                        "The market tested higher prices but rejected them, closing near the low."
                    ),
                    "PA Confluence": (
                        "#d29922",
                        "Price Action Confluence",
                        "Multiple price action signals are aligned on the same candle or within a close cluster of candles. "
                        "Confluence dramatically increases the probability of a signal — "
                        "when an SFP, engulfing candle, and bullish context candle all appear together, "
                        "it indicates the market has made a very decisive statement about direction."
                    ),
                }

                if interp_mode == "Full Scan Summary":
                    # Count PA pattern types
                    all_pa   = interp_df["pa_patterns"].dropna()
                    sfp_bull = all_pa.str.contains("Bullish SFP", na=False).sum()
                    sfp_bear = all_pa.str.contains("Bearish SFP", na=False).sum()
                    engulf   = all_pa.str.contains("Engulfing",   na=False).sum()
                    inside   = all_pa.str.contains("Inside Day",  na=False).sum()
                    context  = all_pa.str.contains("Context",     na=False).sum()
                    conflu   = all_pa.str.contains("Confluence",  na=False).sum()
                    any_pa   = (all_pa != "None").sum()

                    st.markdown(f"""
                    <div style="background:#161b22;border:1px solid #30363d;border-radius:10px;padding:20px 24px;margin-bottom:16px;">
                      <div style="font-weight:700;font-size:1rem;margin-bottom:12px;">Price Action Picture</div>
                      <p style="color:#c9d1d9;font-size:0.9rem;line-height:1.7;margin:0 0 12px 0;">
                        Price action patterns reveal what happened in the most recent session — the actual decisions made
                        by buyers and sellers. They complement technical indicators by showing real-time sentiment.
                      </p>
                      <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;">
                        <div><span style="color:#8b949e;font-size:0.75rem;">BULLISH SFP (TRAPS)</span><br>
                             <span style="font-size:1.4rem;font-weight:700;color:#3fb950;">{sfp_bull}</span></div>
                        <div><span style="color:#8b949e;font-size:0.75rem;">ENGULFING CANDLES</span><br>
                             <span style="font-size:1.4rem;font-weight:700;color:#3fb950;">{engulf}</span></div>
                        <div><span style="color:#8b949e;font-size:0.75rem;">PA CONFLUENCE</span><br>
                             <span style="font-size:1.4rem;font-weight:700;color:#d29922;">{conflu}</span></div>
                      </div>
                    </div>
                    """, unsafe_allow_html=True)

                    st.markdown("#### Price Action Per Stock")
                    for _, row in interp_df.head(15).iterrows():
                        tk       = row["ticker"]
                        pa_str   = str(row.get("pa_patterns","None"))
                        pa_sc    = int(row.get("pa_score",0) or 0)
                        price_v  = float(row.get("price",0))

                        if pa_str == "None" or not pa_str:
                            pa_html = (f'<div style="background:#0d1117;border:1px solid #21262d;border-radius:8px;'
                                      f'padding:14px 18px;margin:6px 0;">'
                                      f'<span style="font-weight:700;color:#e6edf3;">{tk}</span>&nbsp;&nbsp;'
                                      f'<span style="color:#8b949e;font-size:0.85rem;">No significant PA pattern on last candle — '
                                      f'check again after tomorrow\'s session.</span></div>')
                        else:
                            patterns_found = [p.strip() for p in pa_str.split("|")]
                            readings = []
                            for p in patterns_found:
                                if p in pa_explanations:
                                    c, name, expl = pa_explanations[p]
                                    readings.append(f'<span style="color:{c};font-weight:600;">{name}:</span> {expl}')
                            pa_html = (f'<div style="background:#0d1117;border:1px solid #21262d;border-radius:8px;'
                                      f'padding:14px 18px;margin:6px 0;">'
                                      f'<div style="margin-bottom:8px;">'
                                      f'<span style="font-weight:700;color:#e6edf3;">{tk}</span>&nbsp;&nbsp;'
                                      f'{amber(f"PA Score {pa_sc}/5")}</div>'
                                      f'<ul style="margin:0;padding-left:18px;color:#c9d1d9;font-size:0.87rem;line-height:1.8;">'
                                      + "".join(f"<li>{r}</li>" for r in readings) +
                                      f'</ul></div>')
                        st.markdown(pa_html, unsafe_allow_html=True)

                else:
                    row    = interp_df.iloc[0]
                    pa_str = str(row.get("pa_patterns","None"))
                    pa_sc  = int(row.get("pa_score",0) or 0)

                    if pa_str == "None" or not pa_str:
                        st.markdown(f"""
                        <div style="background:#161b22;border:1px solid #30363d;border-radius:10px;padding:20px 24px;">
                          <div style="color:#8b949e;font-size:0.9rem;">
                            No significant price action pattern detected on {interp_ticker}'s last candle.
                            This is not necessarily negative — it simply means the most recent session
                            did not produce a definitive signal. Check again after tomorrow's close.
                          </div>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        patterns_found = [p.strip() for p in pa_str.split("|")]
                        for p in patterns_found:
                            if p in pa_explanations:
                                c, name, expl = pa_explanations[p]
                                st.markdown(f"""
                                <div style="background:#161b22;border:1px solid {c};border-radius:10px;
                                            padding:20px 24px;margin-bottom:12px;">
                                  <div style="color:{c};font-weight:700;font-size:1rem;margin-bottom:10px;">
                                    {interp_ticker} — {name} (PA Score: {pa_sc}/5)
                                  </div>
                                  <p style="color:#c9d1d9;font-size:0.92rem;line-height:1.85;margin:0;">{expl}</p>
                                </div>
                                """, unsafe_allow_html=True)

        # ════════════════════════════════════════════════════════════════
        # VIEW 5 — FUNDAMENTALS (Alpha Vantage)
        # ════════════════════════════════════════════════════════════════
        with st.expander("📊 Fundamentals Interpretation (Alpha Vantage)", expanded=True):
            if "eps_growth_%" not in interp_df.columns or interp_df["eps_growth_%"].isna().all():
                st.markdown("""
                <div style="background:#161b22;border:1px solid #d29922;border-radius:8px;padding:16px 20px;">
                  <b style="color:#d29922;">⚠️ Alpha Vantage key not yet active</b><br>
                  <span style="color:#8b949e;font-size:0.9rem;">
                  Add your key to <code>config.yaml</code> under <code>alpha_vantage_key</code>
                  and run a fresh scan to unlock real EPS data, earnings surprises,
                  revenue growth, analyst targets and PE/PEG ratios for every stock.
                  </span>
                </div>
                """, unsafe_allow_html=True)
            else:
                if interp_mode == "Full Scan Summary":
                    strong_eps = interp_df[
                        pd.to_numeric(interp_df.get("eps_growth_%", pd.Series()), errors="coerce") >= 25
                    ]
                    accel_eps  = interp_df[interp_df.get("eps_accel", pd.Series([False]*len(interp_df))) == True]
                    beat_3plus = interp_df[
                        pd.to_numeric(interp_df.get("consec_beats", pd.Series()), errors="coerce") >= 3
                    ]

                    st.markdown(f"""
                    <div style="background:#161b22;border:1px solid #30363d;border-radius:10px;
                                padding:20px 24px;margin-bottom:16px;">
                      <div style="font-weight:700;font-size:1rem;margin-bottom:12px;">
                        Real Earnings Picture (Alpha Vantage Data)
                      </div>
                      <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;">
                        <div>
                          <span style="color:#8b949e;font-size:0.75rem;">EPS GROWTH ≥25%</span><br>
                          <span style="font-size:1.4rem;font-weight:700;color:#3fb950;">
                            {len(strong_eps)}
                          </span>
                          <div style="color:#8b949e;font-size:0.8rem;">
                            {", ".join(strong_eps["ticker"].head(5).tolist()) or "None"}
                          </div>
                        </div>
                        <div>
                          <span style="color:#8b949e;font-size:0.75rem;">ACCELERATING EPS</span><br>
                          <span style="font-size:1.4rem;font-weight:700;color:#d29922;">
                            {len(accel_eps)}
                          </span>
                          <div style="color:#8b949e;font-size:0.8rem;">
                            {", ".join(accel_eps["ticker"].head(5).tolist()) or "None"}
                          </div>
                        </div>
                        <div>
                          <span style="color:#8b949e;font-size:0.75rem;">3+ CONSEC BEATS</span><br>
                          <span style="font-size:1.4rem;font-weight:700;color:#388bfd;">
                            {len(beat_3plus)}
                          </span>
                          <div style="color:#8b949e;font-size:0.8rem;">
                            {", ".join(beat_3plus["ticker"].head(5).tolist()) or "None"}
                          </div>
                        </div>
                      </div>
                    </div>
                    """, unsafe_allow_html=True)

                    st.markdown("#### EPS Quality Per Stock")
                    for _, row in interp_df.head(15).iterrows():
                        tk      = row["ticker"]
                        eg      = row.get("eps_growth_%")
                        es      = row.get("eps_surprise_%")
                        accel   = bool(row.get("eps_accel", False))
                        beats   = row.get("consec_beats")
                        rev_g   = row.get("rev_growth_%")
                        earn_m  = str(row.get("earn_momentum","–"))
                        eps_sc  = row.get("eps_score", 0)
                        tgt     = row.get("analyst_target")
                        price_v = row.get("price",0)

                        sentences = []

                        if eg is not None:
                            if eg >= 50:
                                sentences.append(f"EPS growing at {eg:+.1f}% YoY — exceptional growth rate. This is the kind of fundamental acceleration that institutional investors pay premium multiples for.")
                            elif eg >= 25:
                                sentences.append(f"EPS growth of {eg:+.1f}% YoY is strong and above the 25% threshold that marks a leading growth stock.")
                            elif eg >= 0:
                                sentences.append(f"EPS growing {eg:+.1f}% YoY — positive but moderate. Look for acceleration in coming quarters.")
                            else:
                                sentences.append(f"EPS declining {eg:.1f}% YoY — earnings are contracting. Strong price momentum with declining earnings is a yellow flag.")

                        if es is not None:
                            if es >= 20:
                                sentences.append(f"Beat estimates by {es:.1f}% last quarter — a massive positive surprise. Institutional re-rating often follows large beats.")
                            elif es >= 5:
                                sentences.append(f"Beat estimates by {es:.1f}% — consistent with a company executing above expectations.")
                            elif es < 0:
                                sentences.append(f"Missed estimates by {abs(es):.1f}% last quarter — a miss can create selling pressure even in uptrending stocks.")

                        if accel:
                            sentences.append("EPS growth is accelerating quarter-over-quarter — the business is gaining momentum, not slowing down.")

                        if beats and beats >= 3:
                            sentences.append(f"{beats} consecutive quarterly beats. Management has a track record of under-promising and over-delivering — a quality signal.")

                        if rev_g is not None and rev_g >= 20:
                            sentences.append(f"Revenue growing {rev_g:+.1f}% YoY — top-line growth is real and substantial, not just cost-cutting driven.")

                        if tgt and price_v:
                            upside = (tgt / price_v - 1) * 100
                            sentences.append(f"Analyst consensus target ${tgt:.2f} implies {upside:+.1f}% upside from current price.")

                        if not sentences:
                            sentences.append("Fundamental data not available for this ticker from Alpha Vantage.")

                        badge_c = "#3fb950" if "Strong" in earn_m else ("#d29922" if "Moderate" in earn_m else "#8b949e")
                        earn_badge = f'<span style="background:{badge_c}22;color:{badge_c};border:1px solid {badge_c};border-radius:4px;padding:2px 8px;font-size:0.75rem;">{earn_m}</span>'

                        st.markdown(
                            f'<div style="background:#0d1117;border:1px solid #21262d;border-radius:8px;'
                            f'padding:14px 18px;margin:6px 0;">'
                            f'<div style="margin-bottom:8px;">'
                            f'<span style="font-weight:700;color:#e6edf3;">{tk}</span>'
                            f'&nbsp;&nbsp;{earn_badge}&nbsp;'
                            f'<span style="color:#8b949e;font-size:0.8rem;">EPS Score: {eps_sc}/15</span></div>'
                            f'<ul style="margin:0;padding-left:18px;color:#c9d1d9;font-size:0.87rem;line-height:1.8;">'
                            + "".join(f"<li>{s}</li>" for s in sentences) +
                            f'</ul></div>',
                            unsafe_allow_html=True
                        )

                else:
                    # Single ticker
                    row     = interp_df.iloc[0]
                    eg      = row.get("eps_growth_%")
                    es      = row.get("eps_surprise_%")
                    accel   = bool(row.get("eps_accel", False))
                    beats   = row.get("consec_beats")
                    rev_g   = row.get("rev_growth_%")
                    earn_m  = str(row.get("earn_momentum","–"))
                    eps_sc  = row.get("eps_score", 0)
                    tgt     = row.get("analyst_target")
                    pe      = row.get("pe_ratio")
                    peg     = row.get("peg_ratio")
                    eps_det = row.get("eps_details","–")
                    price_v = float(row.get("price",0))

                    fund_color = "#3fb950" if "Strong" in earn_m else ("#d29922" if "Moderate" in earn_m else "#f85149")

                    st.markdown(f"""
                    <div style="background:#161b22;border:1px solid {fund_color};
                                border-radius:10px;padding:20px 24px;">
                      <div style="color:{fund_color};font-weight:700;font-size:1rem;margin-bottom:14px;">
                        {interp_ticker} — Fundamental Profile (EPS Score: {eps_sc}/15)
                      </div>
                      <div style="display:grid;grid-template-columns:repeat(2,1fr);gap:20px;">
                        <div>
                          <div style="color:#8b949e;font-size:0.75rem;text-transform:uppercase;margin-bottom:8px;">
                            Earnings Quality
                          </div>
                          <p style="color:#c9d1d9;font-size:0.9rem;line-height:1.85;margin:0;">
                            {"✅" if (eg or 0)>=15 else "⚠️"} <b>EPS Growth YoY:</b>
                            {f"{eg:+.1f}%" if eg is not None else "No data"} —
                            {f"Exceptional growth at {eg:.0f}% — this stock is earning far more than it did a year ago." if (eg or 0)>=50
                             else f"Strong growth at {eg:.0f}% — above the 25% threshold for leading growth stocks." if (eg or 0)>=25
                             else f"Moderate growth of {eg:.0f}%." if (eg or 0)>=0
                             else f"Earnings declining {eg:.0f}% — watch this carefully alongside price action." if eg is not None
                             else "Not available."}<br><br>
                            {"✅" if (es or 0)>=5 else "⚠️"} <b>Last Surprise:</b>
                            {f"{es:+.1f}%" if es is not None else "No data"} —
                            {f"Beat by {es:.1f}% — institutional buyers often accumulate after large beats." if (es or 0)>=20
                             else f"Beat by {es:.1f}% — consistent execution." if (es or 0)>=5
                             else f"Missed by {abs(es or 0):.1f}% — misses can weigh on price even in uptrends." if es is not None
                             else "Not available."}<br><br>
                            {"✅" if accel else "–"} <b>Accelerating:</b>
                            {"Yes — growth rate is increasing quarter over quarter. This is what drives institutional re-rating." if accel
                             else "No acceleration yet. Steady growth but not yet building momentum in the earnings line."}<br><br>
                            {"✅" if (beats or 0)>=3 else "–"} <b>Consecutive Beats:</b>
                            {f"{beats} quarters" if beats else "No data"} —
                            {f"{beats} consecutive beats shows management consistently sets achievable targets and exceeds them — a trust signal for institutional investors." if (beats or 0)>=3
                             else "Less than 3 consecutive beats." if beats is not None else "Not available."}
                          </p>
                        </div>
                        <div>
                          <div style="color:#8b949e;font-size:0.75rem;text-transform:uppercase;margin-bottom:8px;">
                            Valuation & Revenue
                          </div>
                          <p style="color:#c9d1d9;font-size:0.9rem;line-height:1.85;margin:0;">
                            📈 <b>Revenue Growth:</b>
                            {f"{rev_g:+.1f}% YoY — top-line growth confirms the earnings improvement is driven by real business expansion, not just cost cuts." if rev_g is not None else "Not available."}<br><br>
                            💰 <b>PE Ratio:</b> {f"{pe:.1f}x — " if pe else "Not available — "}
                            {f"premium valuation reflecting growth expectations." if (pe or 0)>40
                             else f"reasonable for a growth company." if (pe or 0)>20
                             else f"relatively cheap valuation." if (pe or 0)>0
                             else ""}<br><br>
                            📐 <b>PEG Ratio:</b> {f"{peg:.2f} — " if peg else "Not available — "}
                            {f"below 1.0 = potentially undervalued relative to growth." if (peg or 0)<1 and (peg or 0)>0
                             else f"above 2.0 = growth is priced in." if (peg or 0)>2
                             else ""}<br><br>
                            🎯 <b>Analyst Target:</b>
                            {f"${tgt:.2f} = {(tgt/price_v-1)*100:+.1f}% from current ${price_v:.2f}" if tgt and price_v else "Not available."}
                          </p>
                        </div>
                      </div>
                    </div>
                    """, unsafe_allow_html=True)

        # ════════════════════════════════════════════════════════════════
        # COMPLETE DATA TABLE — every column, labelled + interpreted
        # Only shown for Single Ticker Deep Read
        # ════════════════════════════════════════════════════════════════
        if interp_mode == "Single Ticker Deep Read" and not interp_df.empty:
            with st.expander("📋 Complete Data Table — All Columns with Interpretation", expanded=False):
                row_data = interp_df.iloc[0]

                ORDERED_COLS_DISPLAY = [
                    "rank","ticker","market","theme","price","stage",
                    "perf_1m_%","perf_3m_%","perf_6m_%","rs_3m","rs_6m",
                    "rs_r2500_3m","rs_r2500_6m","rs_r3000g_3m","rs_r3000g_6m","rs_multi_leader",
                    "adr_%","vs_50ma_%","vs_200ma_%","volume","vol_filter","vol_surge_x",
                    "above_50ma","above_200ma","ma50_gt_ma200","near_52wh","pct_off_high_%",
                    "pattern","breaking_out","news_count","sentiment","earn_momentum",
                    "eps_growth_%","eps_surprise_%","eps_accel","consec_beats","rev_growth_%",
                    "eps_score","eps_trend","analyst_target","pe_ratio","peg_ratio","eps_details","next_earnings",
                    "of_bias","of_up_vol_ratio","of_bullish_days","of_consec_up","of_score",
                    "vwap","vwap_upper","vwap_lower","vs_vwap_%","vwap_position","vwap_slope","vwap_score",
                    "ms_structure","ms_hh_hl","ms_bos","ms_swing_high","ms_swing_low",
                    "pa_patterns","pa_engulfing","pa_sfp","pa_inside_day","pa_context","pa_score",
                    "apex_score","scanned_at","market_cap","market_cap_bn","mcap_category",
                    "is_gem","liquidity_score","liquidity_warn","avg_volume_30d",
                    "changes","is_new","delta_score",
                    # Weekly timeframe
                    "weekly_stage","weekly_above_10wma","weekly_above_40wma",
                    "weekly_10gt40","weekly_rs","weekly_base_tight",
                    "weekly_base_depth_%","weekly_hh_hl","weekly_trending_up",
                    "weekly_consec_up_wks","weekly_confirmed","weekly_contradicts",
                    "weekly_score",
                    # Early entry signals
                    "early_entry","early_entry_type","fresh_200ma_cross",
                    "fresh_50ma_cross","pullback_to_50ma","low_adr_base",
                    "early_entry_score","days_since_200ma_cross",
                ]

                def _fmt_cell(col, raw):
                    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                        return "–"
                    if col in ("price","vwap","vwap_upper","vwap_lower",
                               "ms_swing_high","ms_swing_low","analyst_target"):
                        try: return f"${float(raw):.2f}"
                        except: return str(raw)
                    if col in ("perf_1m_%","perf_3m_%","perf_6m_%","vs_50ma_%","vs_200ma_%",
                               "vs_vwap_%","eps_growth_%","eps_surprise_%","rev_growth_%","pct_off_high_%"):
                        try: return f"{float(raw):+.1f}%"
                        except: return str(raw)
                    if col == "adr_%":
                        try: return f"{float(raw):.1f}%"
                        except: return str(raw)
                    if col == "of_bullish_days":
                        try: return f"{float(raw):.0f}%"
                        except: return str(raw)
                    if col in ("rs_3m","rs_6m","rs_r2500_3m","rs_r2500_6m","rs_r3000g_3m","rs_r3000g_6m"):
                        try: return f"{float(raw):.0f}"
                        except: return str(raw)
                    if col in ("vol_surge_x","of_up_vol_ratio"):
                        try: return f"{float(raw):.2f}x"
                        except: return str(raw)
                    if col == "market_cap":
                        try:
                            v = float(raw)
                            return f"${v/1e9:.2f}B" if v>=1e9 else f"${v/1e6:.0f}M"
                        except: return str(raw)
                    if col in ("pe_ratio","peg_ratio","apex_score"):
                        try: return f"{float(raw):.1f}"
                        except: return str(raw)
                    if col == "eps_trend" and isinstance(raw, list):
                        return " → ".join(str(x) for x in raw[:6])
                    return str(raw)

                def _signal_color(col, raw):
                    """Return CSS colour string for a value."""
                    pos_cols = {
                        "stage":       lambda v: "2 ✅" in str(v),
                        "above_50ma":  lambda v: str(v).lower() in ("true","1"),
                        "above_200ma": lambda v: str(v).lower() in ("true","1"),
                        "ma50_gt_ma200":lambda v: str(v).lower() in ("true","1"),
                        "near_52wh":   lambda v: str(v).lower() in ("true","1"),
                        "breaking_out":lambda v: str(v).lower() in ("true","1"),
                        "ms_hh_hl":    lambda v: str(v).lower() in ("true","1"),
                        "ms_bos":      lambda v: str(v).lower() in ("true","1"),
                        "of_bias":     lambda v: "Bullish" in str(v),
                        "vwap_position":lambda v: "Above" in str(v) and "Extended" not in str(v),
                        "ms_structure":lambda v: "Bullish" in str(v),
                        "earn_momentum":lambda v: "Strong" in str(v),
                        "eps_accel":   lambda v: str(v).lower() in ("true","1"),
                        "is_gem":         lambda v: str(v).lower() in ("true","1"),
                        "breaking_out":    lambda v: str(v).lower() in ("true","1"),
                        "rs_multi_leader":    lambda v: str(v).lower() in ("true","1"),
                        "early_entry":         lambda v: str(v).lower() in ("true","1"),
                        "fresh_200ma_cross":   lambda v: str(v).lower() in ("true","1"),
                        "fresh_50ma_cross":    lambda v: str(v).lower() in ("true","1"),
                        "pullback_to_50ma":    lambda v: str(v).lower() in ("true","1"),
                        "low_adr_base":        lambda v: str(v).lower() in ("true","1"),
                        "weekly_confirmed":    lambda v: str(v).lower() in ("true","1"),
                        "weekly_above_40wma":  lambda v: str(v).lower() in ("true","1"),
                        "weekly_10gt40":       lambda v: str(v).lower() in ("true","1"),
                        "weekly_hh_hl":        lambda v: str(v).lower() in ("true","1"),
                        "weekly_base_tight":   lambda v: str(v).lower() in ("true","1"),
                        "weekly_trending_up":  lambda v: str(v).lower() in ("true","1"),
                    }
                    _neg_cols_extended = {
                        "weekly_contradicts":  lambda v: str(v).lower() in ("true","1"),
                    }
                    try:
                        if "weekly_contradicts" in disp.columns:
                            styled = styled.map(
                                lambda v: "color:#f85149;font-weight:700" if str(v).lower() in ("true","1") else "",
                                subset=["weekly_contradicts"]
                            )
                    except Exception:
                        pass
                    neg_cols = {
                        "stage":       lambda v: "4 🔴" in str(v),
                        "above_50ma":  lambda v: str(v).lower() in ("false","0"),
                        "above_200ma": lambda v: str(v).lower() in ("false","0"),
                        "of_bias":     lambda v: "Bearish" in str(v),
                        "vwap_position":lambda v: "Below" in str(v),
                        "ms_structure":lambda v: "Bearish" in str(v),
                        "liquidity_warn":lambda v: str(v).lower() in ("true","1"),
                    }
                    try:
                        if col in pos_cols and pos_cols[col](raw): return "#3fb950"
                        if col in neg_cols and neg_cols[col](raw): return "#f85149"
                    except: pass
                    # Numeric coloring
                    perf_cols = ("perf_1m_%","perf_3m_%","perf_6m_%","vs_50ma_%","vs_200ma_%",
                                 "eps_growth_%","eps_surprise_%","rev_growth_%","delta_score")
                    if col in perf_cols:
                        try:
                            v = float(raw)
                            return "#3fb950" if v > 0 else ("#f85149" if v < 0 else "#8b949e")
                        except: pass
                    if col == "apex_score":
                        try:
                            v = float(raw)
                            return "#3fb950" if v>=70 else ("#d29922" if v>=40 else "#f85149")
                        except: pass
                    if col in ("rs_3m","rs_6m","rs_r2500_3m","rs_r2500_6m","rs_r3000g_3m","rs_r3000g_6m"):
                        try:
                            v = float(raw)
                            return "#3fb950" if v>=100 else ("#d29922" if v>=70 else "#f85149")
                        except: pass
                    return "#c9d1d9"

                # Render the table
                st.markdown("""
                <div style="font-size:0.78rem;color:#8b949e;margin-bottom:8px;">
                  Every field from the scan — what it means and how to read it.
                  <span style="color:#3fb950;">Green</span> = bullish signal &nbsp;|&nbsp;
                  <span style="color:#f85149;">Red</span> = bearish/caution &nbsp;|&nbsp;
                  <span style="color:#d29922;">Amber</span> = neutral/watch
                </div>
                """, unsafe_allow_html=True)

                for col_key in ORDERED_COLS_DISPLAY:
                    raw = row_data.get(col_key)
                    if col_key == "rank":
                        raw = row_data.name if hasattr(row_data, "name") else "–"

                    if raw is None and col_key not in ("delta_score","changes","is_new"):
                        continue  # Skip columns with no data

                    meta   = COLUMN_META.get(col_key, (col_key, "–"))
                    label  = meta[0]
                    interp = meta[1]
                    val    = _fmt_cell(col_key, raw)
                    vcolor = _signal_color(col_key, raw)

                    st.markdown(
                        f'<div style="display:flex;align-items:flex-start;gap:12px;'
                        f'border-bottom:1px solid #21262d;padding:8px 0;">'
                        f'<div style="min-width:160px;flex-shrink:0;">'
                        f'<span style="color:#8b949e;font-size:0.72rem;text-transform:uppercase;'
                        f'letter-spacing:0.06em;">{col_key}</span><br>'
                        f'<span style="color:#d29922;font-size:0.82rem;font-weight:600;">{label}</span>'
                        f'</div>'
                        f'<div style="min-width:120px;flex-shrink:0;">'
                        f'<span style="color:{vcolor};font-size:0.95rem;font-weight:700;">{val}</span>'
                        f'</div>'
                        f'<div style="flex:1;">'
                        f'<span style="color:#8b949e;font-size:0.82rem;line-height:1.6;">{interp}</span>'
                        f'</div>'
                        f'</div>',
                        unsafe_allow_html=True
                    )

                # Excel Download
                st.markdown("---")
                try:
                    xlsx_bytes = build_excel_download(interp_df.iloc[0], interp_ticker)
                    st.download_button(
                        "⬇ Download Excel — Full Deep Read",
                        xlsx_bytes,
                        file_name=f"apexscan_{interp_ticker}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Downloads a fully labelled Excel workbook: Sheet 1 = all data columns with interpretation, Sheet 2 = signal scorecard",
                    )
                except Exception as _xl_err:
                    # Fallback to CSV
                    st.download_button(
                        "⬇ Download CSV — Deep Read",
                        interp_df.to_csv().encode("utf-8"),
                        file_name=f"apexscan_{interp_ticker}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                        mime="text/csv",
                    )
                    st.caption(f"Excel export failed ({_xl_err}) — CSV downloaded instead. Install openpyxl for Excel output.")

        # ── Combined summary for single ticker ─────────────────────────
        if interp_mode == "Single Ticker Deep Read" and not interp_df.empty:
            st.markdown("---")
            st.markdown("#### 🎯 Combined Signal Summary")
            row    = interp_df.iloc[0]
            score  = float(row.get("apex_score",0))
            stage  = str(row.get("stage","?"))
            bias   = str(row.get("of_bias","–"))
            vwap_p = str(row.get("vwap_position","–"))
            ms_s   = str(row.get("ms_structure","–"))
            pa_s   = str(row.get("pa_patterns","None"))

            signals_green = sum([
                "2 ✅" in stage,
                "Bullish" in bias,
                "Above VWAP" in vwap_p and "Extended" not in vwap_p,
                "Bullish" in ms_s,
                "Bullish" in pa_s,
            ])

            if signals_green >= 4:
                verdict_html = f'<span style="color:#3fb950;font-weight:700;font-size:1.1rem;">HIGH CONVICTION SETUP ({signals_green}/5 signals aligned)</span>'
                verdict_text = f"All major signal categories are aligned bullishly for {interp_ticker}. This is the type of setup where multiple independent frameworks agree — the highest conviction scenario."
            elif signals_green >= 3:
                verdict_html = f'<span style="color:#d29922;font-weight:700;font-size:1.1rem;">MODERATE CONVICTION ({signals_green}/5 signals aligned)</span>'
                verdict_text = f"Most signals are positive for {interp_ticker} but some are missing. Worth monitoring closely. Wait for the remaining signals to confirm before committing full position size."
            else:
                verdict_html = f'<span style="color:#f85149;font-weight:700;font-size:1.1rem;">LOW CONVICTION — WAIT ({signals_green}/5 signals aligned)</span>'
                verdict_text = f"Too many signals are not yet aligned for {interp_ticker}. The setup needs more time to develop. Patience here protects capital for when conditions improve."

            st.markdown(f"""
            <div style="background:#161b22;border:1px solid #30363d;border-radius:12px;padding:24px;margin-top:8px;">
              <div style="margin-bottom:12px;">{verdict_html}</div>
              <p style="color:#c9d1d9;font-size:0.92rem;line-height:1.8;margin:0 0 16px 0;">{verdict_text}</p>
              <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:8px;">
                <div style="background:#0d1117;border-radius:6px;padding:10px;text-align:center;">
                  <div style="font-size:1.2rem;">{'✅' if '2 ✅' in stage else '❌'}</div>
                  <div style="color:#8b949e;font-size:0.7rem;margin-top:4px;">STAGE 2</div>
                </div>
                <div style="background:#0d1117;border-radius:6px;padding:10px;text-align:center;">
                  <div style="font-size:1.2rem;">{'✅' if 'Bullish' in bias else '❌'}</div>
                  <div style="color:#8b949e;font-size:0.7rem;margin-top:4px;">ORDER FLOW</div>
                </div>
                <div style="background:#0d1117;border-radius:6px;padding:10px;text-align:center;">
                  <div style="font-size:1.2rem;">{'✅' if 'Above VWAP' in vwap_p and 'Extended' not in vwap_p else '❌'}</div>
                  <div style="color:#8b949e;font-size:0.7rem;margin-top:4px;">ABOVE VWAP</div>
                </div>
                <div style="background:#0d1117;border-radius:6px;padding:10px;text-align:center;">
                  <div style="font-size:1.2rem;">{'✅' if 'Bullish' in ms_s else '❌'}</div>
                  <div style="color:#8b949e;font-size:0.7rem;margin-top:4px;">HH/HL</div>
                </div>
                <div style="background:#0d1117;border-radius:6px;padding:10px;text-align:center;">
                  <div style="font-size:1.2rem;">{'✅' if 'Bullish' in pa_s else '❌'}</div>
                  <div style="color:#8b949e;font-size:0.7rem;margin-top:4px;">PRICE ACTION</div>
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 17 — GUIDE
# ══════════════════════════════════════════════════════════════════════════════

with tabs[16]:
    st.markdown("### 📊 Scan Delta — Momentum Tracker")
    st.caption("Compares scan results over time — stocks consistently improving Apex Score are the ones to watch.")

    _reports_dir  = Path(__file__).resolve().parent / "reports"
    _report_files = sorted(_reports_dir.glob("scan_*.csv"), reverse=True) if _reports_dir.exists() else []

    if len(_report_files) < 2:
        st.info(
            f"Need at least **2 saved scans** to compare. "
            f"Currently {len(_report_files)} scan(s) saved. "
            "Run a scan now, then run another after the next session."
        )
    else:
        _file_labels = [f.stem.replace("scan_","").replace("_"," ") for f in _report_files]
        dc1,dc2,dc3 = st.columns(3)
        with dc1:
            _sel_new = st.selectbox("Newer Scan", _file_labels, index=0, key="delta_new")
        with dc2:
            _sel_old = st.selectbox("Older Scan", _file_labels, index=min(1,len(_file_labels)-1), key="delta_old")
        with dc3:
            _min_chg = st.number_input("Min Score Change to Flag", value=5, min_value=1, max_value=50, key="delta_min")

        _nf = _report_files[_file_labels.index(_sel_new)]
        _of = _report_files[_file_labels.index(_sel_old)]

        if _nf == _of:
            st.warning("Select two different scans to compare.")
        else:
            try:
                _dn = pd.read_csv(_nf)
                _do = pd.read_csv(_of)
                for _d in [_dn, _do]:
                    if "rank" in _d.columns: _d.drop(columns=["rank"], inplace=True, errors="ignore")

                _rows = []
                _all_tks = set(_dn["ticker"].tolist()) | set(_do["ticker"].tolist())

                for _tk in _all_tks:
                    _rn = _dn[_dn["ticker"]==_tk]
                    _ro = _do[_do["ticker"]==_tk]
                    _sn = float(_rn["apex_score"].iloc[0]) if not _rn.empty and "apex_score" in _rn.columns else None
                    _so = float(_ro["apex_score"].iloc[0]) if not _ro.empty and "apex_score" in _ro.columns else None
                    _dlt = round(_sn - _so, 1) if (_sn is not None and _so is not None) else None

                    if _dlt is not None and abs(_dlt) < _min_chg and not _rn.empty and not _ro.empty:
                        continue

                    _r  = _rn.iloc[0] if not _rn.empty else _ro.iloc[0]
                    _status = (
                        "🆕 New Entry"   if _ro.empty else
                        "❌ Dropped Out" if _rn.empty else
                        ("🚀 Strong Rise" if (_dlt or 0) >= 10 else
                         "📈 Rising"      if (_dlt or 0) >= _min_chg else
                         ("📉 Falling"    if (_dlt or 0) <= -_min_chg else "📊 Steady"))
                    )
                    _rows.append({
                        "Ticker":    _tk,
                        "Theme":     str(_r.get("theme","–")),
                        "Status":    _status,
                        "New Score": _sn,
                        "Old Score": _so,
                        "Δ Score":   _dlt,
                        "Stage":     str(_r.get("stage","–")),
                        "3M %":      _r.get("perf_3m_%"),
                        "RS (3m)":   _r.get("rs_3m"),
                        "OF Bias":   str(_r.get("of_bias","–")),
                        "Breaking":  "✅" if str(_r.get("breaking_out","")).lower()=="true" else "–",
                        "Price":     _r.get("price"),
                    })

                if not _rows:
                    st.info("No stocks changed beyond the minimum threshold.")
                else:
                    _ddf = pd.DataFrame(_rows).sort_values("Δ Score", ascending=False, na_position="last")

                    # KPIs
                    _ri = (_ddf["Δ Score"]>0).sum()
                    _fa = (_ddf["Δ Score"]<0).sum()
                    _ne = (_ddf["Status"]=="🆕 New Entry").sum()
                    _dr = (_ddf["Status"]=="❌ Dropped Out").sum()
                    _ac = _ddf["Δ Score"].dropna().mean()
                    v1,v2,v3,v4,v5 = st.columns(5)
                    v1.metric("📈 Rising", _ri)
                    v2.metric("📉 Falling", _fa)
                    v3.metric("🆕 New", _ne)
                    v4.metric("❌ Dropped", _dr)
                    v5.metric("Avg Δ Score", f"{_ac:+.1f}" if not pd.isna(_ac) else "–")

                    st.markdown("---")

                    def _safe_fmt(v, fmt, fallback="–"):
                        """Safely format a value, returning fallback for NaN/None/non-numeric."""
                        if v is None: return fallback
                        try:
                            f = float(v)
                            import math
                            if math.isnan(f): return fallback
                            return fmt.format(f)
                        except (TypeError, ValueError):
                            return fallback

                    def _fmt_delta(df_in):
                        def _c(v):
                            try: return "color:#3fb950;font-weight:700" if float(v)>0 else "color:#f85149;font-weight:700"
                            except: return ""
                        return df_in.style.map(_c, subset=[c for c in ["Δ Score","3M %"] if c in df_in.columns]).format({
                            "New Score": lambda v: _safe_fmt(v, "{:.0f}"),
                            "Old Score": lambda v: _safe_fmt(v, "{:.0f}"),
                            "Δ Score":   lambda v: _safe_fmt(v, "{:+.1f}", "NEW"),
                            "3M %":      lambda v: _safe_fmt(v, "{:+.1f}%"),
                            "RS (3m)":   lambda v: _safe_fmt(v, "{:.0f}"),
                            "Price":     lambda v: _safe_fmt(v, "${:.2f}"),
                        }, na_rep="–")

                    _t1,_t2,_t3,_t4,_t5 = st.tabs(["📈 Risers","📉 Fallers","🆕 New","❌ Dropped","📋 All"])

                    with _t1:
                        _rise = _ddf[_ddf["Δ Score"]>0]
                        if _rise.empty: st.info("No risers.")
                        else:
                            for _,_gr in _rise.head(3).iterrows():
                                _gc = "#3fb950" if float(_gr["Δ Score"] or 0)>=10 else "#d29922"
                                _old_s = f"{_gr['Old Score']:.0f}" if pd.notna(_gr.get("Old Score")) else "–"
                                _new_s = f"{_gr['New Score']:.0f}" if pd.notna(_gr.get("New Score")) else "–"
                                st.markdown(
                                    f'<div style="background:#0d1117;border-left:4px solid {_gc};padding:10px 16px;margin:4px 0;border-radius:4px;">'
                                    f'<b style="color:#e6edf3">{_gr["Ticker"]}</b>'
                                    f'<span style="color:{_gc};font-weight:700;margin-left:12px">{float(_gr["Δ Score"]):+.1f} pts</span>'
                                    f'<span style="color:#8b949e;font-size:.85rem;margin-left:12px">'
                                    f'{_old_s} → {_new_s} | {_gr["Theme"]}</span></div>',
                                    unsafe_allow_html=True
                                )
                            st.dataframe(_fmt_delta(_rise), use_container_width=True, hide_index=True)
                    with _t2:
                        _fall = _ddf[_ddf["Δ Score"]<0]
                        if _fall.empty: st.info("No fallers.")
                        else: st.dataframe(_fmt_delta(_fall), use_container_width=True, hide_index=True)
                    with _t3:
                        _new = _ddf[_ddf["Status"]=="🆕 New Entry"]
                        if _new.empty: st.info("No new entries.")
                        else: st.dataframe(_fmt_delta(_new), use_container_width=True, hide_index=True)
                    with _t4:
                        _drp = _ddf[_ddf["Status"]=="❌ Dropped Out"]
                        if _drp.empty: st.info("No dropped stocks.")
                        else: st.dataframe(_fmt_delta(_drp), use_container_width=True, hide_index=True)
                    with _t5:
                        st.dataframe(_fmt_delta(_ddf), use_container_width=True, hide_index=True)
                        st.download_button(
                            "⬇ Download Delta Report (CSV)",
                            data=_ddf.to_csv(index=False).encode("utf-8"),
                            file_name=f"apexscan_delta_{_sel_new.replace(' ','_')}_vs_{_sel_old.replace(' ','_')}.csv",
                            mime="text/csv",
                        )

                    # Score trajectory chart
                    st.markdown("---")
                    st.markdown("#### 📈 Score Trajectory — Top Movers")
                    _top_m = pd.concat([
                        _ddf.dropna(subset=["Δ Score"]).nlargest(5,"Δ Score"),
                        _ddf.dropna(subset=["Δ Score"]).nsmallest(5,"Δ Score"),
                    ]).drop_duplicates("Ticker")
                    if not _top_m.empty:
                        _cd = []
                        for _,_r in _top_m.iterrows():
                            if pd.notna(_r.get("Old Score")): _cd.append({"Ticker":_r["Ticker"],"Score":_r["Old Score"],"Scan":f"Older ({_sel_old})"})
                            if pd.notna(_r.get("New Score")): _cd.append({"Ticker":_r["Ticker"],"Score":_r["New Score"],"Scan":f"Newer ({_sel_new})"})
                        _cdf = pd.DataFrame(_cd)
                        if not _cdf.empty:
                            _fig_t = px.line(_cdf, x="Scan", y="Score", color="Ticker",
                                title="Apex Score — Before vs After", markers=True)
                            _fig_t.update_layout(paper_bgcolor="#0d1117",plot_bgcolor="#0d1117",
                                font_color="#e6edf3",height=360,
                                yaxis=dict(gridcolor="#21262d",range=[0,100]),
                                xaxis=dict(gridcolor="#21262d"),
                                margin=dict(t=40,b=20,l=20,r=20))
                            st.plotly_chart(_fig_t, use_container_width=True)

            except Exception as _de:
                st.error(f"Error comparing scans: {_de}")


with tabs[18]:
    st.markdown("### ✅ Pre-Buy Checklist")
    st.caption(
        "Run every stock through this 15-point checklist before buying. "
        "Fail items 1–5 = do not buy. Fail 6–9 = reduce size or wait. "
        "Fail 10–12 = wait for better timing."
    )

    if df.empty:
        st.info("Run a Live Scan or Load Last Report first to populate ticker list.")
    else:
        # ══════════════════════════════════════════════════════════════════════
        # FILTER PANEL — pre-screen all scan stocks before opening checklist
        # ══════════════════════════════════════════════════════════════════════
        with st.expander("🔍 Filter Stocks Before Running Checklist", expanded=True):
            _f1, _f2, _f3, _f4 = st.columns(4)
            with _f1:
                _chk_score_min = st.slider(
                    "Min Apex Score", min_value=0, max_value=100,
                    value=40, step=5, key="chk_score_filter",
                    help="Only show stocks scoring above this threshold."
                )
            with _f2:
                _chk_score_max = st.slider(
                    "Max Apex Score", min_value=0, max_value=100,
                    value=100, step=5, key="chk_score_max_filter",
                    help="Only show stocks scoring below this threshold."
                )
            with _f3:
                _chk_stage_filter = st.multiselect(
                    "Stage Filter",
                    ["Stage 2 ✅", "Stage 1 ⏳", "Stage 3 ⚠️", "Stage 4 🔴"],
                    default=["Stage 2 ✅"],
                    key="chk_stage_filter",
                    help="Only show stocks in selected Weinstein stages. Stage 2 = uptrend only."
                )
            with _f4:
                _chk_gem_only = st.checkbox(
                    "💎 Gems Only", value=False, key="chk_gem_only",
                    help="Only show emerging gem stocks (small/mid-cap)."
                )

            # ── Quick-verdict pre-screener ────────────────────────────────────
            st.markdown("##### 🎯 Quick Verdict Pre-Screen")
            st.caption(
                "Runs a fast lightweight check on every stock in the scan to estimate "
                "their verdict category — without opening each one individually."
            )

            _pv1, _pv2, _pv3, _pv4, _pv5 = st.columns(5)
            _verdict_filter = _pv1.radio(
                "Show Only",
                ["All", "✅ Ready", "⏳ Wait", "⚠️ Reduced", "🚫 Skip/Avoid"],
                index=0, key="chk_verdict_filter",
                help="Filter by estimated verdict. 'All' = show everything."
            )

            # ── Pre-screen all tickers against fast rules ──────────────────────
            def _quick_verdict(row_q):
                """
                Fast verdict estimate from scan columns — no live data needed.
                Mirrors the full checklist logic at a high level.
                """
                try:
                    _score  = float(row_q.get("apex_score", 0) or 0)
                    _stage  = str(row_q.get("stage","") or "")
                    _ab200  = str(row_q.get("above_200ma","")).lower() in ("true","1")
                    _ma50g  = str(row_q.get("ma50_gt_ma200","")).lower() in ("true","1")
                    _rs     = float(row_q.get("rs_3m", 0) or 0)
                    _wconf  = str(row_q.get("weekly_confirmed","")).lower() in ("true","1")
                    _wcontr = str(row_q.get("weekly_contradicts","")).lower() in ("true","1")
                    _break  = str(row_q.get("breaking_out","")).lower() in ("true","1")
                    _early  = str(row_q.get("early_entry","")).lower() in ("true","1")
                    _of     = str(row_q.get("of_bias","") or "").lower()
                    _liq    = str(row_q.get("liquidity_warn","")).lower() in ("true","1")
                    _vwap_p = str(row_q.get("vwap_position","") or "").lower()

                    # Hard fails → DO NOT BUY
                    if not (_ab200 and _ma50g):    return "🚫 Do Not Buy", "#f85149"
                    if _wcontr:                    return "🚫 Do Not Buy", "#f85149"
                    if _rs < 0:                    return "🚫 Do Not Buy", "#f85149"
                    if _liq:                       return "🚫 Do Not Buy", "#f85149"

                    # Quality checks
                    _quality_fails = 0
                    if _score < 65:                _quality_fails += 1
                    if "bullish" not in _of:       _quality_fails += 1
                    if not _wconf:                 _quality_fails += 1

                    # Entry timing
                    _timing_ok = _break or _early or "above" in _vwap_p

                    if _quality_fails == 0 and _timing_ok and _score >= 65:
                        return "✅ Ready to Buy", "#3fb950"
                    elif _quality_fails == 0 and not _timing_ok:
                        return "⏳ Wait for Entry", "#388bfd"
                    elif _quality_fails <= 2:
                        return "⚠️ Reduced Size", "#d29922"
                    else:
                        return "🚫 Skip", "#f85149"
                except Exception:
                    return "– Unknown", "#8b949e"

            # Build filtered ticker list
            _all_tickers_df = df.copy() if not df.empty else pd.DataFrame()

            # Apply score filter
            if "apex_score" in _all_tickers_df.columns:
                _sc = pd.to_numeric(_all_tickers_df["apex_score"], errors="coerce")
                _all_tickers_df = _all_tickers_df[
                    (_sc >= _chk_score_min) & (_sc <= _chk_score_max)
                ]

            # Apply stage filter
            if _chk_stage_filter and "stage" in _all_tickers_df.columns:
                _stage_mask = _all_tickers_df["stage"].astype(str).apply(
                    lambda s: any(f.replace("Stage ","").split(" ")[0] in s for f in _chk_stage_filter)
                )
                _all_tickers_df = _all_tickers_df[_stage_mask]

            # Apply gem filter
            if _chk_gem_only and "is_gem" in _all_tickers_df.columns:
                _all_tickers_df = _all_tickers_df[
                    _all_tickers_df["is_gem"].astype(str).str.lower().isin(["true","1"])
                ]

            # Run quick verdict on all filtered rows
            _verdict_data = []
            for _, _qrow in _all_tickers_df.iterrows():
                _qv, _qc = _quick_verdict(_qrow)
                _verdict_data.append({
                    "ticker":      _qrow.get("ticker","–"),
                    "verdict":     _qv,
                    "color":       _qc,
                    "apex_score":  _qrow.get("apex_score","–"),
                    "stage":       str(_qrow.get("stage","–")),
                    "rs_3m":       _qrow.get("rs_3m","–"),
                    "of_bias":     str(_qrow.get("of_bias","–")),
                    "breaking_out":str(_qrow.get("breaking_out","–")),
                    "theme":       str(_qrow.get("theme","–")),
                    "early_entry": str(_qrow.get("early_entry","–")),
                })

            # Apply verdict filter
            _filtered_verdicts = _verdict_data
            if _verdict_filter != "All":
                _vmap = {
                    "✅ Ready":    "✅ Ready to Buy",
                    "⏳ Wait":     "⏳ Wait for Entry",
                    "⚠️ Reduced":  "⚠️ Reduced Size",
                    "🚫 Skip/Avoid":"🚫 Skip",
                }
                _vkey = _vmap.get(_verdict_filter, "")
                _filtered_verdicts = [v for v in _verdict_data
                                       if _vkey.split(" ")[0] in v["verdict"]]

            # ── Summary badges ────────────────────────────────────────────────
            _ready_n   = sum(1 for v in _verdict_data if "Ready"   in v["verdict"])
            _wait_n    = sum(1 for v in _verdict_data if "Wait"    in v["verdict"])
            _reduced_n = sum(1 for v in _verdict_data if "Reduced" in v["verdict"])
            _skip_n    = sum(1 for v in _verdict_data if "Skip" in v["verdict"] or "Do Not" in v["verdict"])

            _b1,_b2,_b3,_b4 = st.columns(4)
            _b1.metric("✅ Ready to Buy",  _ready_n,   help="Pass all hard gates + quality")
            _b2.metric("⏳ Wait for Entry", _wait_n,   help="Good quality, timing not ideal")
            _b3.metric("⚠️ Reduced Size",   _reduced_n, help="1–2 quality issues")
            _b4.metric("🚫 Do Not Buy/Skip",_skip_n,   help="Hard gate failure or too many issues")

            # ── Verdict grid ──────────────────────────────────────────────────
            if _filtered_verdicts:
                st.markdown(f"**{len(_filtered_verdicts)} stocks** match your filters:")

                # Render as a clean colour-coded grid
                _cols_per_row = 4
                for _row_start in range(0, len(_filtered_verdicts), _cols_per_row):
                    _row_items = _filtered_verdicts[_row_start:_row_start+_cols_per_row]
                    _gcols = st.columns(_cols_per_row)
                    for _gi, _gitem in enumerate(_row_items):
                        _gc = _gitem["color"]
                        _gs = _gitem["apex_score"]
                        _gv = _gitem["verdict"]
                        _gcols[_gi].markdown(
                            f'<div style="background:#161b22;border:1px solid {_gc};'
                            f'border-radius:8px;padding:10px;text-align:center;cursor:pointer;">'
                            f'<div style="color:#e6edf3;font-weight:700;font-size:0.95rem;">'
                            f'{_gitem["ticker"]}</div>'
                            f'<div style="color:{_gc};font-size:0.75rem;font-weight:600;margin-top:2px;">'
                            f'{_gv}</div>'
                            f'<div style="color:#8b949e;font-size:0.72rem;margin-top:2px;">'
                            f'Score: {_gs} | {_gitem["stage"].split()[0] if _gitem["stage"] != "–" else "–"}'
                            f'</div></div>',
                            unsafe_allow_html=True
                        )
            else:
                st.info("No stocks match the current filters.")

        st.markdown("---")

        # ── Ticker selector ───────────────────────────────────────────────────
        # Pre-filter the dropdown to only show filtered tickers
        _chk_tickers_filtered = (
            [v["ticker"] for v in _filtered_verdicts]
            if _filtered_verdicts else
            (sorted(df["ticker"].dropna().unique().tolist()) if "ticker" in df.columns else [])
        )

        cb1, cb2 = st.columns([2, 1])
        with cb1:
            _chk_tickers = _chk_tickers_filtered if _chk_tickers_filtered else                            (sorted(df["ticker"].dropna().unique().tolist()) if "ticker" in df.columns else [])
            chk_ticker = st.selectbox(
                "Select Ticker to Evaluate",
                _chk_tickers,
                key="chk_ticker_sel",
                help="Dropdown is pre-filtered by your filter settings above. Change filters to see more stocks."
            )
        with cb2:
            chk_account = st.number_input(
                "Portfolio Size ($)",
                min_value=1000, value=10000, step=1000,
                key="chk_portfolio_size",
                help="Your total portfolio value — used to calculate maximum position size."
            )

        if chk_ticker:
            _crow = df[df["ticker"] == chk_ticker]
            if _crow.empty:
                st.error(f"{chk_ticker} not found in scan results.")
            else:
                row = _crow.iloc[0]

                # ── Live market context (S&P 500 stage + VIX) ────────────────────
                @st.cache_data(ttl=900)   # cache 15 min
                def _get_market_context():
                    try:
                        import yfinance as _yf
                        _spx  = _yf.Ticker("^GSPC").history(period="1y")["Close"]
                        _vix  = _yf.Ticker("^VIX").history(period="5d")["Close"]
                        _spx_cur  = float(_spx.iloc[-1])
                        _spx_50   = float(_spx.rolling(50).mean().iloc[-1])
                        _spx_200  = float(_spx.rolling(200).mean().iloc[-1])
                        _vix_cur  = float(_vix.iloc[-1])
                        _spx_abv50  = _spx_cur > _spx_50
                        _spx_abv200 = _spx_cur > _spx_200
                        _spx_50gt200= _spx_50   > _spx_200
                        if _spx_abv50 and _spx_abv200 and _spx_50gt200:
                            _mkt_stage = "Stage 2 ✅ Uptrend"
                        elif _spx_abv200:
                            _mkt_stage = "Stage 1 ⏳ Basing"
                        elif not _spx_abv200 and not _spx_50gt200:
                            _mkt_stage = "Stage 4 🔴 Downtrend"
                        else:
                            _mkt_stage = "Stage 3 ⚠️ Topping"
                        return {
                            "stage": _mkt_stage, "price": _spx_cur,
                            "ma50": _spx_50, "ma200": _spx_200,
                            "above_50": _spx_abv50, "above_200": _spx_abv200,
                            "ma50gt200": _spx_50gt200, "vix": _vix_cur,
                            "market_ok": _spx_abv200 and _spx_50gt200,
                        }
                    except Exception:
                        return {"stage": "Unknown", "market_ok": True, "vix": None,
                                "above_50": True, "above_200": True, "ma50gt200": True}

                @st.cache_data(ttl=900)
                def _get_sector_strength(sector_theme: str):
                    """Check if the stock's sector ETF is in an uptrend."""
                    _SECTOR_ETFS = {
                        "Information Technology": "XLK",
                        "Healthcare":             "XLV",
                        "Financials":             "XLF",
                        "Consumer Discretionary": "XLY",
                        "Consumer Staples":       "XLP",
                        "Energy":                 "XLE",
                        "Materials":              "XLB",
                        "Industrials":            "XLI",
                        "Utilities":              "XLU",
                        "Real Estate":            "XLRE",
                        "Communication Services": "XLC",
                        "ai_semis":               "SOXX",
                        "cybersecurity":          "CIBR",
                        "biotech_health":         "XBI",
                        "fintech_payments":       "ARKF",
                        "cloud_infra":            "WCLD",
                        "emerging_gems":          "IWM",
                        "space_ev":               "ROKT",
                    }
                    _etf = _SECTOR_ETFS.get(sector_theme)
                    if not _etf:
                        return {"etf": None, "sector_ok": True, "sector_stage": "Unknown", "sector_rs": None}
                    try:
                        import yfinance as _yf
                        _d = _yf.Ticker(_etf).history(period="1y")["Close"]
                        _cur   = float(_d.iloc[-1])
                        _ma50  = float(_d.rolling(50).mean().iloc[-1])
                        _ma200 = float(_d.rolling(200).mean().iloc[-1])
                        _above_both = _cur > _ma50 and _cur > _ma200 and _ma50 > _ma200
                        _above_200  = _cur > _ma200
                        _stage = (
                            "Stage 2 ✅" if (_cur>_ma50>_ma200) else
                            "Stage 1 ⏳" if _above_200 else
                            "Stage 4 🔴" if (_cur<_ma50<_ma200) else
                            "Stage 3 ⚠️"
                        )
                        # Sector RS vs SPY (3m)
                        _spy = _yf.Ticker("SPY").history(period="6mo")["Close"]
                        _etf_ret  = (_cur / float(_d.iloc[-63]) - 1) if len(_d) >= 63 else 0
                        _spy_ret  = (float(_spy.iloc[-1]) / float(_spy.iloc[-63]) - 1) if len(_spy) >= 63 else 0
                        _sec_rs   = round(_etf_ret / abs(_spy_ret) * 100, 1) if _spy_ret != 0 else None
                        return {
                            "etf": _etf, "sector_ok": _above_200,
                            "sector_stage": _stage, "sector_rs": _sec_rs,
                        }
                    except Exception:
                        return {"etf": None, "sector_ok": True, "sector_stage": "Unknown", "sector_rs": None}

                _mkt = _get_market_context()
                # _sec called after field extraction (theme needed first)

                # ── Helper functions ──────────────────────────────────────────
                def _get(col, default=None):
                    v = row.get(col, default)
                    if v is None: return default
                    if isinstance(v, float) and pd.isna(v): return default
                    return v

                def _bool_get(col):
                    v = _get(col)
                    if v is None: return False
                    return str(v).lower() in ("true","1","yes")

                def _num(col, default=0.0):
                    try: return float(_get(col, default) or default)
                    except: return default

                def _str(col, default="–"):
                    v = _get(col)
                    return str(v) if v not in (None,"","None","nan") else default

                # ── Pull all fields ───────────────────────────────────────────
                price          = _num("price")
                stage          = _str("stage")
                above_200ma    = _bool_get("above_200ma")
                ma50_gt_200    = _bool_get("ma50_gt_ma200")
                apex_score     = _num("apex_score")
                rs_3m          = _num("rs_3m", None)
                rs_r2500       = _num("rs_r2500_3m", None)
                rs_r3000g      = _num("rs_r3000g_3m", None)
                rs_multi       = _bool_get("rs_multi_leader")
                of_bias        = _str("of_bias", "Neutral")
                of_ratio       = _num("of_up_vol_ratio", 0)
                of_score       = _num("of_score", 0)
                earn_mom       = _str("earn_momentum", "–")
                eps_accel      = _bool_get("eps_accel")
                eps_growth     = _num("eps_growth_%", None)
                consec_beats   = _num("consec_beats", 0)
                near_52wh      = _bool_get("near_52wh")
                pct_off_high   = _num("pct_off_high_%", -100)
                early_entry    = _bool_get("early_entry")
                early_type     = _str("early_entry_type", "–")
                breaking_out   = _bool_get("breaking_out")
                pattern        = _str("pattern", "–")
                vwap_pos       = _str("vwap_position", "–")
                vwap_slope     = _str("vwap_slope", "–")
                adr_pct        = _num("adr_%", 5)
                next_earn      = _str("next_earnings", "–")
                liq_warn       = _bool_get("liquidity_warn")
                mcap_cat       = _str("mcap_category", "Unknown")
                is_gem         = _bool_get("is_gem")
                weekly_conf    = _bool_get("weekly_confirmed")
                weekly_contra  = _bool_get("weekly_contradicts")
                weekly_stage   = _str("weekly_stage", "–")
                weekly_rs      = _num("weekly_rs", None)
                weekly_tight   = _bool_get("weekly_base_tight")
                weekly_score   = _num("weekly_score", 0)
                vol_filter     = _num("vol_filter", 0)
                theme          = _str("theme", "–")
                changes        = _str("changes","–")

                # Now theme is available — fetch sector strength
                _sec = _get_sector_strength(theme)

                # ── Earnings days-away calculation ────────────────────────────
                earn_days = None
                if next_earn not in ("–","None","nan",""):
                    try:
                        earn_days = (pd.to_datetime(next_earn) - pd.Timestamp.now()).days
                    except Exception:
                        earn_days = None

                # ── Max position size by market cap ───────────────────────────
                _mcap_risk_pct = (
                    0.005 if mcap_cat in ("Micro Cap",)   else
                    0.01  if mcap_cat in ("Small Cap",)   else
                    0.02  if mcap_cat in ("Mid Cap",)     else
                    0.03  if mcap_cat in ("Large Cap",)   else
                    0.05
                )
                _max_risk_dol     = round(chk_account * _mcap_risk_pct, 2)
                _adr_stop_default = round(price * (1 - adr_pct / 100 * 1.25), 2)

                # ── Manual stop override & measured move inputs ───────────────
                st.markdown("---")
                st.markdown("#### ⚙️ Trade Parameters")
                _inp1, _inp2, _inp3 = st.columns(3)
                with _inp1:
                    _manual_stop = st.number_input(
                        "Stop Loss ($) — override",
                        min_value=0.0, max_value=float(price) if price else 9999.0,
                        value=float(_adr_stop_default),
                        step=0.01, format="%.2f",
                        key=f"manual_stop_{chk_ticker}",
                        help=(
                            f"Default = 1.25× ADR below entry (${_adr_stop_default:.2f}). "
                            "Override with your chart-based swing low for a more precise stop."
                        )
                    )
                with _inp2:
                    _base_low = st.number_input(
                        "Base Low ($) — for measured move",
                        min_value=0.0,
                        value=float(round(price * 0.85, 2)),
                        step=0.01, format="%.2f",
                        key=f"base_low_{chk_ticker}",
                        help=(
                            "Enter the lowest price of the base the stock is breaking out from. "
                            "Measured move target = Entry + (Entry − Base Low)."
                        )
                    )
                with _inp3:
                    _risk_override = st.number_input(
                        "Risk % override",
                        min_value=0.1, max_value=5.0,
                        value=float(_mcap_risk_pct * 100),
                        step=0.1, format="%.1f",
                        key=f"risk_pct_{chk_ticker}",
                        help=(
                            f"Default = {_mcap_risk_pct*100:.1f}% for {mcap_cat}. "
                            "Reduce for weaker setups, increase only for A+ setups with perfect checklist."
                        )
                    )

                # Recalculate with overrides
                _stop_price   = _manual_stop
                _stop_dist    = (price - _stop_price) / price if price > 0 else adr_pct/100*1.25
                _max_risk_dol = round(chk_account * _risk_override / 100, 2)
                _max_shares   = int(_max_risk_dol / (price * _stop_dist)) if price > 0 and _stop_dist > 0 else 0
                _pos_value    = round(_max_shares * price, 2)

                # Measured move target (chart-based)
                _base_depth   = price - _base_low
                _mm_target    = round(price + _base_depth, 2)  # measured move
                _rr_mm        = round(_base_depth / (price - _stop_price), 1) if (price - _stop_price) > 0 else 0

                # ADR-based targets + final target selection (computed early so
                # the Mobile Summary Card, which renders before the Trade Plan
                # Summary section, can safely reference _t1_use / _t2_use)
                _t1_adr = round(price * (1 + adr_pct/100 * 4), 2)
                _t2_adr = round(price * (1 + adr_pct/100 * 8), 2)
                _t1_use = min(_mm_target, _t1_adr) if _mm_target > price else _t1_adr
                _t2_use = max(_mm_target, _t2_adr)
                _rr_t1  = round((_t1_use - price) / (price - _stop_price), 1) if price > _stop_price else 0
                _rr_t2  = round((_t2_use - price) / (price - _stop_price), 1) if price > _stop_price else 0

                st.markdown("---")

                # ══════════════════════════════════════════════════════════════
                # BUILD THE 18-POINT CHECKLIST (expanded from 15)
                # ══════════════════════════════════════════════════════════════
                checks = []

                # ── Pre-compute R:R with measured move target ─────────────────
                _rr_ratio = _rr_mm   # chart-based measured move R:R
                _rr_pass  = _rr_ratio >= 2.0

                # ── GROUP 0: MARKET & SECTOR CONTEXT ─────────────────────────
                checks.append({
                    "group":   "🌍 Market & Sector Context — Check Before Everything Else",
                    "num":     0,
                    "label":   "Broad Market in Uptrend (S&P 500)",
                    "pass":    _mkt["market_ok"],
                    "value":   (
                        f"S&P 500: {_mkt['stage']} | "
                        f"VIX: {_mkt['vix']:.1f}" if _mkt.get('vix') else _mkt['stage']
                    ),
                    "why":     (
                        "75% of stocks follow the market direction. Buying individual stocks "
                        "during a market Stage 3/4 downtrend is fighting the tide — even great "
                        "setups fail more often when the index is broken."
                    ),
                    "action":  (
                        "Wait for the S&P 500 to reclaim its 50MA and 200MA before taking "
                        "new long positions. Use this time to build your watchlist."
                        if not _mkt["market_ok"] else
                        "Market environment is supportive. Proceed with individual stock analysis."
                    ),
                    "fatal":   True,
                })
                checks.append({
                    "group":   None,
                    "num":     "0b",
                    "label":   f"Sector Uptrend ({_sec.get('etf','–')})",
                    "pass":    _sec["sector_ok"],
                    "value":   (
                        f"{_sec.get('etf','–')}: {_sec['sector_stage']} | "
                        f"Sector RS: {_sec['sector_rs']:.0f}" if _sec.get('sector_rs') else
                        f"{_sec.get('etf','–')}: {_sec['sector_stage']}"
                    ),
                    "why":     (
                        f"A stock breaking out while its sector ETF ({_sec.get('etf','?')}) "
                        "is in a downtrend has dramatically lower odds. Sector tailwind = "
                        "institutions rotating into this area, lifting all boats."
                    ),
                    "action":  (
                        f"Sector {_sec.get('etf','?')} is weak. Use half-size or wait for "
                        "sector to recover above its 200MA before entering."
                        if not _sec["sector_ok"] else
                        "Sector is supportive — good tailwind for this trade."
                    ),
                    "fatal":   False,
                })

                # ── GROUP 1: NON-NEGOTIABLES (items 1–5) ─────────────────────
                checks.append({
                    "group":   "🚫 Non-Negotiables — Fail any = DO NOT BUY",
                    "num":     1,
                    "label":   "Weekly Stage 2 Confirmed",
                    "pass":    weekly_conf and not weekly_contra,
                    "value":   weekly_stage,
                    "why":     "Weekly chart must show price above 40WMA with 10WMA > 40WMA. Trading against the weekly trend fails 70%+ of the time.",
                    "action":  "Wait for weekly Stage 2 confirmation before entering.",
                    "fatal":   True,
                })
                checks.append({
                    "group":   None,
                    "num":     2,
                    "label":   "Daily Stage 2 Confirmed",
                    "pass":    above_200ma and ma50_gt_200,
                    "value":   stage,
                    "why":     "Price must be above both the 50MA and 200MA, with 50MA > 200MA. This is the only Weinstein stage where you take long positions.",
                    "action":  "Wait for price to reclaim both MAs before entering.",
                    "fatal":   True,
                })
                checks.append({
                    "group":   None,
                    "num":     3,
                    "label":   "RS vs S&P 500 Positive",
                    "pass":    rs_3m is not None and rs_3m > 0,
                    "value":   f"{rs_3m:.0f}" if rs_3m is not None else "–",
                    "why":     "If the stock can't outperform a passive index, there is no reason to own it. RS > 100 = outperforming. The best trades have RS > 150.",
                    "action":  "Wait for RS to turn positive vs the index before buying.",
                    "fatal":   True,
                })
                checks.append({
                    "group":   None,
                    "num":     4,
                    "label":   "No Weekly Contradiction",
                    "pass":    not weekly_contra,
                    "value":   "⚠️ CONTRADICTS" if weekly_contra else "✅ Clear",
                    "why":     "A daily breakout inside a weekly downtrend is the most common retail trap. The weekly trend will eventually win. Avoid completely.",
                    "action":  "Do not trade this setup until the weekly resolves into Stage 2.",
                    "fatal":   True,
                })
                checks.append({
                    "group":   None,
                    "num":     5,
                    "label":   "Adequate Liquidity",
                    "pass":    not liq_warn and vol_filter >= 100_000,
                    "value":   f"{vol_filter:,.0f} shares/day",
                    "why":     "You must be able to exit at your price. Low liquidity = wide spreads + slippage that eats profits before they start.",
                    "action":  "Only trade with limit orders. Consider skipping if volume < 100K/day.",
                    "fatal":   True,
                })

                # ── GROUP 2: QUALITY FILTERS (items 6–9) ─────────────────────
                checks.append({
                    "group":   "📊 Quality Filters — Fail = Reduce Size or Wait",
                    "num":     6,
                    "label":   "Apex Score ≥ 65",
                    "pass":    apex_score >= 65,
                    "value":   f"{apex_score:.0f}/100",
                    "why":     "Below 65 means multiple signals are missing or contradicting. The best trades score 70–90 before they make their move.",
                    "action":  "Wait for score to reach 65+ or reduce size to 50% of normal.",
                    "fatal":   False,
                })
                checks.append({
                    "group":   None,
                    "num":     7,
                    "label":   "Institutional Order Flow Bullish",
                    "pass":    "bullish" in of_bias.lower(),
                    "value":   f"{of_bias} | Up/Down vol: {of_ratio:.2f}x",
                    "why":     "Institutions leave footprints. Persistent up-day volume > down-day volume means large money is accumulating before the move becomes obvious to the crowd.",
                    "action":  "Wait for OF bias to turn Bullish before entering.",
                    "fatal":   False,
                })
                _fund_pass = (
                    "strong" in earn_mom.lower() or
                    eps_accel or
                    (eps_growth is not None and eps_growth >= 25) or
                    consec_beats >= 3
                )
                checks.append({
                    "group":   None,
                    "num":     8,
                    "label":   "Earnings Momentum Positive",
                    "pass":    _fund_pass,
                    "value":   f"{earn_mom} | EPS growth: {eps_growth:+.0f}%" if eps_growth else earn_mom,
                    "why":     "Price follows earnings long-term. Accelerating EPS or consecutive beats signals institutional re-rating. Pure technical setups without fundamental support fail more often.",
                    "action":  "Accept if momentum is Moderate but reduce size. Avoid if declining EPS.",
                    "fatal":   False,
                })
                checks.append({
                    "group":   None,
                    "num":     9,
                    "label":   "Near 52-Week High (< 20% off)",
                    "pass":    pct_off_high >= -20,
                    "value":   f"{pct_off_high:+.1f}% from 52W high",
                    "why":     "Breakouts happen near highs, not from the bottom. Stocks > 30% below their high have massive overhead resistance from prior buyers desperate to break even.",
                    "action":  "Wait for stock to recover closer to highs before buying the breakout.",
                    "fatal":   False,
                })

                # ── GROUP 3: ENTRY TIMING (items 10–12) ──────────────────────
                _entry_pass = early_entry or breaking_out
                checks.append({
                    "group":   "🎯 Entry Timing — Fail = Wait for Better Moment",
                    "num":     10,
                    "label":   "Early Entry or Active Breakout",
                    "pass":    _entry_pass,
                    "value":   (
                        "🚨 BREAKING OUT" if breaking_out else
                        f"✅ {early_type}" if early_entry else
                        "⏳ No signal yet"
                    ),
                    "why":     "Buying extended (10%+ past breakout) dramatically cuts your R:R. Early entry = cheap stock relative to where it's going. Breakout = highest-urgency entry.",
                    "action":  "Wait for a pullback to 50MA, VWAP, or a new base to form before entering.",
                    "fatal":   False,
                })
                _vwap_pass = "above" in vwap_pos.lower() and "extended" not in vwap_pos.lower()
                checks.append({
                    "group":   None,
                    "num":     11,
                    "label":   "Price Above Rising VWAP",
                    "pass":    _vwap_pass,
                    "value":   f"{vwap_pos} | Slope: {vwap_slope}",
                    "why":     "Institutional algorithms anchor to VWAP for large order execution. Being above a rising VWAP means buyers are in control of the auction right now.",
                    "action":  "Wait for a VWAP reclaim on volume — that is the entry trigger.",
                    "fatal":   False,
                })
                _pattern_pass = pattern not in ("–","None","nan","") and "none" not in pattern.lower()
                checks.append({
                    "group":   None,
                    "num":     12,
                    "label":   "Recognisable Chart Pattern",
                    "pass":    _pattern_pass,
                    "value":   pattern,
                    "why":     "Buying random price action is gambling. Buying a well-defined base breakout, cup, or pullback to support is a defined edge with a clear invalidation level.",
                    "action":  "Wait for a proper base to form (3–8 weeks minimum) before entering.",
                    "fatal":   False,
                })

                # ── R:R validation check (between quality and risk groups) ──────
                checks.append({
                    "group":   "📐 Risk:Reward — Minimum 2:1 Required",
                    "num":     "12b",
                    "label":   "Risk:Reward Ratio ≥ 2:1",
                    "pass":    _rr_pass,
                    "value":   (
                        f"{_rr_ratio:.1f}:1 "
                        f"(Entry ${price:.2f} → Target ${_mm_target:.2f} | Stop ${_stop_price:.2f})"
                    ),
                    "why":     (
                        "With a 50% win rate you need at least 2:1 R:R to be profitable. "
                        "Below 2:1 means you need a 67%+ win rate just to break even — "
                        "no trader sustains that consistently. "
                        f"Measured move target = Entry + Base Depth = ${_mm_target:.2f}."
                    ),
                    "action":  (
                        f"Current R:R is {_rr_ratio:.1f}:1. "
                        "Either tighten the stop (move closer to a real swing low) or "
                        "wait for a lower entry to improve the ratio. "
                        "Do not widen your stop to manufacture a better R:R."
                        if not _rr_pass else
                        f"R:R of {_rr_ratio:.1f}:1 is acceptable. Trim 50% at "
                        f"1R gain (${round(price + (price - _stop_price), 2):.2f}), "
                        "let the rest run to the measured move target."
                    ),
                    "fatal":   True,   # R:R < 2:1 is a hard no
                })

                # ── GROUP 4: RISK CONTROLS (items 13–15) ─────────────────────
                _earn_pass = earn_days is None or earn_days > 14
                checks.append({
                    "group":   "⚠️ Risk Controls — Must Define Before Buying",
                    "num":     13,
                    "label":   "Earnings Not Imminent (> 14 days)",
                    "pass":    _earn_pass,
                    "value":   (
                        f"⚠️ In {earn_days} days ({next_earn})" if earn_days is not None and earn_days <= 14
                        else f"✅ {next_earn}" if next_earn != "–"
                        else "✅ No date flagged"
                    ),
                    "why":     "Stocks can gap 10–40% on earnings. Holding through earnings with a full position is not trading — it's gambling on a binary event.",
                    "action":  f"{'Reduce to 50% size before earnings.' if earn_days and earn_days <= 14 else 'Monitor as date approaches.'}",
                    "fatal":   False,
                })
                checks.append({
                    "group":   None,
                    "num":     14,
                    "label":   "Stop Loss Level Defined",
                    "pass":    True,   # always informational
                    "value":   f"${_stop_price:.2f}  ({_stop_dist*100:.1f}% below entry = 1.25× ADR)",
                    "why":     "A stop loss is not optional. It is the price at which you admit you were wrong and exit before a small loss becomes a catastrophic one. Define it BEFORE you buy.",
                    "action":  f"Place stop at ${_stop_price:.2f} (below 1.25× ADR). Move to break-even after +{adr_pct*2:.1f}% gain.",
                    "fatal":   False,
                })
                checks.append({
                    "group":   None,
                    "num":     15,
                    "label":   "Position Size Calculated",
                    "pass":    True,   # always informational
                    "value":   f"{_max_shares} shares × ${price:.2f} = ${_pos_value:,.0f} ({_mcap_risk_pct*100:.1f}% risk)",
                    "why":     f"Maximum risk per trade is {_mcap_risk_pct*100:.1f}% of portfolio for a {mcap_cat}. This keeps any single loss from materially damaging your account.",
                    "action":  f"Max position: {_max_shares} shares (${_pos_value:,.0f}). Max dollar risk: ${_max_risk_dol:,.0f}.",
                    "fatal":   False,
                })

                # ══════════════════════════════════════════════════════════════
                # VERDICT CALCULATION + CONVICTION SCORE
                # ══════════════════════════════════════════════════════════════
                fatal_fails   = [c for c in checks if c.get("fatal") and not c["pass"]]
                quality_fails = [c for c in checks if not c.get("fatal") and not c["pass"]
                                 and str(c.get("num","")).replace("b","").isdigit()
                                 and int(str(c.get("num","0")).replace("b","0")) <= 12]
                all_pass      = len(fatal_fails) == 0 and len(quality_fails) == 0
                partial_pass  = len(fatal_fails) == 0 and len(quality_fails) <= 2
                timing_only   = len(fatal_fails) == 0 and len([
                    c for c in quality_fails
                    if str(c.get("num","0")).replace("b","").isdigit()
                    and int(str(c.get("num","0")).replace("b","0")) <= 9
                ]) == 0

                # ── Conviction Score 0–100 ────────────────────────────────────
                # Measures HOW STRONGLY each check passed, not just whether it did
                _conv = 0
                # Market & sector (max 15)
                if _mkt["market_ok"]:                         _conv += 8
                if _sec["sector_ok"]:                         _conv += 7
                # Weekly (max 15)
                if weekly_conf and not weekly_contra:         _conv += 10
                if weekly_tight:                              _conv += 5
                # Stage & trend (max 10)
                if above_200ma and ma50_gt_200:               _conv += 10
                # RS strength (max 15)
                _rs_val = rs_3m if rs_3m else 0
                if _rs_val >= 150:                            _conv += 15
                elif _rs_val >= 100:                          _conv += 10
                elif _rs_val > 0:                             _conv += 5
                # Order flow (max 10)
                if "strong bullish" in of_bias.lower():       _conv += 10
                elif "bullish" in of_bias.lower():            _conv += 6
                # Fundamentals (max 10)
                if eps_accel:                                 _conv += 5
                if earn_mom and "strong" in earn_mom.lower(): _conv += 5
                # Entry quality (max 10)
                if breaking_out:                              _conv += 10
                elif early_entry:                             _conv += 7
                elif _vwap_pass:                              _conv += 4
                # R:R (max 10)
                if _rr_ratio >= 3.0:                          _conv += 10
                elif _rr_ratio >= 2.0:                        _conv += 6
                elif _rr_ratio >= 1.5:                        _conv += 3
                # Apex score (max 5)
                if apex_score >= 80:                          _conv += 5
                elif apex_score >= 65:                        _conv += 3
                _conviction = min(100, _conv)

                # ── Adjusted position size by conviction ──────────────────────
                # High conviction = full size. Low conviction = reduce.
                _conv_mult = (
                    1.0  if _conviction >= 75 else
                    0.75 if _conviction >= 60 else
                    0.50 if _conviction >= 45 else
                    0.25
                )
                _adj_shares   = int(_max_shares * _conv_mult)
                _adj_value    = round(_adj_shares * price, 2)
                _adj_risk     = round(_adj_shares * price * _stop_dist, 2)

                # ── Time-of-day guidance ──────────────────────────────────────
                _now_est = pd.Timestamp.now("UTC").tz_convert(None) - pd.Timedelta(hours=5)
                _hour    = _now_est.hour
                _minute  = _now_est.minute
                _tod_guidance = (
                    "⏰ PRE-MARKET: Do not chase pre-market moves. Wait for the first 15–30 min "
                    "after open to see where price stabilises before entering."
                    if _hour < 9 or (_hour == 9 and _minute < 30) else
                    "⏰ FIRST 30 MIN (9:30–10:00 AM): High volatility window. "
                    "Wait for the opening range to establish, then buy the first pullback to VWAP. "
                    "Never buy a gap-up open blindly."
                    if _hour == 9 or (_hour == 10 and _minute < 0) else
                    "⏰ PRIME ENTRY WINDOW (10:00 AM–12:00 PM): Best time to enter. "
                    "Direction is established, volume is liquid, institutions are active. "
                    "Enter on a VWAP pullback or breakout continuation."
                    if _hour < 12 else
                    "⏰ MIDDAY DRIFT (12:00–2:00 PM): Avoid new entries. Low volume, "
                    "choppy price action. Breakouts in this window have low follow-through."
                    if _hour < 14 else
                    "⏰ POWER HOUR SETUP (2:00–3:30 PM): Strong stocks accelerate into close. "
                    "If stock is holding above VWAP with volume picking up, enter now. "
                    "Set stop below today's low."
                    if _hour < 15 or (_hour == 15 and _minute < 30) else
                    "⏰ LAST 30 MIN (3:30–4:00 PM): Avoid new entries. Institutional "
                    "window-dressing can cause erratic moves. Wait for next morning."
                    if _hour < 16 else
                    "⏰ AFTER HOURS: Market is closed. Review the setup for next session entry. "
                    "Place a limit order for tomorrow's open if you want to be in early."
                )

                # ── Verdict Banner ────────────────────────────────────────────
                st.markdown("---")

                # Conviction score bar
                _conv_color = "#3fb950" if _conviction >= 75 else "#d29922" if _conviction >= 50 else "#f85149"
                _conv_label = (
                    "A+ Setup" if _conviction >= 85 else
                    "A Setup"  if _conviction >= 75 else
                    "B Setup"  if _conviction >= 60 else
                    "C Setup"  if _conviction >= 45 else
                    "D Setup"
                )
                st.markdown(
                    f'<div style="background:#161b22;border:1px solid #30363d;border-radius:10px;'
                    f'padding:14px 20px;margin-bottom:12px;display:flex;align-items:center;gap:16px;">'
                    f'<div style="flex:1;">'
                    f'<div style="color:#8b949e;font-size:0.75rem;text-transform:uppercase;'
                    f'letter-spacing:.06em;margin-bottom:4px;">CONVICTION SCORE</div>'
                    f'<div style="background:#21262d;border-radius:4px;height:10px;width:100%;">'
                    f'<div style="background:{_conv_color};height:10px;border-radius:4px;'
                    f'width:{_conviction}%;transition:width .5s;"></div></div>'
                    f'</div>'
                    f'<div style="text-align:right;min-width:80px;">'
                    f'<div style="color:{_conv_color};font-size:1.6rem;font-weight:800;">{_conviction}</div>'
                    f'<div style="color:#8b949e;font-size:0.78rem;">{_conv_label}</div>'
                    f'</div></div>',
                    unsafe_allow_html=True
                )

                # Time-of-day guidance
                st.info(_tod_guidance)

                # Main verdict
                if fatal_fails:
                    _fail_list = "  |  ".join(
                        f'#{c["num"]} {c["label"]}' for c in fatal_fails
                    )
                    st.markdown(
                        f'<div style="background:#2a1010;border:2px solid #f85149;border-radius:12px;'
                        f'padding:20px 24px;margin-bottom:16px;">'
                        f'<div style="font-size:1.8rem;font-weight:800;color:#f85149;">🚫 DO NOT BUY</div>'
                        f'<div style="color:#e6edf3;margin-top:8px;font-size:1rem;">'
                        f'<b>{len(fatal_fails)} hard gate(s) failed.</b> No exceptions:<br>'
                        f'{_fail_list}'
                        f'</div></div>',
                        unsafe_allow_html=True
                    )
                elif all_pass:
                    st.markdown(
                        f'<div style="background:#0d2a0d;border:2px solid #3fb950;border-radius:12px;'
                        f'padding:20px 24px;margin-bottom:16px;">'
                        f'<div style="font-size:1.8rem;font-weight:800;color:#3fb950;">✅ READY TO BUY</div>'
                        f'<div style="color:#e6edf3;margin-top:8px;font-size:1rem;">'
                        f'All checks passed. Conviction: <b>{_conviction}/100 ({_conv_label})</b>.<br>'
                        f'Enter <b>{_adj_shares} shares</b> (conviction-adjusted from {_max_shares}) '
                        f'at ${price:.2f}. Stop: <b>${_stop_price:.2f}</b>. '
                        f'Target: <b>${_mm_target:.2f}</b> (R:R {_rr_ratio:.1f}:1).'
                        f'</div></div>',
                        unsafe_allow_html=True
                    )
                elif timing_only:
                    _wait_for = "  |  ".join(c["action"].split(".")[0] for c in quality_fails[:2])
                    st.markdown(
                        f'<div style="background:#1a2a10;border:2px solid #d29922;border-radius:12px;'
                        f'padding:20px 24px;margin-bottom:16px;">'
                        f'<div style="font-size:1.8rem;font-weight:800;color:#d29922;">⏳ WAIT FOR ENTRY</div>'
                        f'<div style="color:#e6edf3;margin-top:8px;font-size:1rem;">'
                        f'Stock quality confirmed. Entry timing not yet ideal.<br>'
                        f'Wait for: {_wait_for}'
                        f'</div></div>',
                        unsafe_allow_html=True
                    )
                elif partial_pass:
                    _fail_list2 = "  |  ".join(
                        f'#{c["num"]} {c["label"]}' for c in quality_fails
                    )
                    st.markdown(
                        f'<div style="background:#1a1a10;border:2px solid #d29922;border-radius:12px;'
                        f'padding:20px 24px;margin-bottom:16px;">'
                        f'<div style="font-size:1.8rem;font-weight:800;color:#d29922;">⚠️ REDUCED SIZE — {_conv_label}</div>'
                        f'<div style="color:#e6edf3;margin-top:8px;font-size:1rem;">'
                        f'{len(quality_fails)} quality issue(s). '
                        f'Conviction-adjusted size: <b>{_adj_shares} shares</b> '
                        f'(${_adj_value:,.0f}, risk ${_adj_risk:,.0f}).<br>'
                        f'Failed: {_fail_list2}'
                        f'</div></div>',
                        unsafe_allow_html=True
                    )
                else:
                    st.markdown(
                        f'<div style="background:#2a1010;border:2px solid #f85149;border-radius:12px;'
                        f'padding:20px 24px;margin-bottom:16px;">'
                        f'<div style="font-size:1.8rem;font-weight:800;color:#f85149;">❌ SKIP</div>'
                        f'<div style="color:#e6edf3;margin-top:8px;font-size:1rem;">'
                        f'{len(quality_fails)} quality checks failed. '
                        f'Better opportunities exist in the scan right now.'
                        f'</div></div>',
                        unsafe_allow_html=True
                    )
                st.markdown("---")

                # Define _verdict_str here so it's available for mobile card AND download
                _verdict_str = (
                    "READY TO BUY"   if all_pass     else
                    "DO NOT BUY"     if fatal_fails  else
                    "WAIT FOR ENTRY" if timing_only  else
                    "REDUCED SIZE"   if partial_pass else
                    "SKIP"
                )

                # ── 📱 Mobile Summary Card ────────────────────────────────────
                with st.expander("📱 Mobile Summary Card (tap to expand)", expanded=False):
                    _vrd_emoji = "✅" if all_pass else "🚫" if fatal_fails else "⏳" if timing_only else "⚠️"
                    _vrd_color = "#3fb950" if all_pass else "#f85149" if fatal_fails else "#d29922"
                    st.markdown(
                        f'<div style="background:#0d1117;border:2px solid {_vrd_color};'
                        f'border-radius:16px;padding:24px;max-width:400px;margin:0 auto;">'

                        # Header
                        f'<div style="text-align:center;margin-bottom:16px;">'
                        f'<div style="font-size:2rem;font-weight:900;color:{_vrd_color};">'
                        f'{_vrd_emoji} {chk_ticker}</div>'
                        f'<div style="color:{_vrd_color};font-size:1rem;font-weight:700;margin-top:4px;">'
                        f'{_verdict_str}</div>'
                        f'<div style="color:#8b949e;font-size:0.82rem;">'
                        f'Conviction: {_conviction}/100 ({_conv_label})</div>'
                        f'</div>'

                        # Divider
                        f'<div style="border-top:1px solid #30363d;margin:12px 0;"></div>'

                        # Key numbers in 2×3 grid
                        f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;'
                        f'margin-bottom:12px;">'
                        f'<div style="background:#161b22;border-radius:8px;padding:10px;text-align:center;">'
                        f'<div style="color:#8b949e;font-size:0.7rem;">ENTRY</div>'
                        f'<div style="color:#e6edf3;font-size:1.3rem;font-weight:800;">${price:.2f}</div>'
                        f'</div>'
                        f'<div style="background:#2a1010;border-radius:8px;padding:10px;text-align:center;">'
                        f'<div style="color:#8b949e;font-size:0.7rem;">STOP</div>'
                        f'<div style="color:#f85149;font-size:1.3rem;font-weight:800;">${_stop_price:.2f}</div>'
                        f'</div>'
                        f'<div style="background:#0d2a0d;border-radius:8px;padding:10px;text-align:center;">'
                        f'<div style="color:#8b949e;font-size:0.7rem;">TARGET 1</div>'
                        f'<div style="color:#3fb950;font-size:1.3rem;font-weight:800;">${_t1_use:.2f}</div>'
                        f'</div>'
                        f'<div style="background:#0d2a0d;border-radius:8px;padding:10px;text-align:center;">'
                        f'<div style="color:#8b949e;font-size:0.7rem;">TARGET 2</div>'
                        f'<div style="color:#3fb950;font-size:1.3rem;font-weight:800;">${_t2_use:.2f}</div>'
                        f'</div>'
                        f'<div style="background:#161b22;border-radius:8px;padding:10px;text-align:center;">'
                        f'<div style="color:#8b949e;font-size:0.7rem;">SHARES</div>'
                        f'<div style="color:#e6edf3;font-size:1.3rem;font-weight:800;">{_adj_shares}</div>'
                        f'</div>'
                        f'<div style="background:#161b22;border-radius:8px;padding:10px;text-align:center;">'
                        f'<div style="color:#8b949e;font-size:0.7rem;">R:R</div>'
                        f'<div style="color:{"#3fb950" if _rr_ratio>=2 else "#d29922"};'
                        f'font-size:1.3rem;font-weight:800;">{_rr_ratio:.1f}:1</div>'
                        f'</div>'
                        f'</div>'

                        # Divider
                        f'<div style="border-top:1px solid #30363d;margin:12px 0;"></div>'

                        # Fatal fails or timing
                        + (
                            f'<div style="color:#f85149;font-size:0.82rem;font-weight:700;">'
                            f'🚫 Failed: {" | ".join(f"#{c["num"]} {c["label"]}" for c in fatal_fails[:2])}'
                            f'</div>'
                            if fatal_fails else
                            f'<div style="color:#3fb950;font-size:0.82rem;">'
                            f'✅ All gates clear · {_conviction}/100 conviction'
                            f'</div>'
                        )

                        # Timing
                        + f'<div style="color:#8b949e;font-size:0.78rem;margin-top:8px;">'
                        + _tod_guidance.split(":")[0] + f'</div>'

                        + f'</div>',
                        unsafe_allow_html=True
                    )

                st.markdown("---")

                # ══════════════════════════════════════════════════════════════
                # CHECKLIST TABLE
                # ══════════════════════════════════════════════════════════════
                current_group = None
                for chk in checks:
                    # Print group header
                    if chk.get("group") and chk["group"] != current_group:
                        current_group = chk["group"]
                        st.markdown(
                            f'<div style="background:#161b22;padding:8px 14px;border-radius:6px;'
                            f'margin:16px 0 8px 0;font-size:0.85rem;font-weight:700;'
                            f'color:#8b949e;letter-spacing:0.05em;">{current_group}</div>',
                            unsafe_allow_html=True
                        )

                    _pass    = chk["pass"]
                    _icon    = "✅" if _pass else ("🚫" if chk.get("fatal") else "❌")
                    _bg      = "#0d2a0d" if _pass else ("#2a1010" if chk.get("fatal") else "#1a1510")
                    _border  = "#3fb950" if _pass else ("#f85149" if chk.get("fatal") else "#d29922")
                    _val_col = "#3fb950" if _pass else ("#f85149" if chk.get("fatal") else "#d29922")

                    st.markdown(
                        f'<div style="background:{_bg};border-left:4px solid {_border};'
                        f'border-radius:0 8px 8px 0;padding:12px 16px;margin:4px 0;'
                        f'display:flex;gap:12px;align-items:flex-start;">'

                        # Number + icon
                        f'<div style="min-width:44px;text-align:center;">'
                        f'<span style="font-size:1.1rem;">{_icon}</span><br>'
                        f'<span style="color:#8b949e;font-size:0.72rem;">#{chk["num"]}</span>'
                        f'</div>'

                        # Label + value
                        f'<div style="flex:1;">'
                        f'<div style="color:#e6edf3;font-weight:700;font-size:0.92rem;">'
                        f'{chk["label"]}</div>'
                        f'<div style="color:{_val_col};font-size:0.88rem;margin-top:2px;">'
                        f'{chk["value"]}</div>'
                        f'</div>'

                        # Why + action
                        f'<div style="flex:2;border-left:1px solid #30363d;padding-left:14px;">'
                        f'<div style="color:#8b949e;font-size:0.80rem;line-height:1.5;">'
                        f'{chk["why"]}</div>'
                        f'{"<div style=\"color:#d29922;font-size:0.78rem;margin-top:4px;\">" + "→ " + chk["action"] + "</div>" if not _pass else ""}'
                        f'</div>'

                        f'</div>',
                        unsafe_allow_html=True
                    )

                # ══════════════════════════════════════════════════════════════
                # TRADE PLAN SUMMARY
                # ══════════════════════════════════════════════════════════════
                st.markdown("---")
                st.markdown("#### 📋 Trade Plan Summary")
                tp1,tp2,tp3 = st.columns(3)

                with tp1:
                    st.markdown(
                        f'<div style="background:#161b22;border:1px solid #30363d;'
                        f'border-radius:10px;padding:16px;">'
                        f'<div style="color:#8b949e;font-size:0.75rem;text-transform:uppercase;'
                        f'letter-spacing:.06em;margin-bottom:8px;">ENTRY</div>'
                        f'<div style="color:#e6edf3;font-size:1.4rem;font-weight:800;">'
                        f'${price:.2f}</div>'
                        f'<div style="color:#8b949e;font-size:0.8rem;margin-top:4px;">'
                        f'Current price — buy at open or on<br>VWAP pullback</div>'
                        f'</div>',
                        unsafe_allow_html=True
                    )

                with tp2:
                    st.markdown(
                        f'<div style="background:#2a1010;border:1px solid #f85149;'
                        f'border-radius:10px;padding:16px;">'
                        f'<div style="color:#8b949e;font-size:0.75rem;text-transform:uppercase;'
                        f'letter-spacing:.06em;margin-bottom:8px;">STOP LOSS</div>'
                        f'<div style="color:#f85149;font-size:1.4rem;font-weight:800;">'
                        f'${_stop_price:.2f}</div>'
                        f'<div style="color:#8b949e;font-size:0.8rem;margin-top:4px;">'
                        f'{_stop_dist*100:.1f}% below entry (1.25× ADR)<br>'
                        f'Max loss: ${_max_risk_dol:,.0f}</div>'
                        f'</div>',
                        unsafe_allow_html=True
                    )

                with tp3:
                    # Chart-based measured move target (primary) / ADR-based target
                    # (secondary) — _t1_use, _t2_use, _rr_t1, _rr_t2 are computed
                    # earlier (right after _rr_mm) so they're available wherever needed.
                    st.markdown(
                        f'<div style="background:#0d2a0d;border:1px solid #3fb950;'
                        f'border-radius:10px;padding:16px;">'
                        f'<div style="color:#8b949e;font-size:0.75rem;text-transform:uppercase;'
                        f'letter-spacing:.06em;margin-bottom:8px;">TARGETS</div>'
                        f'<div style="color:#3fb950;font-size:1.1rem;font-weight:700;">'
                        f'T1: ${_t1_use:.2f} '
                        f'<span style="font-size:0.78rem;color:#8b949e;">'
                        f'(R:R {_rr_t1:.1f}:1 · trim 50%)</span></div>'
                        f'<div style="color:#3fb950;font-size:1.1rem;font-weight:700;margin-top:4px;">'
                        f'T2: ${_t2_use:.2f} '
                        f'<span style="font-size:0.78rem;color:#8b949e;">'
                        f'(R:R {_rr_t2:.1f}:1 · trail stop)</span></div>'
                        f'<div style="color:#388bfd;font-size:0.78rem;margin-top:6px;">'
                        f'📐 Measured move: base depth = ${_base_depth:.2f} → T1 = ${_mm_target:.2f}</div>'
                        f'</div>',
                        unsafe_allow_html=True
                    )

                # ── Position sizing breakdown ──────────────────────────────────
                st.markdown("---")
                st.markdown("#### 💰 Position Sizing — Conviction Adjusted")
                ps1,ps2,ps3,ps4,ps5,ps6 = st.columns(6)
                ps1.metric("Entry Price",          f"${price:.2f}")
                ps2.metric("Max Shares (full)",    f"{_max_shares}",
                           help=f"Full size at {_risk_override:.1f}% risk with manual stop")
                ps3.metric("Adj. Shares",          f"{_adj_shares}",
                           delta=f"{_conv_mult*100:.0f}% of full size",
                           help=f"Conviction {_conviction}/100 → {_conv_mult*100:.0f}% of max size")
                ps4.metric("Position Value",       f"${_adj_value:,.0f}")
                ps5.metric("Max $ Risk",           f"${_adj_risk:,.0f}",
                           help=f"Stop at ${_stop_price:.2f} = {_stop_dist*100:.1f}% below entry")
                ps6.metric("% of Portfolio",       f"{_adj_value/chk_account*100:.1f}%")

                # Conviction breakdown bar
                st.markdown(
                    f'<div style="background:#161b22;border:1px solid #30363d;border-radius:8px;'
                    f'padding:12px 16px;margin-top:8px;font-size:0.82rem;color:#8b949e;">'
                    f'<b style="color:#e6edf3;">Why conviction is {_conviction}/100 ({_conv_label}):</b> '
                    f'Market {"✅" if _mkt["market_ok"] else "❌"} · '
                    f'Sector {"✅" if _sec["sector_ok"] else "❌"} · '
                    f'Weekly {"✅" if weekly_conf else "❌"} · '
                    f'Stage {"✅" if (above_200ma and ma50_gt_200) else "❌"} · '
                    f'RS {f"{rs_3m:.0f}" if rs_3m else "–"} · '
                    f'Flow {of_bias} · '
                    f'R:R {_rr_ratio:.1f}:1 · '
                    f'Apex {apex_score:.0f}'
                    f'</div>',
                    unsafe_allow_html=True
                )

                # ── Direct action buttons ─────────────────────────────────────
                st.markdown("---")
                ac1,ac2,ac3,ac4 = st.columns(4)
                with ac1:
                    if st.button("📊 View Full Deep Dive", key="chk_deepdive",
                                 use_container_width=True):
                        st.session_state["deepdive_ticker"] = chk_ticker
                        st.info(f"Go to 🔍 Stock Deep Dive tab → select {chk_ticker}")
                with ac2:
                    if st.button("📋 Add to Watchlist", key="chk_watchlist",
                                 use_container_width=True):
                        st.session_state["wl_add_ticker"] = chk_ticker
                        st.info(f"Go to 📋 Watchlists tab to add {chk_ticker}")
                with ac3:
                    if st.button("💼 Log as Position", key="chk_portfolio",
                                 use_container_width=True):
                        st.session_state["port_add_ticker"] = chk_ticker
                        st.info(f"Go to 💼 Portfolio Tracker → ➕ Add / Manage to enter {chk_ticker}")
                with ac4:
                    if st.button("👁 Watch This Setup", key="chk_watch_setup",
                                 use_container_width=True,
                                 help="Monitor this setup — get alerted when checklist status improves to Ready"):
                        _chkw = load_chk_watchlist()
                        _existing = [x["ticker"] for x in _chkw]
                        if chk_ticker not in _existing:
                            _chkw.append({
                                "ticker":        chk_ticker,
                                "added":         pd.Timestamp.now().isoformat(),
                                "last_verdict":  _verdict_str,
                                "last_score":    apex_score,
                                "last_conv":     _conviction,
                                "last_checked":  pd.Timestamp.now().isoformat(),
                                "alert_on":      "Ready to Buy",
                                "theme":         theme,
                            })
                            save_chk_watchlist(_chkw)
                            st.success(
                                f"✅ {chk_ticker} added to Setup Monitor. "
                                f"You'll be alerted when status improves to 'Ready to Buy'."
                            )
                        else:
                            st.info(f"{chk_ticker} is already being monitored.")

                # ── Setup Monitor status ───────────────────────────────────────
                _chkw_all = load_chk_watchlist()
                if _chkw_all:
                    _this_watch = [x for x in _chkw_all if x["ticker"] == chk_ticker]
                    if _this_watch:
                        _witem = _this_watch[0]
                        st.markdown(
                            f'<div style="background:#161b22;border:1px solid #388bfd;'
                            f'border-radius:8px;padding:10px 16px;margin-top:8px;'
                            f'font-size:0.82rem;color:#8b949e;">'
                            f'👁 <b style="color:#388bfd;">{chk_ticker} is being monitored.</b> '
                            f'Last verdict: <b>{_witem.get("last_verdict","–")}</b> | '
                            f'Added: {_witem.get("added","–")[:10]}'
                            f'</div>',
                            unsafe_allow_html=True
                        )

                # ── Download trade plan ───────────────────────────────────────
                # (_verdict_str already defined above for mobile card)
                _plan_text = f"""APEXSCAN PRE-BUY CHECKLIST — {chk_ticker}
Generated:       {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")} EST
Portfolio Size:  ${chk_account:,.0f}
Ticker:          {chk_ticker} | {theme} | {mcap_cat}

VERDICT:         {_verdict_str}
CONVICTION:      {_conviction}/100 ({_conv_label})

MARKET CONTEXT
S&P 500 Stage:   {_mkt.get("stage","–")}
VIX:             {f"{_mkt['vix']:.1f}" if _mkt.get("vix") else "–"}
Sector ETF:      {_sec.get("etf","–")} — {_sec.get("sector_stage","–")}
Market OK:       {"YES" if _mkt["market_ok"] else "NO — caution"}
Sector OK:       {"YES" if _sec["sector_ok"] else "NO — weak sector"}

TRADE PLAN
Entry:           ${price:.2f}
Stop Loss:       ${_stop_price:.2f} ({_stop_dist*100:.1f}% below entry)
Target 1:        ${_t1_use:.2f} (R:R {_rr_t1:.1f}:1 — trim 50% here)
Target 2:        ${_t2_use:.2f} (R:R {_rr_t2:.1f}:1 — trail stop)
Measured Move:   ${_mm_target:.2f} (base depth ${_base_depth:.2f})
Shares:          {_adj_shares} (conviction-adjusted) / {_max_shares} max
Position Value:  ${_adj_value:,.0f}
Dollar Risk:     ${_adj_risk:,.0f}
% of Portfolio:  {_adj_value/chk_account*100:.1f}%

TIMING
{_tod_guidance}

CHECKLIST RESULTS ({len([c for c in checks if c["pass"]])}/{len(checks)} passed)
"""
                for chk in checks:
                    _status = "PASS" if chk["pass"] else ("FAIL-FATAL" if chk.get("fatal") else "FAIL")
                    _num_str = f"{chk['num']:02d}" if isinstance(chk['num'], int) else str(chk['num'])
                    _plan_text += f"#{_num_str} [{_status}] {chk['label']}: {chk['value']}\n"
                    if not chk["pass"]:
                        _plan_text += f"    → {chk['action']}\n"

                # ── Log to Trade Journal ─────────────────────────────────────
                if st.button("📓 Log to Trade Journal", key="chk_log_journal",
                             use_container_width=True,
                             help="Save this checklist result to your trade journal for performance tracking"):
                    _journal = load_journal()
                    _journal.append({
                        "date":           pd.Timestamp.now().isoformat(),
                        "ticker":         chk_ticker,
                        "theme":          theme,
                        "mcap_category":  mcap_cat,
                        "verdict":        _verdict_str,
                        "conviction":     _conviction,
                        "conviction_label": _conv_label,
                        "apex_score":     apex_score,
                        "fatal_fails":    len(fatal_fails),
                        "quality_fails":  len(quality_fails),
                        "market_ok":      _mkt["market_ok"],
                        "market_stage":   _mkt.get("stage","–"),
                        "sector_ok":      _sec["sector_ok"],
                        "sector_etf":     _sec.get("etf","–"),
                        "weekly_conf":    weekly_conf,
                        "weekly_contra":  weekly_contra,
                        "stage":          stage,
                        "rs_3m":          rs_3m,
                        "of_bias":        of_bias,
                        "rr_ratio":       _rr_ratio,
                        "entry_price":    price,
                        "stop_price":     _stop_price,
                        "target_1":       _t1_use,
                        "target_2":       _t2_use,
                        "shares":         _adj_shares,
                        "position_value": _adj_value,
                        "dollar_risk":    _adj_risk,
                        "setup_type":     pattern,
                        "early_entry":    early_entry,
                        "breaking_out":   breaking_out,
                        # Outcome fields (filled in later via Trade Journal tab)
                        "outcome":        "",
                        "exit_price":     None,
                        "exit_date":      "",
                        "pnl_pct":        None,
                        "pnl_dollar":     None,
                        "notes":          "",
                    })
                    save_journal(_journal)
                    st.success(
                        f"✅ {chk_ticker} logged to Trade Journal "
                        f"(Conviction: {_conviction}/100, Verdict: {_verdict_str}). "
                        f"Update the outcome in the 📓 Trade Journal tab after the trade."
                    )

                st.download_button(
                    f"⬇ Download Trade Plan — {chk_ticker}",
                    data=_plan_text.encode("utf-8"),
                    file_name=f"apexscan_tradeplan_{chk_ticker}_{pd.Timestamp.now().strftime('%Y%m%d')}.txt",
                    mime="text/plain",
                )


with tabs[19]:
    st.markdown("### 📓 Trade Journal")
    st.caption("Track every trade from checklist to outcome. Builds your personal edge statistics over time.")

    _jnl = load_journal()
    _jnl_j1, _jnl_j2 = st.tabs(["📋 All Entries", "📊 Performance Stats"])

    with _jnl_j1:
        if not _jnl:
            st.info(
                "No journal entries yet. Go to ✅ Pre-Buy Checklist → "
                "run a checklist → click **📓 Log to Trade Journal**."
            )
        else:
            st.markdown(f"**{len(_jnl)} journal entries**")
            # Split: open (no outcome) vs closed (outcome filled)
            _open_entries   = [e for e in _jnl if not e.get("outcome")]
            _closed_entries = [e for e in _jnl if e.get("outcome")]

            _jt1, _jt2 = st.tabs([
                f"⏳ Open / Pending ({len(_open_entries)})",
                f"✅ Closed ({len(_closed_entries)})"
            ])

            with _jt1:
                if not _open_entries:
                    st.info("No open journal entries. Log a trade from the Pre-Buy Checklist.")
                else:
                    for _i, _e in enumerate(_open_entries):
                        _tk  = _e.get("ticker","–")
                        _vrd = _e.get("verdict","–")
                        _con = _e.get("conviction", 0)
                        _ep  = _e.get("entry_price")
                        _sp  = _e.get("stop_price")
                        _t1  = _e.get("target_1")
                        _dt  = str(_e.get("date",""))[:10]
                        _col = "#3fb950" if "READY" in _vrd else "#d29922"

                        with st.expander(
                            f"**{_tk}** — {_vrd} | Conviction: {_con}/100 | {_dt}",
                            expanded=False
                        ):
                            _ei1,_ei2,_ei3,_ei4 = st.columns(4)
                            _ei1.metric("Entry",   f"${_ep:.2f}" if _ep else "–")
                            _ei2.metric("Stop",    f"${_sp:.2f}" if _sp else "–")
                            _ei3.metric("Target 1",f"${_t1:.2f}" if _t1 else "–")
                            _ei4.metric("R:R",     f"{_e.get('rr_ratio',0):.1f}:1")

                            st.markdown("**Update Outcome:**")
                            _oc1,_oc2,_oc3 = st.columns(3)
                            with _oc1:
                                _outcome_sel = st.selectbox(
                                    "Outcome",
                                    ["","Winner — Target Hit","Winner — Partial","Loser — Stop Hit",
                                     "Loser — Manual Exit","Break Even","Still Open"],
                                    key=f"jnl_outcome_{_i}"
                                )
                            with _oc2:
                                _exit_price = st.number_input(
                                    "Exit Price ($)", min_value=0.0,
                                    value=float(_ep or 0), step=0.01, format="%.2f",
                                    key=f"jnl_exit_{_i}"
                                )
                            with _oc3:
                                _exit_notes = st.text_input(
                                    "Notes", placeholder="What happened?",
                                    key=f"jnl_notes_{_i}"
                                )

                            if st.button("💾 Save Outcome", key=f"jnl_save_{_i}"):
                                if _outcome_sel and _exit_price > 0 and _ep:
                                    _pnl_pct = round((_exit_price/_ep - 1)*100, 2)
                                    _pnl_dol = round((_exit_price - _ep) * (_e.get("shares") or 1), 2)
                                    # Update entry in journal
                                    for _je in _jnl:
                                        if (_je.get("ticker")==_tk and
                                                _je.get("date")==_e.get("date")):
                                            _je["outcome"]    = _outcome_sel
                                            _je["exit_price"] = _exit_price
                                            _je["exit_date"]  = pd.Timestamp.now().isoformat()[:10]
                                            _je["pnl_pct"]    = _pnl_pct
                                            _je["pnl_dollar"] = _pnl_dol
                                            _je["notes"]      = _exit_notes
                                    save_journal(_jnl)
                                    st.success(
                                        f"✅ {_tk} outcome saved: {_outcome_sel} | "
                                        f"P&L: ${_pnl_dol:+,.2f} ({_pnl_pct:+.1f}%)"
                                    )
                                    st.rerun()

            with _jt2:
                if not _closed_entries:
                    st.info("No closed entries yet. Close trades by updating their outcome above.")
                else:
                    _cdf = pd.DataFrame(_closed_entries)
                    _cdf["pnl_pct"]    = pd.to_numeric(_cdf["pnl_pct"],    errors="coerce")
                    _cdf["pnl_dollar"] = pd.to_numeric(_cdf["pnl_dollar"], errors="coerce")
                    _cdf["conviction"] = pd.to_numeric(_cdf["conviction"], errors="coerce")

                    # Show table
                    _disp_cols = ["date","ticker","verdict","conviction","entry_price",
                                  "exit_price","pnl_pct","pnl_dollar","outcome","setup_type","notes"]
                    _cdf_disp = _cdf[[c for c in _disp_cols if c in _cdf.columns]].copy()
                    _cdf_disp.columns = [c.replace("_"," ").title() for c in _cdf_disp.columns]

                    def _jc(v):
                        try: return "color:#3fb950;font-weight:700" if float(v)>0 else "color:#f85149;font-weight:700"
                        except: return ""

                    st.dataframe(
                        _cdf_disp.style.map(_jc, subset=[c for c in ["Pnl Pct","Pnl Dollar"] if c in _cdf_disp.columns])
                        .format({
                            "Entry Price":  lambda v: f"${v:.2f}" if pd.notna(v) else "–",
                            "Exit Price":   lambda v: f"${v:.2f}" if pd.notna(v) else "–",
                            "Pnl Pct":      lambda v: f"{v:+.1f}%" if pd.notna(v) else "–",
                            "Pnl Dollar":   lambda v: f"${v:+,.2f}" if pd.notna(v) else "–",
                            "Conviction":   lambda v: f"{v:.0f}/100" if pd.notna(v) else "–",
                        }, na_rep="–"),
                        use_container_width=True, hide_index=True, height=320
                    )
                    st.download_button(
                        "⬇ Export Journal (CSV)",
                        data=_cdf_disp.to_csv(index=False).encode("utf-8"),
                        file_name=f"apexscan_journal_{pd.Timestamp.now().strftime('%Y%m%d')}.csv",
                        mime="text/csv",
                    )

    with _jnl_j2:
        if len(_jnl) < 3:
            st.info("Log at least 3 trades to see performance statistics.")
        else:
            _all_df = pd.DataFrame(_jnl)
            _all_df["pnl_pct"]    = pd.to_numeric(_all_df.get("pnl_pct"), errors="coerce")
            _all_df["pnl_dollar"] = pd.to_numeric(_all_df.get("pnl_dollar"), errors="coerce")
            _all_df["conviction"] = pd.to_numeric(_all_df.get("conviction"), errors="coerce")
            _closed_df = _all_df[_all_df["outcome"].astype(str).str.len() > 0].copy()

            if _closed_df.empty:
                st.info("Close some trades first to see statistics.")
            else:
                _wins   = _closed_df[_closed_df["pnl_pct"] > 0]
                _losses = _closed_df[_closed_df["pnl_pct"] <= 0]
                _wr     = round(len(_wins)/len(_closed_df)*100, 1) if len(_closed_df) > 0 else 0
                _avg_w  = _wins["pnl_pct"].mean()   if not _wins.empty   else 0
                _avg_l  = _losses["pnl_pct"].mean() if not _losses.empty else 0
                _exp    = round((_wr/100 * _avg_w) + ((1-_wr/100) * _avg_l), 2)
                _total  = _closed_df["pnl_dollar"].sum()

                st.markdown("#### 🏆 Your Edge Statistics")
                _s1,_s2,_s3,_s4,_s5 = st.columns(5)
                _s1.metric("Total Trades",    len(_closed_df))
                _s2.metric("Win Rate",        f"{_wr:.1f}%")
                _s3.metric("Avg Winner",      f"{_avg_w:+.1f}%")
                _s4.metric("Avg Loser",       f"{_avg_l:+.1f}%")
                _s5.metric("Expectancy",      f"{_exp:+.2f}%",
                           help="(Win% × Avg Win) + (Loss% × Avg Loss). Positive = you have an edge.")

                # Expectancy interpretation
                _exp_col = "#3fb950" if _exp > 0.5 else "#d29922" if _exp > 0 else "#f85149"
                st.markdown(
                    f'<div style="background:#161b22;border-left:4px solid {_exp_col};'
                    f'padding:12px 16px;border-radius:4px;margin:8px 0;">'
                    f'<b style="color:{_exp_col};">System Expectancy: {_exp:+.2f}% per trade</b><br>'
                    f'<span style="color:#8b949e;font-size:0.85rem;">'
                    f'{"✅ Positive edge — your system makes money over time. Keep following the rules." if _exp > 0.5 else "⚠️ Marginal edge — improve win rate or cut losses faster." if _exp > 0 else "❌ Negative expectancy — your system loses money. Review what is failing."}'
                    f'</span></div>',
                    unsafe_allow_html=True
                )

                st.markdown("---")

                # Performance by conviction bucket
                st.markdown("#### 📊 Win Rate by Conviction Score")
                _conv_buckets = []
                for _label, _lo, _hi in [("A+ (85–100)",85,101),("A (75–84)",75,85),
                                          ("B (60–74)",60,75),("C (45–59)",45,60),("D (<45)",0,45)]:
                    _bucket = _closed_df[(_closed_df["conviction"]>=_lo) & (_closed_df["conviction"]<_hi)]
                    if not _bucket.empty:
                        _bwr = round((_bucket["pnl_pct"]>0).mean()*100, 1)
                        _bexp= round((_bwr/100 * _bucket[_bucket["pnl_pct"]>0]["pnl_pct"].mean() or 0) +
                                     ((1-_bwr/100) * (_bucket[_bucket["pnl_pct"]<=0]["pnl_pct"].mean() or 0)), 2)
                        _conv_buckets.append({
                            "Grade": _label, "Trades": len(_bucket),
                            "Win %": f"{_bwr:.1f}%",
                            "Expectancy": f"{_bexp:+.2f}%",
                            "Avg P&L": f"{_bucket['pnl_pct'].mean():+.1f}%",
                        })
                if _conv_buckets:
                    st.dataframe(pd.DataFrame(_conv_buckets), use_container_width=True, hide_index=True)
                    st.caption(
                        "This table tells you what conviction score threshold to use for full-size trades. "
                        "If A+ setups win 70% and C setups win 35%, trade A+ at full size and skip C entirely."
                    )

                # Performance by setup type
                if "setup_type" in _closed_df.columns:
                    _setup_perf = _closed_df.groupby("setup_type").agg(
                        Trades=("pnl_pct","count"),
                        Win_Rate=("pnl_pct", lambda x: round((x>0).mean()*100,1)),
                        Avg_PnL=("pnl_pct","mean"),
                        Total_PnL=("pnl_dollar","sum"),
                    ).reset_index().sort_values("Avg_PnL",ascending=False)
                    _setup_perf.columns = ["Setup","Trades","Win %","Avg P&L %","Total P&L $"]
                    st.markdown("#### 📋 Performance by Setup Type")
                    st.dataframe(
                        _setup_perf.style.format({
                            "Avg P&L %": "{:+.1f}%",
                            "Total P&L $": "${:+,.2f}",
                            "Win %": "{:.1f}%",
                        }),
                        use_container_width=True, hide_index=True
                    )

                # Cumulative P&L
                _closed_sorted = _closed_df.sort_values("exit_date")
                _closed_sorted["cumulative"] = _closed_sorted["pnl_dollar"].cumsum()
                if not _closed_sorted.empty:
                    import plotly.express as _pxj
                    _fig_j = _pxj.area(
                        _closed_sorted, x="exit_date", y="cumulative",
                        title="Cumulative Journal P&L ($)",
                        color_discrete_sequence=["#3fb950"],
                    )
                    _fig_j.add_hline(y=0, line_dash="dot", line_color="#8b949e")
                    _fig_j.update_layout(
                        paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                        font_color="#e6edf3", height=300,
                        margin=dict(t=40,b=20,l=20,r=20),
                    )
                    _fig_j.update_traces(fill="tozeroy", fillcolor="rgba(63,185,80,0.12)")
                    st.plotly_chart(_fig_j, use_container_width=True)


with tabs[20]:
    st.markdown("### 👁 Setup Monitor")
    st.caption(
        "Stocks you're watching for checklist status improvement. "
        "Get alerted via Telegram when a setup moves from 'Wait' to 'Ready to Buy'."
    )

    _chkw = load_chk_watchlist()

    if not _chkw:
        st.info(
            "No setups being monitored yet. "
            "Go to ✅ Pre-Buy Checklist → run a checklist → click **👁 Watch This Setup**."
        )
    else:
        # ── Check all monitored setups against current scan ───────────────────
        _chkw_rows = []
        for _witem in _chkw:
            _wtk    = _witem.get("ticker","–")
            _wverd  = _witem.get("last_verdict","–")
            _wconv  = _witem.get("last_conv", 0)
            _wadd   = str(_witem.get("added",""))[:10]
            _wdays  = (pd.Timestamp.now() - pd.to_datetime(_wadd)).days if _wadd else "–"

            # Check if in current scan
            _in_scan = not df.empty and _wtk in df.get("ticker", pd.Series()).values
            _new_row  = df[df["ticker"]==_wtk].iloc[0] if _in_scan else None
            _new_score = float(_new_row["apex_score"]) if _new_row is not None and "apex_score" in _new_row else None
            _score_chg = round(_new_score - _witem.get("last_score",0), 1) if _new_score else None

            _chkw_rows.append({
                "Ticker":      _wtk,
                "Theme":       _witem.get("theme","–"),
                "Last Verdict":_wverd,
                "Conviction":  f"{_wconv}/100",
                "Added":       _wadd,
                "Days Watched":_wdays,
                "In Scan":     "✅ Yes" if _in_scan else "❌ Not in last scan",
                "Score Now":   f"{_new_score:.0f}" if _new_score else "–",
                "Score Δ":     f"{_score_chg:+.1f}" if _score_chg else "–",
            })

        _wdf = pd.DataFrame(_chkw_rows)
        def _wc(v):
            try:
                _f = float(str(v).replace("+",""))
                return "color:#3fb950;font-weight:700" if _f>0 else "color:#f85149;font-weight:700"
            except: return ""
        st.dataframe(
            _wdf.style.map(_wc, subset=["Score Δ"]),
            use_container_width=True, hide_index=True
        )

        # ── Manual check button ───────────────────────────────────────────────
        st.markdown("---")
        _mw1,_mw2 = st.columns(2)
        with _mw1:
            if st.button("🔄 Re-check All Setups Against Current Scan",
                         key="recheck_monitor", use_container_width=True):
                _alerts_sent  = 0
                _cur_settings = load_alert_settings()
                _updated      = []
                for _witem in _chkw:
                    _wtk = _witem.get("ticker","–")
                    _in_scan = not df.empty and _wtk in df.get("ticker", pd.Series()).values
                    if _in_scan:
                        _nr = df[df["ticker"]==_wtk].iloc[0]
                        _ns = float(_nr.get("apex_score",0) or 0)
                        _old_s = _witem.get("last_score",0)
                        _witem["last_score"]   = _ns
                        _witem["last_checked"] = pd.Timestamp.now().isoformat()

                        # Fire alert if score jumped ≥ 10 points or verdict improved
                        if _ns - _old_s >= 10:
                            _msg = (
                                "\U0001F4C8 *Setup Improvement Alert*\n\n"
                                f"*{_wtk}* Apex Score: {_old_s:.0f} pts\n"
                                f"Score change: +{_ns - _old_s:.0f}\n"
                                f"Stage: {str(_nr.get(chr(115)+chr(116)+chr(97)+chr(103)+chr(101),chr(45)))}\n"
                                "Check Pre-Buy Checklist now"
                            )
                            if (_cur_settings.get("telegram_token") and
                                    _cur_settings.get("telegram_chat_id")):
                                dispatch_alert(_cur_settings, _msg, f"ApexScan — {_wtk}")
                                _alerts_sent += 1
                    _updated.append(_witem)
                save_chk_watchlist(_updated)
                st.success(
                    f"✅ Re-checked {len(_chkw)} setups. "
                    f"{_alerts_sent} Telegram alert(s) sent."
                )
                st.rerun()

        with _mw2:
            _rm_tk = st.selectbox(
                "Remove from Monitor",
                ["–"] + [x.get("ticker","–") for x in _chkw],
                key="rm_monitor"
            )
            if st.button("🗑 Remove", key="rm_monitor_btn", use_container_width=True):
                if _rm_tk != "–":
                    _chkw = [x for x in _chkw if x.get("ticker") != _rm_tk]
                    save_chk_watchlist(_chkw)
                    st.success(f"Removed {_rm_tk} from monitor.")
                    st.rerun()


with tabs[20]:
    st.markdown("""
### 📖 How to Use ApexScan — Complete Guide

---

#### 🤖 AI Daily Briefing
- Run a **Live Scan** first, then click **Generate Briefing**
- Claude reads your results and writes a plain-English market briefing
- Covers top setups, active breakouts, theme rotation and a risk reminder
- Send directly to Telegram or download as text
- Takes about 10 seconds to generate

#### 📋 Watchlist Manager
- Create named lists: *High Conviction*, *Monitoring*, *Earnings Soon* etc.
- Add tickers manually or paste a comma-separated list to import
- **Scan any watchlist independently** — see scores for just those stocks
- Promote top results from one list to another with one click
- Lists persist between sessions

#### 🔔 Alert Settings
- **Telegram** (recommended): Free, instant, works on any phone
  - Follow the 5-step setup in the tab — takes 5 minutes
  - Alerts fire automatically after every scan
- **Email**: Works with Gmail app passwords
- Alert types: Breakouts, Stop Loss breach, Earnings warnings
- Set minimum score threshold so you only get alerts that matter

---

#### 🏆 Apex Score (0–100)
| Points | Signal |
|---|---|
| 0–40 | 3-month momentum |
| 25 | RS > benchmark |
| 15 | Stage 2 uptrend |
| 10 | Near 52-week high |
| 10 | Active breakout |

#### 📊 RS Score vs S&P 500
- **> 100** 🟢 Beating the market — buy leaders, not laggards
- **70–100** 🟡 Keeping pace
- **< 70** 🔴 Lagging — avoid

#### 📐 Stage Guide
- **Stage 2 ✅** — Only stage worth buying
- **Stage 1 ⏳** — Building base, not ready
- **Stage 3 ⚠️** — Rolling over, be careful
- **Stage 4 🔴** — Downtrend, avoid

#### 🎯 Options Flow
- **Vol/OI > 3x** = unusual — someone is placing a big bet
- Calls = bullish bet, Puts = bearish/hedge
- High notional ($) = more conviction behind the trade
- Combine with high Apex Score for maximum signal

#### 🕵️ Insider Tracker
- 🔥 **Cluster buy** = 2+ insiders buying = strongest possible signal
- C-suite buys (CEO, CFO) matter most
- Insider sells = less meaningful, they sell for many reasons

#### ⚖️ Risk Calculator — The Golden Rule
> **Never risk more than 1–2% of your account on any single trade**

Formula: `Shares = (Account × Risk%) ÷ (Entry − Stop)`

Always set a **minimum 2:1 reward:risk** before entering any trade.

#### ⏱ Backtester
- Entry: Apex Score ≥ threshold + Stage 2
- Exit: Price closes below 50MA or max hold days
- Use 2021–2024 for a full market cycle test (bull + bear + recovery)

---
> ⚠️ ApexScan is for research and education only — not financial advice.

---

#### 🔑 API Key Setup & Verification

**Alpha Vantage (for real EPS data):**
1. Go to **alphavantage.co** → click **Get Free API Key**
2. Open `config.yaml` in VS Code
3. Find the line: `alpha_vantage_key: "YOUR_ALPHA_VANTAGE_KEY_HERE"`
4. Replace the placeholder with your actual key (keep the quote marks)
5. Save the file (Ctrl+S)
6. **Restart the dashboard** — close CMD, reopen it, run the command again
7. The sidebar will show 🟢 Alpha Vantage ✓ Active when it's working

**Important:** After editing config.yaml you MUST restart the dashboard.
Just saving the file is not enough — Streamlit needs to reload.

**Finnhub (for news sentiment):**
Same process — paste your key at `finnhub_key:` in config.yaml.

**Free tier limits:**
- Alpha Vantage free: 25 API calls/day, 5/minute
- Each stock in the scan uses 3 AV calls (earnings + income + overview)
- Cached for 24h so repeated scans don't burn your quota
- With 45 tickers: first scan uses ~45 calls (exceeds daily free limit)
- **Solution:** Add `cache_hours: 168` (1 week) in config.yaml under `alpha_vantage:` to cache aggressively and stay within quota
    """)
